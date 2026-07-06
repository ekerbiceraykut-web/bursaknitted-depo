import sys
import os
import os as _os
LOGO_PATH = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "logo.png")

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QLineEdit, QLabel,
    QComboBox, QDialog, QFormLayout, QDialogButtonBox, QMessageBox,
    QTabWidget, QHeaderView, QFileDialog, QSpinBox, QDoubleSpinBox,
    QTextEdit, QGroupBox, QSplitter, QListWidget, QListWidgetItem,
    QStatusBar, QFrame, QAbstractItemView, QTableView, QInputDialog,
    QCheckBox, QCompleter, QDateEdit, QTreeWidget, QTreeWidgetItem,
    QScrollArea, QGridLayout
)
from PyQt6.QtCore import Qt, QTimer, QSize, QAbstractTableModel, QModelIndex, QVariant, QDate, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QIcon, QBrush, QPixmap


def _excepthook(exctype, value, tb):
    """PyQt6 varsayılan davranışı: bir slot içinde yakalanmayan hata oluşunca
    qFatal() çağrılır ve tüm program anında kapanır (SIGABRT) — kullanıcı sadece
    "program kapandı" görür. Bu hook hatayı gösterir ve programı açık tutar."""
    import traceback
    traceback.print_exception(exctype, value, tb)
    try:
        QMessageBox.critical(QApplication.activeWindow(), "Beklenmeyen Hata",
            f"Bir hata oluştu, son işlem tamamlanmamış olabilir:\n\n{value}")
    except Exception:
        pass


sys.excepthook = _excepthook

# Bağlantı modu: "local" veya "remote"
CONNECTION_MODE = "local"

# Liste döndüren api_client fonksiyonları: uzak çağrı hata verirse _load() metotlarının
# `len(rows)`/`for r in rows` ile çökmemesi için boş liste döndürülür.
_LIST_FUNCS = {
    "get_all_fabrics", "get_all_customers", "get_all_suppliers",
    "get_all_products", "get_all_locations", "get_active_locations",
    "get_locations", "get_all_users", "get_fire_records",
    "get_movements", "get_all_movements", "get_movements_by_range",
    "get_all_orders", "get_all_purchase_orders",
    "get_pending_approval_orders", "get_shippable_orders",
    "get_order_shipments", "get_po_receipts", "get_boyahane_queue",
    "get_po_items_for_order", "get_all_armur_desenleri",
    "get_crm_customers", "get_crm_visits", "get_crm_sales", "get_crm_orders",
    "get_crm_years", "get_stock_snapshots", "get_movement_destinations",
    "get_iplik_cinsleri", "get_iplikler",
}

_REAUTH_IN_PROGRESS = False

def _reauthenticate(parent):
    """Oturum geçersiz hale geldiğinde (sunucu yeniden başlamış olabilir) yeniden
    giriş ister. Başarılı olursa sekmeleri yeniler ve True döner."""
    global _REAUTH_IN_PROGRESS
    if _REAUTH_IN_PROGRESS:
        return False
    _REAUTH_IN_PROGRESS = True
    try:
        dlg = LoginDialog(parent)
        dlg.setWindowModality(Qt.WindowModality.ApplicationModal)
        dlg.err_label.setText("Oturum sona erdi (sunucu yeniden başlamış olabilir).\nLütfen yeniden giriş yapın.")
        dlg.username.setText(CURRENT_USER.get("username", ""))
        dlg.password.setFocus()
        if dlg.exec():
            for w in QApplication.topLevelWidgets():
                if isinstance(w, MainWindow):
                    w._update_user_label()
                    w._rebuild_tabs()
            return True
        return False
    finally:
        _REAUTH_IN_PROGRESS = False

class _DbProxy:
    """db.xxx çağrılarını CONNECTION_MODE'a göre local veya remote'a yönlendirir.
    Uzak moddaki hatalar burada yakalanır, kullanıcıya gösterilir ve uygulamanın
    çökmesi yerine güvenli bir varsayılan değer döndürülür. Oturum süresi dolmuşsa
    (401) yeniden giriş istenir ve işlem otomatik tekrar denenir — kullanıcı az önce
    girdiği veriyi yeniden girmek zorunda kalmaz."""
    def __getattr__(self, name):
        if CONNECTION_MODE != "remote":
            return getattr(db, name)
        import api_client
        attr = getattr(api_client, name)
        if not callable(attr):
            return attr

        def wrapper(*args, **kwargs):
            try:
                return attr(*args, **kwargs)
            except Exception as e:
                parent = QApplication.activeWindow()
                if "Yetkisiz" in str(e) and _reauthenticate(parent):
                    try:
                        return attr(*args, **kwargs)   # yeni token ile tekrar dene
                    except Exception as e2:
                        e = e2
                if "Yetkisiz" not in str(e):
                    QMessageBox.critical(parent, "Bağlantı Hatası",
                        f"Sunucu ile iletişim kurulamadı:\n\n{e}")
                return [] if name in _LIST_FUNCS else None
        return wrapper

_db = _DbProxy()   # tüm kod db yerine _db kullanır ama mevcut kod db değişkenini kullanıyor,
                   # bu yüzden modül seviyesinde db'yi proxy ile değiştiriyoruz

def _get_db():
    return _db

# Giriş yapmış kullanıcı (global)
CURRENT_USER = {"id": 0, "username": "sistem", "full_name": "Sistem", "role": "admin"}

import database as _local_db
db = _local_db   # başlangıçta yerel; login sonrası proxy ile değiştirilir

import order_pdf
from order_pdf import CURRENCY_SYMBOLS

CURRENCY_OPTIONS = ["USD", "EUR", "GBP", "TRY"]


def _make_searchable(cb: QComboBox):
    """QComboBox'u stok-listesi arama kutusuna benzer şekilde aranabilir yapar.

    • Yazılan metne göre öğeler anlık filtrelenir; sonuçlar klavye odağı çalmadan
      combo'nun hemen altında kayan bir QListWidget popup'ta gösterilir.
    • cb.addItem() / cb.clear() Python-düzeyinde override edilir; orijinal C++
      metotlar da çağrılarak currentData() / findData() düzgün çalışmaya devam eder.
    • Seçim sonrası currentIndexChanged doğru index ile ateşlenir.
    """
    cb.setEditable(True)
    cb.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
    cb.setCompleter(None)

    # ── Öğe deposu ve orijinal metot referansları ────────────────────────────
    _orig_add   = cb.addItem   # C++ bound metod — override öncesi sakla
    _orig_clear = cb.clear
    cb._items   = []           # [(text, data), ...]

    # Varsa mevcut öğeleri depoya al
    existing  = [(cb.itemText(i), cb.itemData(i)) for i in range(cb.count())]
    saved_idx = cb.currentIndex()
    _orig_clear()
    for text, data in existing:
        cb._items.append((text, data))
        _orig_add(text, data) if data is not None else _orig_add(text)

    def _addItem(text, data=None):
        cb._items.append((text, data))
        _orig_add(text, data) if data is not None else _orig_add(text)

    def _clearItems():
        cb._items.clear()
        _orig_clear()

    cb.addItem = _addItem
    cb.clear   = _clearItems

    if saved_idx > 0:
        cb.setCurrentIndex(saved_idx)

    # ── Arama popup'u (dialog child widget — ayrı pencere değil) ────────────
    # Qt.WindowType.Popup pencereleri başka pencere aktive olunca otomatik
    # kapanır; bunun yerine popup'u dialog'un child widget'ı yapıyoruz.
    # FocusPolicy.NoFocus → klavye odağı çalmaz; mouse tıklamaları çalışır.
    cb._search_popup = None   # lazy init: _refresh() içinde oluşturulur

    _POPUP_STYLE = (
        "QListWidget{border:1px solid #9E9E9E;background:#fff;outline:0;}"
        "QListWidget::item{padding:5px 10px;}"
        "QListWidget::item:hover{background:#E3F2FD;}"
        "QListWidget::item:selected{background:#1565C0;color:#fff;}"
    )

    def _select(item):
        text = item.text()
        data = item.data(Qt.ItemDataRole.UserRole)
        p = cb._search_popup
        if p:
            p.hide()
        i = cb.findData(data) if data is not None else -1
        if i < 0:
            i = cb.findText(text, Qt.MatchFlag.MatchFixedString)
        prev = cb.currentIndex()
        cb.lineEdit().setText(text)
        if i >= 0 and i != prev:
            cb.setCurrentIndex(i)          # currentIndexChanged ateşlenir
        elif i == prev:
            cb.currentIndexChanged.emit(i) # aynı index seçildi; cascade combo'lar için

    def _refresh():
        txt = cb.lineEdit().text()
        if not txt:
            if cb._search_popup:
                cb._search_popup.hide()
            return

        # Lazy init: popup'u şu anki dialog'un child widget'ı olarak oluştur
        dlg = cb.window()
        if cb._search_popup is None:
            popup = QListWidget(dlg)
            popup.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            popup.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            popup.setStyleSheet(_POPUP_STYLE)
            popup.itemClicked.connect(_select)
            cb._search_popup = popup

        popup = cb._search_popup
        popup.clear()
        low = txt.lower()
        for text, data in cb._items:
            if low in text.lower():
                it = QListWidgetItem(text)
                it.setData(Qt.ItemDataRole.UserRole, data)
                popup.addItem(it)
        if popup.count() == 0:
            popup.hide()
            return

        pos_global = cb.mapToGlobal(cb.rect().bottomLeft())
        pos_local  = dlg.mapFromGlobal(pos_global)
        w     = max(cb.width(), 220)
        row_h = popup.sizeHintForRow(0) or 26
        h     = min(popup.count() * row_h + 6, 240)

        # Dialog alt kenarı taşarsa yukarıda göster
        if pos_local.y() + h > dlg.height():
            top_global = cb.mapToGlobal(cb.rect().topLeft())
            top_local  = dlg.mapFromGlobal(top_global)
            pos_local.setY(top_local.y() - h)

        popup.move(pos_local)
        popup.resize(w, h)
        popup.raise_()
        popup.show()

    cb.lineEdit().textEdited.connect(lambda _: _refresh())
    cb.destroyed.connect(lambda: None)  # popup dlg ile birlikte silinir

# ─────────────────────────────────────────────────────────────────────────────
# Armür Desen Editörü
# ─────────────────────────────────────────────────────────────────────────────

class ArmurGrid(QWidget):
    """Tıklanabilir armür desen ızgarası. Koyu = çözgü üstte, beyaz = atkı üstte."""
    changed = pyqtSignal()

    def __init__(self, rows=8, cols=8, parent=None):
        super().__init__(parent)
        self.rows = rows
        self.cols = cols
        self.grid = set()          # dolu hücreler: {(row, col), ...} (0-tabanlı)
        self.cell_size = 28
        self._update_size()
        self.setMouseTracking(True)

    def _update_size(self):
        self.setFixedSize(self.cols * self.cell_size + 2,
                          self.rows * self.cell_size + 2)

    def resize_grid(self, rows, cols):
        self.rows = rows; self.cols = cols
        self.grid = {(r, c) for r, c in self.grid if r < rows and c < cols}
        self._update_size(); self.update()

    def set_grid(self, filled_cells):
        self.grid = set(map(tuple, filled_cells))
        self.update()

    def get_grid(self):
        return sorted(self.grid)

    def mousePressEvent(self, e):
        col = (e.position().x() - 1) // self.cell_size
        row = (e.position().y() - 1) // self.cell_size
        row = self.rows - 1 - int(row)   # alt=1, üst=rows gibi göster
        col = int(col)
        if 0 <= row < self.rows and 0 <= col < self.cols:
            key = (row, col)
            if key in self.grid:
                self.grid.discard(key)
            else:
                self.grid.add(key)
            self.update(); self.changed.emit()

    def paintEvent(self, e):
        from PyQt6.QtGui import QPainter, QPen
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing, False)
        cs = self.cell_size
        for r in range(self.rows):
            draw_r = self.rows - 1 - r   # ekranda üst=yüksek no
            for c in range(self.cols):
                x = 1 + c * cs; y = 1 + draw_r * cs
                filled = (r, c) in self.grid
                p.fillRect(x, y, cs, cs, QColor("#1A237E") if filled else QColor("#FFFFFF"))
                p.setPen(QPen(QColor("#9E9E9E"), 1))
                p.drawRect(x, y, cs, cs)

        # eksen numaraları
        p.setPen(QColor("#555"))
        font = p.font(); font.setPointSize(7); p.setFont(font)


class ArmurDesignDialog(QDialog):
    """Tek bir armür desenini düzenler / oluşturur."""
    def __init__(self, desen=None, parent=None):
        super().__init__(parent)
        self.desen = desen   # None = yeni
        self.setWindowTitle("Armür Desen Editörü" + (f" — {desen['name']}" if desen else ""))
        self.setMinimumSize(640, 520)
        self._build_ui()
        if desen:
            self._load_desen(desen)
            self._set_locked(True)

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Kilit uyarı çubuğu
        self._lock_bar = QWidget()
        self._lock_bar.setStyleSheet(
            "background:#FFF8E1;border:1px solid #FFB300;border-radius:4px;")
        lock_h = QHBoxLayout(self._lock_bar); lock_h.setContentsMargins(10, 5, 10, 5)
        lock_h.addWidget(QLabel("🔒  Bu desen kilitli — görüntüleme modu"))
        lock_h.addStretch()
        self._btn_unlock = QPushButton("✎  Düzenle")
        self._btn_unlock.setStyleSheet(
            "background:#E65100;color:white;font-weight:bold;border-radius:4px;padding:4px 14px;")
        self._btn_unlock.clicked.connect(lambda: self._set_locked(False))
        lock_h.addWidget(self._btn_unlock)
        self._lock_bar.setVisible(False)
        lay.addWidget(self._lock_bar)

        # Ayarlar satırı
        top = QHBoxLayout()
        top.addWidget(QLabel("Desen Adı:"))
        self.name_edit = QLineEdit(self.desen["name"] if self.desen else "Yeni Desen")
        self.name_edit.setMinimumWidth(180)
        top.addWidget(self.name_edit)
        top.addSpacing(16)
        top.addWidget(QLabel("Çözgü (sütun):"))
        self.cols_spin = QSpinBox(); self.cols_spin.setRange(2, 64); self.cols_spin.setValue(8)
        top.addWidget(self.cols_spin)
        top.addWidget(QLabel("Atkı (satır):"))
        self.rows_spin = QSpinBox(); self.rows_spin.setRange(2, 64); self.rows_spin.setValue(8)
        top.addWidget(self.rows_spin)
        self._btn_apply = QPushButton("Uygula")
        self._btn_apply.clicked.connect(self._apply_size)
        top.addWidget(self._btn_apply)
        top.addStretch()
        lay.addLayout(top)

        # Izgara + etiketler
        grid_wrap = QHBoxLayout()

        # Atkı numaraları (sol)
        self.atki_labels = QWidget()
        self._atki_lbl_lay = QVBoxLayout(self.atki_labels)
        self._atki_lbl_lay.setSpacing(0); self._atki_lbl_lay.setContentsMargins(0,0,0,0)
        grid_wrap.addWidget(self.atki_labels)

        # Izgara
        self.grid_widget = ArmurGrid(8, 8)
        self.grid_widget.changed.connect(self._update_stats)
        scroll = QScrollArea()
        scroll.setWidget(self.grid_widget)
        scroll.setWidgetResizable(False)
        grid_wrap.addWidget(scroll)
        grid_wrap.addStretch()
        lay.addLayout(grid_wrap)

        # Çözgü numaraları (alt)
        self.cozgu_lbl = QLabel()
        self.cozgu_lbl.setStyleSheet("color:#555;font-size:9px;")
        lay.addWidget(self.cozgu_lbl)

        # İstatistikler
        self.stats_lbl = QLabel()
        self.stats_lbl.setStyleSheet("color:#1565C0;font-size:11px;")
        lay.addWidget(self.stats_lbl)

        # Notlar
        lay.addWidget(QLabel("Notlar:"))
        self.notes_edit = QTextEdit()
        self.notes_edit.setMaximumHeight(60)
        lay.addWidget(self.notes_edit)

        # Butonlar
        self._btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        self._btn_box.button(QDialogButtonBox.StandardButton.Save).setText("Kaydet")
        self._btn_box.accepted.connect(self._save)
        self._btn_box.rejected.connect(self.reject)
        lay.addWidget(self._btn_box)

    def _set_locked(self, locked: bool):
        self._lock_bar.setVisible(locked)
        self.name_edit.setReadOnly(locked)
        self.cols_spin.setEnabled(not locked)
        self.rows_spin.setEnabled(not locked)
        self._btn_apply.setEnabled(not locked)
        self.notes_edit.setReadOnly(locked)
        self.grid_widget.setEnabled(not locked)
        save_btn = self._btn_box.button(QDialogButtonBox.StandardButton.Save)
        save_btn.setVisible(not locked)

        self._rebuild_labels()
        self._update_stats()

    def _load_desen(self, d):
        import json
        self.cols_spin.setValue(d.get("sutunlar", 8))
        self.rows_spin.setValue(d.get("satirlar", 8))
        self.grid_widget.resize_grid(d.get("satirlar", 8), d.get("sutunlar", 8))
        try:
            cells = json.loads(d.get("grid", "[]"))
        except Exception:
            cells = []
        self.grid_widget.set_grid(cells)
        self.notes_edit.setPlainText(d.get("notes", ""))
        self._rebuild_labels(); self._update_stats()

    def _apply_size(self):
        r = self.rows_spin.value(); c = self.cols_spin.value()
        self.grid_widget.resize_grid(r, c)
        self._rebuild_labels(); self._update_stats()

    def _rebuild_labels(self):
        r = self.grid_widget.rows; c = self.grid_widget.cols
        cs = self.grid_widget.cell_size
        # Atkı etiketleri (sol, yukarıdan aşağı = büyükten küçüğe)
        for i in reversed(range(self._atki_lbl_lay.count())):
            self._atki_lbl_lay.itemAt(i).widget().deleteLater()
        for row in range(r - 1, -1, -1):
            lbl = QLabel(f"{row+1:2d}")
            lbl.setFixedSize(22, cs)
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            lbl.setStyleSheet("color:#555;font-size:8px;")
            self._atki_lbl_lay.addWidget(lbl)
        # Çözgü etiketleri (alt)
        self.cozgu_lbl.setText("  " + "".join(f"{col+1:^{cs//8}}  " for col in range(c)))

    def _update_stats(self):
        g = self.grid_widget.grid
        total = self.grid_widget.rows * self.grid_widget.cols
        filled = len(g)
        pct = filled / total * 100 if total else 0
        self.stats_lbl.setText(
            f"Çözgü üstte: {filled} hücre ({pct:.1f}%)  |  "
            f"Atkı üstte: {total - filled} hücre ({100-pct:.1f}%)  |  "
            f"Toplam: {total} hücre  ({self.grid_widget.cols} çözgü × {self.grid_widget.rows} atkı)"
        )

    def _save(self):
        import json
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "Hata", "Desen adı zorunlu!"); return
        r = self.grid_widget.rows; c = self.grid_widget.cols
        grid_json = json.dumps(self.grid_widget.get_grid())
        notes = self.notes_edit.toPlainText()
        try:
            if self.desen:
                db.update_armur_desen(self.desen["id"], name, r, c, grid_json, notes)
            else:
                db.add_armur_desen(name, r, c, grid_json, notes)
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e)); return
        self.accept()


class ArmurDesignManagerDialog(QDialog):
    """Armür desen listesi ve yönetimi."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Armür Desen Kütüphanesi")
        self.setMinimumSize(700, 480)
        self._build_ui(); self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Ad", "Çözgü", "Atkı", "Notlar", "Tarih"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        btn_new  = QPushButton("+ Yeni Desen"); btn_new.clicked.connect(self._new)
        btn_edit = QPushButton("✎ Düzenle");    btn_edit.clicked.connect(self._edit)
        btn_del  = QPushButton("✕ Sil")
        btn_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        btn_del.clicked.connect(self._delete)
        btn_close = QPushButton("Kapat"); btn_close.clicked.connect(self.accept)
        for b in (btn_new, btn_edit, btn_del): btn_row.addWidget(b)
        btn_row.addStretch(); btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

    def _load(self):
        rows = db.get_all_armur_desenleri()
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            id_item = QTableWidgetItem(r["name"])
            id_item.setData(Qt.ItemDataRole.UserRole, r["id"])
            self.table.setItem(i, 0, id_item)
            self.table.setItem(i, 1, QTableWidgetItem(str(r.get("sutunlar", ""))))
            self.table.setItem(i, 2, QTableWidgetItem(str(r.get("satirlar", ""))))
            self.table.setItem(i, 3, QTableWidgetItem(r.get("notes", "")))
            self.table.setItem(i, 4, QTableWidgetItem(str(r.get("created_at", ""))[:10]))
        self.table.resizeColumnsToContents()

    def _selected(self):
        row = self.table.currentRow()
        if row < 0: QMessageBox.information(self, "Bilgi", "Desen seçin."); return None
        did = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        return db.get_armur_desen(did)

    def _new(self):
        dlg = ArmurDesignDialog(parent=self)
        if dlg.exec(): self._load()

    def _edit(self):
        d = self._selected()
        if not d: return
        dlg = ArmurDesignDialog(d, self)
        if dlg.exec(): self._load()

    def _delete(self):
        d = self._selected()
        if not d: return
        if QMessageBox.question(self, "Sil", f"'{d['name']}' silinsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_armur_desen(d["id"]); self._load()


# ─────────────────────────────────────────────────────────────────────────────
# Maliyet Hesaplama Modülü
# ─────────────────────────────────────────────────────────────────────────────

def _den_from(value_str, birim):
    """İplik numarasını denye'ye çevirir."""
    try: v = float(value_str)
    except Exception: return 0.0
    if birim == "Den":  return v
    if birim == "dTex": return v * 0.9
    if birim == "Nm":   return 9000 / v if v else 0
    if birim == "Ne":   return 5314.95 / v if v else 0
    return 0.0


from PyQt6.QtCore import QObject, QEvent
from PyQt6.QtWidgets import QAbstractSpinBox, QAbstractScrollArea


class _NoWheelFilter(QObject):
    """Spin box / combo üzerinde fare tekerleği değeri değiştirmesin;
    tekerlek en yakın kaydırma alanına iletilir (sayfa kayar)."""
    def eventFilter(self, obj, ev):
        if ev.type() == QEvent.Type.Wheel:
            p = obj.parent()
            while p is not None and not isinstance(p, QAbstractScrollArea):
                p = p.parent()
            if p is not None:
                QApplication.sendEvent(p.viewport(), ev)
            return True   # widget'ın kendi değeri değişmesin
        return False


_NO_WHEEL_FILTER = _NoWheelFilter()


def _disable_wheel(root):
    """root altındaki tüm spin box ve combo'larda tekerlek ile değer değişimini engeller."""
    for w in root.findChildren((QAbstractSpinBox, QComboBox)):
        w.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        w.installEventFilter(_NO_WHEEL_FILTER)


class MaliyetWidget(QWidget):
    """Ürün diyaloğuna gömülebilir maliyet hesaplama widget'ı."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        _disable_wheel(self)   # fare tekerleği spinbox değerlerini değiştirmesin
        self._recalc()

    def _build_ui(self):
        main = QVBoxLayout(self)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        container = QWidget(); vlay = QVBoxLayout(container)
        scroll.setWidget(container)
        main.addWidget(scroll, 1)

        BIRIMLER = ["Den", "dTex", "Nm", "Ne"]

        grp_genel = QGroupBox("Genel Parametreler")
        fg = QFormLayout(grp_genel); fg.setSpacing(6)
        self.tarak_eni  = QDoubleSpinBox(); self.tarak_eni.setRange(0,500);  self.tarak_eni.setDecimals(1); self.tarak_eni.setSuffix(" cm")
        self.cozgu_sik  = QDoubleSpinBox(); self.cozgu_sik.setRange(0,500);  self.cozgu_sik.setDecimals(1); self.cozgu_sik.setSuffix(" tel/cm")
        self.atki_sik   = QDoubleSpinBox(); self.atki_sik.setRange(0,200);   self.atki_sik.setDecimals(1); self.atki_sik.setSuffix(" atkı/cm")
        self.lbl_toplam_uc    = QLabel("—"); self.lbl_toplam_uc.setStyleSheet("font-weight:bold; color:#1A237E;")
        self.lbl_toplam_desen = QLabel("—"); self.lbl_toplam_desen.setStyleSheet("font-weight:bold; color:#1A237E;")
        fg.addRow("Tarak Eni:", self.tarak_eni)
        fg.addRow("Çözgü Sıklığı (tel/cm):", self.cozgu_sik)
        fg.addRow("Toplam Çözgü Ucu (otomatik):", self.lbl_toplam_uc)
        fg.addRow("Atkı Sıklığı (atkı/cm):", self.atki_sik)
        fg.addRow("Toplam Desen Atım (otomatik):", self.lbl_toplam_desen)
        vlay.addWidget(grp_genel)

        grp_cozgu = QGroupBox("Çözgü İplikleri")
        gc = QGridLayout(grp_cozgu)
        for ci, h in enumerate(["", "İplik Adı", "Numara", "Birim", "Tel/cm", "Toplam Tel Sayısı", "Çekme %", "Fiyat $/kg", "gr/mt", "%", "$/mt"]):
            lb = QLabel(f"<b>{h}</b>"); lb.setAlignment(Qt.AlignmentFlag.AlignCenter); gc.addWidget(lb, 0, ci)
        self.cozgu_rows = []
        for i in range(2):
            ad    = QLineEdit()
            num   = QDoubleSpinBox(); num.setRange(0,99999); num.setDecimals(2)
            bir   = QComboBox(); bir.addItems(BIRIMLER)
            uc_cm = QDoubleSpinBox(); uc_cm.setRange(0,500); uc_cm.setDecimals(1)
            tot_uc = QLabel("—"); tot_uc.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); tot_uc.setStyleSheet("color:#555;padding-right:4px;")
            cek   = QDoubleSpinBox(); cek.setRange(0,50); cek.setDecimals(1); cek.setValue(6.0); cek.setSuffix(" %")
            fiy   = QDoubleSpinBox(); fiy.setRange(0,999); fiy.setDecimals(2); fiy.setSuffix(" $/kg")
            grs   = QLabel("0.00"); grs.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); grs.setStyleSheet("color:#1565C0;padding-right:4px;")
            pct   = QLabel("—"); pct.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); pct.setStyleSheet("color:#7B1FA2;padding-right:4px;")
            dol   = QLabel("0.0000"); dol.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); dol.setStyleSheet("color:#2E7D32;padding-right:4px;")
            for ci, w in enumerate([QLabel(f"Çözgü {i+1}:"), ad, num, bir, uc_cm, tot_uc, cek, fiy, grs, pct, dol]):
                gc.addWidget(w, i+1, ci)
            self.cozgu_rows.append({"ad": ad, "num": num, "bir": bir, "uc_cm": uc_cm, "tot_uc": tot_uc, "cek": cek, "fiy": fiy, "grs": grs, "pct": pct, "dol": dol})
            for w in (num, bir, uc_cm, cek, fiy):
                (w.valueChanged if hasattr(w,'valueChanged') else w.currentIndexChanged).connect(self._recalc)
        def _auto_cozgu2(*_):
            r1 = self.cozgu_rows[0]; r2 = self.cozgu_rows[1]
            if _den_from(str(r2["num"].value()), r2["bir"].currentText()) <= 0: return
            kalan = max(0.0, self.cozgu_sik.value() - r1["uc_cm"].value())
            r2["uc_cm"].blockSignals(True); r2["uc_cm"].setValue(round(kalan, 1)); r2["uc_cm"].blockSignals(False)
            self._recalc()
        self.cozgu_rows[0]["uc_cm"].valueChanged.connect(_auto_cozgu2)
        self.cozgu_sik.valueChanged.connect(_auto_cozgu2)
        self.cozgu_rows[1]["num"].valueChanged.connect(_auto_cozgu2)
        vlay.addWidget(grp_cozgu)

        grp_atki = QGroupBox("Atkı İplikleri")
        ga = QGridLayout(grp_atki)
        for ci, h in enumerate(["", "İplik Adı", "Numara", "Birim", "Atkı Raporu", "Çekme %", "Fiyat $/kg", "atkı/cm", "gr/mt", "%", "$/mt"]):
            lb = QLabel(f"<b>{h}</b>"); lb.setAlignment(Qt.AlignmentFlag.AlignCenter); ga.addWidget(lb, 0, ci)
        self.atki_rows = []
        for i in range(4):
            ad  = QLineEdit()
            num = QDoubleSpinBox(); num.setRange(0,99999); num.setDecimals(2)
            bir = QComboBox(); bir.addItems(BIRIMLER)
            atm = QSpinBox(); atm.setRange(0,256)
            cek = QDoubleSpinBox(); cek.setRange(0,50); cek.setDecimals(1); cek.setValue(8.0); cek.setSuffix(" %")
            fiy = QDoubleSpinBox(); fiy.setRange(0,999); fiy.setDecimals(2); fiy.setSuffix(" $/kg")
            ppm = QLabel("0.00"); ppm.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); ppm.setStyleSheet("color:#555;padding-right:4px;")
            grs = QLabel("0.00"); grs.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); grs.setStyleSheet("color:#1565C0;padding-right:4px;")
            pct = QLabel("—"); pct.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); pct.setStyleSheet("color:#7B1FA2;padding-right:4px;")
            dol = QLabel("0.0000"); dol.setAlignment(Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter); dol.setStyleSheet("color:#2E7D32;padding-right:4px;")
            for ci, w in enumerate([QLabel(f"Atkı {i+1}:"), ad, num, bir, atm, cek, fiy, ppm, grs, pct, dol]):
                ga.addWidget(w, i+1, ci)
            self.atki_rows.append({"ad": ad, "num": num, "bir": bir, "atm": atm, "cek": cek, "fiy": fiy, "ppm": ppm, "grs": grs, "pct": pct, "dol": dol})
            for w in (num, bir, atm, cek, fiy):
                (w.valueChanged if hasattr(w,'valueChanged') else w.currentIndexChanged).connect(self._recalc)
        vlay.addWidget(grp_atki)

        for w in (self.tarak_eni, self.cozgu_sik, self.atki_sik):
            w.valueChanged.connect(self._recalc)

        grp_islem = QGroupBox("İşlem Maliyetleri")
        fi = QFormLayout(grp_islem); fi.setSpacing(6)
        self.fason_dokuma = QDoubleSpinBox(); self.fason_dokuma.setRange(0,9999); self.fason_dokuma.setDecimals(2); self.fason_dokuma.setSuffix(" krş/100 atkı")
        self.usd_kuru     = QDoubleSpinBox(); self.usd_kuru.setRange(0.01,999);   self.usd_kuru.setDecimals(2);     self.usd_kuru.setValue(35.0); self.usd_kuru.setSuffix(" TL/$")
        self.lbl_dokuma_mal = QLabel("—"); self.lbl_dokuma_mal.setStyleSheet("font-weight:bold; color:#1A237E;")
        self.hazirlik_mal   = QDoubleSpinBox(); self.hazirlik_mal.setRange(0,99);  self.hazirlik_mal.setDecimals(4); self.hazirlik_mal.setSuffix(" $/mt")
        self.boya_mal       = QDoubleSpinBox(); self.boya_mal.setRange(0,99);      self.boya_mal.setDecimals(2);     self.boya_mal.setSuffix(" $/kg")
        self.boya_fire      = QDoubleSpinBox(); self.boya_fire.setRange(0,50);     self.boya_fire.setDecimals(1);    self.boya_fire.setSuffix(" %")
        fi.addRow("Fason Dokuma (krş/100 atkı):", self.fason_dokuma)
        fi.addRow("USD Kuru:", self.usd_kuru)
        fi.addRow("Dokuma Maliyeti (hesaplanan, $/mt):", self.lbl_dokuma_mal)
        fi.addRow("Çözgü Hazırlık Maliyeti:", self.hazirlik_mal)
        fi.addRow("Boya / Apre ($/kg kumaş):", self.boya_mal)
        fi.addRow("Boya Firesi:", self.boya_fire)
        for w in (self.fason_dokuma, self.usd_kuru, self.hazirlik_mal, self.boya_mal, self.boya_fire):
            w.valueChanged.connect(self._recalc)
        vlay.addWidget(grp_islem)

        grp_sonuc = QGroupBox("Hesaplama Sonucu")
        fs = QFormLayout(grp_sonuc); fs.setSpacing(6)
        def _rl(color, big=False):
            l = QLabel("—")
            sz = "16px" if big else "13px"
            l.setStyleSheet(f"font-size:{sz};font-weight:bold;color:{color};")
            return l
        self.res_total_grs  = _rl("#2E7D32"); self.res_mat_cost  = _rl("#2E7D32")
        self.res_gri_cost   = _rl("#2E7D32"); self.res_boya_cost = _rl("#2E7D32")
        self.res_total_cost = _rl("#C62828", big=True)
        fs.addRow("Toplam Gramaj (gr/mt):", self.res_total_grs)
        fs.addRow("Malzeme Maliyeti ($/mt):", self.res_mat_cost)
        fs.addRow("Ham Maliyet ($/mt):", self.res_gri_cost)
        fs.addRow("Boya/Apre Maliyeti ($/mt):", self.res_boya_cost)
        fs.addRow("Toplam Maliyet ($/mt):", self.res_total_cost)
        vlay.addWidget(grp_sonuc)
        vlay.addStretch()

        grp_cvt = QGroupBox("Birim Dönüştürücü")
        fcvt = QHBoxLayout(grp_cvt)
        self.cvt_val = QDoubleSpinBox(); self.cvt_val.setRange(0,999999); self.cvt_val.setDecimals(2)
        self.cvt_bir = QComboBox(); self.cvt_bir.addItems(BIRIMLER)
        self.cvt_res = QLabel("—"); self.cvt_res.setStyleSheet("font-size:11px;color:#1565C0;")
        fcvt.addWidget(QLabel("Değer:")); fcvt.addWidget(self.cvt_val)
        fcvt.addWidget(QLabel("Birim:")); fcvt.addWidget(self.cvt_bir)
        fcvt.addWidget(QLabel("→")); fcvt.addWidget(self.cvt_res); fcvt.addStretch()
        self.cvt_val.valueChanged.connect(self._do_convert)
        self.cvt_bir.currentIndexChanged.connect(self._do_convert)
        vlay.addWidget(grp_cvt)

    def load(self, p):
        """Ürün dict'inden verileri yükler (p dict olmalı)."""
        import json
        try: self.tarak_eni.setValue(float(p.get("tarak_eni") or 0))
        except Exception: pass
        try: self.cozgu_sik.setValue(float(p.get("cozgu_sikligi") or 0))
        except Exception: pass
        try: self.atki_sik.setValue(float(p.get("atki_sikligi") or 0))
        except Exception: pass
        for i, row in enumerate(self.cozgu_rows):
            row["ad"].setText(p.get(f"cozgu{i+1}") or "")
        for i, row in enumerate(self.atki_rows):
            row["ad"].setText(p.get(f"atki{i+1}") or "")
        try:
            mj = json.loads(p.get("maliyet_json") or "{}")
        except Exception:
            mj = {}
        if not mj:
            self._recalc(); return
        try: self.fason_dokuma.setValue(float(mj.get("fason_dokuma", 0)))
        except: pass
        try: self.usd_kuru.setValue(float(mj.get("usd_kuru", 35.0)))
        except: pass
        try: self.hazirlik_mal.setValue(float(mj.get("hazirlik_mal", 0)))
        except: pass
        try: self.boya_fire.setValue(float(mj.get("boya_fire", 0)))
        except: pass
        try: self.boya_mal.setValue(float(mj.get("boya_mal", 0)))
        except: pass
        for i, row in enumerate(self.cozgu_rows):
            pre = f"c{i+1}_"
            try:
                if mj.get(pre+"num"):   row["num"].setValue(float(mj[pre+"num"]))
                if mj.get(pre+"bir"):   row["bir"].setCurrentText(mj[pre+"bir"])
                if mj.get(pre+"uc_cm"): row["uc_cm"].setValue(float(mj[pre+"uc_cm"]))
                if mj.get(pre+"cek"):   row["cek"].setValue(float(mj[pre+"cek"]))
                if mj.get(pre+"fiy"):   row["fiy"].setValue(float(mj[pre+"fiy"]))
            except: pass
        for i, row in enumerate(self.atki_rows):
            pre = f"a{i+1}_"
            try:
                if mj.get(pre+"num"): row["num"].setValue(float(mj[pre+"num"]))
                if mj.get(pre+"bir"): row["bir"].setCurrentText(mj[pre+"bir"])
                if mj.get(pre+"atm"): row["atm"].setValue(int(mj[pre+"atm"]))
                if mj.get(pre+"cek"): row["cek"].setValue(float(mj[pre+"cek"]))
                if mj.get(pre+"fiy"): row["fiy"].setValue(float(mj[pre+"fiy"]))
            except: pass
        self._recalc()

    def get_maliyet_json(self):
        import json
        mj = {
            "fason_dokuma": self.fason_dokuma.value(),
            "usd_kuru": self.usd_kuru.value(),
            "hazirlik_mal": self.hazirlik_mal.value(),
            "boya_mal": self.boya_mal.value(),
            "boya_fire": self.boya_fire.value(),
        }
        for i, row in enumerate(self.cozgu_rows):
            pre = f"c{i+1}_"
            mj[pre+"num"]   = row["num"].value()
            mj[pre+"bir"]   = row["bir"].currentText()
            mj[pre+"uc_cm"] = row["uc_cm"].value()
            mj[pre+"cek"]   = row["cek"].value()
            mj[pre+"fiy"]   = row["fiy"].value()
        for i, row in enumerate(self.atki_rows):
            pre = f"a{i+1}_"
            mj[pre+"num"] = row["num"].value()
            mj[pre+"bir"] = row["bir"].currentText()
            mj[pre+"atm"] = row["atm"].value()
            mj[pre+"cek"] = row["cek"].value()
            mj[pre+"fiy"] = row["fiy"].value()
        return json.dumps(mj)

    def _recalc(self):
        tarak     = self.tarak_eni.value()
        cozgu_sik = self.cozgu_sik.value()
        atki_sik  = self.atki_sik.value()
        toplam_uc = tarak * cozgu_sik
        self.lbl_toplam_uc.setText(f"{toplam_uc:.0f} tel  ({tarak:.1f} cm × {cozgu_sik:.1f} tel/cm)")
        desen_atim = max(1, sum(r["atm"].value() for r in self.atki_rows))
        self.lbl_toplam_desen.setText(
            f"{desen_atim}  ({' + '.join(str(r['atm'].value()) for r in self.atki_rows if r['atm'].value()>0) or '0'})"
        )
        aktif_cozgu = [r for r in self.cozgu_rows if _den_from(str(r["num"].value()), r["bir"].currentText()) > 0]
        toplam_girilen_uc_cm = sum(r["uc_cm"].value() for r in aktif_cozgu)
        total_grs = 0.0; total_mat = 0.0
        for row in self.cozgu_rows:
            den = _den_from(str(row["num"].value()), row["bir"].currentText())
            if den <= 0:
                row["tot_uc"].setText("—"); row["grs"].setText("—"); row["dol"].setText("—"); continue
            uc_cm = row["uc_cm"].value()
            if len(aktif_cozgu) == 1:            effective_uc_cm = cozgu_sik
            elif uc_cm > 0:                       effective_uc_cm = uc_cm
            else:                                 effective_uc_cm = max(0.0, cozgu_sik - toplam_girilen_uc_cm)
            total_ends = effective_uc_cm * tarak
            row["tot_uc"].setText(f"{total_ends:.0f}")
            grs = total_ends * den / 9000 * (1 + row["cek"].value() / 100) if total_ends > 0 else 0.0
            mat = grs / 1000 * row["fiy"].value()
            row["grs"].setText(f"{grs:.2f}"); row["dol"].setText(f"{mat:.4f}")
            total_grs += grs; total_mat += mat
        for row in self.atki_rows:
            den = _den_from(str(row["num"].value()), row["bir"].currentText())
            atm = row["atm"].value()
            if den > 0 and atm > 0 and atki_sik > 0 and tarak > 0:
                p_cm = atki_sik * atm / desen_atim
                grs  = p_cm * tarak * den / 9000 * (1 + row["cek"].value() / 100)
            else:
                grs = 0.0
            p_cm_d = atki_sik * atm / desen_atim if atm > 0 else 0
            mat = grs / 1000 * row["fiy"].value()
            row["ppm"].setText(f"{p_cm_d:.2f}"); row["grs"].setText(f"{grs:.2f}"); row["dol"].setText(f"{mat:.4f}")
            total_grs += grs; total_mat += mat
        usd_kuru = max(0.01, self.usd_kuru.value())
        dok = self.fason_dokuma.value() * atki_sik / 100.0 / usd_kuru
        self.lbl_dokuma_mal.setText(f"{dok:.4f} $/mt  ({self.fason_dokuma.value():.2f} krş × {atki_sik:.1f} ÷ {usd_kuru:.2f})")
        haz = self.hazirlik_mal.value(); boya_kg = self.boya_mal.value(); fire_pct = self.boya_fire.value()
        gri  = total_mat + dok + haz
        boya = total_grs / 1000 * (1 + fire_pct / 100) * boya_kg
        top  = gri + boya
        self.res_total_grs.setText(f"{total_grs:.2f} g/mt")
        self.res_mat_cost.setText(f"{total_mat:.4f} $/mt")
        self.res_gri_cost.setText(f"{gri:.4f} $/mt")
        self.res_boya_cost.setText(f"{boya:.4f} $/mt")
        self.res_total_cost.setText(f"{top:.4f} $/mt")
        for row in self.cozgu_rows + self.atki_rows:
            try:
                gv = float(row["grs"].text().replace("—", "0") or 0)
                row["pct"].setText(f"{gv / total_grs * 100:.1f}%" if total_grs > 0 else "—")
            except Exception:
                row["pct"].setText("—")

    def _do_convert(self):
        v = self.cvt_val.value(); b = self.cvt_bir.currentText()
        den = _den_from(str(v), b)
        if den <= 0: self.cvt_res.setText("—"); return
        self.cvt_res.setText(f"Den={den:.2f}  |  dTex={den/0.9:.2f}  |  Nm={9000/den:.2f}  |  Ne={5314.95/den:.2f}")


class ArmurManagerWidget(QWidget):
    """Armür desen kütüphanesi — ürün diyaloğuna gömülebilir."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Ad", "Çözgü", "Atkı", "Notlar", "Tarih"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)
        btn_row = QHBoxLayout()
        btn_new = QPushButton("+ Yeni Desen"); btn_new.clicked.connect(self._new)
        btn_edt = QPushButton("✎ Düzenle");    btn_edt.clicked.connect(self._edit)
        btn_del = QPushButton("✕ Sil"); btn_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        btn_del.clicked.connect(self._delete)
        for b in (btn_new, btn_edt, btn_del): btn_row.addWidget(b)
        btn_row.addStretch(); lay.addLayout(btn_row)

    def _load(self):
        rows = db.get_all_armur_desenleri() or []
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            it = QTableWidgetItem(r["name"]); it.setData(Qt.ItemDataRole.UserRole, r["id"])
            self.table.setItem(i, 0, it)
            self.table.setItem(i, 1, QTableWidgetItem(str(r.get("sutunlar", ""))))
            self.table.setItem(i, 2, QTableWidgetItem(str(r.get("satirlar", ""))))
            self.table.setItem(i, 3, QTableWidgetItem(r.get("notes", "")))
            self.table.setItem(i, 4, QTableWidgetItem(str(r.get("created_at", ""))[:10]))
        self.table.resizeColumnsToContents()

    def _selected(self):
        row = self.table.currentRow()
        if row < 0: QMessageBox.information(self, "Bilgi", "Desen seçin."); return None
        did = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        return db.get_armur_desen(did)

    def _new(self):
        dlg = ArmurDesignDialog(parent=self.window())
        if dlg.exec(): self._load()

    def _edit(self):
        d = self._selected()
        if not d: return
        dlg = ArmurDesignDialog(d, self.window())
        if dlg.exec(): self._load()

    def _delete(self):
        d = self._selected()
        if not d: return
        if QMessageBox.question(self, "Sil", f"'{d['name']}' silinsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_armur_desen(d["id"]); self._load()


class JakarDesenWidget(QWidget):
    """Jakar desen: ayrı JC5 + JPEG/PNG yükleme, önizleme JPEG'den."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._jc5_data   = b"";  self._jc5_ad   = ""
        self._jpeg_data  = b"";  self._jpeg_ad  = ""
        self._cached_pm  = None  # JPEG önizleme cache
        self._build_ui()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _build_ui(self):
        vlay = QVBoxLayout(self); vlay.setSpacing(10)

        top = QHBoxLayout(); top.setSpacing(12)

        # ── JC5 kutusu ──────────────────────────────────────────────────────
        jc5_grp = QGroupBox("JC5 Desen Dosyası")
        jc5_lay = QVBoxLayout(jc5_grp); jc5_lay.setSpacing(6)
        fi1 = QFormLayout(); fi1.setSpacing(4)
        self.lbl_jc5_ad    = QLabel("—"); self.lbl_jc5_ad.setStyleSheet("font-weight:bold;")
        self.lbl_jc5_boyut = QLabel("—")
        fi1.addRow("Dosya:", self.lbl_jc5_ad)
        fi1.addRow("Boyut:", self.lbl_jc5_boyut)
        jc5_lay.addLayout(fi1)
        br1 = QHBoxLayout(); br1.setSpacing(4)
        b_j1 = QPushButton("📂 Yükle");          b_j1.clicked.connect(self._yukle_jc5)
        b_j2 = QPushButton("💾 Kaydet");         b_j2.clicked.connect(self._kaydet_jc5)
        b_j3 = QPushButton("✕ Temizle")
        b_j3.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:4px 8px;")
        b_j3.clicked.connect(self._temizle_jc5)
        for b in (b_j1, b_j2, b_j3): br1.addWidget(b)
        jc5_lay.addLayout(br1)
        top.addWidget(jc5_grp, 1)

        # ── JPEG kutusu ─────────────────────────────────────────────────────
        jpg_grp = QGroupBox("Görsel (JPEG / PNG)")
        jpg_lay = QVBoxLayout(jpg_grp); jpg_lay.setSpacing(6)
        fi2 = QFormLayout(); fi2.setSpacing(4)
        self.lbl_jpg_ad    = QLabel("—"); self.lbl_jpg_ad.setStyleSheet("font-weight:bold;")
        self.lbl_jpg_boyut = QLabel("—")
        fi2.addRow("Dosya:", self.lbl_jpg_ad)
        fi2.addRow("Boyut:", self.lbl_jpg_boyut)
        jpg_lay.addLayout(fi2)
        br2 = QHBoxLayout(); br2.setSpacing(4)
        b_p1 = QPushButton("📂 Yükle");          b_p1.clicked.connect(self._yukle_jpeg)
        b_p2 = QPushButton("💾 Kaydet");         b_p2.clicked.connect(self._kaydet_jpeg)
        b_p3 = QPushButton("✕ Temizle")
        b_p3.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:4px 8px;")
        b_p3.clicked.connect(self._temizle_jpeg)
        for b in (b_p1, b_p2, b_p3): br2.addWidget(b)
        jpg_lay.addLayout(br2)
        top.addWidget(jpg_grp, 1)

        vlay.addLayout(top)

        # ── Önizleme (JPEG gösterilir) ───────────────────────────────────────
        prev_grp = QGroupBox("Önizleme (Görsel)")
        pg = QVBoxLayout(prev_grp)
        self.preview_lbl = QLabel("Henüz görsel yüklenmedi.")
        self.preview_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_lbl.setMinimumHeight(320)
        self.preview_lbl.setWordWrap(True)
        self.preview_lbl.setStyleSheet(
            "background:#f5f5f5; border:1px solid #ccc; border-radius:4px;"
            " color:#888; font-style:italic;"
        )
        pg.addWidget(self.preview_lbl)
        vlay.addWidget(prev_grp, 1)

    # ── JC5 işlemleri ────────────────────────────────────────────────────────

    def _yukle_jc5(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "JC5 Dosyası Seç", "",
            "JC5 Dosyaları (*.jc5);;Tüm Dosyalar (*)"
        )
        if not path: return
        try:
            self._jc5_data = open(path, "rb").read()
            self._jc5_ad   = os.path.basename(path)
            self._guncelle_jc5_info()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dosya okunamadı:\n{e}")

    def _kaydet_jc5(self):
        if not self._jc5_data:
            QMessageBox.information(self, "Bilgi", "Kaydedilecek JC5 dosyası yok.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "JC5 Dosyasını Kaydet", self._jc5_ad or "desen.jc5",
            "JC5 Dosyaları (*.jc5);;Tüm Dosyalar (*)"
        )
        if not path: return
        try:
            open(path, "wb").write(self._jc5_data)
            QMessageBox.information(self, "Başarılı", f"Kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kaydedilemedi:\n{e}")

    def _temizle_jc5(self):
        if self._jc5_data and QMessageBox.question(
            self, "Temizle", "JC5 dosyası silinsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) != QMessageBox.StandardButton.Yes:
            return
        self._jc5_data = b""; self._jc5_ad = ""
        self._guncelle_jc5_info()

    def _guncelle_jc5_info(self):
        if self._jc5_data:
            n = len(self._jc5_data)
            self.lbl_jc5_ad.setText(self._jc5_ad or "—")
            self.lbl_jc5_boyut.setText(
                f"{n/1024/1024:.2f} MB" if n >= 1024*1024 else f"{n/1024:.1f} KB"
            )
        else:
            self.lbl_jc5_ad.setText("—"); self.lbl_jc5_boyut.setText("—")

    # ── JPEG işlemleri ───────────────────────────────────────────────────────

    def _yukle_jpeg(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Görsel Dosyası Seç", "",
            "Resim Dosyaları (*.jpg *.jpeg *.png *.bmp *.tif *.tiff *.gif);;"
            "Tüm Dosyalar (*)"
        )
        if not path: return
        try:
            data = open(path, "rb").read()
            pm = QPixmap(); pm.loadFromData(data)
            if pm.isNull():
                QMessageBox.warning(self, "Uyarı", "Dosya geçerli bir resim değil.")
                return
            self._jpeg_data = data
            self._jpeg_ad   = os.path.basename(path)
            self._cached_pm = pm
            self._guncelle_jpeg_info()
            self._guncelle_onizleme()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dosya okunamadı:\n{e}")

    def _kaydet_jpeg(self):
        if not self._jpeg_data:
            QMessageBox.information(self, "Bilgi", "Kaydedilecek görsel yok.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Görseli Kaydet", self._jpeg_ad or "gorsel.jpg",
            "Resim Dosyaları (*.jpg *.jpeg *.png *.bmp);;Tüm Dosyalar (*)"
        )
        if not path: return
        try:
            open(path, "wb").write(self._jpeg_data)
            QMessageBox.information(self, "Başarılı", f"Kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kaydedilemedi:\n{e}")

    def _temizle_jpeg(self):
        if self._jpeg_data and QMessageBox.question(
            self, "Temizle", "Görsel silinsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) != QMessageBox.StandardButton.Yes:
            return
        self._jpeg_data = b""; self._jpeg_ad = ""; self._cached_pm = None
        self._guncelle_jpeg_info()
        self._guncelle_onizleme()

    def _guncelle_jpeg_info(self):
        if self._jpeg_data:
            n = len(self._jpeg_data)
            self.lbl_jpg_ad.setText(self._jpeg_ad or "—")
            self.lbl_jpg_boyut.setText(
                f"{n/1024/1024:.2f} MB" if n >= 1024*1024 else f"{n/1024:.1f} KB"
            )
        else:
            self.lbl_jpg_ad.setText("—"); self.lbl_jpg_boyut.setText("—")

    def _guncelle_onizleme(self):
        pm = self._cached_pm
        if pm and not pm.isNull():
            w = self.preview_lbl.width() or 640
            h = self.preview_lbl.height() or 360
            self.preview_lbl.setPixmap(
                pm.scaled(w, h, Qt.AspectRatioMode.KeepAspectRatio,
                          Qt.TransformationMode.SmoothTransformation)
            )
            self.preview_lbl.setStyleSheet(
                "background:#fff; border:1px solid #ccc; border-radius:4px;"
            )
        else:
            self.preview_lbl.clear()
            self.preview_lbl.setText("Henüz görsel yüklenmedi.")
            self.preview_lbl.setStyleSheet(
                "background:#f5f5f5; border:1px solid #ccc; border-radius:4px;"
                " color:#888; font-style:italic;"
            )

    def resizeEvent(self, ev):
        super().resizeEvent(ev)
        pm = self._cached_pm
        if pm and not pm.isNull():
            w = self.preview_lbl.width(); h = self.preview_lbl.height()
            if w > 0 and h > 0:
                self.preview_lbl.setPixmap(
                    pm.scaled(w, h, Qt.AspectRatioMode.KeepAspectRatio,
                              Qt.TransformationMode.SmoothTransformation)
                )

    # ── JC5 Parser (statik) ──────────────────────────────────────────────────

    @staticmethod
    def _parse_jc5(data: bytes):
        if len(data) < 110 or data[0:2] != b'\x81\xfe':
            return None
        try:
            import struct, numpy as np
            w = struct.unpack_from('>I', data, 0x21)[0]
            h = struct.unpack_from('>I', data, 0x61)[0]
            if not (0 < w <= 30000 and 0 < h <= 30000):
                return None
            bpr = (w + 7) // 8
            hdr = len(data) - bpr * h
            if not (0 <= hdr <= 512):
                return None
            arr  = np.frombuffer(data[hdr:hdr + bpr * h], dtype=np.uint8).reshape(h, bpr)
            bits = np.unpackbits(arr, axis=1)[:, :w]
            pix  = ((1 - bits) * 255).astype(np.uint8)
            return QPixmap.fromImage(QImage(pix.data, w, h, w, QImage.Format.Format_Grayscale8))
        except Exception:
            return None

    # ── Veri erişimi ──────────────────────────────────────────────────────────

    def load(self, p):
        import base64
        def _dec(key):
            try: return base64.b64decode(p.get(key) or "") if (p.get(key) or "") else b""
            except: return b""
        self._jc5_data  = _dec("jakar_desen_data");  self._jc5_ad  = p.get("jakar_desen_ad")  or ""
        self._jpeg_data = _dec("jakar_jpeg_data");   self._jpeg_ad = p.get("jakar_jpeg_ad")   or ""
        self._cached_pm = None
        if self._jpeg_data:
            pm = QPixmap(); pm.loadFromData(self._jpeg_data)
            self._cached_pm = pm if not pm.isNull() else None
        self._guncelle_jc5_info()
        self._guncelle_jpeg_info()
        self._guncelle_onizleme()

    def get_jc5_b64(self):
        import base64
        return base64.b64encode(self._jc5_data).decode() if self._jc5_data else ""

    def get_jpeg_b64(self):
        import base64
        return base64.b64encode(self._jpeg_data).decode() if self._jpeg_data else ""

    def get_jc5_ad(self):  return self._jc5_ad
    def get_jpeg_ad(self): return self._jpeg_ad


class MaliyetDialog(QDialog):
    """Dokuma kumaşı maliyet hesaplama diyalogu.

    Excel tablosundaki formüller:
      çözgü grs/mt  = toplam_uç × den / 9000 × (1 + çekme/100)
      atkı  grs/mt  = (atım/cm × tarak_eni × den) / 9000 × (1 + çekme/100)
      malzeme $/mt  = grs/mt / 1000 × fiyat_kg
      gri maliyet   = Σmalzeme + dokuma + çözgü_hazırlık
      toplam         = gri + boya × toplam_grs/mt/1000
    """

    def __init__(self, product, parent=None):
        super().__init__(parent)
        self.product = product
        self.setWindowTitle(f"Maliyet Hesaplama — {product.get('product_code','')}")
        self.setMinimumSize(820, 700)
        self._build_ui()
        self._load_from_product()
        self._recalc()

    def _build_ui(self):
        import json
        main = QVBoxLayout(self)

        # ── Ürün bilgisi başlık ───────────────────────────────────
        hdr = QLabel(f"<b>{self.product.get('product_code','')} — {self.product.get('product_name','')}</b>")
        hdr.setStyleSheet("font-size:14px; color:#1A237E; padding:4px;")
        main.addWidget(hdr)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        container = QWidget(); vlay = QVBoxLayout(container)
        scroll.setWidget(container)
        main.addWidget(scroll, 1)

        # ── Genel parametreler ────────────────────────────────────
        grp_genel = QGroupBox("Genel Parametreler")
        fg = QFormLayout(grp_genel); fg.setSpacing(6)
        self.tarak_eni  = QDoubleSpinBox(); self.tarak_eni.setRange(0,500); self.tarak_eni.setDecimals(1); self.tarak_eni.setSuffix(" cm")
        self.cozgu_sik  = QDoubleSpinBox(); self.cozgu_sik.setRange(0,500); self.cozgu_sik.setDecimals(1); self.cozgu_sik.setSuffix(" tel/cm")
        self.atki_sik   = QDoubleSpinBox(); self.atki_sik.setRange(0,200);  self.atki_sik.setDecimals(1); self.atki_sik.setSuffix(" atkı/cm")
        self.lbl_toplam_uc    = QLabel("—"); self.lbl_toplam_uc.setStyleSheet("font-weight:bold; color:#1A237E;")
        self.lbl_toplam_desen = QLabel("—"); self.lbl_toplam_desen.setStyleSheet("font-weight:bold; color:#1A237E;")
        fg.addRow("Tarak Eni:", self.tarak_eni)
        fg.addRow("Çözgü Sıklığı (tel/cm):", self.cozgu_sik)
        fg.addRow("Toplam Çözgü Ucu (otomatik):", self.lbl_toplam_uc)
        fg.addRow("Atkı Sıklığı (atkı/cm):", self.atki_sik)
        fg.addRow("Toplam Desen Atım (otomatik):", self.lbl_toplam_desen)
        note = QLabel("Toplam desen atım = atkı satırlarındaki desen atım değerlerinin toplamı (otomatik hesaplanır).")
        note.setWordWrap(True); note.setStyleSheet("color:#757575; font-size:10px;")
        fg.addRow("", note)
        vlay.addWidget(grp_genel)

        BIRIMLER = ["Den", "dTex", "Nm", "Ne"]

        # ── Çözgüler ─────────────────────────────────────────────
        grp_cozgu = QGroupBox("Çözgü İplikleri")
        gc = QGridLayout(grp_cozgu)
        headers = ["", "İplik Adı", "Numara", "Birim", "Tel/cm", "Toplam Tel Sayısı", "Çekme %", "Fiyat $/kg", "gr/mt", "%", "$/mt"]
        for ci, h in enumerate(headers):
            lbl = QLabel(f"<b>{h}</b>"); lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            gc.addWidget(lbl, 0, ci)
        self.cozgu_rows = []
        for i in range(2):
            lbl    = QLabel(f"Çözgü {i+1}:")
            ad     = QLineEdit()
            num    = QDoubleSpinBox(); num.setRange(0,99999); num.setDecimals(2)
            bir    = QComboBox(); bir.addItems(BIRIMLER)
            uc_cm  = QDoubleSpinBox(); uc_cm.setRange(0,500); uc_cm.setDecimals(1)
            uc_cm.setToolTip("Bu ipliğin cm başına uç sayısı (0 bırakılırsa tek çözgüde otomatik hesaplanır)")
            tot_uc = QLabel("—"); tot_uc.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            tot_uc.setStyleSheet("color:#555; padding-right:4px;")
            cek    = QDoubleSpinBox(); cek.setRange(0,50); cek.setDecimals(1); cek.setValue(6.0); cek.setSuffix(" %")
            fiy    = QDoubleSpinBox(); fiy.setRange(0,999); fiy.setDecimals(2); fiy.setSuffix(" $/kg")
            grs    = QLabel("0.00"); grs.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            grs.setStyleSheet("color:#1565C0; padding-right:4px;")
            pct    = QLabel("—"); pct.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            pct.setStyleSheet("color:#7B1FA2; padding-right:4px;")
            dol    = QLabel("0.0000"); dol.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            dol.setStyleSheet("color:#2E7D32; padding-right:4px;")
            row_w = (lbl, ad, num, bir, uc_cm, tot_uc, cek, fiy, grs, pct, dol)
            for ci, w in enumerate(row_w): gc.addWidget(w, i+1, ci)
            self.cozgu_rows.append({"ad": ad, "num": num, "bir": bir, "uc_cm": uc_cm,
                                    "tot_uc": tot_uc, "cek": cek, "fiy": fiy, "grs": grs, "pct": pct, "dol": dol})
            for w in (num, bir, uc_cm, cek, fiy):
                (w.valueChanged if hasattr(w,'valueChanged') else w.currentIndexChanged).connect(self._recalc)

        # Çözgü 2 uç/cm otomatik doldurma:
        # Çözgü 1 uç/cm veya çözgü sıklığı değişince → Çözgü 2'nin uç/cm = sıklık − Çözgü 1
        def _auto_cozgu2(*_):
            r1 = self.cozgu_rows[0]; r2 = self.cozgu_rows[1]
            den2 = _den_from(str(r2["num"].value()), r2["bir"].currentText())
            if den2 <= 0:
                return   # Çözgü 2 girilmemiş, dokunma
            kalan = max(0.0, self.cozgu_sik.value() - r1["uc_cm"].value())
            r2["uc_cm"].blockSignals(True)
            r2["uc_cm"].setValue(round(kalan, 1))
            r2["uc_cm"].blockSignals(False)
            self._recalc()

        self.cozgu_rows[0]["uc_cm"].valueChanged.connect(_auto_cozgu2)
        self.cozgu_sik.valueChanged.connect(_auto_cozgu2)
        self.cozgu_rows[1]["num"].valueChanged.connect(_auto_cozgu2)
        vlay.addWidget(grp_cozgu)

        # ── Atkılar ───────────────────────────────────────────────
        grp_atki = QGroupBox("Atkı İplikleri")
        ga = QGridLayout(grp_atki)
        headers2 = ["", "İplik Adı", "Numara", "Birim", "Atkı Raporu", "Çekme %", "Fiyat $/kg", "atkı/cm", "gr/mt", "%", "$/mt"]
        for ci, h in enumerate(headers2):
            lbl = QLabel(f"<b>{h}</b>"); lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            ga.addWidget(lbl, 0, ci)
        self.atki_rows = []
        for i in range(4):
            lbl = QLabel(f"Atkı {i+1}:")
            ad  = QLineEdit()
            num = QDoubleSpinBox(); num.setRange(0,99999); num.setDecimals(2)
            bir = QComboBox(); bir.addItems(BIRIMLER)
            atm = QSpinBox(); atm.setRange(0,256)
            cek = QDoubleSpinBox(); cek.setRange(0,50); cek.setDecimals(1); cek.setValue(8.0); cek.setSuffix(" %")
            fiy = QDoubleSpinBox(); fiy.setRange(0,999); fiy.setDecimals(2); fiy.setSuffix(" $/kg")
            ppm = QLabel("0.00"); ppm.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            ppm.setStyleSheet("color:#555; padding-right:4px;")
            grs = QLabel("0.00"); grs.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            grs.setStyleSheet("color:#1565C0; padding-right:4px;")
            pct = QLabel("—"); pct.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            pct.setStyleSheet("color:#7B1FA2; padding-right:4px;")
            dol = QLabel("0.0000"); dol.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            dol.setStyleSheet("color:#2E7D32; padding-right:4px;")
            row_w = (lbl, ad, num, bir, atm, cek, fiy, ppm, grs, pct, dol)
            for ci, w in enumerate(row_w): ga.addWidget(w, i+1, ci)
            self.atki_rows.append({"ad": ad, "num": num, "bir": bir, "atm": atm,
                                   "cek": cek, "fiy": fiy, "ppm": ppm, "grs": grs, "pct": pct, "dol": dol})
            for w in (num, bir, atm, cek, fiy):
                (w.valueChanged if hasattr(w,'valueChanged') else w.currentIndexChanged).connect(self._recalc)
        vlay.addWidget(grp_atki)

        # Genel recalc bağlantıları
        for w in (self.tarak_eni, self.cozgu_sik, self.atki_sik):
            w.valueChanged.connect(self._recalc)

        # ── İşlem Maliyetleri ─────────────────────────────────────
        grp_islem = QGroupBox("İşlem Maliyetleri")
        fi = QFormLayout(grp_islem); fi.setSpacing(6)

        # Fason dokuma: krş/100 atkı — dokuma maliyeti buradan otomatik hesaplanır
        self.fason_dokuma = QDoubleSpinBox()
        self.fason_dokuma.setRange(0, 9999); self.fason_dokuma.setDecimals(2)
        self.fason_dokuma.setSuffix(" krş/100 atkı")
        self.fason_dokuma.setToolTip("Fason dokuma bedeli — 100 atkı başına kuruş")

        self.usd_kuru = QDoubleSpinBox()
        self.usd_kuru.setRange(0.01, 999); self.usd_kuru.setDecimals(2)
        self.usd_kuru.setValue(35.0); self.usd_kuru.setSuffix(" TL/$")
        self.usd_kuru.setToolTip("USD döviz kuru — dokuma maliyetini TL'den $'a çevirmek için")

        self.lbl_dokuma_mal = QLabel("—")
        self.lbl_dokuma_mal.setStyleSheet("font-weight:bold; color:#1A237E;")
        self.lbl_dokuma_mal.setToolTip(
            "= Fason (krş/100 atkı) × Atkı sıklığı (atım/cm) × 100 cm/m ÷ 100 (krş→TL) ÷ USD kuru"
        )

        self.hazirlik_mal = QDoubleSpinBox()
        self.hazirlik_mal.setRange(0, 99); self.hazirlik_mal.setDecimals(4)
        self.hazirlik_mal.setSuffix(" $/mt")
        self.boya_mal = QDoubleSpinBox()
        self.boya_mal.setRange(0, 99); self.boya_mal.setDecimals(2)
        self.boya_mal.setSuffix(" $/kg")

        self.boya_fire = QDoubleSpinBox()
        self.boya_fire.setRange(0, 50); self.boya_fire.setDecimals(1); self.boya_fire.setSuffix(" %")
        self.boya_fire.setToolTip("Boya/apre sürecinde oluşan fire oranı — boya maliyeti bu ağırlık üzerinden hesaplanır")

        fi.addRow("Fason Dokuma (krş/100 atkı):", self.fason_dokuma)
        fi.addRow("USD Kuru:", self.usd_kuru)
        fi.addRow("Dokuma Maliyeti (hesaplanan, $/mt):", self.lbl_dokuma_mal)
        fi.addRow("Çözgü Hazırlık Maliyeti:", self.hazirlik_mal)
        fi.addRow("Boya / Apre ($/kg kumaş):", self.boya_mal)
        fi.addRow("Boya Firesi:", self.boya_fire)
        for w in (self.fason_dokuma, self.usd_kuru, self.hazirlik_mal, self.boya_mal, self.boya_fire):
            w.valueChanged.connect(self._recalc)
        vlay.addWidget(grp_islem)

        # ── Sonuç Özeti ───────────────────────────────────────────
        grp_sonuc = QGroupBox("Hesaplama Sonucu")
        fs = QFormLayout(grp_sonuc); fs.setSpacing(6)
        def _res_lbl(color="#000", big=False):
            sz = "16px" if big else "13px"
            l = QLabel("—"); l.setStyleSheet(f"font-size:{sz}; font-weight:bold; color:{color};")
            return l
        self.res_total_grs   = _res_lbl("#2E7D32")
        self.res_mat_cost    = _res_lbl("#2E7D32")
        self.res_gri_cost    = _res_lbl("#2E7D32")
        self.res_boya_cost   = _res_lbl("#2E7D32")
        self.res_total_cost  = _res_lbl("#C62828", big=True)
        fs.addRow("Toplam Gramaj (gr/mt):", self.res_total_grs)
        fs.addRow("Malzeme Maliyeti ($/mt):", self.res_mat_cost)
        fs.addRow("Ham Maliyet ($/mt):", self.res_gri_cost)
        fs.addRow("Boya/Apre Maliyeti ($/mt):", self.res_boya_cost)
        fs.addRow("Toplam Maliyet ($/mt):", self.res_total_cost)
        vlay.addWidget(grp_sonuc)
        vlay.addStretch()

        # ── Birim çevirici ────────────────────────────────────────
        grp_cvt = QGroupBox("Birim Dönüştürücü")
        fcvt = QHBoxLayout(grp_cvt)
        self.cvt_val  = QDoubleSpinBox(); self.cvt_val.setRange(0,999999); self.cvt_val.setDecimals(2)
        self.cvt_bir  = QComboBox(); self.cvt_bir.addItems(BIRIMLER)
        self.cvt_res  = QLabel("—"); self.cvt_res.setStyleSheet("font-size:11px;color:#1565C0;")
        fcvt.addWidget(QLabel("Değer:")); fcvt.addWidget(self.cvt_val)
        fcvt.addWidget(QLabel("Birim:")); fcvt.addWidget(self.cvt_bir)
        fcvt.addWidget(QLabel("→")); fcvt.addWidget(self.cvt_res)
        fcvt.addStretch()
        self.cvt_val.valueChanged.connect(self._do_convert)
        self.cvt_bir.currentIndexChanged.connect(self._do_convert)
        vlay.addWidget(grp_cvt)

        # ── Alt Butonlar ──────────────────────────────────────────
        bot = QHBoxLayout()
        btn_save = QPushButton("💾 Maliyet Parametrelerini Ürüne Kaydet")
        btn_save.setStyleSheet("background:#1565C0;color:white;font-weight:bold;border-radius:4px;padding:8px 16px;")
        btn_save.clicked.connect(self._save_to_product)
        btn_close = QPushButton("Kapat"); btn_close.clicked.connect(self.reject)
        bot.addWidget(btn_save); bot.addStretch(); bot.addWidget(btn_close)
        main.addLayout(bot)

    def _load_from_product(self):
        import json
        p = self.product
        # Teknik alanlar
        try: self.tarak_eni.setValue(float(p.get("tarak_eni") or 0))
        except Exception: pass
        try: self.cozgu_sik.setValue(float(p.get("cozgu_sikligi") or 0))
        except Exception: pass
        try: self.atki_sik.setValue(float(p.get("atki_sikligi") or 0))
        except Exception: pass
        # İplik adları
        for i, row in enumerate(self.cozgu_rows):
            key = f"cozgu{i+1}"
            row["ad"].setText(p.get(key) or "")
        for i, row in enumerate(self.atki_rows):
            key = f"atki{i+1}"
            row["ad"].setText(p.get(key) or "")
        # Kaydedilmiş maliyet JSON
        try:
            mj = json.loads(p.get("maliyet_json") or "{}")
        except Exception:
            mj = {}
        if mj:
            # desen_atim artık satırlardan otomatik hesaplanıyor, yüklemeye gerek yok
            self.fason_dokuma.setValue(float(mj.get("fason_dokuma", 0)))
            self.usd_kuru.setValue(float(mj.get("usd_kuru", 35.0)))
            self.hazirlik_mal.setValue(float(mj.get("hazirlik_mal", 0)))
            self.boya_fire.setValue(float(mj.get("boya_fire", 0)))
            self.boya_mal.setValue(float(mj.get("boya_mal", 0)))
            for i, row in enumerate(self.cozgu_rows):
                pre = f"c{i+1}_"
                if mj.get(pre+"num"):   row["num"].setValue(float(mj[pre+"num"]))
                if mj.get(pre+"bir"):   row["bir"].setCurrentText(mj[pre+"bir"])
                if mj.get(pre+"uc_cm"): row["uc_cm"].setValue(float(mj[pre+"uc_cm"]))
                if mj.get(pre+"cek"):   row["cek"].setValue(float(mj[pre+"cek"]))
                if mj.get(pre+"fiy"):   row["fiy"].setValue(float(mj[pre+"fiy"]))
            for i, row in enumerate(self.atki_rows):
                pre = f"a{i+1}_"
                if mj.get(pre+"num"): row["num"].setValue(float(mj[pre+"num"]))
                if mj.get(pre+"bir"): row["bir"].setCurrentText(mj[pre+"bir"])
                if mj.get(pre+"atm"): row["atm"].setValue(int(mj[pre+"atm"]))
                if mj.get(pre+"cek"): row["cek"].setValue(float(mj[pre+"cek"]))
                if mj.get(pre+"fiy"): row["fiy"].setValue(float(mj[pre+"fiy"]))

    def _recalc(self):
        tarak     = self.tarak_eni.value()
        cozgu_sik = self.cozgu_sik.value()
        atki_sik  = self.atki_sik.value()

        # Toplam çözgü ucu = tarak eni × çözgü sıklığı (her zaman)
        toplam_uc = tarak * cozgu_sik
        self.lbl_toplam_uc.setText(f"{toplam_uc:.0f} tel  ({tarak:.1f} cm × {cozgu_sik:.1f} tel/cm)")

        # Toplam desen atım = tüm atkı satırlarındaki desen atım değerlerinin toplamı
        desen_atim = sum(r["atm"].value() for r in self.atki_rows)
        desen_atim = max(1, desen_atim)
        self.lbl_toplam_desen.setText(
            f"{desen_atim}  ({' + '.join(str(r['atm'].value()) for r in self.atki_rows if r['atm'].value() > 0) or '0'})"
        )

        # Her çözgü satırının uç/cm değeri: 0 girilmişse otomatik hesapla
        # Kural: hangi çözgü satırlarının aktif olduğunu belirle (num > 0)
        aktif_cozgu = [r for r in self.cozgu_rows if _den_from(str(r["num"].value()), r["bir"].currentText()) > 0]
        girmis_uc   = [r["uc_cm"].value() for r in aktif_cozgu]
        toplam_girilen_uc_cm = sum(girmis_uc)

        total_grs = 0.0; total_mat = 0.0

        # Çözgü hesapları
        for i, row in enumerate(self.cozgu_rows):
            den = _den_from(str(row["num"].value()), row["bir"].currentText())
            cek = row["cek"].value()
            fiy = row["fiy"].value()
            if den <= 0:
                row["tot_uc"].setText("—")
                row["grs"].setText("—")
                row["dol"].setText("—")
                continue

            uc_cm = row["uc_cm"].value()
            if len(aktif_cozgu) == 1:
                # Tek çözgü: tüm çözgü sıklığını kullan
                effective_uc_cm = cozgu_sik
            elif uc_cm > 0:
                # İki çözgü, uç/cm elle girilmiş
                effective_uc_cm = uc_cm
            else:
                # İki çözgü ama bu satır için uç/cm girilmemiş: kalan payı kullan
                effective_uc_cm = max(0.0, cozgu_sik - toplam_girilen_uc_cm)

            total_ends = effective_uc_cm * tarak
            row["tot_uc"].setText(f"{total_ends:.0f}")

            grs = total_ends * den / 9000 * (1 + cek / 100) if total_ends > 0 else 0.0
            mat = grs / 1000 * fiy
            row["grs"].setText(f"{grs:.2f}")
            row["dol"].setText(f"{mat:.4f}")
            total_grs += grs; total_mat += mat

        # Atkı hesapları
        for row in self.atki_rows:
            den  = _den_from(str(row["num"].value()), row["bir"].currentText())
            atm  = row["atm"].value()         # desende kaç atım
            cek  = row["cek"].value()
            fiy  = row["fiy"].value()
            if den > 0 and atm > 0 and atki_sik > 0 and tarak > 0:
                p_cm = atki_sik * atm / desen_atim
                grs  = p_cm * tarak * den / 9000 * (1 + cek / 100)
            else:
                grs = 0.0
            p_cm_disp = atki_sik * atm / desen_atim if atm > 0 and desen_atim > 0 else 0
            mat  = grs / 1000 * fiy
            row["ppm"].setText(f"{p_cm_disp:.2f}")
            row["grs"].setText(f"{grs:.2f}")
            row["dol"].setText(f"{mat:.4f}")
            total_grs += grs; total_mat += mat

        # Dokuma maliyeti = fason (krş/100 atkı) × atkı sıklığı (atım/cm) × 100 cm/m
        #                   ÷ 100 (krş→TL) ÷ USD kuru (TL/$)
        usd_kuru = max(0.01, self.usd_kuru.value())
        dok = self.fason_dokuma.value() * atki_sik / 100.0 / usd_kuru   # $/mt
        self.lbl_dokuma_mal.setText(
            f"{dok:.4f} $/mt  "
            f"({self.fason_dokuma.value():.2f} krş × {atki_sik:.1f} atkı/cm ÷ {usd_kuru:.2f} TL/$)"
        )

        haz      = self.hazirlik_mal.value()
        boya_kg  = self.boya_mal.value()
        fire_pct = self.boya_fire.value()
        gri  = total_mat + dok + haz
        # Boya maliyeti: fire sonrası ağırlık = grs/mt × (1 + fire%)
        boya = total_grs / 1000 * (1 + fire_pct / 100) * boya_kg
        top  = gri + boya

        self.res_total_grs.setText(f"{total_grs:.2f} g/mt")
        self.res_mat_cost.setText(f"{total_mat:.4f} $/mt")
        self.res_gri_cost.setText(f"{gri:.4f} $/mt")
        self.res_boya_cost.setText(f"{boya:.4f} $/mt")
        self.res_total_cost.setText(f"{top:.4f} $/mt")
        for row in self.cozgu_rows + self.atki_rows:
            try:
                gv = float(row["grs"].text().replace("—", "0") or 0)
                row["pct"].setText(f"{gv / total_grs * 100:.1f}%" if total_grs > 0 else "—")
            except Exception:
                row["pct"].setText("—")

    def _do_convert(self):
        v = self.cvt_val.value(); b = self.cvt_bir.currentText()
        den = _den_from(str(v), b)
        if den <= 0: self.cvt_res.setText("—"); return
        nm   = 9000 / den
        ne   = 5314.95 / den
        dtex = den / 0.9
        self.cvt_res.setText(
            f"Den={den:.2f}  |  dTex={dtex:.2f}  |  Nm={nm:.2f}  |  Ne={ne:.2f}"
        )

    def _save_to_product(self):
        import json
        mj = {
            "desen_atim": sum(r["atm"].value() for r in self.atki_rows),
            "fason_dokuma": self.fason_dokuma.value(),
            "usd_kuru": self.usd_kuru.value(),
            "hazirlik_mal": self.hazirlik_mal.value(),
            "boya_mal": self.boya_mal.value(),
            "boya_fire": self.boya_fire.value(),
        }
        for i, row in enumerate(self.cozgu_rows):
            pre = f"c{i+1}_"
            mj[pre+"num"]   = row["num"].value()
            mj[pre+"bir"]   = row["bir"].currentText()
            mj[pre+"uc_cm"] = row["uc_cm"].value()
            mj[pre+"cek"] = row["cek"].value()
            mj[pre+"fiy"] = row["fiy"].value()
        for i, row in enumerate(self.atki_rows):
            pre = f"a{i+1}_"
            mj[pre+"num"] = row["num"].value()
            mj[pre+"bir"] = row["bir"].currentText()
            mj[pre+"atm"] = row["atm"].value()
            mj[pre+"cek"] = row["cek"].value()
            mj[pre+"fiy"] = row["fiy"].value()
        try:
            p = self.product
            db.update_product(
                p["id"], p["product_code"], p.get("product_name",""), p.get("composition",""),
                p.get("width",""), p.get("gramaj",""), p.get("shrinkage",""),
                p.get("price",0), p.get("supplier",""), p.get("active",1), p.get("reference_code",""),
                p.get("cozgu1",""), p.get("cozgu2",""),
                p.get("atki1",""), p.get("atki2",""), p.get("atki3",""), p.get("atki4",""),
                p.get("dokuma_tipi",""), str(self.cozgu_sik.value()),
                p.get("tarak_no",""), str(self.tarak_eni.value()),
                str(self.atki_sik.value()), p.get("orgu_desen",""),
                json.dumps(mj)
            )
            QMessageBox.information(self, "Kaydedildi", "Maliyet parametreleri ürüne kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))


FABRIC_TYPE_COLORS = {"HAM": QColor("#5D4037"), "PFD": QColor("#00695C"),
                      "BOYALI": QColor("#545454"), "İPLİĞİ BOYALI": QColor("#EF6C00"),
                      "BASKILI": QColor("#6A1B9A")}

PRINT_TYPES = ["RONJAN", "ROTASYON", "DİJİTAL", "FLOK", "VARAK", "GLITTER"]

GREY   = "#545454"
GREY_D = "#3A3A3A"   # koyu gri (çerçeve, hover)
GREY_L = "#6B6B6B"   # açık gri (hover)

COLORS = {
    "primary":      GREY,
    "primary_light": GREY_L,
    "success": "#2E7D32",
    "danger": "#C62828",
    "warning": "#F57F17",
    "bg": "#F5F5F5",
    "white": "#FFFFFF",
    "border": "#BDBDBD",
    "text": "#212121",
    "subtext": "#757575",
    "row_alt": "#F8F9FA",
    "row_hover": "#EEEEEE",
    "header_bg": GREY,
    "header_fg": "#FFFFFF",
}

STYLESHEET = f"""
QMainWindow, QWidget {{
    background-color: {COLORS['bg']};
    color: {COLORS['text']};
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}}
QPushButton {{
    background-color: {COLORS['primary']};
    color: white;
    border: none;
    border-radius: 4px;
    padding: 7px 16px;
    font-weight: bold;
    min-height: 30px;
}}
QPushButton:hover {{
    background-color: {COLORS['primary_light']};
}}
QPushButton:pressed {{
    background-color: {GREY_D};
}}
QPushButton.success {{
    background-color: {COLORS['success']};
}}
QPushButton.success:hover {{
    background-color: #388E3C;
}}
QPushButton.danger {{
    background-color: {COLORS['danger']};
}}
QPushButton.danger:hover {{
    background-color: #D32F2F;
}}
QPushButton.warning {{
    background-color: {COLORS['warning']};
    color: #212121;
}}
QLineEdit, QComboBox, QDoubleSpinBox, QTextEdit {{
    border: 1px solid {COLORS['border']};
    border-radius: 4px;
    padding: 5px 8px;
    background: white;
    min-height: 28px;
}}
QLineEdit:focus, QComboBox:focus, QDoubleSpinBox:focus, QTextEdit:focus {{
    border: 2px solid {GREY};
}}
QTableWidget {{
    background: white;
    gridline-color: #E0E0E0;
    border: 1px solid {COLORS['border']};
    border-radius: 4px;
    selection-background-color: {COLORS['row_hover']};
    selection-color: {COLORS['text']};
}}
QTableWidget::item {{
    padding: 4px 8px;
}}
QHeaderView::section {{
    background-color: {COLORS['header_bg']};
    color: {COLORS['header_fg']};
    font-weight: bold;
    padding: 6px 8px;
    border: none;
    border-right: 1px solid {GREY_L};
}}
QTabWidget::pane {{
    border: 1px solid {COLORS['border']};
    border-radius: 4px;
    background: white;
}}
QTabBar::tab {{
    background: #E0E0E0;
    color: {COLORS['text']};
    padding: 8px 20px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
    min-width: 100px;
}}
QTabBar::tab:selected {{
    background: {GREY};
    color: white;
    font-weight: bold;
}}
QGroupBox {{
    font-weight: bold;
    border: 1px solid {COLORS['border']};
    border-radius: 4px;
    margin-top: 8px;
    padding-top: 8px;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    padding: 0 6px;
    color: {COLORS['primary']};
}}
QStatusBar {{
    background: #F0F0F0;
    color: #212121;
    font-size: 13px;
    min-height: 28px;
    border-top: 1px solid #BDBDBD;
}}
QLabel.title {{
    font-size: 16px;
    font-weight: bold;
    color: {COLORS['primary']};
}}
QLabel.stat {{
    font-size: 15px;
    font-weight: bold;
    color: {COLORS['primary']};
}}
"""


class _MobileAccessDialog(QDialog):
    """WiFi erişim bilgisi."""
    def __init__(self, parent, ip, port):
        super().__init__(parent)
        self.setWindowTitle("Mobil Erişim — WiFi")
        self.setMinimumWidth(380)
        lay = QVBoxLayout(self)
        lay.setSpacing(12)

        info = QLabel(
            f"<b>WiFi Yerel Erişim Açıldı</b><br><br>"
            f"Aynı WiFi ağındaki telefon / tablet tarayıcısına:<br><br>"
            f"<span style='font-size:18px; color:#545454; font-family:monospace'>"
            f"http://{ip}:{port}</span><br><br>"
            f"<span style='color:#757575; font-size:12px'>"
            f"İnternetten her yerden erişmek için:<br>"
            f"Menü → 📱 Mobil Erişim → İnternetten Erişim Aç (ngrok)</span>"
        )
        info.setWordWrap(True)
        info.setStyleSheet("background:#E3F2FD; padding:16px; border-radius:6px;")
        lay.addWidget(info)

        url = f"http://{ip}:{port}"
        btn_copy = QPushButton("📋 Adresi Kopyala")
        btn_copy.clicked.connect(lambda: (QApplication.clipboard().setText(url),
                                          btn_copy.setText("✓ Kopyalandı!")))
        btn_close = QPushButton("Tamam")
        btn_close.clicked.connect(self.accept)
        row = QHBoxLayout()
        row.addWidget(btn_copy); row.addWidget(btn_close)
        lay.addLayout(row)


class _NgrokSetupDialog(QDialog):
    """ngrok token kurulum dialog'u."""
    def __init__(self, parent, current_token=""):
        super().__init__(parent)
        self.setWindowTitle("ngrok Token Kurulumu")
        self.setMinimumWidth(480)
        self.token = ""
        lay = QVBoxLayout(self)
        lay.setSpacing(12)

        steps = QLabel(
            "<b>Kurulum adımları (sadece bir kez):</b><br><br>"
            "1. <a href='https://ngrok.com/signup'>ngrok.com/signup</a> adresinden <b>ücretsiz hesap</b> açın<br>"
            "2. Giriş yapın → <b>Your Authtoken</b> sayfasına gidin<br>"
            "3. Token'ı kopyalayıp aşağıya yapıştırın<br><br>"
            "<span style='color:#2E7D32'>✓ Ücretsiz — aylık 1GB, sınırsız kullanıcı</span>"
        )
        steps.setWordWrap(True)
        steps.setOpenExternalLinks(True)
        steps.setStyleSheet("background:#FFF8E1; padding:14px; border-radius:6px;")
        lay.addWidget(steps)

        form = QFormLayout()
        self.token_edit = QLineEdit(current_token)
        self.token_edit.setPlaceholderText("2abc123xyz... şeklinde token yapıştırın")
        form.addRow("ngrok Token:", self.token_edit)
        lay.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._save)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _save(self):
        t = self.token_edit.text().strip()
        if len(t) < 10:
            QMessageBox.warning(self, "Hata", "Geçerli bir token giriniz!")
            return
        self.token = t
        self.accept()


class _NgrokActiveDialog(QDialog):
    """ngrok aktif — URL ve QR göster."""
    def __init__(self, parent, url):
        super().__init__(parent)
        self.setWindowTitle("İnternetten Erişim Açıldı 🌍")
        self.setMinimumWidth(420)
        lay = QVBoxLayout(self)
        lay.setSpacing(12)

        # HTTPS URL göster
        info = QLabel(
            f"<b>🌍 İnternetten erişim AÇIK</b><br><br>"
            f"Herhangi bir telefon veya bilgisayar tarayıcısından:<br><br>"
            f"<span style='font-size:16px; color:#545454; font-family:monospace'>{url}</span><br><br>"
            f"<span style='color:#757575; font-size:12px'>"
            f"⚠ Program kapatılınca veya durdurulunca erişim kapanır.<br>"
            f"Her açılışta URL değişebilir.</span>"
        )
        info.setWordWrap(True)
        info.setStyleSheet("background:#E8F5E9; padding:16px; border-radius:6px;")
        lay.addWidget(info)

        # QR kod (online servis üzerinden)
        try:
            from PyQt6.QtNetwork import QNetworkAccessManager
        except Exception:
            pass

        qr_label = QLabel()
        qr_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=180x180&data={url}"
        # QR kodu indirip göster
        try:
            import urllib.request
            data = urllib.request.urlopen(qr_url, timeout=5).read()
            pix = QPixmap()
            pix.loadFromData(data)
            qr_label.setPixmap(pix)
            qr_label.setToolTip("Telefon kamerasıyla tarayın")
        except Exception:
            qr_label.setText("(QR kod için internet gerekli)")
        lay.addWidget(qr_label)

        btn_copy = QPushButton("📋 Adresi Kopyala")
        btn_copy.clicked.connect(lambda: (QApplication.clipboard().setText(url),
                                          btn_copy.setText("✓ Kopyalandı!")))
        btn_close = QPushButton("Tamam")
        btn_close.clicked.connect(self.accept)
        row = QHBoxLayout()
        row.addWidget(btn_copy); row.addWidget(btn_close)
        lay.addLayout(row)


# ── Excel sütun eşleştirme ───────────────────────────────────────

# (alan_anahtarı, etiket, zorunlu_mu)
CUSTOMER_FIELDS = [
    ("name",    "Müşteri Adı", True),
    ("code",    "Kodu",        False),
    ("phone",   "Telefon",     False),
    ("address", "Adres",       False),
    ("tax_no",  "Vergi No",    False),
]
CUSTOMER_AUTO = {
    "name":    ["ad","isim","müşteri","musteri","name","unvan","firma"],
    "code":    ["kod","code"],
    "phone":   ["tel","phone","gsm","cep"],
    "address": ["adres","address"],
    "tax_no":  ["vergi","tax"],
}

SUPPLIER_FIELDS = [
    ("name",    "Tedarikçi Adı", True),
    ("code",    "Kodu",          False),
    ("phone",   "Telefon",       False),
    ("address", "Adres",         False),
    ("tax_no",  "Vergi No",      False),
]
SUPPLIER_AUTO = {
    "name":    ["ad","isim","tedarikçi","tedarikci","müşteri","musteri","name","unvan","firma"],
    "code":    ["kod","code"],
    "phone":   ["tel","phone","gsm","cep"],
    "address": ["adres","address"],
    "tax_no":  ["vergi","tax"],
}

PRODUCT_FIELDS = [
    ("product_code",   "Ürün Kodu",          True),
    ("reference_code", "Açıklama",            False),
    ("product_name",   "Ürün Adı/Bilgisi",   False),
    ("composition",    "Kompozisyon",        False),
    ("width",          "En",                 False),
    ("gramaj",         "Gramaj",             False),
    ("shrinkage",      "Çekme",              False),
    ("price",          "Fiyat",              False),
    ("supplier",       "Tedarikçi/Fason",    False),
]
PRODUCT_AUTO = {
    "product_code":   ["kumas","ürün kodu","urun kodu","product_code","kod"],
    "reference_code": ["kumas_kodu","kumaş kodu","kumas kodu","referans"],
    "product_name":   ["kumas_on_adi","kumas_adi","ürün adı","urun adi","ürün bilgisi","product_name","ad"],
    "composition":    ["kumas_bilesimi","bilesim","bileşim","kompozisyon","composition"],
    "width":          ["kumas_en","en","width"],
    "gramaj":         ["kumas_gramaj","gramaj","gsm"],
    "shrinkage":      ["kumas_cekme","cekme","çekme","shrinkage"],
    "price":          ["kumas_fiy","fiyat","price","birim_fiyat"],
    "supplier":       ["kumas_fason_id","fason","tedarikci","tedarikçi","supplier"],
}


def _excel_val_to_str(val):
    import pandas as pd
    if val is None or (hasattr(pd, "isna") and pd.isna(val)):
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def _customer_tax_no(customer_id):
    customer = db.get_customer(customer_id)
    return dict(customer).get("tax_no", "") if customer else ""


def _wire_header_persistence(table, key, default_fn=None):
    """Sütun sırası ve genişliği kalıcı: kapatıp açınca aynı kalır.
    default_fn: kayıtlı durum yoksa (ilk çalıştırma) uygulanacak varsayılan
    sütun genişlikleri; verilmezse table.resizeColumnsToContents() kullanılır."""
    if table.property("hdr_wired"):
        return
    from PyQt6.QtCore import QSettings
    hdr = table.header() if isinstance(table, QTreeWidget) else table.horizontalHeader()
    settings = QSettings("BursaKnitted", "DepoTakip")
    st = settings.value(key)
    restored = False
    if st is not None:
        restored = hdr.restoreState(st)
    if not restored:
        if default_fn:
            default_fn()
        elif isinstance(table, QTreeWidget):
            for col in range(table.columnCount()):
                table.resizeColumnToContents(col)
        else:
            table.resizeColumnsToContents()
    # sectionResized, sürükleyerek yeniden boyutlandırma sırasında (mouseMoveEvent
    # içinden) art arda tetiklenir; o anda hdr.saveState() çağırmak macOS'ta
    # QHeaderView::resizeSection içinde yeniden girilebilirlik (reentrancy) sorunu
    # yaratıp SIGSEGV'e yol açabiliyor. Kayıt işlemini QTimer ile geciktirip
    # event-loop'a döndükten sonra yapıyoruz.
    save_timer = QTimer(table)
    save_timer.setSingleShot(True)
    save_timer.setInterval(300)
    def _save_hdr():
        QSettings("BursaKnitted", "DepoTakip").setValue(key, hdr.saveState())
    save_timer.timeout.connect(_save_hdr)
    hdr.sectionMoved.connect(lambda *a: save_timer.start())
    hdr.sectionResized.connect(lambda *a: save_timer.start())
    table.setProperty("hdr_wired", True)


class ExcelColumnMapDialog(QDialog):
    """Excel dosyasından sütun eşleştirmesi yapıp kayıt listesi üretir.
    fields: [(alan_anahtarı, etiket, zorunlu_mu), ...]
    auto_keywords: {alan_anahtarı: [başlık eşleştirme anahtar kelimeleri]}
    Kabul edildikten sonra self.records doldurulur."""

    def __init__(self, path, fields, auto_keywords=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Excel Sütun Eşleştirme")
        self.setMinimumSize(760, 540)
        self.path = path
        self.fields = fields
        self.auto_keywords = auto_keywords or {}
        self.df = None
        self.records = []
        self._build_ui()
        self._load_preview()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        self.chk_header = QCheckBox("İlk satır sütun başlığı içeriyor")
        self.chk_header.setChecked(True)
        self.chk_header.toggled.connect(self._load_preview)
        lay.addWidget(self.chk_header)

        lay.addWidget(QLabel("Önizleme (ilk 5 satır):"))
        self.preview = QTableWidget()
        self.preview.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.preview.verticalHeader().setVisible(False)
        self.preview.setMaximumHeight(170)
        lay.addWidget(self.preview)

        lay.addWidget(QLabel("Hangi sütun hangi alana karşılık geliyor?"))
        form = QFormLayout(); form.setSpacing(6)
        self.combos = {}
        for key, label, required in self.fields:
            combo = QComboBox()
            form.addRow((label + " *:") if required else (label + ":"), combo)
            self.combos[key] = combo
        lay.addLayout(form)

        self.info_label = QLabel("")
        self.info_label.setStyleSheet("color:#757575;")
        lay.addWidget(self.info_label)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _load_preview(self):
        import pandas as pd
        header = 0 if self.chk_header.isChecked() else None
        try:
            df = pd.read_excel(self.path, header=header)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel okunamadı:\n{e}")
            return

        if header is None:
            df.columns = [f"Sütun {i+1}" for i in range(len(df.columns))]
        else:
            df.columns = [str(c).strip() for c in df.columns]
        self.df = df
        cols = list(df.columns)

        # Önizleme tablosu
        self.preview.setColumnCount(len(cols))
        self.preview.setHorizontalHeaderLabels([str(c) for c in cols])
        n = min(5, len(df))
        self.preview.setRowCount(n)
        for r in range(n):
            for c in range(len(cols)):
                self.preview.setItem(r, c, QTableWidgetItem(_excel_val_to_str(df.iat[r, c])))
        self.preview.resizeColumnsToContents()

        # Eşleştirme kutularını doldur
        for combo in self.combos.values():
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("— Kullanılmıyor —", "")
            for c in cols:
                combo.addItem(str(c), c)
            combo.blockSignals(False)

        # Otomatik eşleştirme: önce tam eşleşme, sonra anahtar kelime içeriği
        if header == 0:
            used = set()
            lower_cols = {c: str(c).strip().lower() for c in cols}
            for pass_exact in (True, False):
                for key, combo in self.combos.items():
                    if combo.currentData():
                        continue
                    for kw in self.auto_keywords.get(key, []):
                        match = None
                        for c, lc in lower_cols.items():
                            if c in used:
                                continue
                            if (pass_exact and lc == kw) or (not pass_exact and kw in lc):
                                match = c; break
                        if match:
                            idx = combo.findData(match)
                            if idx >= 0:
                                combo.setCurrentIndex(idx)
                                used.add(match)
                            break

        self.info_label.setText(f"Toplam {len(df)} satır bulundu.")

    def _on_accept(self):
        if self.df is None:
            return self.reject()
        mapping = {}
        for key, label, required in self.fields:
            col = self.combos[key].currentData()
            if required and not col:
                return QMessageBox.warning(self, "Hata", f"'{label}' sütunu seçilmelidir.")
            mapping[key] = col

        records = []
        for _, row in self.df.iterrows():
            rec = {key: (_excel_val_to_str(row.get(col, "")) if col else "") for key, col in mapping.items()}
            if all(rec.get(key) for key, label, required in self.fields if required):
                records.append(rec)
        self.records = records
        self.accept()


# ══ CRM modülü (Excel tablolarına göre) ═════════════════════════
CRM_PAZARLAMACILAR = ["AYKUT EKERBİÇER", "AHMET AYAT", "SEYHAN ZÜNBÜL", "İNAN ACAR", "SERTAÇ İŞLER"]
CRM_MUSTERI_TURLERI = ["MEVCUT MÜŞTERİ", "HEDEF MÜŞTERİ", "POTANSİYEL MÜŞTERİ"]
CRM_MUSTERI_TIPLERI = ["MEVCUT", "POTANSİYEL", "HEDEF"]
CRM_DOVIZLER = ["USD", "EUR", "TL", "GBP"]
# 5 Ways hedefleri (aylık) — referans için
CRM_HEDEFLER = {
    "ziyaret": "Haftada 3 / Ayda 12 (yeni + mevcut)",
    "sip_tutar_usd": "150.000 USD / ay",
    "sip_adet": "15 / ay",
    "sip_metre": "45.000 MT / ay",
    "ort_sip_usd": "10.000 USD",
    "kar_orani": "% 30",
}


def _crm_money(v):
    """Para/sayı metnini (₺, $, £, '1.266,53' vb.) float'a çevirir."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    import re as _re
    s = _re.sub(r"[^\d,.\-]", "", str(v).strip())
    if not s or s in ("-", ".", ","):
        return 0.0
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _crm_iso(v):
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    if v is None:
        return ""
    return str(v).strip()


def crm_parse_excel(path):
    """CRM Excel dosyasını (6 sayfa) okuyup toplu import için sözlük döndürür."""
    import re as _re
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"customers": [], "visits": [], "sales": [], "orders": []}

    def _s(v):
        return str(v).strip() if v is not None else ""

    if "müşteri listesi" in wb.sheetnames:
        for row in wb["müşteri listesi"].iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3 or not row[2]:
                continue
            out["customers"].append({"musteri_turu": _s(row[0]), "pazarlamaci": _s(row[1]),
                                     "firma": _s(row[2])})

    for sheet, durum in [("gerçekleşen ziyaret", "GERÇEKLEŞEN"), ("planlanan ziyaret", "PLANLANAN")]:
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if not row or not any(row) or len(row) < 4:
                continue
            if not (row[1] or row[2]):
                continue
            out["visits"].append({"durum": durum, "tarih": _crm_iso(row[0]), "pazarlamaci": _s(row[1]),
                                  "musteri": _s(row[2]), "musteri_tipi": _s(row[3]),
                                  "notlar": _s(row[4]) if len(row) > 4 else ""})

    if "FİİLİ SATIŞLAR" in wb.sheetnames:
        for row in wb["FİİLİ SATIŞLAR"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            paz = _s(row[6]) if len(row) > 6 else ""
            is_iade = 0
            up = paz.upper()
            if up.endswith("İADE") or up.endswith("IADE"):
                is_iade = 1
                paz = _re.sub(r"\s*[İI]ADE$", "", paz).strip()
            out["sales"].append({"musteri": _s(row[0]), "metre": _crm_money(row[1]),
                                 "tutar": _crm_money(row[2]), "doviz": _s(row[3]) or "USD",
                                 "usd_tutar": _crm_money(row[4]), "ay": _crm_iso(row[5]),
                                 "pazarlamaci": paz,
                                 "musteri_tipi": _s(row[7]) if len(row) > 7 else "", "is_iade": is_iade})

    if "SİPARİŞLER" in wb.sheetnames:
        for row in wb["SİPARİŞLER"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            g = lambda i: row[i] if len(row) > i else None
            out["orders"].append({"musteri": _s(g(0)), "pazarlamaci": _s(g(1)),
                                  "musteri_tipi": _s(g(2)), "tarih": _crm_iso(g(3)), "kod": _s(g(4)),
                                  "renk": _s(g(5)), "miktar": _crm_money(g(6)),
                                  "maliyet_fiyati": _crm_money(g(7)), "satis_fiyati": _crm_money(g(8)),
                                  "kar_orani": _crm_money(g(9)), "teorik_kar_usd": _crm_money(g(10)),
                                  "kar_tl": _crm_money(g(11)), "vade": _s(g(12)),
                                  "ciro": _crm_money(g(13)), "ciro_usd": _crm_money(g(14)),
                                  "usd_kuru": _crm_money(g(15))})
    return out


class _SortItem(QTableWidgetItem):
    """Sayısal sıralama için UserRole'daki değere göre karşılaştırır."""
    def __lt__(self, other):
        a = self.data(Qt.ItemDataRole.UserRole)
        b = other.data(Qt.ItemDataRole.UserRole)
        if a is not None and b is not None:
            try:
                return float(a) < float(b)
            except Exception:
                pass
        return super().__lt__(other)


class _LineChart(QWidget):
    """Bağımlılıksız çizgi grafiği — her seri (pazarlamacı) renkli bir çizgi."""
    _PALET = ["#1565C0", "#E65100", "#2E7D32", "#6A1B9A", "#00838F", "#C62828", "#5D4037"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.series = []      # [(ad, [değerler], renk)]
        self.xlabels = []
        self.title = ""
        self.kind = "money"
        self.setMinimumHeight(380)

    def set_data(self, series, xlabels, kind="money", title=""):
        out = []
        for i, (ad, vals) in enumerate(series):
            out.append((ad, vals, self._PALET[i % len(self._PALET)]))
        self.series = out
        self.xlabels = xlabels
        self.kind = kind
        self.title = title
        self.update()

    def _fmt(self, v):
        if self.kind == "pct":
            return f"{v:.0f}%"
        if abs(v) >= 1_000_000:
            return f"{v/1_000_000:.1f}M"
        if abs(v) >= 1_000:
            return f"{v/1_000:.0f}K"
        return f"{v:.0f}"

    def paintEvent(self, ev):
        from PyQt6.QtGui import QPainter, QPen, QPolygonF
        from PyQt6.QtCore import QPointF
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        p.fillRect(0, 0, W, H, QColor("#FFFFFF"))
        L, R, T, B = 64, 24, 36, 70          # kenar boşlukları
        pw, ph = W - L - R, H - T - B
        if not self.series or not self.xlabels or pw <= 10 or ph <= 10:
            p.setPen(QColor("#999")); p.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter,
                                                  "Veri yok"); return
        n = len(self.xlabels)
        vmax = 0.0
        for _, vals, _c in self.series:
            for v in vals:
                if v is not None and v > vmax:
                    vmax = v
        if vmax <= 0:
            vmax = 1.0
        # üst sınırı yuvarla
        import math
        mag = 10 ** int(math.floor(math.log10(vmax)))
        vmax = math.ceil(vmax / mag) * mag

        # başlık
        if self.title:
            f = p.font(); f.setBold(True); f.setPointSize(11); p.setFont(f)
            p.setPen(QColor("#222")); p.drawText(L, 6, pw, 22,
                Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, self.title)
            f.setBold(False); f.setPointSize(9); p.setFont(f)

        # ızgara + y ekseni
        p.setPen(QColor("#888"))
        steps = 4
        for s in range(steps + 1):
            yv = vmax * s / steps
            y = T + ph - (yv / vmax) * ph
            p.setPen(QPen(QColor("#ECECEC"), 1)); p.drawLine(int(L), int(y), int(L + pw), int(y))
            p.setPen(QColor("#777")); p.drawText(0, int(y) - 8, L - 6, 16,
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter, self._fmt(yv))

        # x ekseni etiketleri
        xs = [L + (pw * i / (n - 1) if n > 1 else pw / 2) for i in range(n)]
        p.setPen(QColor("#555"))
        for i, lbl in enumerate(self.xlabels):
            p.drawText(int(xs[i]) - 24, T + ph + 6, 48, 16,
                       Qt.AlignmentFlag.AlignCenter, lbl)

        # seriler
        for ad, vals, col in self.series:
            pen = QPen(QColor(col)); pen.setWidth(2); p.setPen(pen)
            poly = QPolygonF()
            for i, v in enumerate(vals):
                if v is None:
                    continue
                y = T + ph - (v / vmax) * ph
                poly.append(QPointF(xs[i], y))
            if poly.count() > 1:
                p.drawPolyline(poly)
            p.setBrush(QBrush(QColor(col)))
            for i, v in enumerate(vals):
                if v is None:
                    continue
                y = T + ph - (v / vmax) * ph
                p.drawEllipse(QPointF(xs[i], y), 3, 3)

        # gösterge (legend)
        lx, ly = L, T + ph + 28
        f = p.font(); f.setPointSize(9); p.setFont(f)
        for ad, vals, col in self.series:
            p.setBrush(QBrush(QColor(col))); p.setPen(QColor(col))
            p.drawEllipse(QPointF(lx + 6, ly + 7), 4, 4)
            p.setPen(QColor("#333"))
            tw = p.fontMetrics().horizontalAdvance(ad) + 22
            p.drawText(lx + 16, ly, tw, 16, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, ad)
            lx += 16 + tw + 8
            if lx > L + pw - 80:
                lx = L; ly += 18
        p.end()


class CRMView(QWidget):
    """CRM — Müşteri Listesi, Ziyaretler, Fiili Satışlar, Siparişler, Analiz (5 Ways)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 6)

        # Üst araç çubuğu — Excel içe aktarma
        top = QHBoxLayout()
        title = QLabel("👥 CRM — Müşteri İlişkileri Yönetimi")
        title.setStyleSheet("font-size:15px; font-weight:bold;")
        top.addWidget(title)
        top.addStretch()
        btn_imp = QPushButton("📥 Excel'den İçe Aktar")
        btn_imp.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;border-radius:4px;padding:6px 14px;")
        btn_imp.clicked.connect(self._import_excel)
        top.addWidget(btn_imp)
        lay.addLayout(top)

        self.tabs = QTabWidget()
        self.tabs.addTab(self._build_customers_tab(), "🏢 Müşteri Listesi")
        self.tabs.addTab(self._build_visits_tab(), "📍 Ziyaretler")
        self.tabs.addTab(self._build_sales_tab(), "💵 Fiili Satışlar")
        self.tabs.addTab(self._build_orders_tab(), "🧾 Siparişler")
        self.tabs.addTab(self._build_analysis_tab(), "📊 Analiz (5 Ways)")
        self.tabs.addTab(self._build_chart_tab(), "📈 Grafik")
        self.tabs.currentChanged.connect(lambda i: self.refresh())
        lay.addWidget(self.tabs)

    def refresh(self):
        idx = self.tabs.currentIndex()
        try:
            [self._load_customers, self._load_visits, self._load_sales,
             self._load_orders, self._load_analysis, self._load_chart][idx]()
        except Exception:
            pass

    def showEvent(self, e):
        super().showEvent(e)
        self.refresh()

    # ── ortak yardımcılar ─────────────────────────────────────────
    def _mk_table(self, headers):
        t = QTableWidget()
        t.setColumnCount(len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.verticalHeader().setVisible(False)
        t.setSortingEnabled(True)
        t.horizontalHeader().setSectionsMovable(True)
        return t

    def _paz_combo(self, current="", with_all=False):
        cb = QComboBox()
        cb.setEditable(True)
        if with_all:
            cb.addItem("— Tümü —", "")
        for p in CRM_PAZARLAMACILAR:
            cb.addItem(p, p)
        if current and cb.findData(current) < 0:
            cb.addItem(current, current)
        if current:
            cb.setCurrentText(current)
        elif not with_all:
            cb.setCurrentIndex(-1)
        return cb

    # ── Sekme 1: Müşteri Listesi ──────────────────────────────────
    def _build_customers_tab(self):
        w = QWidget(); v = QVBoxLayout(w)
        top = QHBoxLayout()
        self.cu_search = QLineEdit(); self.cu_search.setPlaceholderText("Firma / pazarlamacı ara...")
        self.cu_search.textChanged.connect(self._load_customers)
        self.cu_paz = self._paz_combo(with_all=True)
        self.cu_paz.setEditable(False)
        self.cu_paz.currentIndexChanged.connect(self._load_customers)
        b_add = QPushButton("+ Yeni"); b_add.clicked.connect(self._add_customer)
        b_edit = QPushButton("✎ Düzenle"); b_edit.clicked.connect(self._edit_customer)
        b_del = QPushButton("✕ Sil"); b_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 12px;")
        b_del.clicked.connect(self._delete_customer)
        top.addWidget(QLabel("Ara:")); top.addWidget(self.cu_search)
        top.addWidget(QLabel("Pazarlamacı:")); top.addWidget(self.cu_paz); top.addStretch()
        for b in (b_add, b_edit, b_del): top.addWidget(b)
        v.addLayout(top)
        self.cu_table = self._mk_table(["Müşteri Türü", "Pazarlamacı", "Firma"])
        self.cu_table.horizontalHeader().setStretchLastSection(True)
        self.cu_table.doubleClicked.connect(self._edit_customer)
        v.addWidget(self.cu_table)
        self.cu_count = QLabel(); v.addWidget(self.cu_count)
        return w

    def _load_customers(self):
        rows = db.get_crm_customers(search=self.cu_search.text().strip(),
                                    pazarlamaci=self.cu_paz.currentData() or "") or []
        t = self.cu_table; t.setSortingEnabled(False); t.setRowCount(len(rows))
        for i, r in enumerate(rows):
            it = QTableWidgetItem(r["musteri_turu"] or "")
            it.setData(Qt.ItemDataRole.UserRole, r["id"])
            t.setItem(i, 0, it)
            t.setItem(i, 1, QTableWidgetItem(r["pazarlamaci"] or ""))
            t.setItem(i, 2, QTableWidgetItem(r["firma"] or ""))
        t.setSortingEnabled(True)
        self.cu_count.setText(f"{len(rows)} müşteri")

    def _customer_dialog(self, c=None):
        dlg = QDialog(self); dlg.setWindowTitle("Müşteri" + (" Düzenle" if c else " Ekle"))
        dlg.setMinimumWidth(380); lay = QVBoxLayout(dlg); form = QFormLayout()
        dlg.turu = QComboBox(); dlg.turu.addItems(CRM_MUSTERI_TURLERI)
        if c: dlg.turu.setCurrentText(c["musteri_turu"] or CRM_MUSTERI_TURLERI[0])
        dlg.paz = self._paz_combo(c["pazarlamaci"] if c else "")
        dlg.firma = QLineEdit(c["firma"] if c else "")
        form.addRow("Müşteri Türü:", dlg.turu)
        form.addRow("Pazarlamacı:", dlg.paz)
        form.addRow("Firma *:", dlg.firma)
        lay.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject); lay.addWidget(bb)
        return dlg

    def _add_customer(self):
        d = self._customer_dialog()
        if d.exec():
            if not d.firma.text().strip():
                return QMessageBox.warning(self, "Hata", "Firma zorunlu!")
            db.add_crm_customer(d.turu.currentText(), d.paz.currentText().strip(), d.firma.text())
            self._load_customers()

    def _edit_customer(self):
        cid = self._row_id(self.cu_table)
        if not cid: return
        c = self._find(db.get_crm_customers(), cid)
        if not c: return
        d = self._customer_dialog(c)
        if d.exec():
            db.update_crm_customer(cid, d.turu.currentText(), d.paz.currentText().strip(), d.firma.text())
            self._load_customers()

    def _delete_customer(self):
        cid = self._row_id(self.cu_table)
        if not cid: return
        if self._confirm_del():
            db.delete_crm_customer(cid); self._load_customers()

    # ── Sekme 2: Ziyaretler ───────────────────────────────────────
    def _build_visits_tab(self):
        w = QWidget(); v = QVBoxLayout(w)
        top = QHBoxLayout()
        self.vi_search = QLineEdit(); self.vi_search.setPlaceholderText("Müşteri / pazarlamacı ara...")
        self.vi_search.textChanged.connect(self._load_visits)
        self.vi_durum = QComboBox()
        self.vi_durum.addItem("— Tümü —", ""); self.vi_durum.addItem("Gerçekleşen", "GERÇEKLEŞEN")
        self.vi_durum.addItem("Planlanan", "PLANLANAN")
        self.vi_durum.currentIndexChanged.connect(self._load_visits)
        self.vi_paz = QComboBox(); self.vi_paz.addItem("— Tümü —", "")
        for p in CRM_PAZARLAMACILAR: self.vi_paz.addItem(p, p)
        self.vi_paz.currentIndexChanged.connect(self._load_visits)
        self.vi_group = QComboBox()
        self.vi_group.addItem("Detay (gruplama yok)", "")
        self.vi_group.addItem("Aya göre", "ay")
        self.vi_group.addItem("Pazarlamacıya göre", "pazarlamaci")
        self.vi_group.addItem("Müşteriye göre", "musteri")
        self.vi_group.addItem("Müşteri Tipine göre", "musteri_tipi")
        self.vi_group.currentIndexChanged.connect(self._load_visits)
        b_add = QPushButton("+ Yeni"); b_add.clicked.connect(self._add_visit)
        b_edit = QPushButton("✎ Düzenle"); b_edit.clicked.connect(self._edit_visit)
        b_del = QPushButton("✕ Sil"); b_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 12px;")
        b_del.clicked.connect(self._delete_visit)
        top.addWidget(QLabel("Ara:")); top.addWidget(self.vi_search)
        top.addWidget(QLabel("Durum:")); top.addWidget(self.vi_durum)
        top.addWidget(QLabel("Pazarlamacı:")); top.addWidget(self.vi_paz)
        top.addWidget(QLabel("Grupla:")); top.addWidget(self.vi_group); top.addStretch()
        for b in (b_add, b_edit, b_del): top.addWidget(b)
        v.addLayout(top)
        self.vi_table = self._mk_table(["Durum", "Tarih", "Pazarlamacı", "Müşteri", "Müşteri Tipi", "Notlar"])
        self.vi_table.setSortingEnabled(False)   # alt toplam satırı sabit kalsın
        self.vi_table.horizontalHeader().setStretchLastSection(True)
        self.vi_table.doubleClicked.connect(self._edit_visit)
        v.addWidget(self.vi_table)
        self.vi_count = QLabel(); v.addWidget(self.vi_count)
        return w

    def _load_visits(self):
        rows = db.get_crm_visits(search=self.vi_search.text().strip(),
                                 durum=self.vi_durum.currentData() or "",
                                 pazarlamaci=self.vi_paz.currentData() or "") or []
        group = self.vi_group.currentData() or ""
        if group:
            self._load_visits_grouped(rows, group)
            return
        # ── Detay görünüm + alt toplam ────────────────────────────
        t = self.vi_table
        t.setColumnCount(6)
        t.setHorizontalHeaderLabels(["Durum", "Tarih", "Pazarlamacı", "Müşteri", "Müşteri Tipi", "Notlar"])
        t.setRowCount(len(rows) + (1 if rows else 0))
        musteriler = set()
        for i, r in enumerate(rows):
            musteriler.add((r["musteri"] or "").strip().upper())
            d_item = QTableWidgetItem("📍 Gerçekleşen" if r["durum"] == "GERÇEKLEŞEN" else "🗓 Planlanan")
            d_item.setData(Qt.ItemDataRole.UserRole, r["id"])
            d_item.setForeground(QBrush(QColor("#2E7D32" if r["durum"] == "GERÇEKLEŞEN" else "#1565C0")))
            t.setItem(i, 0, d_item)
            t.setItem(i, 1, QTableWidgetItem(r["tarih"] or ""))
            t.setItem(i, 2, QTableWidgetItem(r["pazarlamaci"] or ""))
            t.setItem(i, 3, QTableWidgetItem(r["musteri"] or ""))
            t.setItem(i, 4, QTableWidgetItem(r["musteri_tipi"] or ""))
            t.setItem(i, 5, QTableWidgetItem(r["notlar"] or ""))
        if rows:
            self._put_total_row(t, len(rows),
                [f"TOPLAM ({len(rows)} ziyaret)", "", "", f"{len(musteriler)} tekil müşteri", "", ""])
        self.vi_count.setText(f"{len(rows)} ziyaret")

    def _load_visits_grouped(self, rows, group):
        labels = {"ay": "Ay", "pazarlamaci": "Pazarlamacı", "musteri": "Müşteri",
                  "musteri_tipi": "Müşteri Tipi"}
        t = self.vi_table
        t.setColumnCount(4)
        t.setHorizontalHeaderLabels([labels[group], "Ziyaret Adedi", "Tekil Müşteri", "Gerçekleşen / Planlanan"])
        groups = {}
        for r in rows:
            key = (r["tarih"] or "")[:7] if group == "ay" else (r[group] or "—")
            g = groups.setdefault(key, {"adet": 0, "musteriler": set(), "gerc": 0, "plan": 0})
            g["adet"] += 1
            g["musteriler"].add((r["musteri"] or "").strip().upper())
            if r["durum"] == "GERÇEKLEŞEN": g["gerc"] += 1
            else: g["plan"] += 1
        keys = sorted(groups.keys())
        tot_adet = 0; tot_m = set(); tot_g = 0; tot_p = 0
        t.setRowCount(len(keys) + (1 if keys else 0))
        for i, key in enumerate(keys):
            g = groups[key]
            tot_adet += g["adet"]; tot_m |= g["musteriler"]; tot_g += g["gerc"]; tot_p += g["plan"]
            t.setItem(i, 0, QTableWidgetItem(key or "—"))
            t.setItem(i, 1, self._num(g["adet"]))
            t.setItem(i, 2, self._num(len(g["musteriler"])))
            t.setItem(i, 3, QTableWidgetItem(f"{g['gerc']} / {g['plan']}"))
        if keys:
            self._put_total_row(t, len(keys),
                ["TOPLAM", self._num(tot_adet), self._num(len(tot_m)), f"{tot_g} / {tot_p}"])
        t.resizeColumnsToContents()
        self.vi_count.setText(f"{len(keys)} grup • {len(rows)} ziyaret")

    def _visit_dialog(self, vrow=None):
        dlg = QDialog(self); dlg.setWindowTitle("Ziyaret" + (" Düzenle" if vrow else " Ekle"))
        dlg.setMinimumWidth(400); lay = QVBoxLayout(dlg); form = QFormLayout()
        dlg.durum = QComboBox(); dlg.durum.addItem("Gerçekleşen", "GERÇEKLEŞEN"); dlg.durum.addItem("Planlanan", "PLANLANAN")
        if vrow:
            idx = dlg.durum.findData(vrow["durum"]); dlg.durum.setCurrentIndex(idx if idx >= 0 else 0)
        dlg.tarih = QDateEdit(); dlg.tarih.setCalendarPopup(True); dlg.tarih.setDisplayFormat("dd.MM.yyyy")
        dlg.tarih.setDate(QDate.fromString(vrow["tarih"], "yyyy-MM-dd") if vrow and vrow["tarih"] else QDate.currentDate())
        dlg.paz = self._paz_combo(vrow["pazarlamaci"] if vrow else "")
        dlg.musteri = QLineEdit(vrow["musteri"] if vrow else "")
        dlg.tipi = QComboBox(); dlg.tipi.setEditable(True); dlg.tipi.addItems(CRM_MUSTERI_TIPLERI)
        if vrow: dlg.tipi.setCurrentText(vrow["musteri_tipi"] or "")
        else: dlg.tipi.setCurrentIndex(-1)
        dlg.notlar = QLineEdit(vrow["notlar"] if vrow else "")
        form.addRow("Durum:", dlg.durum)
        form.addRow("Tarih:", dlg.tarih)
        form.addRow("Pazarlamacı:", dlg.paz)
        form.addRow("Müşteri:", dlg.musteri)
        form.addRow("Müşteri Tipi:", dlg.tipi)
        form.addRow("Notlar:", dlg.notlar)
        lay.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject); lay.addWidget(bb)
        return dlg

    def _add_visit(self):
        d = self._visit_dialog()
        if d.exec():
            db.add_crm_visit(d.durum.currentData(), d.tarih.date().toString("yyyy-MM-dd"),
                             d.paz.currentText().strip(), d.musteri.text(), d.tipi.currentText().strip(),
                             d.notlar.text())
            self._load_visits()

    def _edit_visit(self):
        vid = self._row_id(self.vi_table)
        if not vid: return
        vrow = self._find(db.get_crm_visits(), vid)
        if not vrow: return
        d = self._visit_dialog(vrow)
        if d.exec():
            db.update_crm_visit(vid, d.durum.currentData(), d.tarih.date().toString("yyyy-MM-dd"),
                                d.paz.currentText().strip(), d.musteri.text(),
                                d.tipi.currentText().strip(), d.notlar.text())
            self._load_visits()

    def _delete_visit(self):
        vid = self._row_id(self.vi_table)
        if not vid: return
        if self._confirm_del():
            db.delete_crm_visit(vid); self._load_visits()

    # ── Sekme 3: Fiili Satışlar ───────────────────────────────────
    def _build_sales_tab(self):
        w = QWidget(); v = QVBoxLayout(w)
        top = QHBoxLayout()
        self.sa_search = QLineEdit(); self.sa_search.setPlaceholderText("Müşteri / pazarlamacı ara...")
        self.sa_search.textChanged.connect(self._load_sales)
        self.sa_year = QComboBox(); self.sa_year.currentIndexChanged.connect(self._load_sales)
        self.sa_paz = QComboBox(); self.sa_paz.addItem("— Tümü —", "")
        for p in CRM_PAZARLAMACILAR: self.sa_paz.addItem(p, p)
        self.sa_paz.currentIndexChanged.connect(self._load_sales)
        self.sa_group = QComboBox()
        self.sa_group.addItem("Detay (gruplama yok)", "")
        self.sa_group.addItem("Aya göre", "ay")
        self.sa_group.addItem("Yıla göre", "yil")
        self.sa_group.addItem("Pazarlamacıya göre", "pazarlamaci")
        self.sa_group.addItem("Müşteriye göre", "musteri")
        self.sa_group.currentIndexChanged.connect(self._load_sales)
        b_add = QPushButton("+ Yeni"); b_add.clicked.connect(self._add_sale)
        b_edit = QPushButton("✎ Düzenle"); b_edit.clicked.connect(self._edit_sale)
        b_del = QPushButton("✕ Sil"); b_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 12px;")
        b_del.clicked.connect(self._delete_sale)
        top.addWidget(QLabel("Ara:")); top.addWidget(self.sa_search)
        top.addWidget(QLabel("Yıl:")); top.addWidget(self.sa_year)
        top.addWidget(QLabel("Pazarlamacı:")); top.addWidget(self.sa_paz)
        top.addWidget(QLabel("Grupla:")); top.addWidget(self.sa_group); top.addStretch()
        for b in (b_add, b_edit, b_del): top.addWidget(b)
        v.addLayout(top)
        self.sa_table = self._mk_table(["Müşteri", "Metre", "Tutar", "Döviz", "USD Tutar", "Ay", "Pazarlamacı", "Müşteri Tipi", "İade"])
        self.sa_table.setSortingEnabled(False)   # alt toplam satırı sabit kalsın
        self.sa_table.doubleClicked.connect(self._edit_sale)
        v.addWidget(self.sa_table)
        self.sa_count = QLabel(); v.addWidget(self.sa_count)
        return w

    def _refresh_year_combo(self, combo):
        cur = combo.currentData()
        combo.blockSignals(True); combo.clear()
        combo.addItem("— Tümü —", "")
        for y in (db.get_crm_years() or []):
            combo.addItem(y, y)
        if cur:
            i = combo.findData(cur)
            if i >= 0: combo.setCurrentIndex(i)
        combo.blockSignals(False)

    @staticmethod
    def _sale_net(r):
        """İade ise metre/tutar/usd net olarak (negatif) döner."""
        metre, tutar, usd = r["metre"] or 0, r["tutar"] or 0, r["usd_tutar"] or 0
        if r["is_iade"]:
            return -abs(metre), -abs(tutar), -abs(usd)
        return metre, tutar, usd

    def _load_sales(self):
        if self.sa_year.count() == 0:
            self._refresh_year_combo(self.sa_year)
        rows = db.get_crm_sales(search=self.sa_search.text().strip(),
                                pazarlamaci=self.sa_paz.currentData() or "",
                                year=self.sa_year.currentData() or "") or []
        group = self.sa_group.currentData() or ""
        if group:
            self._load_sales_grouped(rows, group)
            return
        # ── Detay görünüm + net alt toplam ────────────────────────
        t = self.sa_table
        t.setColumnCount(9)
        t.setHorizontalHeaderLabels(["Müşteri", "Metre", "Tutar", "Döviz", "USD Tutar", "Ay",
                                     "Pazarlamacı", "Müşteri Tipi", "İade"])
        sum_metre = sum_usd = 0.0
        t.setRowCount(len(rows) + (1 if rows else 0))
        for i, r in enumerate(rows):
            mn, tn, un = self._sale_net(r)
            sum_metre += mn; sum_usd += un
            it = QTableWidgetItem(r["musteri"] or "")
            it.setData(Qt.ItemDataRole.UserRole, r["id"])
            t.setItem(i, 0, it)
            t.setItem(i, 1, self._num(r["metre"]))
            t.setItem(i, 2, self._num(r["tutar"]))
            t.setItem(i, 3, QTableWidgetItem(r["doviz"] or ""))
            t.setItem(i, 4, self._num(r["usd_tutar"]))
            t.setItem(i, 5, QTableWidgetItem(r["ay"] or ""))
            t.setItem(i, 6, QTableWidgetItem(r["pazarlamaci"] or ""))
            t.setItem(i, 7, QTableWidgetItem(r["musteri_tipi"] or ""))
            iade = QTableWidgetItem("↩ İade" if r["is_iade"] else "")
            if r["is_iade"]: iade.setForeground(QBrush(QColor("#C62828")))
            t.setItem(i, 8, iade)
        if rows:
            cells = [f"TOPLAM ({len(rows)} kayıt, net)", self._num(sum_metre), "", "",
                     self._num(sum_usd), "", "", "", ""]
            self._put_total_row(t, len(rows), cells)
        self.sa_count.setText(f"{len(rows)} satış kaydı")

    def _load_sales_grouped(self, rows, group):
        labels = {"ay": "Ay", "yil": "Yıl", "pazarlamaci": "Pazarlamacı", "musteri": "Müşteri"}
        t = self.sa_table
        t.setColumnCount(4)
        t.setHorizontalHeaderLabels([labels[group], "Kayıt Adedi", "Metre (net)", "USD Tutar (net)"])
        groups = {}
        for r in rows:
            if group == "ay":
                key = (r["ay"] or "")[:7]
            elif group == "yil":
                key = (r["ay"] or "")[:4]
            else:
                key = (r[group] or "—")
            g = groups.setdefault(key, {"adet": 0, "metre": 0.0, "usd": 0.0})
            mn, tn, un = self._sale_net(r)
            g["adet"] += 1; g["metre"] += mn; g["usd"] += un
        keys = sorted(groups.keys())
        tot = {"adet": 0, "metre": 0.0, "usd": 0.0}
        t.setRowCount(len(keys) + (1 if keys else 0))
        for i, key in enumerate(keys):
            g = groups[key]
            for k in tot: tot[k] += g[k]
            t.setItem(i, 0, QTableWidgetItem(key or "—"))
            t.setItem(i, 1, self._num(g["adet"]))
            t.setItem(i, 2, self._num(g["metre"]))
            t.setItem(i, 3, self._num(g["usd"]))
        if keys:
            self._put_total_row(t, len(keys),
                ["TOPLAM", self._num(tot["adet"]), self._num(tot["metre"]), self._num(tot["usd"])])
        t.resizeColumnsToContents()
        self.sa_count.setText(f"{len(keys)} grup • {len(rows)} satış kaydı")

    def _sale_dialog(self, s=None):
        dlg = QDialog(self); dlg.setWindowTitle("Fiili Satış" + (" Düzenle" if s else " Ekle"))
        dlg.setMinimumWidth(400); lay = QVBoxLayout(dlg); form = QFormLayout()
        dlg.musteri = QLineEdit(s["musteri"] if s else "")
        dlg.metre = QDoubleSpinBox(); dlg.metre.setRange(-9999999, 9999999); dlg.metre.setDecimals(2)
        dlg.metre.setValue(float(s["metre"]) if s else 0)
        dlg.tutar = QDoubleSpinBox(); dlg.tutar.setRange(-99999999, 99999999); dlg.tutar.setDecimals(2)
        dlg.tutar.setValue(float(s["tutar"]) if s else 0)
        dlg.doviz = QComboBox(); dlg.doviz.addItems(CRM_DOVIZLER)
        if s: dlg.doviz.setCurrentText(s["doviz"] or "USD")
        dlg.usd = QDoubleSpinBox(); dlg.usd.setRange(-99999999, 99999999); dlg.usd.setDecimals(2)
        dlg.usd.setValue(float(s["usd_tutar"]) if s else 0)
        dlg.ay = QDateEdit(); dlg.ay.setCalendarPopup(True); dlg.ay.setDisplayFormat("MM.yyyy")
        dlg.ay.setDate(QDate.fromString(s["ay"], "yyyy-MM-dd") if s and s["ay"] else QDate.currentDate())
        dlg.paz = self._paz_combo(s["pazarlamaci"] if s else "")
        dlg.tipi = QComboBox(); dlg.tipi.setEditable(True); dlg.tipi.addItems(CRM_MUSTERI_TIPLERI)
        if s: dlg.tipi.setCurrentText(s["musteri_tipi"] or "")
        else: dlg.tipi.setCurrentIndex(-1)
        dlg.iade = QCheckBox("İade kaydı (analizde düşülür)")
        if s and s["is_iade"]: dlg.iade.setChecked(True)
        form.addRow("Müşteri:", dlg.musteri)
        form.addRow("Metre:", dlg.metre)
        form.addRow("Tutar:", dlg.tutar)
        form.addRow("Döviz:", dlg.doviz)
        form.addRow("USD Tutar:", dlg.usd)
        form.addRow("Ay:", dlg.ay)
        form.addRow("Pazarlamacı:", dlg.paz)
        form.addRow("Müşteri Tipi:", dlg.tipi)
        form.addRow("", dlg.iade)
        lay.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject); lay.addWidget(bb)
        return dlg

    def _add_sale(self):
        d = self._sale_dialog()
        if d.exec():
            db.add_crm_sale(d.musteri.text(), d.metre.value(), d.tutar.value(), d.doviz.currentText(),
                            d.usd.value(), d.ay.date().toString("yyyy-MM-01"), d.paz.currentText().strip(),
                            d.tipi.currentText().strip(), 1 if d.iade.isChecked() else 0)
            self._refresh_year_combo(self.sa_year); self._load_sales()

    def _edit_sale(self):
        sid = self._row_id(self.sa_table)
        if not sid: return
        s = self._find(db.get_crm_sales(), sid)
        if not s: return
        d = self._sale_dialog(s)
        if d.exec():
            db.update_crm_sale(sid, d.musteri.text(), d.metre.value(), d.tutar.value(), d.doviz.currentText(),
                               d.usd.value(), d.ay.date().toString("yyyy-MM-01"), d.paz.currentText().strip(),
                               d.tipi.currentText().strip(), 1 if d.iade.isChecked() else 0)
            self._load_sales()

    def _delete_sale(self):
        sid = self._row_id(self.sa_table)
        if not sid: return
        if self._confirm_del():
            db.delete_crm_sale(sid); self._load_sales()

    # ── Sekme 4: Siparişler ───────────────────────────────────────
    def _build_orders_tab(self):
        w = QWidget(); v = QVBoxLayout(w)
        top = QHBoxLayout()
        self.or_search = QLineEdit(); self.or_search.setPlaceholderText("Müşteri / kod / pazarlamacı ara...")
        self.or_search.textChanged.connect(self._load_orders)
        self.or_year = QComboBox(); self.or_year.currentIndexChanged.connect(self._load_orders)
        self.or_group = QComboBox()
        self.or_group.addItem("Detay (gruplama yok)", "")
        self.or_group.addItem("Aya göre", "ay")
        self.or_group.addItem("Müşteriye göre", "musteri")
        self.or_group.addItem("Pazarlamacıya göre", "pazarlamaci")
        self.or_group.currentIndexChanged.connect(self._load_orders)
        b_add = QPushButton("+ Yeni"); b_add.clicked.connect(self._add_order)
        b_edit = QPushButton("✎ Düzenle"); b_edit.clicked.connect(self._edit_order)
        b_del = QPushButton("✕ Sil"); b_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 12px;")
        b_del.clicked.connect(self._delete_order)
        top.addWidget(QLabel("Ara:")); top.addWidget(self.or_search)
        top.addWidget(QLabel("Yıl:")); top.addWidget(self.or_year)
        top.addWidget(QLabel("Grupla:")); top.addWidget(self.or_group); top.addStretch()
        for b in (b_add, b_edit, b_del): top.addWidget(b)
        v.addLayout(top)
        self.or_table = self._mk_table(["Müşteri", "Pazarlamacı", "Tarih", "Kod", "Renk", "Miktar",
                                        "Maliyet", "Satış", "Kâr %", "Teorik Kâr USD", "Vade", "Ciro", "Ciro USD", "Kur"])
        self.or_table.setSortingEnabled(False)   # alt toplam satırı sabit kalsın
        self.or_table.doubleClicked.connect(self._edit_order)
        v.addWidget(self.or_table)
        self.or_count = QLabel(); v.addWidget(self.or_count)
        return w

    @staticmethod
    def _order_agg_init():
        return {"adet": 0, "miktar": 0.0, "mal_x_q": 0.0, "sat_x_q": 0.0,
                "teorik": 0.0, "kar_tl": 0.0, "ciro": 0.0, "ciro_usd": 0.0}

    @staticmethod
    def _order_agg_add(a, r):
        q = r["miktar"] or 0
        a["adet"] += 1
        a["miktar"] += q
        a["mal_x_q"] += (r["maliyet_fiyati"] or 0) * q
        a["sat_x_q"] += (r["satis_fiyati"] or 0) * q
        a["teorik"] += r["teorik_kar_usd"] or 0
        a["kar_tl"] += r["kar_tl"] or 0
        a["ciro"] += r["ciro"] or 0
        a["ciro_usd"] += r["ciro_usd"] or 0

    @staticmethod
    def _order_agg_derive(a):
        q = a["miktar"] or 0
        cu = a["ciro_usd"] or 0
        return {
            "maliyet": (a["mal_x_q"] / q) if q else 0.0,      # miktar ağırlıklı
            "satis": (a["sat_x_q"] / q) if q else 0.0,        # miktar ağırlıklı
            "kar_orani": (a["teorik"] / cu) if cu else 0.0,   # ciro (USD) ağırlıklı
        }

    def _load_orders(self):
        if self.or_year.count() == 0:
            self._refresh_year_combo(self.or_year)
        rows = db.get_crm_orders(search=self.or_search.text().strip(),
                                 year=self.or_year.currentData() or "") or []
        group = self.or_group.currentData() or ""
        if group:
            self._load_orders_grouped(rows, group)
            return
        # ── Detay görünüm + alt toplam ────────────────────────────
        t = self.or_table
        t.setColumnCount(15)
        t.setHorizontalHeaderLabels(["Müşteri", "Pazarlamacı", "Tarih", "Kod", "Renk", "Miktar",
                                     "Maliyet", "Satış", "Kâr %", "Teorik Kâr USD", "Kâr TL",
                                     "Vade", "Ciro", "Ciro USD", "Kur"])
        agg = self._order_agg_init()
        t.setRowCount(len(rows) + (1 if rows else 0))
        for i, r in enumerate(rows):
            self._order_agg_add(agg, r)
            it = QTableWidgetItem(r["musteri"] or "")
            it.setData(Qt.ItemDataRole.UserRole, r["id"])
            t.setItem(i, 0, it)
            t.setItem(i, 1, QTableWidgetItem(r["pazarlamaci"] or ""))
            t.setItem(i, 2, QTableWidgetItem(r["tarih"] or ""))
            t.setItem(i, 3, QTableWidgetItem(r["kod"] or ""))
            t.setItem(i, 4, QTableWidgetItem(r["renk"] or ""))
            t.setItem(i, 5, self._num(r["miktar"]))
            t.setItem(i, 6, self._num(r["maliyet_fiyati"]))
            t.setItem(i, 7, self._num(r["satis_fiyati"]))
            t.setItem(i, 8, self._num((r["kar_orani"] or 0) * 100, suffix="%"))
            t.setItem(i, 9, self._num(r["teorik_kar_usd"]))
            t.setItem(i, 10, self._num(r["kar_tl"]))
            t.setItem(i, 11, QTableWidgetItem(r["vade"] or ""))
            t.setItem(i, 12, self._num(r["ciro"]))
            t.setItem(i, 13, self._num(r["ciro_usd"]))
            t.setItem(i, 14, self._num(r["usd_kuru"]))
        if rows:
            d = self._order_agg_derive(agg)
            cells = ["TOPLAM ({} sipariş)".format(agg["adet"]), "", "", "", "",
                     self._num(agg["miktar"]), self._num(d["maliyet"]), self._num(d["satis"]),
                     self._num(d["kar_orani"] * 100, suffix="%"), self._num(agg["teorik"]),
                     self._num(agg["kar_tl"]), "",
                     self._num(agg["ciro"]), self._num(agg["ciro_usd"]), ""]
            self._put_total_row(t, len(rows), cells)
        self.or_count.setText(f"{len(rows)} sipariş")

    def _load_orders_grouped(self, rows, group):
        labels = {"ay": "Ay", "musteri": "Müşteri", "pazarlamaci": "Pazarlamacı"}
        t = self.or_table
        t.setColumnCount(9)
        t.setHorizontalHeaderLabels([labels[group], "Sipariş Adedi", "Miktar", "Ağ. Maliyet",
                                     "Ağ. Satış", "Ağ. Kâr %", "Teorik Kâr USD", "Kâr TL", "Ciro USD"])
        groups = {}
        for r in rows:
            if group == "ay":
                key = (r["tarih"] or "")[:7]      # YYYY-MM
            else:
                key = (r[group] or "—")
            groups.setdefault(key, self._order_agg_init())
            self._order_agg_add(groups[key], r)
        keys = sorted(groups.keys())
        grand = self._order_agg_init()
        t.setRowCount(len(keys) + (1 if keys else 0))
        for i, key in enumerate(keys):
            a = groups[key]; d = self._order_agg_derive(a)
            for k in grand: grand[k] += a[k]
            t.setItem(i, 0, QTableWidgetItem(key or "—"))
            t.setItem(i, 1, self._num(a["adet"]))
            t.setItem(i, 2, self._num(a["miktar"]))
            t.setItem(i, 3, self._num(d["maliyet"]))
            t.setItem(i, 4, self._num(d["satis"]))
            t.setItem(i, 5, self._num(d["kar_orani"] * 100, suffix="%"))
            t.setItem(i, 6, self._num(a["teorik"]))
            t.setItem(i, 7, self._num(a["kar_tl"]))
            t.setItem(i, 8, self._num(a["ciro_usd"]))
        if keys:
            gd = self._order_agg_derive(grand)
            cells = ["TOPLAM", self._num(grand["adet"]), self._num(grand["miktar"]),
                     self._num(gd["maliyet"]), self._num(gd["satis"]),
                     self._num(gd["kar_orani"] * 100, suffix="%"), self._num(grand["teorik"]),
                     self._num(grand["kar_tl"]), self._num(grand["ciro_usd"])]
            self._put_total_row(t, len(keys), cells)
        t.resizeColumnsToContents()
        self.or_count.setText(f"{len(keys)} grup • {len(rows)} sipariş")

    def _put_total_row(self, t, row, cells):
        """Alt toplam satırını koyu + gri zeminle yazar. cells: str veya QTableWidgetItem."""
        f = QFont(); f.setBold(True)
        for c, val in enumerate(cells):
            it = val if isinstance(val, QTableWidgetItem) else QTableWidgetItem(str(val))
            it.setFont(f)
            it.setBackground(QBrush(QColor("#ECEFF1")))
            if c == 0:
                it.setForeground(QBrush(QColor("#1A237E")))
            t.setItem(row, c, it)

    def _order_dialog(self, o=None):
        dlg = QDialog(self); dlg.setWindowTitle("Sipariş" + (" Düzenle" if o else " Ekle"))
        dlg.setMinimumWidth(440); lay = QVBoxLayout(dlg); form = QFormLayout()
        def sb(rng, dec, val):
            s = QDoubleSpinBox(); s.setRange(-rng, rng); s.setDecimals(dec); s.setValue(float(val) if val else 0); return s
        dlg.musteri = QLineEdit(o["musteri"] if o else "")
        dlg.paz = self._paz_combo(o["pazarlamaci"] if o else "")
        dlg.tipi = QComboBox(); dlg.tipi.setEditable(True); dlg.tipi.addItems(CRM_MUSTERI_TIPLERI)
        if o: dlg.tipi.setCurrentText(o["musteri_tipi"] or "")
        else: dlg.tipi.setCurrentIndex(-1)
        dlg.tarih = QDateEdit(); dlg.tarih.setCalendarPopup(True); dlg.tarih.setDisplayFormat("dd.MM.yyyy")
        dlg.tarih.setDate(QDate.fromString(o["tarih"], "yyyy-MM-dd") if o and o["tarih"] else QDate.currentDate())
        dlg.kod = QLineEdit(o["kod"] if o else "")
        dlg.renk = QLineEdit(o["renk"] if o else "")
        dlg.miktar = sb(99999999, 2, o["miktar"] if o else 0)
        dlg.maliyet = sb(9999999, 4, o["maliyet_fiyati"] if o else 0)
        dlg.satis = sb(9999999, 4, o["satis_fiyati"] if o else 0)
        dlg.kar = sb(100, 4, o["kar_orani"] if o else 0); dlg.kar.setToolTip("Otomatik: (satış − maliyet) / maliyet")
        dlg.teorik = sb(99999999, 2, o["teorik_kar_usd"] if o else 0)
        dlg.kartl = sb(999999999, 2, o["kar_tl"] if o else 0)
        dlg.vade = QLineEdit(o["vade"] if o else "")
        dlg.ciro = sb(999999999, 2, o["ciro"] if o else 0)
        dlg.cirousd = sb(999999999, 2, o["ciro_usd"] if o else 0)
        dlg.kur = sb(10000, 4, o["usd_kuru"] if o else 0)

        # Otomatik hesaplanan alanlar — salt okunur + soluk zemin
        auto_style = "background:#F1F3F4; color:#37474F;"
        for wd, tip in [(dlg.kar, "Otomatik: (satış − maliyet) / maliyet"),
                        (dlg.teorik, "Otomatik: birim kâr × miktar × (ciro USD / ciro)"),
                        (dlg.kartl, "Otomatik: teorik kâr USD × kur"),
                        (dlg.ciro, "Otomatik: satış × miktar")]:
            wd.setReadOnly(True); wd.setButtonSymbols(QDoubleSpinBox.ButtonSymbols.NoButtons)
            wd.setStyleSheet(auto_style); wd.setToolTip(tip)
        dlg.cirousd.setToolTip("Otomatik = satış × miktar (USD). Döviz farklıysa elle düzeltebilirsiniz.")

        dlg._busy = False
        def _recalc(full=True):
            if dlg._busy:
                return
            dlg._busy = True
            try:
                mal, sat, q, kur = dlg.maliyet.value(), dlg.satis.value(), dlg.miktar.value(), dlg.kur.value()
                birim = sat - mal
                dlg.kar.setValue((birim / mal) if mal else 0)
                dlg.ciro.setValue(sat * q)
                if full:                       # maliyet/satış/miktar değişti → ciro USD'yi USD varsayımıyla tazele
                    dlg.cirousd.setValue(sat * q)
                c, cu = dlg.ciro.value(), dlg.cirousd.value()
                ratio = (cu / c) if c else 1.0
                dlg.teorik.setValue(birim * q * ratio)
                dlg.kartl.setValue(dlg.teorik.value() * kur)
            finally:
                dlg._busy = False
        dlg._recalc = _recalc
        for wd in (dlg.maliyet, dlg.satis, dlg.miktar):
            wd.valueChanged.connect(lambda _=None: _recalc(True))
        dlg.kur.valueChanged.connect(lambda _=None: _recalc(False))
        dlg.cirousd.valueChanged.connect(lambda _=None: _recalc(False))

        for lbl, wd in [("Müşteri:", dlg.musteri), ("Pazarlamacı:", dlg.paz), ("Müşteri Tipi:", dlg.tipi),
                        ("Tarih:", dlg.tarih), ("Kod:", dlg.kod), ("Renk:", dlg.renk), ("Miktar:", dlg.miktar),
                        ("Maliyet Fiyatı:", dlg.maliyet), ("Satış Fiyatı:", dlg.satis), ("Kâr Oranı (oto):", dlg.kar),
                        ("Teorik Kâr USD (oto):", dlg.teorik), ("Kâr TL (oto):", dlg.kartl), ("Vade:", dlg.vade),
                        ("Ciro (oto):", dlg.ciro), ("Ciro USD:", dlg.cirousd), ("USD Kuru:", dlg.kur)]:
            form.addRow(lbl, wd)
        lay.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject); lay.addWidget(bb)
        return dlg

    def _order_vals(self, d):
        return (d.musteri.text(), d.paz.currentText().strip(), d.tipi.currentText().strip(),
                d.tarih.date().toString("yyyy-MM-dd"), d.kod.text(), d.renk.text(), d.miktar.value(),
                d.maliyet.value(), d.satis.value(), d.kar.value(), d.teorik.value(), d.kartl.value(),
                d.vade.text(), d.ciro.value(), d.cirousd.value(), d.kur.value())

    def _add_order(self):
        d = self._order_dialog()
        if d.exec():
            db.add_crm_order(*self._order_vals(d))
            self._refresh_year_combo(self.or_year); self._load_orders()

    def _edit_order(self):
        oid = self._row_id(self.or_table)
        if not oid: return
        o = self._find(db.get_crm_orders(), oid)
        if not o: return
        d = self._order_dialog(o)
        if d.exec():
            db.update_crm_order(oid, *self._order_vals(d)); self._load_orders()

    def _delete_order(self):
        oid = self._row_id(self.or_table)
        if not oid: return
        if self._confirm_del():
            db.delete_crm_order(oid); self._load_orders()

    # ── Sekme 5: Analiz (Aylık × Pazarlamacı) ─────────────────────
    _AY_ADLARI = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz",
                  "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    # (ways, tanım, hedef, anahtar, tip)  tip: int|money|pct
    _AN_SATIRLAR = [
        ("1", "AYLIK ZİYARET EDİLEN YENİ MÜŞTERİ SAYISI",   "Haftada 3 / Ayda 12", "visit_yeni",     "int"),
        ("",  "AYLIK ZİYARET EDİLEN MEVCUT MÜŞTERİ SAYISI", "Haftada 3 / Ayda 12", "visit_mevcut",   "int"),
        ("2", "YENİ MÜŞTERİ SİPARİŞ ADEDİ",                 "",                    "yeni_sip_adet",  "int"),
        ("",  "YENİ MÜŞTERİ SİPARİŞ METRESİ",               "% 10 (4500 MT)",      "yeni_sip_metre", "money"),
        ("",  "AYLIK TOPLAM SİPARİŞ METRESİ - GERÇEKLEŞEN", "",                    "fiili_metre",    "money"),
        ("",  "AYLIK TOPLAM SİP TUTARI (USD)",              "150.000 USD",         "fiili_usd",      "money"),
        ("3", "AYLIK ALINAN SİPARİŞ ADEDİ",                 "15",                  "sip_adet",       "int"),
        ("",  "AYLIK ALINAN SİPARİŞ METRESİ",               "45.000 MT/AY",        "sip_metre",      "money"),
        ("4", "AYLIK ALINAN ORT SİP TUTARI (USD)",          "10.000 USD",          "ort_ciro_usd",   "money"),
        ("",  "AYLIK ALINAN SİPARİŞİN TEORİK KAR TUTARI (USD)", "",                "teorik_kar_usd", "money"),
        ("5", "AYLIK ALINAN SİPARİŞİN TEORİK ORT KAR ORANI (%)", "% 30",           "ort_kar_orani",  "pct"),
    ]
    # ay arka plan renkleri (Ocak yeşil, Şubat turuncu, Mart sarı …)
    _AN_AY_RENK = ["#E8F5E9", "#FFF3E0", "#FFFDE7", "#E3F2FD", "#F3E5F5", "#FCE4EC",
                   "#E0F7FA", "#F1F8E9", "#FFF8E1", "#EDE7F6", "#E1F5FE", "#FBE9E7"]

    def _build_analysis_tab(self):
        w = QWidget(); v = QVBoxLayout(w)
        top = QHBoxLayout()
        top.addWidget(QLabel("Pazarlamacı:"))
        self.an_paz = QComboBox()
        self.an_paz.addItem("📊 Toplam (Tümü)", "")
        for p in CRM_PAZARLAMACILAR:
            self.an_paz.addItem(p, p)
        self.an_paz.currentIndexChanged.connect(self._load_analysis)
        top.addWidget(self.an_paz)
        top.addWidget(QLabel("   Yıl:"))
        self.an_year = QComboBox(); self.an_year.currentIndexChanged.connect(self._load_analysis)
        top.addWidget(self.an_year); top.addStretch()
        v.addLayout(top)
        info = QLabel("Her hücrede üstte değer, altta bir önceki aya göre değişim (▲ artış / ▼ azalış).")
        info.setStyleSheet("color:#555; font-size:11px;")
        v.addWidget(info)
        self.an_table = QTableWidget()
        self.an_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.an_table.setSortingEnabled(False)
        self.an_table.setWordWrap(False)
        self.an_table.verticalHeader().setVisible(False)
        self.an_table.setAlternatingRowColors(True)
        v.addWidget(self.an_table)
        return w

    @staticmethod
    def _an_combine(scopes):
        keys = ("visit_yeni", "visit_mevcut", "yeni_sip_adet", "yeni_sip_metre", "sip_adet",
                "sip_metre", "fiili_metre", "fiili_usd", "sip_ciro_usd", "teorik_kar_usd")
        tot = {k: 0.0 for k in keys}
        for s in scopes:
            for k in keys:
                tot[k] += s.get(k, 0) or 0
        tot["ort_ciro_usd"] = tot["sip_ciro_usd"] / tot["sip_adet"] if tot["sip_adet"] else 0
        tot["ort_kar_orani"] = tot["teorik_kar_usd"] / tot["sip_ciro_usd"] if tot["sip_ciro_usd"] else 0
        return tot

    @staticmethod
    def _an_val(scope, key):
        if key == "ort_kar_orani":
            return (scope.get("ort_kar_orani", 0) or 0) * 100
        return scope.get(key, 0) or 0

    @staticmethod
    def _an_fmt(val, kind):
        if kind == "pct":
            return f"{val:,.0f}%"
        return f"{val:,.0f}"

    @staticmethod
    def _an_delta(curr, prev, kind):
        """Bir önceki aya göre değişim metni + renk. pct için puan, diğerleri için %."""
        if kind == "pct":
            d = curr - prev
            if abs(d) < 0.5:
                return ("◦ 0", "#9E9E9E")
            return (f"{'▲' if d > 0 else '▼'} {abs(d):.0f}p", "#2E7D32" if d > 0 else "#C62828")
        if prev == 0:
            return ("▲ yeni", "#2E7D32") if curr else ("", None)
        d = (curr - prev) / prev * 100.0
        if abs(d) < 0.5:
            return ("◦ %0", "#9E9E9E")
        return (f"{'▲' if d > 0 else '▼'} %{abs(d):.0f}", "#2E7D32" if d > 0 else "#C62828")

    def _load_analysis(self):
        if self.an_year.count() == 0:
            self._refresh_year_combo(self.an_year)
            if self.an_year.count() > 1:
                self.an_year.setCurrentIndex(1)   # en güncel yıl
        data = db.get_crm_analysis(year=self.an_year.currentData() or "") or {}
        # Uzak modda JSON, ay anahtarlarını (1,2,…) metne ("1","2",…) çevirir; int'e normalize et
        for p in list(data.keys()):
            mo = data[p].get("months") or {}
            data[p]["months"] = {int(k): v for k, v in mo.items()}
        reps = [p for p in CRM_PAZARLAMACILAR if p in data] or sorted(data.keys())

        # Seçili pazarlamacı (veya Toplam) için ay-kapsam ve yıl-kapsam fonksiyonları
        sel = self.an_paz.currentData() or ""
        if sel:   # tek pazarlamacı
            d = data.get(sel) or {"months": {}, "yearly": {}}
            month_scope = lambda m: d["months"].get(m, {})
            year_scope = d.get("yearly", {})
        else:     # Toplam (tüm pazarlamacılar birleşik)
            month_scope = lambda m: self._an_combine([data[p]["months"].get(m, {}) for p in reps])
            year_scope = self._an_combine([data[p]["yearly"] for p in reps])

        # Veri olan ayları bul
        nz = ("visit_yeni", "visit_mevcut", "yeni_sip_adet", "sip_adet", "sip_metre",
              "fiili_metre", "fiili_usd", "teorik_kar_usd")
        months = [m for m in range(1, 13) if any(month_scope(m).get(k, 0) for k in nz)]
        if not months:
            months = list(range(1, 13))

        headers = ["5W", "TANIMLAR", "HEDEFLER"] + [self._AY_ADLARI[m - 1] for m in months] + ["YILLIK"]
        t = self.an_table
        t.setColumnCount(len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.setRowCount(len(self._AN_SATIRLAR))
        bold = QFont(); bold.setBold(True)
        gray = QBrush(QColor("#ECEFF1"))

        for ri, (ways, tanim, hedef, key, kind) in enumerate(self._AN_SATIRLAR):
            wi = QTableWidgetItem(ways); wi.setFont(bold)
            wi.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if ways: wi.setBackground(QBrush(QColor("#C8E6C9")))
            t.setItem(ri, 0, wi)
            ti = QTableWidgetItem(tanim); ti.setFont(bold)
            t.setItem(ri, 1, ti)
            hi = QTableWidgetItem(hedef); hi.setForeground(QBrush(QColor("#1565C0")))
            t.setItem(ri, 2, hi)
            # aylık değerler + bir önceki aya göre % değişim
            month_vals = [self._an_val(month_scope(m), key) for m in months]
            for mi, m in enumerate(months):
                val = month_vals[mi]
                txt = self._an_fmt(val, kind)
                color = None
                if mi > 0:
                    dtxt, color = self._an_delta(val, month_vals[mi - 1], kind)
                    if dtxt:
                        txt += "\n" + dtxt
                it = QTableWidgetItem(txt)
                it.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if color:
                    it.setForeground(QBrush(QColor(color)))
                t.setItem(ri, 3 + mi, it)
            # YILLIK
            yi = QTableWidgetItem(self._an_fmt(self._an_val(year_scope, key), kind))
            yi.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            yi.setBackground(gray); yi.setFont(bold)
            t.setItem(ri, 3 + len(months), yi)

        t.resizeColumnsToContents()
        t.setColumnWidth(1, 300)
        t.setColumnWidth(2, 120)
        t.resizeRowsToContents()
        t.verticalHeader().setVisible(False)

    # ── Sekme 6: Grafik (trend) ───────────────────────────────────
    def _build_chart_tab(self):
        w = QWidget(); v = QVBoxLayout(w)
        top = QHBoxLayout()
        top.addWidget(QLabel("Yıl:"))
        self.ch_year = QComboBox(); self.ch_year.currentIndexChanged.connect(self._load_chart)
        top.addWidget(self.ch_year)
        top.addWidget(QLabel("   Metrik:"))
        self.ch_metric = QComboBox()
        for ways, tanim, hedef, key, kind in self._AN_SATIRLAR:
            self.ch_metric.addItem(tanim, key)
        di = self.ch_metric.findData("fiili_usd")
        if di >= 0:
            self.ch_metric.setCurrentIndex(di)   # varsayılan: Aylık Toplam Sip. Tutarı
        self.ch_metric.currentIndexChanged.connect(self._load_chart)
        top.addWidget(self.ch_metric)
        self.ch_toplam = QCheckBox("Toplam çizgisi"); self.ch_toplam.setChecked(True)
        self.ch_toplam.toggled.connect(self._load_chart)
        top.addWidget(self.ch_toplam)
        top.addStretch()
        v.addLayout(top)
        self.ch_chart = _LineChart()
        v.addWidget(self.ch_chart, 1)
        return w

    def _load_chart(self):
        if self.ch_year.count() == 0:
            self._refresh_year_combo(self.ch_year)
            if self.ch_year.count() > 1:
                self.ch_year.setCurrentIndex(1)
        data = db.get_crm_analysis(year=self.ch_year.currentData() or "") or {}
        for p in list(data.keys()):
            mo = data[p].get("months") or {}
            data[p]["months"] = {int(k): v for k, v in mo.items()}
        reps = [p for p in CRM_PAZARLAMACILAR if p in data] or sorted(data.keys())
        key = self.ch_metric.currentData() or "fiili_usd"
        kind = next((k for wy, tn, hd, mk, k in self._AN_SATIRLAR if mk == key), "money")

        # veri olan aylar
        nz = ("visit_yeni", "visit_mevcut", "yeni_sip_adet", "sip_adet", "sip_metre",
              "fiili_metre", "fiili_usd", "teorik_kar_usd")
        def mcomb(m):
            return self._an_combine([data[p]["months"].get(m, {}) for p in reps])
        months = [m for m in range(1, 13) if any(mcomb(m).get(k, 0) for k in nz)] or list(range(1, 13))
        xlabels = [self._AY_ADLARI[m - 1][:3] for m in months]

        series = []
        for p in reps:
            vals = [self._an_val(data[p]["months"].get(m, {}), key) for m in months]
            series.append((p.split()[0], vals))
        if self.ch_toplam.isChecked():
            tvals = [self._an_val(mcomb(m), key) for m in months]
            series.append(("TOPLAM", tvals))
        title = self.ch_metric.currentText() + " — " + (self.ch_year.currentData() or "Tüm yıllar")
        self.ch_chart.set_data(series, xlabels, kind, title)

    # ── ortak küçük yardımcılar ───────────────────────────────────
    def _num(self, val, suffix=""):
        val = val or 0
        it = _SortItem(f"{val:,.2f}{suffix}" if not float(val).is_integer() else f"{val:,.0f}{suffix}")
        it.setData(Qt.ItemDataRole.UserRole, float(val))
        it.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        return it

    def _row_id(self, table):
        r = table.currentRow()
        if r < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir satır seçin.")
            return None
        it = table.item(r, 0)
        return it.data(Qt.ItemDataRole.UserRole) if it else None

    def _find(self, rows, rid):
        for r in (rows or []):
            if r["id"] == rid:
                return dict(r)
        return None

    def _confirm_del(self):
        return QMessageBox.question(self, "Sil", "Kayıt silinsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "CRM Excel Dosyası", "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        try:
            data = crm_parse_excel(path)
        except Exception as e:
            return QMessageBox.critical(self, "Hata", f"Excel okunamadı:\n{e}")
        n = {k: len(v) for k, v in data.items()}
        msg = (f"Okunan kayıtlar:\n• Müşteri: {n['customers']}\n• Ziyaret: {n['visits']}\n"
               f"• Fiili Satış: {n['sales']}\n• Sipariş: {n['orders']}\n\n"
               "Mevcut CRM verisinin ÜZERİNE yazılsın mı?\n"
               "(Evet = önce mevcut CRM verisi silinir, sonra bunlar eklenir)")
        ret = QMessageBox.question(self, "İçe Aktar", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel)
        if ret == QMessageBox.StandardButton.Cancel:
            return
        replace = (ret == QMessageBox.StandardButton.Yes)
        try:
            counts = db.crm_import_bulk(customers=data["customers"], visits=data["visits"],
                                        sales=data["sales"], orders=data["orders"], replace=replace)
        except Exception as e:
            return QMessageBox.critical(self, "Hata", f"İçe aktarma başarısız:\n{e}")
        self._refresh_year_combo(self.sa_year); self._refresh_year_combo(self.or_year)
        self._refresh_year_combo(self.an_year)
        self.refresh()
        QMessageBox.information(self, "Tamam",
            f"İçe aktarıldı:\n• Müşteri: {counts.get('customers',0)}\n• Ziyaret: {counts.get('visits',0)}\n"
            f"• Fiili Satış: {counts.get('sales',0)}\n• Sipariş: {counts.get('orders',0)}")


class CustomerManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Müşteri Yönetimi")
        self.setMinimumSize(680, 460)
        self._build_ui(); self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Arama
        top = QHBoxLayout()
        self.search = QLineEdit(); self.search.setPlaceholderText("Müşteri adı, kodu, telefon...")
        self.search.textChanged.connect(lambda: self._load(self.search.text()))
        top.addWidget(QLabel("Ara:")); top.addWidget(self.search); top.addStretch()
        lay.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Müşteri Adı", "Kodu", "Telefon", "Adres", "Vergi No", "Durum"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for i in (1,2,3,4,5): hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        btn_add   = QPushButton("+ Yeni Müşteri");  btn_add.clicked.connect(self._add)
        btn_edit  = QPushButton("✎ Düzenle");        btn_edit.clicked.connect(self._edit)
        btn_del   = QPushButton("✕ Sil")
        btn_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        btn_del.clicked.connect(self._delete)
        btn_excel = QPushButton("📥 Excel'den İçe Aktar")
        btn_excel.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;border-radius:4px;padding:6px 14px;")
        btn_excel.clicked.connect(self._import_excel)
        btn_close = QPushButton("Kapat"); btn_close.clicked.connect(self.accept)
        for b in (btn_add, btn_edit, btn_del, btn_excel):
            btn_row.addWidget(b)
        btn_row.addStretch(); btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

    def _load(self, search=""):
        rows = db.get_all_customers(search=search, active_only=False)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            name_item = QTableWidgetItem(r["name"] or "")
            name_item.setData(Qt.ItemDataRole.UserRole, r["id"])
            self.table.setItem(i, 0, name_item)
            self.table.setItem(i, 1, QTableWidgetItem(r["code"] or ""))
            self.table.setItem(i, 2, QTableWidgetItem(r["phone"] or ""))
            self.table.setItem(i, 3, QTableWidgetItem(r["address"] or ""))
            self.table.setItem(i, 4, QTableWidgetItem(r["tax_no"] or ""))
            s = QTableWidgetItem("✅ Aktif" if r["active"] else "⛔ Pasif")
            s.setForeground(QBrush(QColor("#2E7D32" if r["active"] else "#C62828")))
            self.table.setItem(i, 5, s)
        self.table.setSortingEnabled(True)   # başlığa tıklayınca sıralar

    def _selected_id(self):
        row = self.table.currentRow()
        if row < 0: QMessageBox.information(self,"Bilgi","Müşteri seçin."); return None
        return self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)

    def _customer_dialog(self, c=None):
        dlg = QDialog(self); dlg.setWindowTitle("Müşteri" + (" Düzenle" if c else " Ekle"))
        dlg.setMinimumWidth(360); lay = QVBoxLayout(dlg); form = QFormLayout(); form.setSpacing(8)
        dlg.name   = QLineEdit(c["name"]    if c else "")
        dlg.code   = QLineEdit(c["code"]    if c else "")
        dlg.phone  = QLineEdit(c["phone"]   if c else "")
        dlg.addr   = QLineEdit(c["address"] if c else "")
        dlg.tax_no = QLineEdit(c["tax_no"]  if c else "")
        form.addRow("Müşteri Adı *:", dlg.name)
        form.addRow("Kodu:",          dlg.code)
        form.addRow("Telefon:",       dlg.phone)
        form.addRow("Adres:",         dlg.addr)
        form.addRow("Vergi No:",      dlg.tax_no)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        return dlg

    def _add(self):
        dlg = self._customer_dialog()
        if dlg.exec():
            if not dlg.name.text().strip():
                return QMessageBox.warning(self,"Hata","Müşteri adı zorunlu!")
            db.add_customer(dlg.name.text(), dlg.code.text(), dlg.phone.text(), dlg.addr.text(), dlg.tax_no.text())
            self._load(self.search.text())

    def _edit(self):
        cid = self._selected_id()
        if not cid: return
        c = db.get_customer(cid)
        dlg = self._customer_dialog(c)
        if dlg.exec():
            db.update_customer(cid, dlg.name.text(), dlg.code.text(),
                               dlg.phone.text(), dlg.addr.text(), dlg.tax_no.text())
            self._load(self.search.text())

    def _delete(self):
        cid = self._selected_id()
        if not cid: return
        if QMessageBox.question(self,"Sil","Müşteri silinsin mi?",
                QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_customer(cid); self._load(self.search.text())

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self,"Müşteri Excel Dosyası","","Excel (*.xlsx *.xls)")
        if not path: return
        try:
            dlg = ExcelColumnMapDialog(path, CUSTOMER_FIELDS, CUSTOMER_AUTO, self)
            if not dlg.exec():
                return
            records = dlg.records
            if not records:
                return QMessageBox.warning(self,"Hata","Eşleştirilen sütunda geçerli veri bulunamadı.")
            db.import_customers_bulk(records)
            self._load(self.search.text())
            QMessageBox.information(self,"Başarılı",f"{len(records)} müşteri içe aktarıldı.")
        except Exception as e:
            QMessageBox.critical(self,"Hata",f"İçe aktarma hatası:\n{e}")


class SupplierManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Tedarikçi Yönetimi")
        self.setMinimumSize(680, 460)
        self._build_ui(); self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Arama
        top = QHBoxLayout()
        self.search = QLineEdit(); self.search.setPlaceholderText("Tedarikçi adı, kodu, telefon...")
        self.search.textChanged.connect(lambda: self._load(self.search.text()))
        top.addWidget(QLabel("Ara:")); top.addWidget(self.search); top.addStretch()
        lay.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Tedarikçi Adı", "Kodu", "Telefon", "Adres", "Vergi No", "E-posta", "Durum"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)   # sütunlar elle genişletilebilir
        hdr.setStretchLastSection(False)
        hdr.setSectionsMovable(True)                                   # sütunlar sürüklenip taşınabilir
        def _supplier_default_widths():
            for i, wdt in enumerate([220, 90, 110, 240, 110, 180, 80]):
                self.table.setColumnWidth(i, wdt)
        _wire_header_persistence(self.table, "suppliers_header", _supplier_default_widths)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        btn_add   = QPushButton("+ Yeni Tedarikçi");  btn_add.clicked.connect(self._add)
        btn_edit  = QPushButton("✎ Düzenle");          btn_edit.clicked.connect(self._edit)
        btn_del   = QPushButton("✕ Sil")
        btn_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        btn_del.clicked.connect(self._delete)
        btn_excel = QPushButton("📥 Excel'den İçe Aktar")
        btn_excel.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;border-radius:4px;padding:6px 14px;")
        btn_excel.clicked.connect(self._import_excel)
        btn_close = QPushButton("Kapat"); btn_close.clicked.connect(self.accept)
        for b in (btn_add, btn_edit, btn_del, btn_excel):
            btn_row.addWidget(b)
        btn_row.addStretch(); btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

    def _load(self, search=""):
        rows = db.get_all_suppliers(search=search, active_only=False)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            name_item = QTableWidgetItem(r["name"] or "")
            name_item.setData(Qt.ItemDataRole.UserRole, r["id"])
            self.table.setItem(i, 0, name_item)
            self.table.setItem(i, 1, QTableWidgetItem(r["code"] or ""))
            self.table.setItem(i, 2, QTableWidgetItem(r["phone"] or ""))
            self.table.setItem(i, 3, QTableWidgetItem(r["address"] or ""))
            self.table.setItem(i, 4, QTableWidgetItem(r["tax_no"] or ""))
            self.table.setItem(i, 5, QTableWidgetItem(dict(r).get("email") or ""))
            s = QTableWidgetItem("✅ Aktif" if r["active"] else "⛔ Pasif")
            s.setForeground(QBrush(QColor("#2E7D32" if r["active"] else "#C62828")))
            self.table.setItem(i, 6, s)
        self.table.setSortingEnabled(True)   # başlığa tıklayınca sıralar

    def _selected_id(self):
        row = self.table.currentRow()
        if row < 0: QMessageBox.information(self,"Bilgi","Tedarikçi seçin."); return None
        return self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)

    def _supplier_dialog(self, s=None):
        dlg = QDialog(self); dlg.setWindowTitle("Tedarikçi" + (" Düzenle" if s else " Ekle"))
        dlg.setMinimumWidth(360); lay = QVBoxLayout(dlg); form = QFormLayout(); form.setSpacing(8)
        dlg.name   = QLineEdit(s["name"]    if s else "")
        dlg.code   = QLineEdit(s["code"]    if s else "")
        dlg.phone  = QLineEdit(s["phone"]   if s else "")
        dlg.addr   = QLineEdit(s["address"] if s else "")
        dlg.tax_no = QLineEdit(s["tax_no"]  if s else "")
        dlg.email  = QLineEdit(dict(s).get("email", "") if s else "")
        form.addRow("Tedarikçi Adı *:", dlg.name)
        form.addRow("Kodu:",            dlg.code)
        form.addRow("Telefon:",         dlg.phone)
        form.addRow("Adres:",           dlg.addr)
        form.addRow("Vergi No:",        dlg.tax_no)
        form.addRow("E-posta:",         dlg.email)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        return dlg

    def _add(self):
        dlg = self._supplier_dialog()
        if dlg.exec():
            if not dlg.name.text().strip():
                return QMessageBox.warning(self,"Hata","Tedarikçi adı zorunlu!")
            db.add_supplier(dlg.name.text(), dlg.code.text(), dlg.phone.text(), dlg.addr.text(),
                            dlg.tax_no.text(), dlg.email.text().strip())
            self._load(self.search.text())

    def _edit(self):
        sid = self._selected_id()
        if not sid: return
        s = db.get_supplier(sid)
        dlg = self._supplier_dialog(s)
        if dlg.exec():
            db.update_supplier(sid, dlg.name.text(), dlg.code.text(),
                               dlg.phone.text(), dlg.addr.text(), dlg.tax_no.text(),
                               email=dlg.email.text().strip())
            self._load(self.search.text())

    def _delete(self):
        sid = self._selected_id()
        if not sid: return
        if QMessageBox.question(self,"Sil","Tedarikçi silinsin mi?",
                QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_supplier(sid); self._load(self.search.text())

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self,"Tedarikçi Excel Dosyası","","Excel (*.xlsx *.xls)")
        if not path: return
        try:
            dlg = ExcelColumnMapDialog(path, SUPPLIER_FIELDS, SUPPLIER_AUTO, self)
            if not dlg.exec():
                return
            records = dlg.records
            if not records:
                return QMessageBox.warning(self,"Hata","Eşleştirilen sütunda geçerli veri bulunamadı.")
            db.import_suppliers_bulk(records)
            self._load(self.search.text())
            QMessageBox.information(self,"Başarılı",f"{len(records)} tedarikçi içe aktarıldı.")
        except Exception as e:
            QMessageBox.critical(self,"Hata",f"İçe aktarma hatası:\n{e}")


class IplikWidget(QWidget):
    """İplik giriş formu (İplik Kataloğu'nda kullanılır)."""
    changed = pyqtSignal()
    NUMARA = ["DN", "NM", "NE", "DTEX"]
    _NUM_MAP = {"DN": "Den", "DTEX": "dTex", "NM": "Nm", "NE": "Ne"}
    PARLAKLIK = ["Yarım Mat", "Parlak", "Full Mat", "Süper Parlak"]
    CEKIM = ["Ring", "Vortex", "Open End", "Compact", "Karde", "Penye", "Siro",
             "Tex", "FDY", "IMG", "Soft IMG", "Single Cover"]
    MAX_CINS = 6

    def __init__(self, parent=None):
        super().__init__(parent)
        try:
            self._cins_list = db.get_iplik_cinsleri() or []
        except Exception:
            self._cins_list = list(getattr(_local_db, "IPLIK_CINSLERI_BASE", []))
        self._build_ui()
        _disable_wheel(self)

    def _build_ui(self):
        v = QVBoxLayout(self)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        inner = QWidget(); scroll.setWidget(inner); v.addWidget(scroll, 1)
        vl = QVBoxLayout(inner)

        # 1) Numara sistemine kadar olan alanlar (Kat sayısı üst seviyede — dialogda)
        form = QFormLayout(); form.setSpacing(8)
        self.iplik_no = QLineEdit()
        self.flament  = QLineEdit()
        self.numara   = QComboBox(); self.numara.addItems(self.NUMARA)
        form.addRow("İplik No:", self.iplik_no)
        form.addRow("Flament Sayısı:", self.flament)
        form.addRow("Numara Sistemi:", self.numara)
        vl.addLayout(form)

        # 2) İplik Cinsi & Oranları — numara sisteminin hemen altında
        grp = QGroupBox("İplik Cinsi & Oranları — toplam %100 olmalı")
        gl = QVBoxLayout(grp)
        grid = QGridLayout(); gl.addLayout(grid)
        self._rows = []
        for i in range(self.MAX_CINS):
            lbl = QLabel(f"İplik Cinsi {i+1}:")
            cb = QComboBox()   # seçmeli açılır liste (elle yazılamaz; yeni için buton)
            oran = QDoubleSpinBox(); oran.setRange(0, 100); oran.setDecimals(1); oran.setSuffix(" %")
            oran.valueChanged.connect(self._recompute)
            grid.addWidget(lbl, i, 0); grid.addWidget(cb, i, 1); grid.addWidget(oran, i, 2)
            self._rows.append({"lbl": lbl, "cins": cb, "oran": oran})
        self._refresh_cins_combos()
        r2 = QHBoxLayout()
        b_new = QPushButton("+ Yeni İplik Cinsi"); b_new.clicked.connect(self._add_cins_type)
        self.total_lbl = QLabel()
        r2.addWidget(b_new); r2.addStretch(); r2.addWidget(self.total_lbl)
        gl.addLayout(r2)
        vl.addWidget(grp)

        # 3) Kalan alanlar
        form2 = QFormLayout(); form2.setSpacing(8)
        self.parlaklik = QComboBox(); self.parlaklik.addItem("— Seçiniz —", "")
        for x in self.PARLAKLIK: self.parlaklik.addItem(x, x)
        self.cekim = QComboBox(); self.cekim.addItem("— Seçiniz —", "")
        for x in self.CEKIM: self.cekim.addItem(x, x)
        self.tur = QLineEdit()
        self.likra = QLineEdit(); self.likra.setPlaceholderText("DN cinsinden — girilirse 1/3'ü hesaba katılır")
        self.fiyat = QDoubleSpinBox(); self.fiyat.setRange(0, 999999); self.fiyat.setDecimals(2); self.fiyat.setSuffix(" $/kg")
        form2.addRow("Parlaklık:", self.parlaklik)
        form2.addRow("Çekim Sistemi:", self.cekim)
        form2.addRow("Tur Sayısı:", self.tur)
        form2.addRow("Likra (DN):", self.likra)
        form2.addRow("Fiyat:", self.fiyat)
        vl.addLayout(form2)

        self.denye_lbl = QLabel("—"); self.denye_lbl.setStyleSheet("font-weight:bold; color:#1A237E; font-size:13px;")
        f3 = QFormLayout(); f3.addRow("Toplam Denye (likra 1/3 dahil):", self.denye_lbl)
        vl.addLayout(f3)
        vl.addStretch()

        for w in (self.iplik_no, self.likra):
            w.textChanged.connect(self._update_denye)
        self.numara.currentIndexChanged.connect(self._update_denye)
        # Tüm alanlar değişince 'changed' yayınla (ad otomatik güncellensin)
        for w in (self.iplik_no, self.flament, self.tur, self.likra):
            w.textChanged.connect(lambda *a: self.changed.emit())
        for cb in (self.numara, self.parlaklik, self.cekim):
            cb.currentIndexChanged.connect(lambda *a: self.changed.emit())
        self.fiyat.valueChanged.connect(lambda *a: self.changed.emit())
        for r in self._rows:
            r["cins"].currentIndexChanged.connect(lambda *a: self.changed.emit())
            r["oran"].valueChanged.connect(lambda *a: self.changed.emit())
        self._recompute()

    def build_name(self):
        """Alanlardan otomatik iplik adı/kodu üretir (boşlukla ayrılmış düz metin)."""
        parts = []
        no = self.iplik_no.text().strip(); fl = self.flament.text().strip()
        if no and fl:
            parts.append(f"{no}/{fl}")
        elif no or fl:
            parts.append(no or fl)
        num = self.numara.currentText().strip()
        if num:
            parts.append(num)
        for i in self._active_indices():
            r = self._rows[i]
            c = (r["cins"].currentData() or "").strip(); o = r["oran"].value()
            if c:
                parts.append(f"%{o:.0f} {c.upper()}")
        par = (self.parlaklik.currentData() or "").strip()
        if par:
            parts.append(par.upper())
        cek = (self.cekim.currentData() or "").strip()
        if cek:
            parts.append(cek.upper())
        tur = self.tur.text().strip()
        if tur:
            parts.append(f"{tur} TUR")
        lik = self.likra.text().strip()
        if lik:
            parts.append(f"{lik} LİKRA")
        return " ".join(parts)

    def _active_indices(self):
        """Görünür/aktif kompozisyon satırlarının indeksleri.
        İlk satır her zaman; sonraki satır ancak bir önceki satır dolu (oran>0) VE
        toplam < 100 ise açılır → en fazla bir boş satır görünür."""
        idxs = []
        total = 0.0; stop = False
        for i, r in enumerate(self._rows):
            if i == 0:
                active = True
            elif stop:
                active = False
            else:
                prev = self._rows[i - 1]["oran"].value()
                active = (prev > 0) and (total < 99.999)
            if active:
                idxs.append(i)
                o = r["oran"].value()
                total += o
                if o <= 0 or total >= 99.999:
                    stop = True
            else:
                stop = True
        return idxs

    def _recompute(self):
        active = set(self._active_indices())
        total = 0.0
        for i, r in enumerate(self._rows):
            vis = i in active
            for w in (r["lbl"], r["cins"], r["oran"]):
                w.setVisible(vis)
            if vis:
                total += r["oran"].value()
        if abs(total - 100) < 0.01:
            self.total_lbl.setText(f"Toplam: %{total:.0f}  ✓")
            self.total_lbl.setStyleSheet("color:#2E7D32; font-weight:bold;")
        else:
            self.total_lbl.setText(f"Toplam: %{total:.0f}" + ("  (100 olmalı)" if total > 0 else ""))
            self.total_lbl.setStyleSheet("color:#C62828; font-weight:bold;")
        self._update_denye()

    def denye_value(self):
        no = self.iplik_no.text().strip().replace(",", ".")
        base = _den_from(no, self._NUM_MAP.get(self.numara.currentText(), "Den"))
        try: lik = float((self.likra.text() or "0").replace(",", "."))
        except Exception: lik = 0
        return base + lik / 3.0

    def _update_denye(self):
        toplam = self.denye_value()
        self.denye_lbl.setText(f"{toplam:,.1f} DN" if toplam else "—")

    def _refresh_cins_combos(self):
        """Tüm iplik cinsi açılır listelerini güncel listeyle doldurur (seçimi korur)."""
        for r in self._rows:
            cb = r["cins"]
            cur = cb.currentText()
            cb.blockSignals(True)
            cb.clear()
            cb.addItem("— Seçiniz —", "")
            for c in self._cins_list:
                cb.addItem(c, c)
            if cur:
                idx = cb.findText(cur)
                if idx < 0:            # kayıtlı ama listede yoksa yine göster
                    cb.addItem(cur, cur); idx = cb.count() - 1
                cb.setCurrentIndex(idx)
            else:
                cb.setCurrentIndex(0)
            cb.blockSignals(False)

    def _add_cins_type(self):
        name, ok = QInputDialog.getText(self, "Yeni İplik Cinsi", "İplik cinsi adı:")
        if not ok or not name.strip():
            return
        name = name.strip()
        try:
            db.add_iplik_cinsi(name)
            self._cins_list = db.get_iplik_cinsleri() or self._cins_list
        except Exception:
            if name not in self._cins_list:
                self._cins_list.append(name)
        self._refresh_cins_combos()

    def get_dict(self):
        comp = []
        for i in self._active_indices():
            r = self._rows[i]
            c = (r["cins"].currentData() or "").strip(); o = r["oran"].value()
            if c or o:
                comp.append({"cins": c, "oran": o})
        return {
            "iplik_no": self.iplik_no.text().strip(),
            "flament": self.flament.text().strip(),
            "numara_sistemi": self.numara.currentText(),
            "parlaklik": self.parlaklik.currentData() or "",
            "cekim_sistemi": self.cekim.currentData() or "",
            "tur_sayisi": self.tur.text().strip(),
            "likra_dn": self.likra.text().strip(),
            "fiyat": self.fiyat.value(),
            "kompozisyon": comp,
        }

    def load_dict(self, d):
        d = d or {}
        self.iplik_no.setText(str(d.get("iplik_no", "")))
        self.flament.setText(str(d.get("flament", "")))
        i = self.numara.findText(d.get("numara_sistemi", "")); self.numara.setCurrentIndex(i if i >= 0 else 0)
        pi = self.parlaklik.findData(d.get("parlaklik", "")); self.parlaklik.setCurrentIndex(pi if pi >= 0 else 0)
        ci = self.cekim.findData(d.get("cekim_sistemi", "")); self.cekim.setCurrentIndex(ci if ci >= 0 else 0)
        self.tur.setText(str(d.get("tur_sayisi", "")))
        self.likra.setText(str(d.get("likra_dn", "")))
        try: self.fiyat.setValue(float(d.get("fiyat") or 0))
        except Exception: pass
        for j, item in enumerate((d.get("kompozisyon") or [])[:self.MAX_CINS]):
            cb = self._rows[j]["cins"]; cins = item.get("cins", "")
            if cins:
                idx = cb.findText(cins)
                if idx < 0:
                    cb.addItem(cins, cins); idx = cb.count() - 1
                cb.setCurrentIndex(idx)
            try: self._rows[j]["oran"].setValue(float(item.get("oran") or 0))
            except Exception: pass
        self._recompute()

class IplikEntryDialog(QDialog):
    """İplik tanımı — Kat sayısı + her kat için ayrı iplik formu (sekme). Ad otomatik."""
    MAX_KAT = 6

    def __init__(self, parent=None, iplik=None):
        super().__init__(parent)
        self.iplik = dict(iplik) if iplik else None
        self.setWindowTitle("İplik" + (" Düzenle" if iplik else " Ekle"))
        self.setMinimumSize(600, 660)
        lay = QVBoxLayout(self)

        form = QFormLayout()
        self.ad = QLineEdit()
        self.ad.setReadOnly(True)
        self.ad.setStyleSheet("background:#F1F3F4; color:#1A237E; font-weight:bold;")
        self.ad.setToolTip("Otomatik oluşturulur — alanları doldurdukça güncellenir")
        self.kat_spin = QSpinBox(); self.kat_spin.setRange(1, self.MAX_KAT); self.kat_spin.setValue(1)
        self.kat_spin.valueChanged.connect(self._rebuild_kat_tabs)
        form.addRow("İplik Adı / Kodu (oto):", self.ad)
        form.addRow("Kat Sayısı:", self.kat_spin)
        lay.addLayout(form)

        self.kat_tabs = QTabWidget()
        lay.addWidget(self.kat_tabs, 1)
        self._plies = []   # her kat için IplikWidget

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.button(QDialogButtonBox.StandardButton.Ok).setText("Kaydet")
        bb.accepted.connect(self._accept); bb.rejected.connect(self.reject)
        lay.addWidget(bb)

        # Mevcut kaydı yükle
        katlar = []
        if self.iplik:
            import json
            try:
                d = json.loads(self.iplik.get("data_json") or "{}")
            except Exception:
                d = {}
            katlar = d.get("katlar") or ([d] if d.get("iplik_no") or d.get("kompozisyon") else [])
        n = max(1, len(katlar))
        self.kat_spin.blockSignals(True); self.kat_spin.setValue(n); self.kat_spin.blockSignals(False)
        self._rebuild_kat_tabs()
        for i, kd in enumerate(katlar):
            if i < len(self._plies):
                self._plies[i].load_dict(kd)
        self._refresh_name()

    def _rebuild_kat_tabs(self):
        n = self.kat_spin.value()
        while len(self._plies) < n:
            w = IplikWidget()
            w.changed.connect(self._refresh_name)
            self._plies.append(w)
            self.kat_tabs.addTab(w, f"Kat {len(self._plies)}")
        while len(self._plies) > n:
            self.kat_tabs.removeTab(len(self._plies) - 1)
            self._plies.pop()
        self._refresh_name()

    def _refresh_name(self):
        n = self.kat_spin.value()
        names = [self._plies[i].build_name() for i in range(min(n, len(self._plies)))]
        names = [x for x in names if x]
        ad = " + ".join(names)
        if n > 1:
            ad = (ad + "  " if ad else "") + f"{n} KAT"
        self.ad.setText(ad)

    def _accept(self):
        if not self.ad.text().strip():
            return QMessageBox.warning(self, "Eksik", "En az bir iplik bilgisi girin (ad otomatik oluşur).")
        self.accept()

    def result_data(self):
        import json
        n = self.kat_spin.value()
        katlar = [self._plies[i].get_dict() for i in range(min(n, len(self._plies)))]
        toplam = sum(self._plies[i].denye_value() for i in range(min(n, len(self._plies))))
        data = {"kat": n, "katlar": katlar, "toplam_denye": f"{toplam:,.1f} DN" if toplam else ""}
        return self.ad.text().strip(), json.dumps(data, ensure_ascii=False)


class IplikManagementDialog(QDialog):
    """İplik Kataloğu — ürün kataloğu gibi liste + ekle/düzenle/sil."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🧵 İplik Kataloğu")
        self.setMinimumSize(880, 520)
        self._build_ui(); self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        top = QHBoxLayout()
        self.search = QLineEdit(); self.search.setPlaceholderText("İplik adı / kodu ara...")
        self.search.textChanged.connect(lambda: self._load(self.search.text()))
        top.addWidget(QLabel("Ara:")); top.addWidget(self.search); top.addStretch()
        lay.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ["İplik Adı / Kodu", "Kat", "Kompozisyon (katlar)", "Toplam Denye", "Fiyat ($/kg)"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setSectionsMovable(True)
        hdr.setStretchLastSection(True)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        b_add = QPushButton("+ Yeni İplik"); b_add.clicked.connect(self._add)
        b_edit = QPushButton("✎ Düzenle"); b_edit.clicked.connect(self._edit)
        b_del = QPushButton("✕ Sil")
        b_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        b_del.clicked.connect(self._delete)
        b_close = QPushButton("Kapat"); b_close.clicked.connect(self.accept)
        for b in (b_add, b_edit, b_del): btn_row.addWidget(b)
        btn_row.addStretch(); btn_row.addWidget(b_close)
        lay.addLayout(btn_row)

    def _load(self, search=""):
        import json
        rows = db.get_iplikler(search=search) or []
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            r = dict(r)
            try: d = json.loads(r.get("data_json") or "{}")
            except Exception: d = {}
            katlar = d.get("katlar") or ([d] if d.get("kompozisyon") else [])
            kat_ozet = []; fiyatlar = []
            for kd in katlar:
                komp = " ".join(f"%{c.get('oran',0):.0f} {c.get('cins','')}"
                                for c in (kd.get("kompozisyon") or []))
                parca = " ".join(x for x in [str(kd.get("iplik_no", "")), komp] if x.strip())
                if parca.strip():
                    kat_ozet.append(parca)
                fv = kd.get("fiyat") or 0
                if fv:
                    fiyatlar.append(f"{float(fv):,.2f}")
            it = QTableWidgetItem(r.get("ad", "") or "")
            it.setData(Qt.ItemDataRole.UserRole, r.get("id"))
            self.table.setItem(i, 0, it)
            self.table.setItem(i, 1, QTableWidgetItem(str(d.get("kat", len(katlar) or 1))))
            self.table.setItem(i, 2, QTableWidgetItem("  +  ".join(kat_ozet)))
            self.table.setItem(i, 3, QTableWidgetItem(str(d.get("toplam_denye", ""))))
            self.table.setItem(i, 4, QTableWidgetItem(" + ".join(fiyatlar)))
        self.table.resizeColumnsToContents()

    def _selected_id(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "İplik seçin."); return None
        return self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)

    def _add(self):
        dlg = IplikEntryDialog(self)
        if dlg.exec():
            ad, data = dlg.result_data()
            db.add_iplik(ad, data)
            self._load(self.search.text())

    def _edit(self):
        iid = self._selected_id()
        if not iid: return
        iplik = db.get_iplik(iid)
        if not iplik: return
        dlg = IplikEntryDialog(self, iplik)
        if dlg.exec():
            ad, data = dlg.result_data()
            db.update_iplik(iid, ad, data)
            self._load(self.search.text())

    def _delete(self):
        iid = self._selected_id()
        if not iid: return
        if QMessageBox.question(self, "Sil", "İplik silinsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_iplik(iid); self._load(self.search.text())


class _ProductNameConfirmDialog(QDialog):
    """Ürün adı onayı. Onaylamazsa (Yeni Öner) mevcut öneri kara listeye alınır
    ve yeni bir ad önerilir; kullanıcı elle de yazabilir."""
    def __init__(self, parent, suggested):
        super().__init__(parent)
        self.setWindowTitle("Ürün Adı Onayı")
        self.setMinimumWidth(380)
        self._suggested = suggested   # o an önerilen (otomatik) ad
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("Ürün adını onaylayın, beğenmezseniz yeni öneri isteyin\n"
                             "ya da elle yazın:"))
        self.edit = QLineEdit(suggested)
        lay.addWidget(self.edit)
        row = QHBoxLayout()
        b_new = QPushButton("🎲 Beğenmedim / Yeni Öner")
        b_new.clicked.connect(self._reject_and_new)
        row.addWidget(b_new); row.addStretch()
        lay.addLayout(row)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.button(QDialogButtonBox.StandardButton.Ok).setText("✅ Onayla")
        bb.button(QDialogButtonBox.StandardButton.Cancel).setText("İptal")
        bb.accepted.connect(self._accept); bb.rejected.connect(self.reject)
        lay.addWidget(bb)

    def _reject_and_new(self):
        # Mevcut öneriyi kalıcı olarak kara listeye al, tekrar önerme
        if self._suggested:
            try:
                db.reject_product_name(self._suggested)
            except Exception:
                pass
        try:
            self._suggested = db.generate_product_name()
        except Exception:
            self._suggested = ""
        self.edit.setText(self._suggested)

    def _accept(self):
        if not self.edit.text().strip():
            return QMessageBox.warning(self, "Eksik", "Ürün adı boş olamaz.")
        self.accept()

    def result_name(self):
        return self.edit.text().strip()


class ProductManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ürün Kataloğu Yönetimi")
        self.setMinimumSize(920, 560)
        self._build_ui()
        self._load()
        self._load_numune()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        self.tab_widget = QTabWidget()

        # ── Sekme 1: Aktif Ürünler ────────────────────────────────
        tab_aktif = QWidget()
        v1 = QVBoxLayout(tab_aktif)
        top1 = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("Ürün kodu veya adı...")
        self.search.textChanged.connect(lambda: self._load(self.search.text()))
        top1.addWidget(QLabel("Ara:")); top1.addWidget(self.search, 1); top1.addStretch()
        v1.addLayout(top1)
        self.table = QTableWidget()
        self._setup_table(self.table)
        v1.addWidget(self.table)
        self.tab_widget.addTab(tab_aktif, "📦 Aktif Ürünler")

        # ── Sekme 2: Numuneler ────────────────────────────────────
        tab_numune = QWidget()
        v2 = QVBoxLayout(tab_numune)
        top2 = QHBoxLayout()
        self.numune_search = QLineEdit()
        self.numune_search.setPlaceholderText("Numune kodu veya adı...")
        self.numune_search.textChanged.connect(lambda: self._load_numune(self.numune_search.text()))
        top2.addWidget(QLabel("Ara:")); top2.addWidget(self.numune_search, 1); top2.addStretch()
        v2.addLayout(top2)
        self.numune_table = QTableWidget()
        self._setup_table(self.numune_table)
        v2.addWidget(self.numune_table)
        self.tab_widget.addTab(tab_numune, "🔶 Numuneler")

        self.tab_widget.currentChanged.connect(self._on_tab_change)
        lay.addWidget(self.tab_widget)

        btn_row = QHBoxLayout()
        btn_add  = QPushButton("+ Yeni Ürün");  btn_add.clicked.connect(self._add)
        btn_edit = QPushButton("✎ Düzenle");     btn_edit.clicked.connect(self._edit)
        btn_del  = QPushButton("✕ Sil")
        btn_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        btn_del.clicked.connect(self._delete)
        self.btn_convert = QPushButton("🔄 Aktife Dönüştür")
        self.btn_convert.setStyleSheet(
            "background:#E65100;color:white;font-weight:bold;border-radius:4px;padding:6px 14px;")
        self.btn_convert.clicked.connect(self._convert_numune)
        self.btn_convert.setVisible(False)
        self.btn_excel = QPushButton("📥 Excel'den İçe Aktar")
        self.btn_excel.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;border-radius:4px;padding:6px 14px;")
        self.btn_excel.clicked.connect(self._import_excel)
        btn_close = QPushButton("Kapat"); btn_close.clicked.connect(self.accept)
        for b in (btn_add, btn_edit, btn_del, self.btn_convert, self.btn_excel):
            btn_row.addWidget(b)
        btn_row.addStretch(); btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

    def _setup_table(self, tbl):
        tbl.setColumnCount(10)
        tbl.setHorizontalHeaderLabels(
            ["Ürün Kodu", "Numune Kodu", "Açıklama", "Ürün Adı/Bilgisi", "Kompozisyon", "En", "Gramaj",
             "Fiyat", "Tedarikçi/Fason", "Durum"])
        tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.verticalHeader().setVisible(False)
        hdr = tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(False)
        hdr.setSectionsMovable(True)

    def _on_tab_change(self, idx):
        self.btn_convert.setVisible(idx == 1)
        self.btn_excel.setVisible(idx == 0)

    def _fill_table(self, tbl, rows, header_key):
        tbl.setSortingEnabled(False)
        tbl.setRowCount(len(rows))
        italic_font = QFont(); italic_font.setItalic(True)
        for i, r in enumerate(rows):
            is_numune = (r["product_status"] if "product_status" in r.keys() else "AKTİF") == "NUMUNE"
            numune_code = (r["numune_code"] if "numune_code" in r.keys() else "") or ""
            code_item = QTableWidgetItem(r["product_code"] or "")
            code_item.setData(Qt.ItemDataRole.UserRole, r["id"])
            tbl.setItem(i, 0, code_item)
            tbl.setItem(i, 1, QTableWidgetItem(numune_code))
            tbl.setItem(i, 2, QTableWidgetItem(r["reference_code"] or ""))
            tbl.setItem(i, 3, QTableWidgetItem(r["product_name"] or ""))
            tbl.setItem(i, 4, QTableWidgetItem(r["composition"] or ""))
            tbl.setItem(i, 5, QTableWidgetItem(r["width"] or ""))
            tbl.setItem(i, 6, QTableWidgetItem(r["gramaj"] or ""))
            price = r["price"] or 0
            price_item = _FireSortItem(f"{price:,.2f}" if price else "")
            price_item.setData(Qt.ItemDataRole.UserRole, price)
            tbl.setItem(i, 7, price_item)
            tbl.setItem(i, 8, QTableWidgetItem(r["supplier"] or ""))
            if is_numune:
                s = QTableWidgetItem("🔶 Numune")
                s.setForeground(QBrush(QColor("#E65100")))
            elif r["active"]:
                s = QTableWidgetItem("✅ Aktif")
                s.setForeground(QBrush(QColor("#2E7D32")))
            else:
                s = QTableWidgetItem("⛔ Pasif")
                s.setForeground(QBrush(QColor("#C62828")))
            tbl.setItem(i, 9, s)
            if is_numune:
                for col in range(tbl.columnCount()):
                    cell = tbl.item(i, col)
                    if cell:
                        cell.setFont(italic_font)
                        cell.setForeground(QBrush(QColor("#E65100")))
            for col in range(tbl.columnCount()):
                cell = tbl.item(i, col)
                if cell and cell.text():
                    cell.setToolTip(cell.text())
        tbl.setSortingEnabled(True)
        _wire_header_persistence(tbl, header_key)

    def _load(self, search=""):
        all_rows = db.get_all_products(search=search, active_only=False)
        rows = [r for r in all_rows
                if (r["product_status"] if "product_status" in r.keys() else "AKTİF") != "NUMUNE"]
        self._fill_table(self.table, rows, "products_header_v2")

    def _load_numune(self, search=""):
        rows = db.get_all_products(search=search, active_only=False, status_filter="NUMUNE")
        self._fill_table(self.numune_table, rows, "numune_header_v2")

    def _refresh_all(self):
        self._load(self.search.text())
        self._load_numune(self.numune_search.text())

    def _selected_id(self):
        tbl = self.numune_table if self.tab_widget.currentIndex() == 1 else self.table
        row = tbl.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Ürün seçin.")
            return None
        return tbl.item(row, 0).data(Qt.ItemDataRole.UserRole)

    def _product_dialog(self, p=None):
        if p is not None and not isinstance(p, dict):
            p = dict(p)  # sqlite3.Row → dict
        dlg = QDialog(self)
        dlg.setWindowTitle("Ürün" + (" Düzenle" if p else " Ekle"))
        dlg.setMinimumSize(860, 640)
        lay = QVBoxLayout(dlg)

        tabs = QTabWidget()

        # ── Sekme 1: Genel ────────────────────────────────────────
        tab1 = QWidget(); f1 = QFormLayout(tab1); f1.setSpacing(8)
        cur_status = (p.get("product_status") or "AKTİF") if p else "AKTİF"
        dlg.product_status = QComboBox()
        dlg.product_status.addItems(["AKTİF", "NUMUNE"])
        dlg.product_status.setCurrentText(cur_status)

        dlg.code     = QLineEdit(p["product_code"] if p else "")
        dlg.ref      = QLineEdit(p.get("reference_code","") if p else "")
        dlg.name     = QLineEdit(p.get("product_name","") if p else "")
        if not p:   # yeni ürün → otomatik benzersiz ad öner
            try:
                dlg.name.setText(db.generate_product_name())
            except Exception:
                pass
        # Adın yanına "yeni ad üret" düğmesi
        _name_w = QWidget(); _nh = QHBoxLayout(_name_w); _nh.setContentsMargins(0,0,0,0); _nh.setSpacing(6)
        _btn_gen = QPushButton("🎲"); _btn_gen.setFixedWidth(34)
        _btn_gen.setToolTip("Otomatik yeni ad üret")
        def _regen_name():
            try:
                dlg.name.setText(db.generate_product_name())
            except Exception:
                pass
        _btn_gen.clicked.connect(_regen_name)
        _nh.addWidget(dlg.name, 1); _nh.addWidget(_btn_gen)
        dlg.comp     = QLineEdit(p.get("composition","") if p else "")
        dlg.width    = QLineEdit(p.get("width","") if p else "")
        dlg.gramaj   = QLineEdit(p.get("gramaj","") if p else "")
        dlg.shrink   = QLineEdit(p.get("shrinkage","") if p else "")
        dlg.price    = QDoubleSpinBox(); dlg.price.setRange(0, 999999); dlg.price.setDecimals(2)
        dlg.price.setValue(float(p.get("price") or 0) if p else 0)
        dlg.price_currency = QComboBox(); dlg.price_currency.addItems(["USD", "EUR", "TL", "GBP"])
        dlg.price_currency.setCurrentText(p.get("price_currency","USD") if p else "USD")
        dlg.price_currency.setFixedWidth(72)
        _price_w = QWidget(); _ph = QHBoxLayout(_price_w); _ph.setContentsMargins(0,0,0,0); _ph.setSpacing(6)
        _ph.addWidget(dlg.price, 1); _ph.addWidget(dlg.price_currency)
        dlg.supplier = QLineEdit(p.get("supplier","") if p else "")

        def _on_status_change(txt):
            is_numune = txt == "NUMUNE"
            dlg.code.setReadOnly(is_numune)
            if is_numune:
                dlg.code.clear()
                dlg.code.setPlaceholderText("Otomatik atanacak (NMN-XXX)")
                dlg.code.setStyleSheet("color:#888; background:#FFF8E1;")
            else:
                dlg.code.setPlaceholderText("")
                dlg.code.setStyleSheet("")
        dlg.product_status.currentTextChanged.connect(_on_status_change)
        if cur_status == "NUMUNE":
            _on_status_change("NUMUNE")

        f1.addRow("Ürün Durumu:", dlg.product_status)
        f1.addRow("Ürün Kodu *:", dlg.code)
        f1.addRow("Ürün Adı/Bilgisi:", _name_w)
        f1.addRow("Kompozisyon:", dlg.comp)
        f1.addRow("En (cm):", dlg.width)
        f1.addRow("Gramaj (gr/m2):", dlg.gramaj)
        f1.addRow("Çekme %:", dlg.shrink)
        f1.addRow("Fiyat:", _price_w)
        f1.addRow("Tedarikçi/Fason:", dlg.supplier)
        f1.addRow("Açıklama:", dlg.ref)
        tabs.addTab(tab1, "Genel")

        # ── Sekme 2: Teknik Özellikler ───────────────────────────
        tab2 = QWidget(); f2 = QFormLayout(tab2); f2.setSpacing(8)
        def _le(key, ph=""): le = QLineEdit(p.get(key,"") if p else ""); le.setPlaceholderText(ph); return le
        def _dsb(key, suffix="", rng=(0,9999), dec=1):
            sb = QDoubleSpinBox(); sb.setRange(*rng); sb.setDecimals(dec)
            if suffix: sb.setSuffix(suffix)
            try: sb.setValue(float(p.get(key) or 0) if p else 0)
            except: pass
            return sb
        def _row2(le, extra_lbl, extra_w):
            """İplik adı + yan alan (uç/cm veya desen atım) satırı."""
            w = QWidget(); h = QHBoxLayout(w); h.setContentsMargins(0,0,0,0); h.setSpacing(6)
            h.addWidget(le, 1); h.addWidget(QLabel(extra_lbl)); h.addWidget(extra_w)
            return w

        # Çözgü iplikleri — iplik adı + tel/cm
        dlg.cozgu1    = _le("cozgu1", "İplik adı")
        dlg.cozgu1_uc = QDoubleSpinBox(); dlg.cozgu1_uc.setRange(0,500); dlg.cozgu1_uc.setDecimals(1); dlg.cozgu1_uc.setSuffix(" tel/cm"); dlg.cozgu1_uc.setFixedWidth(120)
        dlg.cozgu2    = _le("cozgu2", "İplik adı")
        dlg.cozgu2_uc = QDoubleSpinBox(); dlg.cozgu2_uc.setRange(0,500); dlg.cozgu2_uc.setDecimals(1); dlg.cozgu2_uc.setSuffix(" tel/cm"); dlg.cozgu2_uc.setFixedWidth(120)

        # Atkı iplikleri — iplik adı + atkı raporu
        dlg.atki1 = _le("atki1", "İplik adı"); dlg.atki1_atm = QSpinBox(); dlg.atki1_atm.setRange(0,256); dlg.atki1_atm.setSuffix(" atkı"); dlg.atki1_atm.setFixedWidth(100)
        dlg.atki2 = _le("atki2", "İplik adı"); dlg.atki2_atm = QSpinBox(); dlg.atki2_atm.setRange(0,256); dlg.atki2_atm.setSuffix(" atkı"); dlg.atki2_atm.setFixedWidth(100)
        dlg.atki3 = _le("atki3", "İplik adı"); dlg.atki3_atm = QSpinBox(); dlg.atki3_atm.setRange(0,256); dlg.atki3_atm.setSuffix(" atkı"); dlg.atki3_atm.setFixedWidth(100)
        dlg.atki4 = _le("atki4", "İplik adı"); dlg.atki4_atm = QSpinBox(); dlg.atki4_atm.setRange(0,256); dlg.atki4_atm.setSuffix(" atkı"); dlg.atki4_atm.setFixedWidth(100)

        dlg.dokuma_tipi   = _le("dokuma_tipi"); dlg.tarak_no    = _le("tarak_no")
        dlg.orgu_desen    = _le("orgu_desen")
        dlg.teknik_aciklama = QLineEdit(p.get("teknik_aciklama","") if p else "")
        dlg.tarak_eni_sb  = _dsb("tarak_eni",    " cm",      (0, 500), 1)
        dlg.cozgu_sik_sb  = _dsb("cozgu_sikligi"," tel/cm",  (0, 500), 1)
        dlg.atki_sik_sb   = _dsb("atki_sikligi", " atkı/cm", (0, 200), 1)

        f2.addRow("Çözgü 1:", _row2(dlg.cozgu1, "Tel/cm:", dlg.cozgu1_uc))
        f2.addRow("Çözgü 2:", _row2(dlg.cozgu2, "Tel/cm:", dlg.cozgu2_uc))
        f2.addRow("Atkı 1:",  _row2(dlg.atki1,  "Atkı Raporu:", dlg.atki1_atm))
        f2.addRow("Atkı 2:",  _row2(dlg.atki2,  "Atkı Raporu:", dlg.atki2_atm))
        f2.addRow("Atkı 3:",  _row2(dlg.atki3,  "Atkı Raporu:", dlg.atki3_atm))
        f2.addRow("Atkı 4:",  _row2(dlg.atki4,  "Atkı Raporu:", dlg.atki4_atm))
        for w, lbl in [
            (dlg.dokuma_tipi,     "Dokuma Tipi:"),
            (dlg.tarak_no,        "Tarak No:"),
            (dlg.tarak_eni_sb,    "Tarak Eni (cm):"),
            (dlg.cozgu_sik_sb,    "Çözgü Sıklığı (tel/cm):"),
            (dlg.atki_sik_sb,     "Atkı Sıklığı (atkı/cm):"),
            (dlg.orgu_desen,      "Örgü/Desen:"),
            (dlg.teknik_aciklama, "Açıklama:"),
        ]:
            f2.addRow(lbl, w)
        tabs.addTab(tab2, "Teknik Özellikler")

        # ── Sekme 3: Maliyet Hesaplama ───────────────────────────
        dlg.maliyet_w = MaliyetWidget()
        if p:
            dlg.maliyet_w.load(p)
        tabs.addTab(dlg.maliyet_w, "Maliyet Hesaplama")

        # ── Teknik ↔ Maliyet tam senkronizasyon ──────────────────
        mw = dlg.maliyet_w

        # Maliyet yüklendikten sonra uç/cm ve atm değerlerini teknik sekmeye yansıt
        dlg.cozgu1_uc.setValue(mw.cozgu_rows[0]["uc_cm"].value())
        dlg.cozgu2_uc.setValue(mw.cozgu_rows[1]["uc_cm"].value())
        for i, sb in enumerate([dlg.atki1_atm, dlg.atki2_atm, dlg.atki3_atm, dlg.atki4_atm]):
            sb.setValue(mw.atki_rows[i]["atm"].value())

        def _t2m_num(*_):
            """Teknik sayısal alanlar → Maliyet spinbox'ları."""
            for src, dst in [(dlg.tarak_eni_sb, mw.tarak_eni),
                             (dlg.cozgu_sik_sb, mw.cozgu_sik),
                             (dlg.atki_sik_sb,  mw.atki_sik),
                             (dlg.cozgu1_uc,    mw.cozgu_rows[0]["uc_cm"]),
                             (dlg.cozgu2_uc,    mw.cozgu_rows[1]["uc_cm"])]:
                dst.blockSignals(True); dst.setValue(src.value()); dst.blockSignals(False)
            for i, sb in enumerate([dlg.atki1_atm, dlg.atki2_atm, dlg.atki3_atm, dlg.atki4_atm]):
                mw.atki_rows[i]["atm"].blockSignals(True)
                mw.atki_rows[i]["atm"].setValue(sb.value())
                mw.atki_rows[i]["atm"].blockSignals(False)
            mw._recalc()

        def _t2m_txt(row_idx, is_cozgu):
            """Teknik metin alanları → Maliyet iplik adı alanları."""
            def _inner(text):
                rows = mw.cozgu_rows if is_cozgu else mw.atki_rows
                rows[row_idx]["ad"].blockSignals(True)
                rows[row_idx]["ad"].setText(text)
                rows[row_idx]["ad"].blockSignals(False)
            return _inner

        def _m2t_num(*_):
            """Maliyet spinbox'ları → Teknik sayısal alanlar."""
            for src, dst in [(mw.tarak_eni, dlg.tarak_eni_sb),
                             (mw.cozgu_sik, dlg.cozgu_sik_sb),
                             (mw.atki_sik,  dlg.atki_sik_sb),
                             (mw.cozgu_rows[0]["uc_cm"], dlg.cozgu1_uc),
                             (mw.cozgu_rows[1]["uc_cm"], dlg.cozgu2_uc)]:
                dst.blockSignals(True); dst.setValue(src.value()); dst.blockSignals(False)
            for i, sb in enumerate([dlg.atki1_atm, dlg.atki2_atm, dlg.atki3_atm, dlg.atki4_atm]):
                sb.blockSignals(True); sb.setValue(mw.atki_rows[i]["atm"].value()); sb.blockSignals(False)

        def _m2t_txt(row_idx, is_cozgu):
            """Maliyet iplik adı → Teknik metin alanları."""
            teknik_fields = [dlg.cozgu1, dlg.cozgu2] if is_cozgu else [dlg.atki1, dlg.atki2, dlg.atki3, dlg.atki4]
            def _inner(text):
                teknik_fields[row_idx].blockSignals(True)
                teknik_fields[row_idx].setText(text)
                teknik_fields[row_idx].blockSignals(False)
            return _inner

        # Sayısal ↔ sayısal bağlantılar
        for w in (dlg.tarak_eni_sb, dlg.cozgu_sik_sb, dlg.atki_sik_sb, dlg.cozgu1_uc, dlg.cozgu2_uc):
            w.valueChanged.connect(_t2m_num)
        for sb in (dlg.atki1_atm, dlg.atki2_atm, dlg.atki3_atm, dlg.atki4_atm):
            sb.valueChanged.connect(_t2m_num)
        for w in (mw.tarak_eni, mw.cozgu_sik, mw.atki_sik,
                  mw.cozgu_rows[0]["uc_cm"], mw.cozgu_rows[1]["uc_cm"]):
            w.valueChanged.connect(_m2t_num)
        for row in mw.atki_rows:
            row["atm"].valueChanged.connect(_m2t_num)

        # Metin ↔ metin bağlantıları (çözgü adları)
        dlg.cozgu1.textChanged.connect(_t2m_txt(0, True))
        dlg.cozgu2.textChanged.connect(_t2m_txt(1, True))
        mw.cozgu_rows[0]["ad"].textChanged.connect(_m2t_txt(0, True))
        mw.cozgu_rows[1]["ad"].textChanged.connect(_m2t_txt(1, True))
        # Metin ↔ metin bağlantıları (atkı adları)
        for i, le in enumerate([dlg.atki1, dlg.atki2, dlg.atki3, dlg.atki4]):
            le.textChanged.connect(_t2m_txt(i, False))
        for i, row in enumerate(mw.atki_rows):
            row["ad"].textChanged.connect(_m2t_txt(i, False))

        # ── Sekme 4: Armür Desen ─────────────────────────────────
        dlg.armur_w = ArmurManagerWidget()
        tabs.addTab(dlg.armur_w, "Armür Desen")

        dlg.jakar_w = JakarDesenWidget()
        if p:
            dlg.jakar_w.load(p)
        tabs.addTab(dlg.jakar_w, "Jakar Desen")

        lay.addWidget(tabs)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        def _try_accept():
            is_numune = dlg.product_status.currentText() == "NUMUNE"
            if not is_numune and not dlg.code.text().strip():
                QMessageBox.warning(dlg, "Eksik Alan", "Ürün kodu zorunludur.\nLütfen Genel sekmesinde ürün kodunu girin.")
                tabs.setCurrentIndex(0)
                dlg.code.setFocus()
                return
            dlg.accept()
        btns.button(QDialogButtonBox.StandardButton.Ok).clicked.connect(_try_accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        return dlg

    def _collect_product(self, dlg, pid=None, active=1):
        """Dialog alanlarından ürün verisini toplar ve db'ye yazar."""
        product_status = dlg.product_status.currentText()
        code = dlg.code.text().strip()
        if product_status != "NUMUNE" and not code:
            QMessageBox.warning(self, "Hata", "Ürün kodu zorunlu!")
            return False
        mw = dlg.maliyet_w
        kwargs = dict(
            product_name=dlg.name.text(), composition=dlg.comp.text(),
            width=dlg.width.text(), gramaj=dlg.gramaj.text(),
            shrinkage=dlg.shrink.text(), price=dlg.price.value(),
            supplier=dlg.supplier.text(), reference_code=dlg.ref.text(),
            cozgu1=dlg.cozgu1.text(), cozgu2=dlg.cozgu2.text(),
            atki1=dlg.atki1.text(), atki2=dlg.atki2.text(),
            atki3=dlg.atki3.text(), atki4=dlg.atki4.text(),
            dokuma_tipi=dlg.dokuma_tipi.text(),
            tarak_no=dlg.tarak_no.text(),
            orgu_desen=dlg.orgu_desen.text(),
            cozgu_sikligi=str(dlg.cozgu_sik_sb.value()),
            tarak_eni=str(dlg.tarak_eni_sb.value()),
            atki_sikligi=str(dlg.atki_sik_sb.value()),
            maliyet_json=mw.get_maliyet_json(),
            teknik_aciklama=dlg.teknik_aciklama.text(),
            price_currency=dlg.price_currency.currentText(),
            jakar_desen_ad=dlg.jakar_w.get_jc5_ad(),
            jakar_desen_data=dlg.jakar_w.get_jc5_b64(),
            jakar_jpeg_ad=dlg.jakar_w.get_jpeg_ad(),
            jakar_jpeg_data=dlg.jakar_w.get_jpeg_b64(),
            product_status=product_status,
        )
        try:
            if pid:
                db.update_product(pid, code, active=active, **kwargs)
            else:
                db.add_product(code, **kwargs)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kaydedilemedi:\n{e}")
            return False
        return True

    def _add(self):
        dlg = self._product_dialog()
        if not dlg.exec():
            return
        # Ürün adı onayı: onaylanmazsa reddedilen ad kara listeye alınır ve yeni ad önerilir
        cd = _ProductNameConfirmDialog(self, dlg.name.text().strip())
        if not cd.exec():
            return   # iptal → kayıt yok
        dlg.name.setText(cd.result_name())
        if self._collect_product(dlg):
            self._refresh_all()

    def _edit(self):
        pid = self._selected_id()
        if not pid: return
        p = db.get_product(pid)
        dlg = self._product_dialog(p)
        if dlg.exec() and self._collect_product(dlg, pid=pid, active=p["active"]):
            self._refresh_all()

    def _delete(self):
        pid = self._selected_id()
        if not pid: return
        if QMessageBox.question(self,"Sil","Ürün silinsin mi?",
                QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_product(pid); self._refresh_all()

    def _convert_numune(self):
        pid = self._selected_id()
        if not pid: return
        p = db.get_product(pid)
        if not p:
            return
        if isinstance(p, dict):
            p_status = p.get("product_status") or "AKTİF"
            p_code   = p.get("product_code") or ""
        else:
            p = dict(p)
            p_status = p.get("product_status") or "AKTİF"
            p_code   = p.get("product_code") or ""
        if p_status != "NUMUNE":
            QMessageBox.information(self, "Bilgi", f"'{p_code}' ürünü zaten aktif.\nSadece NUMUNE ürünler dönüştürülebilir.")
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("Numune → Aktif Dönüşüm")
        dlg.setMinimumWidth(360)
        lay = QVBoxLayout(dlg)
        form = QFormLayout()
        lbl_old = QLabel(p_code)
        lbl_old.setStyleSheet("color:#E65100;font-weight:bold;")
        form.addRow("Numune Kodu:", lbl_old)
        new_code_edit = QLineEdit()
        new_code_edit.setPlaceholderText("Yeni ürün kodu (ör. BK-1234)")
        form.addRow("Yeni Ürün Kodu *:", new_code_edit)
        lay.addLayout(form)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("✅ Aktife Dönüştür")
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        if not dlg.exec():
            return
        new_code = new_code_edit.text().strip().upper()
        if not new_code:
            QMessageBox.warning(self, "Eksik Alan", "Yeni ürün kodu boş olamaz.")
            return
        try:
            db.convert_numune_to_aktif(pid, new_code)
            self._refresh_all()
            QMessageBox.information(self, "Dönüştürüldü",
                f"Numune ürün aktife dönüştürüldü:\n{p_code}  →  {new_code}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dönüşüm başarısız:\n{e}")

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self,"Ürün Excel Dosyası","","Excel (*.xlsx *.xls)")
        if not path: return
        try:
            dlg = ExcelColumnMapDialog(path, PRODUCT_FIELDS, PRODUCT_AUTO, self)
            if not dlg.exec():
                return
            records = dlg.records
            if not records:
                return QMessageBox.warning(self,"Hata","Eşleştirilen sütunda geçerli veri bulunamadı.")
            n = db.import_products_bulk(records)
            self._refresh_all()
            QMessageBox.information(self,"Başarılı",f"{n} ürün içe aktarıldı/güncellendi.")
        except Exception as e:
            QMessageBox.critical(self,"Hata",f"İçe aktarma hatası:\n{e}")



class LocationManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Raf / Lokasyon Tanımlamaları")
        self.setMinimumSize(640, 480)
        self._build_ui()
        self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Tablo
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Lokasyon Adı", "Grup", "Açıklama", "Durum", ""])
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        btn_add  = QPushButton("+ Yeni Lokasyon"); btn_add.clicked.connect(self._add)
        btn_edit = QPushButton("✎ Düzenle");       btn_edit.clicked.connect(self._edit)
        btn_del  = QPushButton("✕ Sil")
        btn_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        btn_del.clicked.connect(self._delete)
        btn_sync = QPushButton("🔄 Stoktan Senkronize Et")
        btn_sync.setStyleSheet("background:#37474F;color:white;border-radius:4px;padding:6px 14px;")
        btn_sync.clicked.connect(self._sync)
        btn_close = QPushButton("Kapat"); btn_close.clicked.connect(self.accept)
        for b in (btn_add, btn_edit, btn_del, btn_sync):
            btn_row.addWidget(b)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

    def _load(self):
        rows = db.get_all_locations()
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            name_item = QTableWidgetItem(r["name"])
            name_item.setData(Qt.ItemDataRole.UserRole, r["id"])
            self.table.setItem(i, 0, name_item)
            grp_item = QTableWidgetItem(r["group_name"])
            grp_item.setForeground(QBrush(QColor("#545454") if r["group_name"] == "DEPO" else QColor("#545454")))
            grp_item.setFont(QFont("", -1, QFont.Weight.Bold))
            self.table.setItem(i, 1, grp_item)
            self.table.setItem(i, 2, QTableWidgetItem(r["description"] or ""))
            status = QTableWidgetItem("✅ Aktif" if r["active"] else "⛔ Pasif")
            status.setForeground(QBrush(QColor("#2E7D32") if r["active"] else QColor("#C62828")))
            self.table.setItem(i, 3, status)
        self.table.setSortingEnabled(True)   # başlığa tıklayınca sıralar

    def _selected_id(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Bir lokasyon seçin.")
            return None
        return self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)

    def _add(self):
        dlg = self._loc_dialog()
        if dlg.exec():
            name, grp, desc = dlg._get()
            if not name:
                return QMessageBox.warning(self, "Hata", "Lokasyon adı boş olamaz!")
            try:
                db.add_location(name, grp, desc)
                self._load()
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Eklenemedi:\n{e}")

    def _edit(self):
        lid = self._selected_id()
        if not lid: return
        rows = db.get_all_locations()
        loc = next((r for r in rows if r["id"] == lid), None)
        if not loc: return
        dlg = self._loc_dialog(loc)
        if dlg.exec():
            name, grp, desc = dlg._get()
            active = dlg.active_cb.isChecked()
            db.update_location(lid, name, grp, desc, active)
            self._load()

    def _delete(self):
        lid = self._selected_id()
        if not lid: return
        loc = next((r for r in db.get_all_locations() if r["id"] == lid), None)
        if not loc: return

        # Lokasyonda aktif stok varsa silme — stoklar hiçbir şekilde silinmez
        fabrics = db.get_all_fabrics(location=loc["name"])
        if fabrics:
            toplam_mt = sum((f["meter"] or 0) for f in fabrics)
            QMessageBox.warning(
                self, "Lokasyon Silinemez",
                f"<b>{loc['name']}</b> lokasyonunda <b>{len(fabrics)} stok kaydı</b> var "
                f"({toplam_mt:,.0f} mt).<br><br>"
                f"Stoklar silinemez — önce bu kayıtları çıkış/transfer ile "
                f"<b>başka lokasyona taşıyın</b>, sonra lokasyonu silebilirsiniz.")
            return

        if QMessageBox.question(self, "Sil", f"<b>{loc['name']}</b> lokasyonu silinsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            try:
                db.delete_location(lid)
            except Exception as e:
                QMessageBox.critical(self, "Silinemedi", str(e))
            self._load()

    def _sync(self):
        db.sync_locations()
        self._load()
        QMessageBox.information(self, "Tamam", "Stok lokasyonları senkronize edildi.")

    def _loc_dialog(self, loc=None):
        dlg = QDialog(self)
        dlg.setWindowTitle("Lokasyon Ekle" if not loc else "Lokasyon Düzenle")
        dlg.setMinimumWidth(340)
        lay = QVBoxLayout(dlg)
        form = QFormLayout(); form.setSpacing(8)
        dlg.name_edit = QLineEdit(loc["name"] if loc else "")
        dlg.grp_cb = QComboBox()
        dlg.grp_cb.addItems(["DEPO", "DIŞ DEPO"])
        if loc:
            idx = dlg.grp_cb.findText(loc["group_name"])
            if idx >= 0: dlg.grp_cb.setCurrentIndex(idx)
        dlg.desc_edit = QLineEdit(loc["description"] if loc else "")
        dlg.active_cb = QCheckBox("Aktif")
        dlg.active_cb.setChecked(loc["active"] if loc else True)
        form.addRow("Lokasyon Adı *:", dlg.name_edit)
        form.addRow("Grup:", dlg.grp_cb)
        form.addRow("Açıklama:", dlg.desc_edit)
        form.addRow("", dlg.active_cb)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        dlg._get = lambda: (dlg.name_edit.text().strip(), dlg.grp_cb.currentText(), dlg.desc_edit.text().strip())
        return dlg


class EmailSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("E-posta Rapor Ayarları")
        self.setMinimumWidth(480)
        self._build_ui()
        self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(10)

        # Gmail notu
        note = QLabel(
            "<b>Gmail kullanıyorsanız:</b> Normal şifre değil, "
            "<a href='https://myaccount.google.com/apppasswords'>Google Uygulama Şifresi</a> gereklidir.<br>"
            "Google Hesabım → Güvenlik → 2 Adımlı Doğrulama → Uygulama Şifreleri"
        )
        note.setWordWrap(True)
        note.setOpenExternalLinks(True)
        note.setStyleSheet("background:#FFF8E1;padding:10px;border-radius:6px;font-size:12px;")
        lay.addWidget(note)

        form = QFormLayout(); form.setSpacing(8)

        self.smtp_host = QLineEdit()
        self.smtp_port = QLineEdit()
        self.smtp_user = QLineEdit(); self.smtp_user.setPlaceholderText("gönderen@gmail.com")
        self.smtp_pass = QLineEdit(); self.smtp_pass.setEchoMode(QLineEdit.EchoMode.Password)
        self.smtp_pass.setPlaceholderText("Uygulama şifresi (16 karakter)")
        self.from_addr = QLineEdit(); self.from_addr.setPlaceholderText("gönderen@gmail.com (boş = smtp_user)")
        self.to_addrs  = QLineEdit(); self.to_addrs.setPlaceholderText("alici@gmail.com, diger@gmail.com")

        self.send_hour = QComboBox()
        for h in range(24):
            self.send_hour.addItem(f"{h:02d}:00", h)

        self.send_enabled = QCheckBox("Her gün otomatik rapor gönder")

        form.addRow("SMTP Sunucu:", self.smtp_host)
        form.addRow("SMTP Port:", self.smtp_port)
        form.addRow("Kullanıcı Adı:", self.smtp_user)
        form.addRow("Şifre:", self.smtp_pass)
        form.addRow("Gönderen:", self.from_addr)
        form.addRow("Alıcılar:", self.to_addrs)
        form.addRow("Gönderim Saati:", self.send_hour)
        form.addRow("", self.send_enabled)
        lay.addLayout(form)

        btn_row = QHBoxLayout()
        btn_test = QPushButton("📤 Şimdi Test Gönder")
        btn_test.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;border-radius:4px;padding:7px 14px;")
        btn_test.clicked.connect(self._test_send)
        btn_save = QPushButton("Kaydet")
        btn_cancel = QPushButton("İptal")
        btn_cancel.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:7px 14px;")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save)
        btn_row.addWidget(btn_test); btn_row.addStretch()
        btn_row.addWidget(btn_save); btn_row.addWidget(btn_cancel)
        lay.addLayout(btn_row)

    def _load(self):
        import email_report as er
        cfg = er.get_email_config()
        self.smtp_host.setText(cfg["smtp_host"])
        self.smtp_port.setText(str(cfg["smtp_port"]))
        self.smtp_user.setText(cfg["smtp_user"])
        self.smtp_pass.setText(cfg["smtp_pass"])
        self.from_addr.setText(cfg["from_addr"])
        self.to_addrs.setText(cfg["to_addrs"])
        idx = self.send_hour.findData(cfg["send_hour"])
        if idx >= 0: self.send_hour.setCurrentIndex(idx)
        self.send_enabled.setChecked(cfg["send_enabled"])

    def _save(self):
        import email_report as er
        er.save_email_config(
            smtp_host    = self.smtp_host.text().strip(),
            smtp_port    = self.smtp_port.text().strip() or "587",
            smtp_user    = self.smtp_user.text().strip(),
            smtp_pass    = self.smtp_pass.text(),
            from_addr    = self.from_addr.text().strip(),
            to_addrs     = self.to_addrs.text().strip(),
            send_hour    = self.send_hour.currentData(),
            send_enabled = self.send_enabled.isChecked(),
        )
        QMessageBox.information(self, "Kaydedildi", "E-posta ayarları kaydedildi.")
        self.accept()

    def _test_send(self):
        self._save_silent()
        import email_report as er
        try:
            n = er.send_report(test=True)
            QMessageBox.information(self, "Gönderildi ✓",
                f"Test raporu {n} alıcıya gönderildi!\nGelen kutunuzu kontrol edin.")
        except Exception as e:
            QMessageBox.critical(self, "Gönderilemedi", str(e))

    def _save_silent(self):
        import email_report as er
        er.save_email_config(
            smtp_host    = self.smtp_host.text().strip(),
            smtp_port    = self.smtp_port.text().strip() or "587",
            smtp_user    = self.smtp_user.text().strip(),
            smtp_pass    = self.smtp_pass.text(),
            from_addr    = self.from_addr.text().strip(),
            to_addrs     = self.to_addrs.text().strip(),
            send_hour    = self.send_hour.currentData(),
            send_enabled = self.send_enabled.isChecked(),
        )


class CompanySettingsDialog(QDialog):
    """Sipariş formlarında kullanılan firma/banka bilgileri ve varsayılan
    sözleşme şartnamesi metni — settings tablosunda saklanır."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Şirket / Banka Ayarları")
        self.setMinimumSize(560, 600)
        self._build_ui()
        self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(10)

        form = QFormLayout(); form.setSpacing(8)
        self.company_name = QLineEdit()
        self.company_address = QLineEdit()
        self.company_phone = QLineEdit()
        self.company_tax = QLineEdit()
        self.company_origin = QLineEdit()
        self.company_website = QLineEdit()
        self.company_email_info = QLineEdit()
        self.company_email_planlama = QLineEdit()
        form.addRow("Firma Unvanı:", self.company_name)
        form.addRow("Adres:", self.company_address)
        form.addRow("Telefon:", self.company_phone)
        form.addRow("Vergi Numarası:", self.company_tax)
        form.addRow("Menşei:", self.company_origin)
        form.addRow("Web Sitesi:", self.company_website)
        form.addRow("E-posta (Info):", self.company_email_info)
        form.addRow("E-posta (Planlama):", self.company_email_planlama)
        lay.addLayout(form)

        lay.addWidget(QLabel("Banka Bilgileri:"))
        self.bank_info = QTextEdit()
        self.bank_info.setMinimumHeight(140)
        lay.addWidget(self.bank_info)

        lay.addWidget(QLabel("Varsayılan Sözleşme Şartnamesi:"))
        self.contract_template = QTextEdit()
        self.contract_template.setMinimumHeight(220)
        lay.addWidget(self.contract_template)

        btn_row = QHBoxLayout()
        btn_save = QPushButton("Kaydet")
        btn_save.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;border-radius:4px;padding:7px 14px;")
        btn_cancel = QPushButton("İptal")
        btn_cancel.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:7px 14px;")
        btn_cancel.clicked.connect(self.reject)
        btn_save.clicked.connect(self._save)
        btn_row.addStretch()
        btn_row.addWidget(btn_save); btn_row.addWidget(btn_cancel)
        lay.addLayout(btn_row)

    def _load(self):
        cfg = db.get_company_settings() or {}
        self.company_name.setText(cfg.get("name", ""))
        self.company_address.setText(cfg.get("address", ""))
        self.company_phone.setText(cfg.get("phone", ""))
        self.company_tax.setText(cfg.get("tax", ""))
        self.company_origin.setText(cfg.get("origin", ""))
        self.company_website.setText(cfg.get("website", ""))
        self.company_email_info.setText(cfg.get("email_info", ""))
        self.company_email_planlama.setText(cfg.get("email_planlama", ""))
        self.bank_info.setPlainText(cfg.get("bank_info", ""))
        self.contract_template.setPlainText(cfg.get("contract_template", ""))

    def _save(self):
        db.save_company_settings(
            name=self.company_name.text().strip(),
            address=self.company_address.text().strip(),
            phone=self.company_phone.text().strip(),
            tax=self.company_tax.text().strip(),
            origin=self.company_origin.text().strip(),
            website=self.company_website.text().strip(),
            email_info=self.company_email_info.text().strip(),
            email_planlama=self.company_email_planlama.text().strip(),
            bank_info=self.bank_info.toPlainText(),
            contract_template=self.contract_template.toPlainText(),
        )
        QMessageBox.information(self, "Kaydedildi", "Şirket / banka ayarları kaydedildi.")
        self.accept()


class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Giriş Yap")
        self.setMinimumWidth(360)
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.WindowTitleHint)
        self._build_ui()
        self._load_server_setting()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        if _os.path.exists(LOGO_PATH):
            logo = QLabel()
            pix  = QPixmap(LOGO_PATH)
            logo.setPixmap(pix.scaledToWidth(220, Qt.TransformationMode.SmoothTransformation))
            logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(logo)

        title = QLabel("Depo Takip Sistemine Giriş")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("font-size:15px; font-weight:bold; color:#545454;")
        layout.addWidget(title)

        # ── Bağlantı modu ──────────────────────────────────────
        conn_box = QGroupBox("Bağlantı")
        conn_lay = QVBoxLayout(conn_box)
        conn_lay.setSpacing(6)

        self.rb_local  = QPushButton("💻 Bu Bilgisayar (yerel)")
        self.rb_remote = QPushButton("🌐 Sunucuya Bağlan")
        self.rb_local.setCheckable(True);  self.rb_local.setChecked(True)
        self.rb_remote.setCheckable(True); self.rb_remote.setChecked(False)
        self.rb_local.setMinimumHeight(36)
        self.rb_remote.setMinimumHeight(36)
        self.rb_local.clicked.connect(lambda: self._set_mode("local"))
        self.rb_remote.clicked.connect(lambda: self._set_mode("remote"))

        self.server_url = QLineEdit()
        self.server_url.setPlaceholderText("http://192.168.1.x:5060")
        self.server_url.setVisible(False)

        conn_lay.addWidget(self.rb_local)
        conn_lay.addWidget(self.rb_remote)
        conn_lay.addWidget(self.server_url)
        layout.addWidget(conn_box)
        # Başlangıç stilleri — server_url tanımlandıktan sonra
        self._set_mode("local")

        # ── Giriş formu ────────────────────────────────────────
        form = QFormLayout(); form.setSpacing(8)
        self.username = QLineEdit(); self.username.setPlaceholderText("Kullanıcı adı")
        self.password = QLineEdit(); self.password.setEchoMode(QLineEdit.EchoMode.Password)
        self.password.setPlaceholderText("Şifre")
        form.addRow("Kullanıcı:", self.username)
        form.addRow("Şifre:",     self.password)
        layout.addLayout(form)

        self.err_label = QLabel("")
        self.err_label.setStyleSheet("color:#C62828; font-size:12px;")
        self.err_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.err_label)

        btn = QPushButton("Giriş Yap")
        btn.setMinimumHeight(38)
        btn.clicked.connect(self._login)
        layout.addWidget(btn)

        self.password.returnPressed.connect(self._login)
        self.username.returnPressed.connect(lambda: self.password.setFocus())

    def _set_mode(self, mode):
        self.rb_local.setChecked(mode == "local")
        self.rb_remote.setChecked(mode == "remote")
        self.server_url.setVisible(mode == "remote")
        ACTIVE   = "background:#545454; color:white; font-weight:bold; text-align:left; padding:8px 12px; border-radius:4px;"
        INACTIVE = "background:white; color:#545454; font-weight:normal; text-align:left; padding:8px 12px; border-radius:4px; border:1px solid #BDBDBD;"
        if mode == "local":
            self.rb_local.setStyleSheet(ACTIVE)
            self.rb_remote.setStyleSheet(INACTIVE)
        else:
            self.rb_remote.setStyleSheet(ACTIVE)
            self.rb_local.setStyleSheet(INACTIVE)

    def _load_server_setting(self):
        """Daha önce kullanılan sunucu adresini yükle."""
        try:
            import json
            cfg_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "config.json")
            cfg = json.load(open(cfg_path)) if _os.path.exists(cfg_path) else {}
            saved_url = cfg.get("server_url", "")
            if saved_url:
                self.server_url.setText(saved_url)
                self._set_mode("remote")
        except Exception:
            pass

    def _save_server_setting(self, url):
        try:
            import json
            cfg_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "config.json")
            cfg = json.load(open(cfg_path)) if _os.path.exists(cfg_path) else {}
            cfg["server_url"] = url
            with open(cfg_path, "w") as f: json.dump(cfg, f, indent=2)
        except Exception:
            pass

    def _login(self):
        global CURRENT_USER, CONNECTION_MODE
        self.err_label.setText("")

        if self.rb_remote.isChecked():
            # ── Sunucuya bağlan ─────────────────────────────────
            url = self.server_url.text().strip()
            if not url:
                self.err_label.setText("Sunucu adresi giriniz!")
                return
            import api_client
            api_client.configure(url)
            self.err_label.setText("Bağlanıyor...")
            QApplication.processEvents()
            if not api_client.ping():
                self.err_label.setText(f"Sunucuya bağlanılamadı:\n{url}\n\nAdres ve sunucunun açık olduğundan emin olun.")
                return
            try:
                user = api_client.login(self.username.text().strip(), self.password.text())
                CONNECTION_MODE = "remote"
                CURRENT_USER.update(user)
                # db'yi proxy'ye çevir — __main__ modülünü kullan (import main değil!)
                import sys
                _mod = sys.modules['__main__']
                _mod.db = _mod._db
                self._save_server_setting(url)
                self.accept()
            except Exception as e:
                self.err_label.setText(str(e))
                self.password.clear(); self.password.setFocus()
        else:
            # ── Yerel bağlantı ──────────────────────────────────
            user = db.authenticate(self.username.text().strip(), self.password.text())
            if user:
                CONNECTION_MODE = "local"
                CURRENT_USER.update(user)
                self._save_server_setting("")  # yerel ise kayıtlı URL'yi temizle
                self.accept()
            else:
                self.err_label.setText("Kullanıcı adı veya şifre hatalı!")
                self.password.clear(); self.password.setFocus()


class UserManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Kullanıcı Yönetimi")
        self.setMinimumSize(580, 400)
        self._build_ui()
        self._load()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Tablo
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Kullanıcı Adı", "Ad Soyad", "Rol", "Durum", "Kayıt Tarihi"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        # Butonlar
        btn_row = QHBoxLayout()
        btn_add  = QPushButton("+ Yeni Kullanıcı"); btn_add.clicked.connect(self._add)
        btn_edit = QPushButton("✏ Düzenle");         btn_edit.clicked.connect(self._edit)
        btn_pw   = QPushButton("🔑 Şifre Değiştir"); btn_pw.clicked.connect(self._change_pw)
        btn_tog  = QPushButton("⏸ Aktif/Pasif");    btn_tog.clicked.connect(self._toggle)
        btn_del  = QPushButton("✕ Sil")
        btn_del.setStyleSheet("background:#757575; color:white; border-radius:4px; padding:6px 14px;")
        btn_del.clicked.connect(self._delete)
        for b in (btn_add, btn_edit, btn_pw, btn_tog, btn_del):
            btn_row.addWidget(b)
        btn_row.addStretch()
        layout.addLayout(btn_row)

    def _load(self):
        users = db.get_all_users()
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(users))
        for i, u in enumerate(users):
            uname_item = QTableWidgetItem(u["username"])
            uname_item.setData(Qt.ItemDataRole.UserRole, u["id"])
            self.table.setItem(i, 0, uname_item)
            self.table.setItem(i, 1, QTableWidgetItem(u["full_name"] or ""))
            self.table.setItem(i, 2, QTableWidgetItem(u["role"]))
            status = QTableWidgetItem("✅ Aktif" if u["active"] else "⛔ Pasif")
            status.setForeground(QBrush(QColor("#2E7D32") if u["active"] else QColor("#C62828")))
            self.table.setItem(i, 3, status)
            self.table.setItem(i, 4, QTableWidgetItem(str(u["created_at"])[:10]))
        self.table.setSortingEnabled(True)   # başlığa tıklayınca sıralar

    def _selected_id(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Kullanıcı seçin.")
            return None
        return self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)

    def _add(self):
        dlg = QDialog(self); dlg.setWindowTitle("Yeni Kullanıcı"); dlg.setMinimumWidth(320)
        lay = QVBoxLayout(dlg); form = QFormLayout(); form.setSpacing(8)
        uname = QLineEdit(); fname = QLineEdit()
        pw1 = QLineEdit(); pw1.setEchoMode(QLineEdit.EchoMode.Password)
        pw2 = QLineEdit(); pw2.setEchoMode(QLineEdit.EchoMode.Password)
        role_cb = QComboBox()
        role_cb.addItems(["kullanici", "admin", "planlama", "satışçı", "depo-sevkiyat",
                          "muhasebe", "fason-takip", "kartela", "kalite-kontrol"])
        form.addRow("Kullanıcı Adı *:", uname)
        form.addRow("Ad Soyad:", fname)
        form.addRow("Şifre *:", pw1)
        form.addRow("Şifre (tekrar):", pw2)
        form.addRow("Rol:", role_cb)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        if dlg.exec():
            if not uname.text().strip():
                return QMessageBox.warning(self, "Hata", "Kullanıcı adı boş olamaz!")
            if pw1.text() != pw2.text():
                return QMessageBox.warning(self, "Hata", "Şifreler eşleşmiyor!")
            if len(pw1.text()) < 4:
                return QMessageBox.warning(self, "Hata", "Şifre en az 4 karakter olmalı!")
            try:
                db.add_user(uname.text(), fname.text(), pw1.text(), role_cb.currentText())
                self._load()
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kullanıcı eklenemedi:\n{e}")

    def _edit(self):
        uid = self._selected_id()
        if not uid:
            return
        row = self.table.currentRow()
        cur_fname = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        cur_role  = self.table.item(row, 2).text() if self.table.item(row, 2) else ""
        cur_uname = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
        dlg = QDialog(self); dlg.setWindowTitle(f"Kullanıcı Düzenle — {cur_uname}")
        dlg.setMinimumWidth(320)
        lay = QVBoxLayout(dlg); form = QFormLayout(); form.setSpacing(8)
        fname = QLineEdit(cur_fname)
        role_cb = QComboBox()
        role_cb.addItems(["kullanici", "admin", "planlama", "satışçı", "depo-sevkiyat",
                          "muhasebe", "fason-takip", "kartela", "kalite-kontrol"])
        idx = role_cb.findText(cur_role)
        if idx >= 0:
            role_cb.setCurrentIndex(idx)
        form.addRow("Ad Soyad:", fname)
        form.addRow("Rol:", role_cb)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        if dlg.exec():
            db.update_user(uid, fname.text().strip(), role_cb.currentText())
            self._load()

    def _change_pw(self):
        uid = self._selected_id()
        if not uid: return
        pw, ok = QInputDialog.getText(self, "Şifre Değiştir", "Yeni şifre:", QLineEdit.EchoMode.Password)
        if ok and pw:
            if len(pw) < 4:
                return QMessageBox.warning(self, "Hata", "Şifre en az 4 karakter!")
            db.update_user_password(uid, pw)
            QMessageBox.information(self, "Tamam", "Şifre güncellendi.")

    def _toggle(self):
        uid = self._selected_id()
        if not uid: return
        if uid == CURRENT_USER.get("id"):
            return QMessageBox.warning(self, "Uyarı", "Kendi hesabınızı pasif yapamazsınız!")
        db.toggle_user_active(uid); self._load()

    def _delete(self):
        uid = self._selected_id()
        if not uid: return
        if uid == CURRENT_USER.get("id"):
            return QMessageBox.warning(self, "Uyarı", "Kendi hesabınızı silemezsiniz!")
        if QMessageBox.question(self, "Sil", "Kullanıcı silinsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_user(uid); self._load()


class FabricDialog(QDialog):
    def __init__(self, parent=None, fabric=None):
        super().__init__(parent)
        self.fabric = fabric
        self.setWindowTitle("Kumaş Ekle" if not fabric else "Kumaş Düzenle")
        self.setMinimumWidth(300)
        self._build_ui()
        if fabric:
            self._populate(fabric)
        self.adjustSize()
        self.resize(int(self.width() * 0.6), self.height())

    def _build_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(10)

        # Ürün kodu — katalogdan sadece seçim (aranabilir, serbest metin kabul edilmez)
        self.product_code = QComboBox()
        _make_searchable(self.product_code)
        self.product_code.currentIndexChanged.connect(self._on_product_change)

        # Ürün açıklaması — serbest metin, elle girilir
        self.product_name = QLineEdit()

        self.color = QLineEdit()
        self.lab_no = QLineEdit()
        self.lab_no.setPlaceholderText("Lab dip onay numarası (opsiyonel)")
        self._load_products()

        # Hedef Lokasyon — iki kademeli (grup → depo)
        self.loc_group_combo = QComboBox()   # 1. kademe: grup seçimi (DEPO / DIŞ DEPO …)
        self.location_combo  = QComboBox()   # 2. kademe: ilgili gruptaki lokasyonlar
        _make_searchable(self.location_combo)
        self.loc_group_combo.currentIndexChanged.connect(self._on_loc_group_change)

        # Satın alma lokasyonu — müşteri listesinden seçim
        self.entry_loc = QComboBox()
        _make_searchable(self.entry_loc)
        self._load_locations()
        self._load_entry_customers()

        self.meter = QDoubleSpinBox()
        self.meter.setRange(0, 999999)
        self.meter.setDecimals(2)
        self.kg = QDoubleSpinBox()
        self.kg.setRange(0, 999999)
        self.kg.setDecimals(2)
        self.piece_count = QLineEdit()

        self.birim_fiyat = QDoubleSpinBox()
        self.birim_fiyat.setRange(0, 9999999)
        self.birim_fiyat.setDecimals(2)
        self.birim_fiyat.setSuffix(" $")
        self.birim_fiyat.setStyleSheet("border: 1px solid #BDBDBD;")

        # Kumaş tipi — zorunlu
        self.fabric_type = QComboBox()
        self.fabric_type.addItem("— Seçiniz —", "")
        for t in ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI"]:
            self.fabric_type.addItem(t, t)
        self.fabric_type.setStyleSheet("border: 1px solid #BDBDBD;")

        # Baskı tipi — sadece Kumaş Tipi = BASKILI iken görünür
        self.print_type = QComboBox()
        self.print_type.addItem("— Seçiniz —", "")
        for t in PRINT_TYPES:
            self.print_type.addItem(t, t)
        self.print_type.setStyleSheet("border: 1px solid #BDBDBD;")

        self.zemin_rengi = QLineEdit()
        self.baski_desen_no = QLineEdit()

        self.fabric_type.currentIndexChanged.connect(self._update_print_fields)
        self.print_type.currentIndexChanged.connect(self._update_print_fields)

        self.lot = QLineEdit()
        self.lot.setPlaceholderText("Boş bırakılırsa otomatik verilir (LOT-20260610-001)")

        self.description = QTextEdit()
        self.description.setMaximumHeight(70)

        form.addRow("Ürün Kodu *:", self.product_code)
        form.addRow("Ürün Açıklaması:", self.product_name)
        form.addRow("Kumaş Tipi *:", self.fabric_type)
        form.addRow("Baskı Tipi *:", self.print_type)
        form.addRow("Renk:", self.color)
        form.addRow("Zemin Rengi:", self.zemin_rengi)
        form.addRow("Lab No:", self.lab_no)
        form.addRow("Baskı Desen No:", self.baski_desen_no)

        form.addRow("Satın Alma Lokasyonu:", self.entry_loc)

        form.addRow("Depo Türü *:", self.loc_group_combo)
        form.addRow("Hedef Lokasyon *:", self.location_combo)
        form.addRow("Lot:", self.lot)
        form.addRow("Metre:", self.meter)
        form.addRow("Kilo:", self.kg)
        form.addRow("Top/Adet:", self.piece_count)
        form.addRow("Birim Fiyat ($/mt):", self.birim_fiyat)
        form.addRow("Açıklama:", self.description)

        layout.addLayout(form)
        self._form = form
        self._update_print_fields()

        # Zorunlu alan notu
        note = QLabel("* işaretli alanlar zorunludur")
        note.setStyleSheet("color:#9E9E9E; font-size:11px;")
        layout.addWidget(note)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _load_products(self, select_code="", select_name=""):
        """Ürün kodu kombosunu kataloğa göre doldurur. select_code katalogda
        yoksa (eski kayıt), kaybolmaması için listeye eklenir."""
        self._products = {}
        self.product_code.blockSignals(True)
        self.product_code.clear()
        self.product_code.addItem("— Seçiniz —", "")
        self.product_code.addItem("➕ Yeni ürün ekle...", "__NEW__")
        for r in db.get_all_products(active_only=True):
            code = r["product_code"]
            name = r["product_name"] or ""
            is_numune = (r["product_status"] if "product_status" in r.keys() else "AKTİF") == "NUMUNE"
            self._products[code] = name
            label = f"{code} — {name}" if name else code
            if is_numune:
                label = "🔶 " + label
            self.product_code.addItem(label, code)
        if select_code:
            idx = self.product_code.findData(select_code)
            if idx < 0:
                self._products[select_code] = select_name
                label = f"{select_code} — {select_name}" if select_name else select_code
                self.product_code.addItem(label, select_code)
                idx = self.product_code.count() - 1
            self.product_code.setCurrentIndex(idx)
        self.product_code.blockSignals(False)
        self._on_product_change()

    def _on_product_change(self, idx=None):
        code = self.product_code.currentData()
        if code == "__NEW__":
            self._add_new_product()
            return

    def _add_new_product(self):
        """Katalogda olmayan bir ürün kodu için hızlı ekleme."""
        dlg = QDialog(self)
        dlg.setWindowTitle("Yeni Ürün Ekle")
        dlg.setMinimumWidth(320)
        lay = QVBoxLayout(dlg)
        form = QFormLayout(); form.setSpacing(8)
        code_edit = QLineEdit()
        name_edit = QLineEdit()
        form.addRow("Ürün Kodu *:", code_edit)
        form.addRow("Ürün Adı/Bilgisi:", name_edit)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        if dlg.exec():
            code = code_edit.text().strip().upper()
            if not code:
                QMessageBox.warning(self, "Hata", "Ürün kodu zorunlu!")
                self.product_code.setCurrentIndex(0)
                return
            try:
                db.add_product(code, name_edit.text().strip())
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Ürün eklenemedi:\n{e}")
                self.product_code.setCurrentIndex(0)
                return
            self._load_products(select_code=code)
        else:
            self.product_code.setCurrentIndex(0)

    def _load_locations(self):
        locs = db.get_active_locations()
        self._all_locs = sorted(locs, key=lambda x: (x["group_name"] or "", x["name"] or ""))

        # 1. kademe: benzersiz grup isimleri
        groups = list(dict.fromkeys(l["group_name"] or "Diğer" for l in self._all_locs))
        self.loc_group_combo.blockSignals(True)
        self.loc_group_combo.clear()
        self.loc_group_combo.addItem("— Depo Türü Seçiniz —", "")
        for g in groups:
            self.loc_group_combo.addItem(g, g)
        self.loc_group_combo.blockSignals(False)

        # 2. kademe başlangıçta boş
        self.location_combo.clear()
        self.location_combo.addItem("— Önce depo türü seçiniz —", "")

    def _on_loc_group_change(self):
        grp = self.loc_group_combo.currentData() or ""
        self.location_combo.clear()
        if not grp:
            self.location_combo.addItem("— Önce depo türü seçiniz —", "")
            return
        self.location_combo.addItem("— Lokasyon Seçiniz —", "")
        for l in self._all_locs:
            if (l["group_name"] or "Diğer") == grp:
                self.location_combo.addItem(l["name"], l["name"])

    def _load_entry_customers(self):
        self.entry_loc.clear()
        self.entry_loc.addItem("— Seçiniz —", "")
        for c in db.get_all_customers():
            label = c["name"] + (f" ({c['code']})" if c.get("code") else "")
            self.entry_loc.addItem(label, c["name"])

    def _selected_location(self):
        return self.location_combo.currentData() or ""

    def _selected_location_group(self):
        return self.loc_group_combo.currentData() or ""

    def _update_print_fields(self):
        """BASKILI / Baskı Tipi seçimine göre Renk, Zemin Rengi, Lab No,
        Baskı Tipi ve Baskı Desen No alanlarının görünürlüğünü ayarlar."""
        is_baskili = self.fabric_type.currentData() == "BASKILI"
        pt = self.print_type.currentData() if is_baskili else ""
        is_ronjan = pt == "RONJAN"
        for widget, visible in (
            (self.print_type, is_baskili),
            (self.color, not is_baskili),
            (self.zemin_rengi, is_ronjan),
            (self.lab_no, (not is_baskili) or is_ronjan),
            (self.baski_desen_no, is_baskili and bool(pt)),
        ):
            label = self._form.labelForField(widget)
            if label:
                label.setVisible(visible)
            widget.setVisible(visible)

    def _populate(self, f):
        self._load_products(select_code=f["product_code"] or "", select_name=f["product_name"] or "")
        self.product_name.setText(f["product_name"] or "")
        self.color.setText(f["color"] or "")
        self.lab_no.setText(dict(f).get("lab_no") or "")
        loc_val = f["location"] or ""
        if loc_val:
            # Grubunu bul, önce grup seç ki 2. kademe dolsun
            matched_grp = next(
                (l["group_name"] or "Diğer" for l in self._all_locs if l["name"] == loc_val),
                None)
            if matched_grp:
                gi = self.loc_group_combo.findData(matched_grp)
                if gi >= 0:
                    self.loc_group_combo.setCurrentIndex(gi)  # tetikler _on_loc_group_change
            else:
                # Listede yok — gruba göre ekleyelim
                self.location_combo.addItem(loc_val, loc_val)
            idx = self.location_combo.findData(loc_val)
            if idx >= 0:
                self.location_combo.setCurrentIndex(idx)
        entry = dict(f).get("entry_location") or ""
        if entry:
            i = self.entry_loc.findData(entry)
            if i < 0:
                self.entry_loc.addItem(entry, entry)
                i = self.entry_loc.count() - 1
            self.entry_loc.setCurrentIndex(i)
        self.meter.setValue(f["meter"] or 0)
        self.kg.setValue(f["kg"] or 0)
        self.piece_count.setText(f["piece_count"] or "")
        self.birim_fiyat.setValue(f["birim_fiyat"] or 0)
        ft_idx = self.fabric_type.findData(f["fabric_type"] or "")
        if ft_idx >= 0:
            self.fabric_type.setCurrentIndex(ft_idx)
        pt_idx = self.print_type.findData(dict(f).get("print_type") or "")
        if pt_idx >= 0:
            self.print_type.setCurrentIndex(pt_idx)
        self.zemin_rengi.setText(dict(f).get("zemin_rengi") or "")
        self.baski_desen_no.setText(dict(f).get("baski_desen_no") or "")
        self.lot.setText(f["lot"] or "")
        self.description.setPlainText(f["description"] or "")
        self._update_print_fields()

    def _validate(self):
        errors = []
        if not self.product_code.currentData():
            errors.append("• Ürün kodu katalogdan seçilmelidir")
        if not self.location_combo.currentData():
            errors.append("• Hedef lokasyon seçilmelidir")
        if not self.fabric_type.currentData():
            errors.append("• Kumaş tipi seçilmelidir (Ham / PFD / Boyalı / İpliği Boyalı / Baskılı)")
            self.fabric_type.setStyleSheet("border: 2px solid #C62828; border-radius:4px;")
        else:
            self.fabric_type.setStyleSheet("")
        if self.fabric_type.currentData() == "BASKILI" and not self.print_type.currentData():
            errors.append("• Baskı tipi seçilmelidir (Ronjan / Rotasyon / Dijital / Flok / Varak / Glitter)")
            self.print_type.setStyleSheet("border: 2px solid #C62828; border-radius:4px;")
        else:
            self.print_type.setStyleSheet("border: 1px solid #BDBDBD;")
        if errors:
            QMessageBox.warning(self, "Eksik Bilgi", "\n".join(errors))
            return
        self.accept()

    def get_data(self):
        fabric_type = self.fabric_type.currentData() or ""
        is_baskili = fabric_type == "BASKILI"
        print_type = (self.print_type.currentData() or "") if is_baskili else ""
        is_ronjan = print_type == "RONJAN"
        return {
            "product_code": (self.product_code.currentData() or "").strip().upper(),
            "product_name": self.product_name.text().strip(),
            "color": self.color.text().strip().upper() if not is_baskili else "",
            "lab_no": self.lab_no.text().strip() if ((not is_baskili) or is_ronjan) else "",
            "location": self._selected_location(),
            "entry_location": self.entry_loc.currentData() or self._selected_location(),
            "fabric_type": fabric_type,
            "print_type": print_type,
            "zemin_rengi": self.zemin_rengi.text().strip().upper() if is_ronjan else "",
            "baski_desen_no": self.baski_desen_no.text().strip() if (is_baskili and print_type) else "",
            "lot": self.lot.text().strip(),
            "meter": self.meter.value(),
            "kg": self.kg.value(),
            "piece_count": self.piece_count.text().strip(),
            "birim_fiyat": self.birim_fiyat.value(),
            "description": self.description.toPlainText().strip(),
        }


class OrderItemDialog(QDialog):
    """Sipariş kalemi ekle/düzenle — tek bir ürün/renk/kumaş tipi satırı."""

    FABRIC_TYPES = ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI"]

    def __init__(self, parent=None, item=None, currency="USD"):
        super().__init__(parent)
        self.currency = currency
        self.setWindowTitle("Sipariş Kalemi Ekle" if not item else "Sipariş Kalemini Düzenle")
        self.setMinimumSize(460, 540)
        self._build_ui()
        if item:
            self._populate(item)

    def _build_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(10)

        self.product_code = QComboBox()
        _make_searchable(self.product_code)
        self._load_products()
        self.product_code.currentIndexChanged.connect(self._on_product_change)
        form.addRow("Ürün Kodu *:", self.product_code)

        self.composition = QLineEdit()
        self.width = QLineEdit()
        self.gramaj = QLineEdit()
        form.addRow("Kompozisyon:", self.composition)
        form.addRow("En:", self.width)
        form.addRow("Gramaj:", self.gramaj)

        self.fabric_type = QComboBox()
        self.fabric_type.addItem("— Seçiniz —", "")
        for t in self.FABRIC_TYPES:
            self.fabric_type.addItem(t, t)
        self.fabric_type.setStyleSheet("border: 1px solid #BDBDBD;")
        form.addRow("Kumaş Tipi *:", self.fabric_type)

        # Baskı tipi — sadece Kumaş Tipi = BASKILI iken görünür
        self.print_type = QComboBox()
        self.print_type.addItem("— Seçiniz —", "")
        for t in PRINT_TYPES:
            self.print_type.addItem(t, t)
        self.print_type.setStyleSheet("border: 1px solid #BDBDBD;")
        form.addRow("Baskı Tipi *:", self.print_type)

        self.color = QLineEdit()
        self.zemin_rengi = QLineEdit()
        self.lab_no = QLineEdit()
        self.baski_desen_no = QLineEdit()
        form.addRow("Renk:", self.color)
        form.addRow("Zemin Rengi:", self.zemin_rengi)
        form.addRow("Lab No:", self.lab_no)
        form.addRow("Baskı Desen No:", self.baski_desen_no)

        self.fabric_type.currentIndexChanged.connect(self._update_print_fields)
        self.print_type.currentIndexChanged.connect(self._update_print_fields)

        self.description = QLineEdit()
        self.description.setPlaceholderText("Opsiyonel not (örn. numune ile aynı ton)")
        form.addRow("Açıklama:", self.description)

        self.meter = QDoubleSpinBox()
        self.meter.setRange(0, 999999)
        self.meter.setDecimals(2)
        self.meter.valueChanged.connect(self._update_total)
        form.addRow("Metre:", self.meter)

        kg_row = QWidget()
        kg_lay = QHBoxLayout(kg_row)
        kg_lay.setContentsMargins(0, 0, 0, 0)
        self.kg = QDoubleSpinBox()
        self.kg.setRange(0, 999999)
        self.kg.setDecimals(2)
        btn_calc_kg = QPushButton("🧮 Hesapla")
        btn_calc_kg.setToolTip("Kilo = Gramaj × (En / 100) × Metre / 1000")
        btn_calc_kg.clicked.connect(self._calc_kg)
        kg_lay.addWidget(self.kg, 1)
        kg_lay.addWidget(btn_calc_kg)
        form.addRow("Kilo:", kg_row)

        symbol = CURRENCY_SYMBOLS.get(self.currency, "$")
        self.sale_price = QDoubleSpinBox()
        self.sale_price.setRange(0, 9999999)
        self.sale_price.setDecimals(2)
        self.sale_price.setSuffix(f" {symbol}")
        self.sale_price.setStyleSheet("border: 1px solid #BDBDBD;")
        self.sale_price.valueChanged.connect(self._update_total)
        form.addRow(f"Satış Fiyatı ({symbol}/mt):", self.sale_price)

        self.total_lbl = QLabel(f"0.00 {symbol}")
        self.total_lbl.setStyleSheet("font-weight:bold; color:#1A237E;")
        form.addRow("Tutar:", self.total_lbl)

        layout.addLayout(form)
        self._form = form
        self._update_print_fields()

        note = QLabel("* işaretli alanlar zorunludur")
        note.setStyleSheet("color:#9E9E9E; font-size:11px;")
        layout.addWidget(note)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    # ── Ürün ─────────────────────────────────────────────────────
    def _load_products(self, select_code="", select_name=""):
        self._products = {}
        self.product_code.blockSignals(True)
        self.product_code.clear()
        self.product_code.addItem("— Seçiniz —", "")
        self.product_code.addItem("➕ Yeni ürün ekle...", "__NEW__")
        for r in db.get_all_products(active_only=True):
            code = r["product_code"]
            name = r["product_name"] or ""
            is_numune = (r["product_status"] if "product_status" in r.keys() else "AKTİF") == "NUMUNE"
            self._products[code] = name
            label = f"{code} — {name}" if name else code
            if is_numune:
                label = "🔶 " + label
            self.product_code.addItem(label, code)
        if select_code:
            idx = self.product_code.findData(select_code)
            if idx < 0:
                self._products[select_code] = select_name
                label = f"{select_code} — {select_name}" if select_name else select_code
                self.product_code.addItem(label, select_code)
                idx = self.product_code.count() - 1
            self.product_code.setCurrentIndex(idx)
        self.product_code.blockSignals(False)
        self._on_product_change()

    def _on_product_change(self, idx=None):
        code = self.product_code.currentData()
        if code == "__NEW__":
            self._add_new_product()
            return
        if code:
            product = db.get_product_by_code(code)
            if product:
                p = dict(product)
                self.composition.setText(p.get("composition") or "")
                self.width.setText(p.get("width") or "")
                self.gramaj.setText(p.get("gramaj") or "")

    def _add_new_product(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Yeni Ürün Ekle")
        dlg.setMinimumWidth(320)
        lay = QVBoxLayout(dlg)
        form = QFormLayout(); form.setSpacing(8)
        code_edit = QLineEdit()
        name_edit = QLineEdit()
        form.addRow("Ürün Kodu *:", code_edit)
        form.addRow("Ürün Adı/Bilgisi:", name_edit)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        if dlg.exec():
            code = code_edit.text().strip().upper()
            if not code:
                QMessageBox.warning(self, "Hata", "Ürün kodu zorunlu!")
                self.product_code.setCurrentIndex(0)
                return
            try:
                db.add_product(code, name_edit.text().strip())
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Ürün eklenemedi:\n{e}")
                self.product_code.setCurrentIndex(0)
                return
            self._load_products(select_code=code)
        else:
            self.product_code.setCurrentIndex(0)

    # ── Hesaplamalar ─────────────────────────────────────────────
    def _calc_kg(self):
        try:
            gramaj = float(str(self.gramaj.text()).replace(",", ".").strip())
            en = float(str(self.width.text()).replace(",", ".").strip())
            metre = self.meter.value()
            kg = gramaj * (en / 100) * metre / 1000
            self.kg.setValue(round(kg, 2))
        except ValueError:
            QMessageBox.warning(self, "Hesaplanamadı",
                "Kilo hesaplamak için Gramaj, En ve Metre alanları sayısal olmalıdır.")

    def _update_total(self):
        symbol = CURRENCY_SYMBOLS.get(self.currency, "$")
        total = self.meter.value() * self.sale_price.value()
        self.total_lbl.setText(f"{total:,.2f} {symbol}")

    def _update_print_fields(self):
        """BASKILI / Baskı Tipi seçimine göre Renk, Zemin Rengi, Lab No,
        Baskı Tipi ve Baskı Desen No alanlarının görünürlüğünü ayarlar."""
        is_baskili = self.fabric_type.currentData() == "BASKILI"
        pt = self.print_type.currentData() if is_baskili else ""
        is_ronjan = pt == "RONJAN"
        for widget, visible in (
            (self.print_type, is_baskili),
            (self.color, not is_baskili),
            (self.zemin_rengi, is_ronjan),
            (self.lab_no, (not is_baskili) or is_ronjan),
            (self.baski_desen_no, is_baskili and bool(pt)),
        ):
            label = self._form.labelForField(widget)
            if label:
                label.setVisible(visible)
            widget.setVisible(visible)

    def _populate(self, item):
        self._load_products(select_code=item.get("product_code") or "", select_name=item.get("product_name") or "")
        self.composition.setText(item.get("composition") or "")
        self.width.setText(item.get("width") or "")
        self.gramaj.setText(item.get("gramaj") or "")
        ft_idx = self.fabric_type.findData(item.get("fabric_type") or "")
        if ft_idx >= 0:
            self.fabric_type.setCurrentIndex(ft_idx)
        pt_idx = self.print_type.findData(item.get("print_type") or "")
        if pt_idx >= 0:
            self.print_type.setCurrentIndex(pt_idx)
        self.color.setText(item.get("color") or "")
        self.zemin_rengi.setText(item.get("zemin_rengi") or "")
        self.lab_no.setText(item.get("lab_no") or "")
        self.baski_desen_no.setText(item.get("baski_desen_no") or "")
        self.description.setText(item.get("description") or "")
        self.meter.setValue(item.get("meter") or 0)
        self.kg.setValue(item.get("kg") or 0)
        self.sale_price.setValue(item.get("sale_price") or 0)
        self._update_total()
        self._update_print_fields()

    def _validate(self):
        errors = []
        code = self.product_code.currentData()
        if not code or code == "__NEW__":
            errors.append("• Ürün kodu seçilmelidir")
        if not self.fabric_type.currentData():
            errors.append("• Kumaş tipi seçilmelidir (Ham / PFD / Boyalı / İpliği Boyalı / Baskılı)")
            self.fabric_type.setStyleSheet("border: 2px solid #C62828; border-radius:4px;")
        else:
            self.fabric_type.setStyleSheet("border: 1px solid #BDBDBD;")
        if self.fabric_type.currentData() == "BASKILI" and not self.print_type.currentData():
            errors.append("• Baskı tipi seçilmelidir (Ronjan / Rotasyon / Dijital / Flok / Varak / Glitter)")
            self.print_type.setStyleSheet("border: 2px solid #C62828; border-radius:4px;")
        else:
            self.print_type.setStyleSheet("border: 1px solid #BDBDBD;")
        if errors:
            QMessageBox.warning(self, "Eksik Bilgi", "\n".join(errors))
            return
        self.accept()

    def get_data(self):
        fabric_type = self.fabric_type.currentData() or ""
        is_baskili = fabric_type == "BASKILI"
        print_type = (self.print_type.currentData() or "") if is_baskili else ""
        is_ronjan = print_type == "RONJAN"
        return {
            "product_code": (self.product_code.currentData() or "").strip().upper(),
            "product_name": self._products.get(self.product_code.currentData(), ""),
            "composition": self.composition.text().strip(),
            "width": self.width.text().strip(),
            "gramaj": self.gramaj.text().strip(),
            "fabric_type": fabric_type,
            "color": self.color.text().strip().upper() if not is_baskili else "",
            "lab_no": self.lab_no.text().strip() if ((not is_baskili) or is_ronjan) else "",
            "print_type": print_type,
            "zemin_rengi": self.zemin_rengi.text().strip().upper() if is_ronjan else "",
            "baski_desen_no": self.baski_desen_no.text().strip() if (is_baskili and print_type) else "",
            "description": self.description.text().strip(),
            "meter": self.meter.value(),
            "kg": self.kg.value(),
            "sale_price": self.sale_price.value(),
        }


class OrderDialog(QDialog):
    """Satış siparişi açma/düzenleme — 3 sekme: sipariş bilgileri, sipariş
    kalemleri (çoklu ürün/renk/kumaş tipi), sözleşme şartnamesi. Kaydedilince
    status='PLANLAMA BEKLİYOR' ile orders tablosuna yazılır."""

    PAYMENT_METHODS = ["Peşin", "Havale/EFT", "30 Gün Vade", "60 Gün Vade",
                       "90 Gün Vade", "120 Gün Vade", "Akreditif (L/C)", "Diğer"]
    DELIVERY_TERMS = ["FOB", "CIF", "EXW", "DAP", "Diğer"]

    ITEM_COLS = ["Ürün Kodu", "Kompozisyon", "En", "Gramaj", "Kumaş Tipi", "Renk",
                 "Lab No", "Açıklama", "Metre", "Kilo", "Birim Fiyat", "Tutar",
                 "Baskı Tipi", "Zemin Rengi", "Baskı Desen No"]

    def __init__(self, parent=None, order=None):
        super().__init__(parent)
        self.order = order
        self._items = list(order.get("items", [])) if order else []
        self.setWindowTitle("Yeni Sipariş" if not order else f"Sipariş Düzenle — {order['order_no']}")
        self.setMinimumSize(740, 660)
        self._build_ui()
        if order:
            self._populate(order)
        else:
            self.order_date.setDate(QDate.currentDate())
            self.delivery_date.setDate(QDate.currentDate().addDays(30))
            self._refresh_items_table()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # ── Sekme 1: Sipariş Bilgileri ──────────────────────────────
        tab1 = QWidget()
        form1 = QFormLayout(tab1)
        form1.setSpacing(10)

        self.customer = QComboBox()
        _make_searchable(self.customer)
        self._load_customers()
        self.customer.currentIndexChanged.connect(self._on_customer_change)
        form1.addRow("Müşteri *:", self.customer)

        self.customer_ref = QLineEdit()
        self.customer_ref.setPlaceholderText("Müşterinin kendi sipariş/PO numarası (opsiyonel)")
        form1.addRow("Müşteri Referans:", self.customer_ref)

        self.order_no_edit = QLineEdit()
        self.order_no_edit.setReadOnly(True)
        self.order_no_edit.setText("Kaydedince atanacak")
        self.order_no_edit.setStyleSheet("color:#757575;")
        form1.addRow("Sipariş No:", self.order_no_edit)

        self.order_date = QDateEdit()
        self.order_date.setCalendarPopup(True)
        self.order_date.setDisplayFormat("dd.MM.yyyy")
        form1.addRow("Sipariş Tarihi:", self.order_date)

        self.delivery_date = QDateEdit()
        self.delivery_date.setCalendarPopup(True)
        self.delivery_date.setDisplayFormat("dd.MM.yyyy")
        form1.addRow("Termin *:", self.delivery_date)

        self.currency = QComboBox()
        for code in CURRENCY_OPTIONS:
            self.currency.addItem(f"{code} ({CURRENCY_SYMBOLS[code]})", code)
        form1.addRow("Para Birimi *:", self.currency)

        self.payment_method = QComboBox()
        self.payment_method.setEditable(True)
        self.payment_method.addItem("")
        for p in self.PAYMENT_METHODS:
            self.payment_method.addItem(p)
        form1.addRow("Ödeme Şekli:", self.payment_method)

        self.delivery_terms = QComboBox()
        self.delivery_terms.setEditable(True)
        self.delivery_terms.addItem("")
        for d in self.DELIVERY_TERMS:
            self.delivery_terms.addItem(d)
        form1.addRow("Teslimat Şartları:", self.delivery_terms)

        self.delivery_address = QLineEdit()
        form1.addRow("Teslimat Adresi:", self.delivery_address)

        self.sales_rep = QComboBox()
        self._load_sales_reps()
        form1.addRow("Siparişi Alan:", self.sales_rep)

        self.notes = QTextEdit()
        self.notes.setMaximumHeight(70)
        form1.addRow("Notlar:", self.notes)

        self.tabs.addTab(tab1, "📋 Sipariş Bilgileri")

        # ── Sekme 2: Sipariş Kalemleri ───────────────────────────────
        tab2 = QWidget()
        lay2 = QVBoxLayout(tab2)

        toolbar2 = QHBoxLayout()
        btn_add_item = QPushButton("+ Kalem Ekle")
        btn_add_item.setStyleSheet("background:#2E7D32; color:white; font-weight:bold; border-radius:4px; padding:6px 14px;")
        btn_add_item.clicked.connect(self._add_item)
        btn_edit_item = QPushButton("✎ Düzenle")
        btn_edit_item.clicked.connect(self._edit_item)
        btn_del_item = QPushButton("✕ Kaldır")
        btn_del_item.setStyleSheet("background:#757575; color:white; border-radius:4px; padding:6px 14px;")
        btn_del_item.clicked.connect(self._remove_item)
        toolbar2.addWidget(btn_add_item)
        toolbar2.addWidget(btn_edit_item)
        toolbar2.addWidget(btn_del_item)
        toolbar2.addStretch()
        lay2.addLayout(toolbar2)

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(len(self.ITEM_COLS))
        self.items_table.setHorizontalHeaderLabels(self.ITEM_COLS)
        self.items_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.items_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.items_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.items_table.verticalHeader().setVisible(False)
        hdr = self.items_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        self.items_table.doubleClicked.connect(self._edit_item)
        lay2.addWidget(self.items_table)

        self.items_total_lbl = QLabel("Genel Toplam: 0.00 $")
        self.items_total_lbl.setStyleSheet("font-weight:bold; color:#1A237E; font-size:13px;")
        self.items_total_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)
        lay2.addWidget(self.items_total_lbl)

        self.tabs.addTab(tab2, "📦 Sipariş Kalemleri")

        # Para birimi değişince kalemler tablosu yeniden formatlanır —
        # items_table oluşturulduktan sonra bağlanır (erken tetiklenmesin)
        self.currency.currentIndexChanged.connect(self._refresh_items_table)

        note = QLabel("* işaretli alanlar zorunludur")
        note.setStyleSheet("color:#9E9E9E; font-size:11px;")
        layout.addWidget(note)

        btn_row = QHBoxLayout()
        if self.order:
            btn_pdf = QPushButton("📄 PDF Al")
            btn_pdf.clicked.connect(self._export_pdf)
            btn_row.addWidget(btn_pdf)
        btn_row.addStretch()
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)
        btn_row.addWidget(buttons)
        layout.addLayout(btn_row)

    # ── Para Birimi / Kalemler ───────────────────────────────────
    def _current_currency(self):
        return self.currency.currentData() or "USD"

    def _refresh_items_table(self):
        symbol = CURRENCY_SYMBOLS.get(self._current_currency(), "$")
        self.items_table.setRowCount(len(self._items))
        grand_total = 0.0
        for i, it in enumerate(self._items):
            meter = it.get("meter") or 0
            sale_price = it.get("sale_price") or 0
            total = meter * sale_price
            grand_total += total
            values = [
                it.get("product_code") or "",
                it.get("composition") or "",
                it.get("width") or "",
                it.get("gramaj") or "",
                it.get("fabric_type") or "",
                it.get("color") or "",
                it.get("lab_no") or "",
                it.get("description") or "",
                f"{meter:,.2f}",
                f"{(it.get('kg') or 0):,.2f}",
                f"{sale_price:,.2f} {symbol}",
                f"{total:,.2f} {symbol}",
                it.get("print_type") or "",
                it.get("zemin_rengi") or "",
                it.get("baski_desen_no") or "",
            ]
            for j, v in enumerate(values):
                cell = QTableWidgetItem(v)
                if 8 <= j <= 11:
                    cell.setTextAlignment(int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight))
                if j == 4:
                    cell.setForeground(QBrush(FABRIC_TYPE_COLORS.get(it.get("fabric_type", ""), QColor("#333333"))))
                self.items_table.setItem(i, j, cell)
        self.items_total_lbl.setText(f"Genel Toplam: {grand_total:,.2f} {symbol}")

    def _add_item(self):
        dlg = OrderItemDialog(self, currency=self._current_currency())
        if dlg.exec():
            self._items.append(dlg.get_data())
            self._refresh_items_table()

    def _edit_item(self):
        row = self.items_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir kalem seçin.")
            return
        dlg = OrderItemDialog(self, item=self._items[row], currency=self._current_currency())
        if dlg.exec():
            self._items[row] = dlg.get_data()
            self._refresh_items_table()

    def _remove_item(self):
        row = self.items_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir kalem seçin.")
            return
        if QMessageBox.question(self, "Kaldır", "Bu kalem kaldırılsın mı?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            del self._items[row]
            self._refresh_items_table()

    # ── PDF ──────────────────────────────────────────────────────
    def _export_pdf(self):
        company = db.get_company_settings()
        data = self.get_data()
        order = {**data, "order_no": self.order.get("order_no", ""),
                 "customer_tax_no": _customer_tax_no(data.get("customer_id"))}
        default_name = f"{order['order_no']}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Sipariş Sözleşmesi PDF", default_name, "PDF (*.pdf)")
        if not path:
            return
        try:
            order_pdf.generate_order_pdf(order, company, path)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{e}")
            return
        QMessageBox.information(self, "PDF Oluşturuldu", f"Kaydedildi:\n{path}")

    # ── Müşteri ──────────────────────────────────────────────────
    def _load_sales_reps(self, select_name=None):
        self.sales_rep.blockSignals(True)
        self.sales_rep.clear()
        self.sales_rep.addItem("— Seçiniz —", "")
        for u in (db.get_all_users() or []):
            full = u.get("full_name","") or u.get("username","")
            self.sales_rep.addItem(full, full)
        if select_name:
            idx = self.sales_rep.findData(select_name)
            if idx < 0:
                self.sales_rep.addItem(select_name, select_name)
                idx = self.sales_rep.count() - 1
            self.sales_rep.setCurrentIndex(idx)
        self.sales_rep.blockSignals(False)

    def _load_customers(self, select_id=None):
        self._customers = {}
        self.customer.blockSignals(True)
        self.customer.clear()
        self.customer.addItem("— Seçiniz —", "")
        self.customer.addItem("➕ Yeni müşteri ekle...", "__NEW__")
        for c in db.get_all_customers():
            cid = str(c["id"])
            self._customers[cid] = c["name"]
            label = c["name"] + (f" ({c['code']})" if c["code"] else "")
            self.customer.addItem(label, cid)
        if select_id is not None:
            sid = str(select_id)
            idx = self.customer.findData(sid)
            if idx < 0:
                c = db.get_customer(select_id)
                if c:
                    self._customers[sid] = c["name"]
                    label = c["name"] + (f" ({c['code']})" if c["code"] else "")
                    self.customer.addItem(label, sid)
                    idx = self.customer.count() - 1
            if idx >= 0:
                self.customer.setCurrentIndex(idx)
        self.customer.blockSignals(False)

    def _on_customer_change(self, idx=None):
        if self.customer.currentData() == "__NEW__":
            self._add_new_customer()

    def _add_new_customer(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Yeni Müşteri Ekle")
        dlg.setMinimumWidth(320)
        lay = QVBoxLayout(dlg)
        form = QFormLayout(); form.setSpacing(8)
        name_edit = QLineEdit()
        code_edit = QLineEdit()
        form.addRow("Müşteri Adı *:", name_edit)
        form.addRow("Kodu:", code_edit)
        lay.addLayout(form)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        if dlg.exec():
            name = name_edit.text().strip()
            if not name:
                QMessageBox.warning(self, "Hata", "Müşteri adı zorunlu!")
                self.customer.setCurrentIndex(0)
                return
            try:
                cid = db.add_customer(name, code_edit.text().strip())
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Müşteri eklenemedi:\n{e}")
                self.customer.setCurrentIndex(0)
                return
            self._load_customers(select_id=cid)
        else:
            self.customer.setCurrentIndex(0)

    # ── Doldurma / Doğrulama / Veri ─────────────────────────────────
    def _populate(self, o):
        self._load_customers(select_id=o.get("customer_id"))
        self.customer_ref.setText(o.get("customer_ref") or "")
        self.order_no_edit.setText(o.get("order_no") or "")

        d = QDate.fromString(str(o.get("order_date") or ""), "yyyy-MM-dd")
        self.order_date.setDate(d if d.isValid() else QDate.currentDate())

        d = QDate.fromString(str(o.get("delivery_date") or ""), "yyyy-MM-dd")
        self.delivery_date.setDate(d if d.isValid() else QDate.currentDate().addDays(30))

        cur_idx = self.currency.findData(o.get("currency") or "USD")
        self.currency.setCurrentIndex(cur_idx if cur_idx >= 0 else 0)

        idx = self.payment_method.findText(o.get("payment_method") or "")
        if idx >= 0:
            self.payment_method.setCurrentIndex(idx)
        else:
            self.payment_method.setEditText(o.get("payment_method") or "")

        idx = self.delivery_terms.findText(o.get("delivery_terms") or "")
        if idx >= 0:
            self.delivery_terms.setCurrentIndex(idx)
        else:
            self.delivery_terms.setEditText(o.get("delivery_terms") or "")

        self.delivery_address.setText(o.get("delivery_address") or "")
        self._load_sales_reps(select_name=o.get("sales_rep") or "")
        self.notes.setPlainText(o.get("notes") or "")
        self._refresh_items_table()

    def _validate(self):
        errors = []
        cust = self.customer.currentData()
        if not cust or cust == "__NEW__":
            errors.append("• Müşteri seçilmelidir")
        if not self._items:
            errors.append("• En az bir sipariş kalemi eklenmelidir")
        if errors:
            QMessageBox.warning(self, "Eksik Bilgi", "\n".join(errors))
            return
        self.accept()

    def get_data(self):
        customer_id = self.customer.currentData()
        return {
            "customer_id": int(customer_id) if customer_id and customer_id != "__NEW__" else None,
            "customer_name": self._customers.get(customer_id, ""),
            "customer_ref": self.customer_ref.text().strip(),
            "currency": self.currency.currentData() or "USD",
            "payment_method": self.payment_method.currentText().strip(),
            "delivery_terms": self.delivery_terms.currentText().strip(),
            "delivery_address": self.delivery_address.text().strip(),
            "delivery_date": self.delivery_date.date().toString("yyyy-MM-dd"),
            "order_date": self.order_date.date().toString("yyyy-MM-dd"),
            "sales_rep": self.sales_rep.currentData() or "",
            "notes": self.notes.toPlainText().strip(),
            "items": self._items,
        }


class CellEditDialog(QDialog):
    """Tek bir alanı düzenleyen küçük pencere — çift tıklamayla açılır.
    count > 1 ise toplu düzenleme: aynı değer seçili tüm kayıtlara yazılır."""
    def __init__(self, parent, fabric, field, label, count=1):
        super().__init__(parent)
        self.fabric = fabric
        self.field = field
        self.setWindowTitle(f"{label} Düzenle" if count == 1 else f"{label} — Toplu Düzenle")
        self.setMinimumWidth(360)

        layout = QVBoxLayout(self)
        if count > 1:
            info = QLabel(f"⚠ <b>{count} kayıt</b> birden güncellenecek — "
                          f"girilen değer seçili tüm satırlara yazılır.")
            info.setStyleSheet("background:#FFF3E0; padding:6px; border-radius:4px; color:#E65100;")
        else:
            info = QLabel(f"<b>{fabric['product_code']}</b> — {fabric['color']}  "
                          f"<span style='color:#555'>({fabric['location']})</span>")
            info.setStyleSheet("background:#FFF8E1; padding:6px; border-radius:4px;")
        layout.addWidget(info)

        form = QFormLayout()
        cur = fabric[field]

        if field == "fabric_type":
            self.widget = QComboBox()
            self.widget.addItem("— Seçiniz —", "")
            for t in ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI"]:
                self.widget.addItem(t, t)
            idx = self.widget.findData(cur or "")
            if idx >= 0:
                self.widget.setCurrentIndex(idx)
        elif field == "print_type":
            self.widget = QComboBox()
            self.widget.addItem("— Seçiniz —", "")
            for t in PRINT_TYPES:
                self.widget.addItem(t, t)
            idx = self.widget.findData(cur or "")
            if idx >= 0:
                self.widget.setCurrentIndex(idx)
        elif field == "location":
            locs = db.get_active_locations()
            self._rafs = sorted(l["name"] for l in locs if l["group_name"] == "DEPO")
            dis = sorted(l["name"] for l in locs if l["group_name"] != "DEPO")
            self.depo = QComboBox()
            self.depo.addItem("— Seçiniz —", "")
            self.depo.addItem("DEPO", "__DEPO__")
            for n in dis:
                self.depo.addItem(n, n)
            _make_searchable(self.depo)
            self.raf = QComboBox()
            self.raf.addItem("— Raf Seçiniz —", "")
            for n in self._rafs:
                self.raf.addItem(n, n)
            _make_searchable(self.raf)
            self.raf_label = QLabel("Raf:")
            cur_loc = cur or ""
            if cur_loc in self._rafs:
                self.depo.setCurrentIndex(self.depo.findData("__DEPO__"))
                self.raf.setCurrentIndex(self.raf.findData(cur_loc))
            elif cur_loc:
                i = self.depo.findData(cur_loc)
                if i < 0:
                    self.depo.addItem(cur_loc, cur_loc)
                    i = self.depo.count() - 1
                self.depo.setCurrentIndex(i)
            self.depo.currentIndexChanged.connect(self._on_depo_change)
            self.widget = self.depo
        elif field == "entry_location":
            self.widget = QComboBox()
            locs = db.get_active_locations()
            for l in sorted(locs, key=lambda x: (x["group_name"] != "DEPO", x["name"])):
                self.widget.addItem(l["name"], l["name"])
            _make_searchable(self.widget)
            idx = self.widget.findData(cur or "")
            if idx < 0 and cur:
                self.widget.addItem(cur, cur)
                idx = self.widget.count() - 1
            if idx >= 0:
                self.widget.setCurrentIndex(idx)
        elif field in ("meter", "kg", "birim_fiyat"):
            self.widget = QDoubleSpinBox()
            self.widget.setRange(0, 9999999)
            self.widget.setDecimals(2)
            if field == "birim_fiyat":
                self.widget.setSuffix(" $")
            self.widget.setValue(cur or 0)
        else:
            self.widget = QLineEdit(str(cur or ""))

        form.addRow(f"{label}:", self.widget)
        if field == "location":
            form.addRow(self.raf_label, self.raf)
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        if field == "location":
            self._on_depo_change(0)
        self.widget.setFocus()

    def _on_depo_change(self, idx):
        is_depo = self.depo.currentData() == "__DEPO__"
        self.raf_label.setVisible(is_depo)
        self.raf.setVisible(is_depo)

    def _validate(self):
        v = self.value()
        if self.field == "product_code" and not str(v).strip():
            QMessageBox.warning(self, "Eksik Bilgi", "Ürün kodu boş olamaz.")
            return
        if self.field == "location" and not v:
            QMessageBox.warning(self, "Eksik Bilgi", "Lokasyon (DEPO ise raf) seçilmelidir.")
            return
        if self.field == "fabric_type" and not v:
            QMessageBox.warning(self, "Eksik Bilgi", "Kumaş tipi seçilmelidir.")
            return
        self.accept()

    def value(self):
        if self.field in ("fabric_type", "entry_location", "print_type"):
            return self.widget.currentData() or ""
        if self.field == "location":
            d = self.depo.currentData() or ""
            return (self.raf.currentData() or "") if d == "__DEPO__" else d
        if self.field in ("meter", "kg", "birim_fiyat"):
            return self.widget.value()
        text = self.widget.text().strip()
        if self.field in ("product_code", "color", "zemin_rengi"):
            return text.upper()
        return text


class MovementDialog(QDialog):
    def __init__(self, parent, fabric, movement_type):
        super().__init__(parent)
        self.fabric = fabric
        self.movement_type = movement_type
        self._manual_pct = None   # fire oranı elle girildiyse yüzdesi
        self._skip_fire = False   # kullanıcı fire takibi istemedi
        label = "GİRİŞ" if movement_type == "GİRİŞ" else "ÇIKIŞ"
        self.setWindowTitle(f"Stok {label} — {fabric['product_code']} / {fabric['color']}")
        self.setMinimumWidth(440)
        self._build_ui(fabric)

    def _build_ui(self, fabric):
        layout = QVBoxLayout(self)

        bg = "#E3F2FD" if self.movement_type == "GİRİŞ" else "#FFEBEE"
        pieces = dict(fabric).get("piece_count") or "-"
        info = QLabel(f"<b>{fabric['product_name'] or fabric['product_code']}</b>  |  "
                      f"Renk: {fabric['color']}  |  Lokasyon: {fabric['location']}<br>"
                      f"Mevcut: <b>{fabric['meter']:.2f} mt</b>  /  <b>{fabric['kg']:.2f} kg</b>"
                      f"  /  <b>{pieces} top</b>")
        info.setStyleSheet(f"background:{bg}; padding:8px; border-radius:4px;")
        layout.addWidget(info)

        form = QFormLayout(); form.setSpacing(10)
        self._form = form

        self.meter = QDoubleSpinBox(); self.meter.setRange(0, 999999); self.meter.setDecimals(2)
        self.kg    = QDoubleSpinBox(); self.kg.setRange(0, 999999);    self.kg.setDecimals(2)
        self.piece_count = QLineEdit()
        self.notes = QLineEdit()

        # Dış depodan müşteriye sevk = boyahane fire takibi
        self._is_dis_depo = False
        if self.movement_type == "ÇIKIŞ":
            dis = {l["name"] for l in db.get_active_locations() if l["group_name"] != "DEPO"}
            self._is_dis_depo = (fabric["location"] or "") in dis

        if self._is_dis_depo:
            self.pre_meter = QDoubleSpinBox(); self.pre_meter.setRange(0, 999999); self.pre_meter.setDecimals(2)
            self.pre_kg    = QDoubleSpinBox(); self.pre_kg.setRange(0, 999999);    self.pre_kg.setDecimals(2)
            self.lbl_pre_m = QLabel("Çıkış Öncesi:")
            self.lbl_pre_k = QLabel("Çıkış Öncesi:")

            mrow = QWidget(); ml = QHBoxLayout(mrow); ml.setContentsMargins(0, 0, 0, 0)
            ml.addWidget(self.meter, 1); ml.addWidget(self.lbl_pre_m); ml.addWidget(self.pre_meter, 1)
            krow = QWidget(); kl = QHBoxLayout(krow); kl.setContentsMargins(0, 0, 0, 0)
            kl.addWidget(self.kg, 1); kl.addWidget(self.lbl_pre_k); kl.addWidget(self.pre_kg, 1)
            form.addRow("Metre:", mrow)
            form.addRow("Kilo:", krow)

            self.lbl_fire = QLabel("Fire:")
            self.fire_label = QLabel("—")
            self.fire_label.setStyleSheet("color:#C62828; font-weight:bold;")
            form.addRow(self.lbl_fire, self.fire_label)

            for w in (self.meter, self.kg, self.pre_meter, self.pre_kg):
                w.valueChanged.connect(self._update_fire_label)
        else:
            form.addRow("Metre:", self.meter)
            form.addRow("Kilo:", self.kg)

        form.addRow("Top/Adet:", self.piece_count)

        # Çıkışta kumaş tipi / baskı / renk / lab / parti bilgileri
        if self.movement_type == "ÇIKIŞ":
            self.out_fabric_type = QComboBox()
            self.out_fabric_type.addItem("— Seçiniz —", "")
            for t in ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI"]:
                self.out_fabric_type.addItem(t, t)
            ft_idx = self.out_fabric_type.findData(fabric["fabric_type"] or "")
            if ft_idx >= 0:
                self.out_fabric_type.setCurrentIndex(ft_idx)

            self.out_print_type = QComboBox()
            self.out_print_type.addItem("— Seçiniz —", "")
            for t in PRINT_TYPES:
                self.out_print_type.addItem(t, t)
            pt_idx = self.out_print_type.findData(dict(fabric).get("print_type") or "")
            if pt_idx >= 0:
                self.out_print_type.setCurrentIndex(pt_idx)

            self.out_color = QLineEdit()
            self.lab_no = QLineEdit()
            self.out_zemin_rengi = QLineEdit()
            self.out_zemin_rengi.setText(dict(fabric).get("zemin_rengi") or "")
            self.out_baski_desen_no = QLineEdit()
            self.out_baski_desen_no.setText(dict(fabric).get("baski_desen_no") or "")
            self.parti_no = QLineEdit()

            form.addRow("Kumaş Tipi:", self.out_fabric_type)
            form.addRow("Baskı Tipi:", self.out_print_type)
            form.addRow("Renk:", self.out_color)
            form.addRow("Zemin Rengi:", self.out_zemin_rengi)
            form.addRow("Lab No:", self.lab_no)
            form.addRow("Baskı Desen No:", self.out_baski_desen_no)
            form.addRow("Parti No:", self.parti_no)

            self.out_fabric_type.currentIndexChanged.connect(self._update_out_print_fields)
            self.out_print_type.currentIndexChanged.connect(self._update_out_print_fields)

        form.addRow("Not:", self.notes)

        # Çıkış ise hedef seçimi
        if self.movement_type == "ÇIKIŞ":
            self.dest_type = QComboBox()
            self.dest_type.addItems(["Müşteri", "Lokasyon", "Diğer"])
            self.dest_type.currentIndexChanged.connect(self._on_dest_type_change)

            self.dest_customer = QComboBox()
            _make_searchable(self.dest_customer)
            self.lbl_customer = QLabel("Müşteri:")
            self._load_customers()

            # Lokasyon — iki kademeli: DEPO / dış depolar, DEPO seçilirse raf
            self.dest_depo = QComboBox()
            _make_searchable(self.dest_depo)
            self.lbl_depo = QLabel("Lokasyon:")
            self.dest_raf = QComboBox()
            _make_searchable(self.dest_raf)
            self.lbl_raf = QLabel("Raf:")
            self._load_locations()
            self.dest_depo.currentIndexChanged.connect(self._on_dest_depo_change)

            self.dest_other = QLineEdit()
            self.dest_other.setPlaceholderText("Hedef açıklayınız...")
            self.lbl_other = QLabel("Diğer:")

            form.addRow("Çıkış Hedefi:", self.dest_type)
            form.addRow(self.lbl_customer, self.dest_customer)
            form.addRow(self.lbl_depo, self.dest_depo)
            form.addRow(self.lbl_raf, self.dest_raf)
            form.addRow(self.lbl_other, self.dest_other)
            self._on_dest_type_change(0)
            self._update_out_print_fields()

        layout.addLayout(form)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._on_ok)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        if self._is_dis_depo:
            self._update_fire_visibility()

    def _fire_active(self):
        """Fire takibi: dış depodan yapılan her çıkışta (hedef ne olursa olsun)."""
        return self._is_dis_depo and self.movement_type == "ÇIKIŞ"

    def _update_fire_visibility(self):
        on = self._fire_active()
        for w in (self.lbl_pre_m, self.pre_meter, self.lbl_pre_k, self.pre_kg,
                  self.lbl_fire, self.fire_label):
            w.setVisible(on)

    def _update_fire_label(self):
        if not self._is_dis_depo:
            return
        pre_m, m = self.pre_meter.value(), self.meter.value()
        pre_k, k = self.pre_kg.value(), self.kg.value()
        parts = []
        if pre_m > 0:
            if pre_m < m:
                self.fire_label.setText("⚠ Çıkış öncesi metre, çıkış metresinden küçük olamaz")
                return
            parts.append(f"{pre_m - m:,.2f} mt (%{(pre_m - m) / pre_m * 100:.1f})")
        if pre_k > 0:
            if pre_k < k:
                self.fire_label.setText("⚠ Çıkış öncesi kilo, çıkış kilosundan küçük olamaz")
                return
            parts.append(f"{pre_k - k:,.2f} kg (%{(pre_k - k) / pre_k * 100:.1f})")
        self.fire_label.setText("  |  ".join(parts) if parts else "—")

    def _on_ok(self):
        if self._fire_active() and not self._skip_fire:
            pre_m, pre_k = self.pre_meter.value(), self.pre_kg.value()
            m, k = self.meter.value(), self.kg.value()
            if pre_m <= 0 and pre_k <= 0:
                reply = QMessageBox.question(
                    self, "Fire Oranı",
                    "Çıkış öncesi metre/kilo girilmedi.\n\nFire oranını elle girmek ister misiniz?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes:
                    from PyQt6.QtWidgets import QInputDialog
                    pct, ok = QInputDialog.getDouble(
                        self, "Fire Oranı", "Fire oranı (%):", 0, 0, 99.9, 1)
                    if not ok:
                        return
                    factor = 1 - pct / 100
                    if m > 0:
                        self.pre_meter.setValue(m / factor)
                    if k > 0:
                        self.pre_kg.setValue(k / factor)
                    self._manual_pct = pct
                    style = "color:#C62828; font-weight:bold; border:1px solid #C62828; border-radius:3px;"
                    self.pre_meter.setStyleSheet(style)
                    self.pre_kg.setStyleSheet(style)
                    self._update_fire_label()
                    return   # hesaplananı göster, kullanıcı tekrar OK'a basınca kaydedilir
                else:
                    self._skip_fire = True   # fire takibi olmadan normal çıkış
            elif (m > 0 and pre_m > 0 and pre_m < m) or (k > 0 and pre_k > 0 and pre_k < k):
                QMessageBox.warning(self, "Hatalı Değer",
                                    "Çıkış öncesi miktar, çıkış miktarından küçük olamaz.")
                return
        self.accept()

    def _load_customers(self):
        self.dest_customer.clear()
        self.dest_customer.addItem("— Seçiniz —", "")
        for c in db.get_all_customers():
            label = c["name"] + (f" ({c['code']})" if c["code"] else "")
            self.dest_customer.addItem(label, str(c["id"]))

    def _load_locations(self):
        locs = db.get_active_locations()
        rafs        = sorted(l["name"] for l in locs if l["group_name"] == "DEPO")
        dis_depolar = sorted(l["name"] for l in locs if l["group_name"] != "DEPO")

        self.dest_depo.clear()
        self.dest_depo.addItem("— Seçiniz —", "")
        self.dest_depo.addItem("DEPO", "__DEPO__")
        for name in dis_depolar:
            self.dest_depo.addItem(name, name)

        self.dest_raf.clear()
        self.dest_raf.addItem("— Raf Seçiniz —", "")
        for name in rafs:
            self.dest_raf.addItem(name, name)

    def _on_dest_type_change(self, idx):
        t = self.dest_type.currentText()
        for w in (self.lbl_customer, self.dest_customer):
            w.setVisible(t == "Müşteri")
        for w in (self.lbl_depo, self.dest_depo):
            w.setVisible(t == "Lokasyon")
        for w in (self.lbl_other, self.dest_other):
            w.setVisible(t == "Diğer")
        self._on_dest_depo_change(0)
        if self._is_dis_depo:
            self._update_fire_visibility()

    def _on_dest_depo_change(self, idx):
        show_raf = (self.dest_type.currentText() == "Lokasyon"
                    and self.dest_depo.currentData() == "__DEPO__")
        self.lbl_raf.setVisible(show_raf)
        self.dest_raf.setVisible(show_raf)
        if not show_raf:
            self.dest_raf.setCurrentIndex(0)

    def _update_out_print_fields(self):
        """BASKILI / Baskı Tipi seçimine göre Renk, Zemin Rengi, Lab No,
        Baskı Tipi ve Baskı Desen No alanlarının görünürlüğünü ayarlar."""
        is_baskili = self.out_fabric_type.currentData() == "BASKILI"
        pt = self.out_print_type.currentData() if is_baskili else ""
        is_ronjan = pt == "RONJAN"
        for widget, visible in (
            (self.out_print_type, is_baskili),
            (self.out_color, not is_baskili),
            (self.out_zemin_rengi, is_ronjan),
            (self.lab_no, (not is_baskili) or is_ronjan),
            (self.out_baski_desen_no, is_baskili and bool(pt)),
        ):
            label = self._form.labelForField(widget)
            if label:
                label.setVisible(visible)
            widget.setVisible(visible)

    def get_data(self):
        d = {
            "meter":       self.meter.value(),
            "kg":          self.kg.value(),
            "piece_count": self.piece_count.text().strip(),
            "notes":       self.notes.text().strip(),
            "destination": "", "destination_type": "",
            "fire_active": False, "pre_meter": 0.0, "pre_kg": 0.0,
            "fire_pct": 0.0, "fire_manual": False,
            "out_color": "", "lab_no": "", "parti_no": "",
            "out_fabric_type": "", "out_print_type": "",
            "out_zemin_rengi": "", "out_baski_desen_no": "",
        }
        if self.movement_type == "ÇIKIŞ":
            fabric_type = self.out_fabric_type.currentData() or ""
            is_baskili = fabric_type == "BASKILI"
            print_type = (self.out_print_type.currentData() or "") if is_baskili else ""
            is_ronjan = print_type == "RONJAN"
            d["out_fabric_type"] = fabric_type
            d["out_print_type"] = print_type
            d["out_color"] = self.out_color.text().strip().upper() if not is_baskili else ""
            d["out_zemin_rengi"] = self.out_zemin_rengi.text().strip().upper() if is_ronjan else ""
            d["lab_no"] = self.lab_no.text().strip() if ((not is_baskili) or is_ronjan) else ""
            d["out_baski_desen_no"] = self.out_baski_desen_no.text().strip() if (is_baskili and print_type) else ""
            d["parti_no"] = self.parti_no.text().strip()
            t = self.dest_type.currentText()
            d["destination_type"] = t
            if t == "Müşteri":
                d["destination"] = self.dest_customer.currentText() if self.dest_customer.currentData() else ""
            elif t == "Lokasyon":
                depo = self.dest_depo.currentData() or ""
                d["destination"] = (self.dest_raf.currentData() or "") if depo == "__DEPO__" else depo
            else:
                d["destination"] = self.dest_other.text().strip()

            if self._fire_active() and not self._skip_fire:
                pre_m, pre_k = self.pre_meter.value(), self.pre_kg.value()
                if pre_m > 0 or pre_k > 0:
                    d["fire_active"] = True
                    d["pre_meter"] = pre_m
                    d["pre_kg"] = pre_k
                    base = pre_m if pre_m > 0 else pre_k
                    out  = d["meter"] if pre_m > 0 else d["kg"]
                    d["fire_pct"] = max(0.0, (base - out) / base * 100) if base > 0 else 0.0
                    d["fire_manual"] = self._manual_pct is not None
        return d


class MultiCikisDialog(QDialog):
    """Birden çok kumaş satırını (aynı ürün/farklı lot) tek hedefe toplu çıkış."""
    def __init__(self, parent, fabrics):
        super().__init__(parent)
        self.fabrics = [dict(f) for f in fabrics]
        self.setWindowTitle(f"Toplu Çıkış — {len(self.fabrics)} kalem")
        self.setMinimumSize(780, 500)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        form = QFormLayout(); form.setSpacing(8)
        self.dest_type = QComboBox(); self.dest_type.addItems(["Müşteri", "Lokasyon", "Diğer"])
        self.dest_type.currentIndexChanged.connect(self._on_type)
        self.dest_customer = QComboBox(); _make_searchable(self.dest_customer)
        self.dest_customer.addItem("— Seçiniz —", "")
        for c in (db.get_all_customers() or []):
            self.dest_customer.addItem(c["name"], c["name"])
        self.dest_loc = QComboBox(); _make_searchable(self.dest_loc)
        self.dest_loc.addItem("— Seçiniz —", "")
        for l in (db.get_active_locations() or []):
            self.dest_loc.addItem(l["name"], l["name"])
        self.dest_other = QLineEdit(); self.dest_other.setPlaceholderText("Hedef açıklayınız...")
        self.notes = QLineEdit()
        self.lbl_customer = QLabel("Müşteri:"); self.lbl_loc = QLabel("Lokasyon:"); self.lbl_other = QLabel("Diğer:")
        form.addRow("Çıkış Hedefi:", self.dest_type)
        form.addRow(self.lbl_customer, self.dest_customer)
        form.addRow(self.lbl_loc, self.dest_loc)
        form.addRow(self.lbl_other, self.dest_other)
        form.addRow("Not:", self.notes)
        lay.addLayout(form)

        cols = ["Ürün Kodu", "Lot", "Renk", "Lokasyon", "Mevcut mt", "Mevcut kg", "Çıkış mt", "Çıkış kg"]
        self.tbl = QTableWidget(); self.tbl.setColumnCount(len(cols))
        self.tbl.setHorizontalHeaderLabels(cols)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setRowCount(len(self.fabrics))
        self._spins = []
        for i, f in enumerate(self.fabrics):
            av_m = f.get("meter") or 0; av_k = f.get("kg") or 0
            for c, txt in enumerate([f.get("product_code") or "", f.get("lot") or "",
                                     f.get("color") or "", f.get("location") or "",
                                     f"{av_m:,.2f}", f"{av_k:,.2f}"]):
                it = QTableWidgetItem(txt)
                if c >= 4:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.tbl.setItem(i, c, it)
            ms = QDoubleSpinBox(); ms.setRange(0, av_m if av_m > 0 else 999999); ms.setDecimals(2); ms.setValue(av_m)
            ks = QDoubleSpinBox(); ks.setRange(0, av_k if av_k > 0 else 999999); ks.setDecimals(2); ks.setValue(av_k)
            self.tbl.setCellWidget(i, 6, ms); self.tbl.setCellWidget(i, 7, ks)
            self._spins.append((ms, ks))
        self.tbl.resizeColumnsToContents()
        lay.addWidget(self.tbl)

        row = QHBoxLayout()
        b_full = QPushButton("Tümünü tam çıkış"); b_full.clicked.connect(self._fill_full)
        b_zero = QPushButton("Sıfırla"); b_zero.clicked.connect(self._fill_zero)
        row.addWidget(b_full); row.addWidget(b_zero); row.addStretch()
        lay.addLayout(row)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.button(QDialogButtonBox.StandardButton.Ok).setText("Çıkışı Yap")
        bb.accepted.connect(self._accept); bb.rejected.connect(self.reject)
        lay.addWidget(bb)
        self._on_type()

    def _on_type(self):
        t = self.dest_type.currentText()
        for w in (self.lbl_customer, self.dest_customer): w.setVisible(t == "Müşteri")
        for w in (self.lbl_loc, self.dest_loc): w.setVisible(t == "Lokasyon")
        for w in (self.lbl_other, self.dest_other): w.setVisible(t == "Diğer")

    def _fill_full(self):
        for i, f in enumerate(self.fabrics):
            self._spins[i][0].setValue(f.get("meter") or 0)
            self._spins[i][1].setValue(f.get("kg") or 0)

    def _fill_zero(self):
        for ms, ks in self._spins:
            ms.setValue(0); ks.setValue(0)

    def _dest(self):
        t = self.dest_type.currentText()
        if t == "Müşteri":
            return (self.dest_customer.currentData() or "", t)
        if t == "Lokasyon":
            return (self.dest_loc.currentData() or "", t)
        return (self.dest_other.text().strip(), t)

    def _accept(self):
        dest, t = self._dest()
        if not dest:
            return QMessageBox.warning(self, "Eksik", "Lütfen çıkış hedefi seçin.")
        if not any(ms.value() > 0 or ks.value() > 0 for ms, ks in self._spins):
            return QMessageBox.warning(self, "Eksik", "En az bir satıra çıkış miktarı girin.")
        self.accept()

    def result_data(self):
        dest, t = self._dest()
        rows = [(self.fabrics[i]["id"], ms.value(), ks.value())
                for i, (ms, ks) in enumerate(self._spins)]
        return dest, t, self.notes.text().strip(), rows


class SplitCikisDialog(QDialog):
    """Tek lottan birden çok hedefe parçalı çıkış (ör. 800 müşteri + 100 depo + 50 kartela).
    Fire, toplam çıkış metresine göre hesaplanır."""
    def __init__(self, parent, fabric):
        super().__init__(parent)
        self.fabric = dict(fabric)
        self.setWindowTitle(f"Çıkış — {self.fabric.get('product_code','')} / {self.fabric.get('color','')}")
        self.setMinimumSize(720, 560)
        locs = db.get_active_locations() or []
        dis = {l["name"] for l in locs if l["group_name"] != "DEPO"}
        self._is_dis_depo = (self.fabric.get("location") or "") in dis
        self._loc_names = {l["name"] for l in locs}   # lokasyon adları → transfer
        self._cust_names = {c["name"] for c in (db.get_all_customers() or [])}
        self._dest_options = []
        for c in (db.get_all_customers() or []):
            self._dest_options.append((f"👤 {c['name']}", ("Müşteri", c["name"])))
        for l in locs:
            self._dest_options.append((f"🏭 {l['name']}", ("Lokasyon", l["name"])))
        self._rows = []
        self._lines = []
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        f = self.fabric
        av_m = f.get("meter") or 0; av_k = f.get("kg") or 0
        info = QLabel(f"<b>{f.get('product_name') or f.get('product_code')}</b>  |  Renk: {f.get('color','')}  |  "
                     f"Lot: {f.get('lot','')}  |  Lokasyon: {f.get('location','')}<br>"
                     f"Mevcut: <b>{av_m:,.2f} mt / {av_k:,.2f} kg</b>")
        info.setStyleSheet("background:#FFEBEE; padding:8px; border-radius:4px;")
        lay.addWidget(info)

        form = QFormLayout(); form.setSpacing(8)
        self._form = form
        self.out_fabric_type = QComboBox(); self.out_fabric_type.addItem("— Seçiniz —", "")
        for t in ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI"]:
            self.out_fabric_type.addItem(t, t)
        i = self.out_fabric_type.findData(f.get("fabric_type") or "")
        if i >= 0: self.out_fabric_type.setCurrentIndex(i)
        self.out_print_type = QComboBox(); self.out_print_type.addItem("— Seçiniz —", "")
        for t in PRINT_TYPES: self.out_print_type.addItem(t, t)
        pt_i = self.out_print_type.findData(f.get("print_type") or "")
        if pt_i >= 0: self.out_print_type.setCurrentIndex(pt_i)
        self.out_color = QLineEdit()
        self.out_zemin_rengi = QLineEdit(f.get("zemin_rengi") or "")
        self.lab_no = QLineEdit()
        self.out_baski_desen_no = QLineEdit(f.get("baski_desen_no") or "")
        self.parti_no = QLineEdit()
        form.addRow("Kumaş Tipi:", self.out_fabric_type)
        form.addRow("Baskı Tipi:", self.out_print_type)
        form.addRow("Renk:", self.out_color)
        form.addRow("Zemin Rengi:", self.out_zemin_rengi)
        form.addRow("Lab No:", self.lab_no)
        form.addRow("Baskı Desen No:", self.out_baski_desen_no)
        form.addRow("Parti No:", self.parti_no)
        # Kumaş/baskı tipine göre ilgili satırları göster/gizle
        self.out_fabric_type.currentIndexChanged.connect(self._update_out_fields)
        self.out_print_type.currentIndexChanged.connect(self._update_out_fields)
        if self._is_dis_depo:
            self.pre_meter = QDoubleSpinBox(); self.pre_meter.setRange(0, 999999); self.pre_meter.setDecimals(2); self.pre_meter.setValue(av_m)
            self.pre_kg = QDoubleSpinBox(); self.pre_kg.setRange(0, 999999); self.pre_kg.setDecimals(2); self.pre_kg.setValue(av_k)
            self.fire_label = QLabel("—"); self.fire_label.setStyleSheet("color:#C62828; font-weight:bold;")
            form.addRow("Çıkış Öncesi Metre:", self.pre_meter)
            form.addRow("Çıkış Öncesi Kilo:", self.pre_kg)
            form.addRow("Fire (toplam çıkışa göre):", self.fire_label)
            self.pre_meter.valueChanged.connect(self._update_totals)
            self.pre_kg.valueChanged.connect(self._update_totals)
        self.notes = QLineEdit()
        form.addRow("Not:", self.notes)
        lay.addLayout(form)

        lay.addWidget(QLabel("<b>Çıkış Hedefleri</b> — aynı lottan birden çok hedefe parçalı çıkış:"))
        self.tbl = QTableWidget(); self.tbl.setColumnCount(3)
        self.tbl.setHorizontalHeaderLabels(["Hedef (Müşteri / Lokasyon / serbest)", "Metre", "Kilo"])
        self.tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setMinimumHeight(160)
        lay.addWidget(self.tbl)

        brow = QHBoxLayout()
        b_add = QPushButton("+ Hedef Satırı"); b_add.clicked.connect(self._add_row)
        b_del = QPushButton("- Satır Sil"); b_del.clicked.connect(self._del_row)
        self.total_lbl = QLabel(""); self.total_lbl.setStyleSheet("font-weight:bold; color:#1565C0;")
        brow.addWidget(b_add); brow.addWidget(b_del); brow.addStretch(); brow.addWidget(self.total_lbl)
        lay.addLayout(brow)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.button(QDialogButtonBox.StandardButton.Ok).setText("Çıkışı Yap")
        bb.accepted.connect(self._accept); bb.rejected.connect(self.reject)
        lay.addWidget(bb)

        self._add_row(); self._add_row()
        self._update_totals()
        self._update_out_fields()

    def _update_out_fields(self):
        """BASKILI/Baskı Tipi seçimine göre Renk, Zemin Rengi, Lab No,
        Baskı Tipi ve Baskı Desen No satırlarını göster/gizle."""
        is_baskili = self.out_fabric_type.currentData() == "BASKILI"
        pt = self.out_print_type.currentData() if is_baskili else ""
        is_ronjan = pt == "RONJAN"
        for widget, visible in (
            (self.out_print_type, is_baskili),
            (self.out_color, not is_baskili),
            (self.out_zemin_rengi, is_ronjan),
            (self.lab_no, (not is_baskili) or is_ronjan),
            (self.out_baski_desen_no, is_baskili and bool(pt)),
        ):
            label = self._form.labelForField(widget)
            if label:
                label.setVisible(visible)
            widget.setVisible(visible)

    def _add_row(self):
        r = self.tbl.rowCount(); self.tbl.insertRow(r)
        hedef = QComboBox(); hedef.setEditable(True); _make_searchable(hedef)
        hedef.addItem("", None)
        for lbl, data in self._dest_options:
            hedef.addItem(lbl, data)
        hedef.setCurrentIndex(0)
        ms = QDoubleSpinBox(); ms.setRange(0, 999999); ms.setDecimals(2); ms.valueChanged.connect(self._update_totals)
        ks = QDoubleSpinBox(); ks.setRange(0, 999999); ks.setDecimals(2); ks.valueChanged.connect(self._update_totals)
        self.tbl.setCellWidget(r, 0, hedef); self.tbl.setCellWidget(r, 1, ms); self.tbl.setCellWidget(r, 2, ks)
        self._rows.append((hedef, ms, ks))

    def _del_row(self):
        if self.tbl.rowCount() > 1:
            self.tbl.removeRow(self.tbl.rowCount() - 1)
            self._rows.pop()
            self._update_totals()

    def _update_totals(self):
        tot_m = sum(ms.value() for _, ms, _ in self._rows)
        tot_k = sum(ks.value() for _, _, ks in self._rows)
        self.total_lbl.setText(f"Toplam çıkış: {tot_m:,.2f} mt / {tot_k:,.2f} kg")
        if self._is_dis_depo:
            pm = self.pre_meter.value()
            fire = pm - tot_m
            pct = (fire / pm * 100) if pm > 0 else 0
            self.fire_label.setText(f"{fire:,.2f} mt  (%{pct:.1f})")

    def _parse_hedef(self, combo):
        # Hedef adını al (combo data'sından ya da yazılan metinden)
        data = combo.currentData()
        dest = data[1] if data else combo.currentText().strip()
        for pre in ("👤 ", "🏭 "):
            if dest.startswith(pre):
                dest = dest[len(pre):]
        dest = dest.strip()
        if not dest:
            return (None, None)
        # Tipi HER ZAMAN isme göre belirle: lokasyon → transfer, müşteri → düş
        if dest in self._loc_names:
            return ("Lokasyon", dest)
        if dest in self._cust_names:
            return ("Müşteri", dest)
        return ("Diğer", dest)

    def _accept(self):
        lines = []
        for hedef, ms, ks in self._rows:
            if ms.value() > 0 or ks.value() > 0:
                t, dest = self._parse_hedef(hedef)
                if not dest:
                    return QMessageBox.warning(self, "Eksik", "Miktar girilen satırda hedef seçin/yazın.")
                lines.append((t, dest, ms.value(), ks.value()))
        if not lines:
            return QMessageBox.warning(self, "Eksik", "En az bir hedefe çıkış miktarı girin.")
        tot_m = sum(l[2] for l in lines)
        limit = self.pre_meter.value() if self._is_dis_depo else (self.fabric.get("meter") or 0)
        if limit > 0 and tot_m - limit > 0.01:
            return QMessageBox.warning(self, "Hata",
                f"Toplam çıkış ({tot_m:,.0f} mt) mevcut/çıkış öncesi miktarı ({limit:,.0f} mt) aşıyor.")
        self._lines = lines
        self.accept()

    def result(self):
        ft = self.out_fabric_type.currentData() or ""
        is_bask = ft == "BASKILI"
        pt = (self.out_print_type.currentData() or "") if is_bask else ""
        out = {"out_fabric_type": ft, "out_print_type": pt,
               "out_color": self.out_color.text().strip().upper() if not is_bask else "",
               "out_zemin_rengi": self.out_zemin_rengi.text().strip().upper(),
               "lab_no": self.lab_no.text().strip(),
               "out_baski_desen_no": self.out_baski_desen_no.text().strip(),
               "parti_no": self.parti_no.text().strip(),
               "notes": self.notes.text().strip()}
        fire = None
        if self._is_dis_depo:
            pm, pk = self.pre_meter.value(), self.pre_kg.value()
            tot_m = sum(l[2] for l in self._lines); tot_k = sum(l[3] for l in self._lines)
            if pm > tot_m + 0.01 or pk > tot_k + 0.01:
                pct = ((pm - tot_m) / pm * 100) if pm > 0 else 0
                fire = {"pre_m": pm, "pre_k": pk, "out_m": tot_m, "out_k": tot_k, "pct": pct}
        return self._lines, out, fire


TYPE_COLORS = {"GİRİŞ": "#2E7D32", "SATINALMA GİRİŞİ": "#1565C0",
               "ÇIKIŞ": "#C62828", "SİLME": "#880E4F"}


def _fill_movement_table(table, movements, show_product=False):
    """Hareket tablosunu doldur. show_product=True ise ürün sütunu eklenir.
    Sütunlar elle ayarlanabilir, başlıklar taşınabilir, tıklayınca sıralanır."""
    if show_product:
        cols = ["Tarih", "Tür", "Ürün Kodu", "Renk", "Lot", "Satın Alma Lok.", "Lokasyon", "Metre", "Kilo", "Top/Adet", "Hedef", "Kullanıcı", "Not"]
    else:
        cols = ["Tarih", "Tür", "Metre", "Kilo", "Top/Adet", "Hedef", "Kullanıcı", "Not"]

    table.setColumnCount(len(cols))
    table.setHorizontalHeaderLabels(cols)
    hdr = table.horizontalHeader()
    hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)   # elle genişlik
    hdr.setStretchLastSection(True)
    hdr.setSectionsMovable(True)   # başlıklar sürüklenerek yer değiştirir
    table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    table.setAlternatingRowColors(True)
    table.verticalHeader().setVisible(False)
    table.setSortingEnabled(False)   # doldururken kapalı
    table.setRowCount(len(movements))

    for i, m in enumerate(movements):
        m = dict(m)   # sqlite3.Row'da .get yok; dict'e çevirince iki kaynak da çalışır
        col = 0
        def _set(val, align=None, color=None, bold=False, sort_key=None):
            nonlocal col
            text = str(val) if val else ""
            item = _FireSortItem(text)
            if sort_key is not None:
                item.setData(Qt.ItemDataRole.UserRole, sort_key)
            if text:
                item.setToolTip(text)   # sığmayan yazılar üzerine gelince okunur
            if align:
                item.setTextAlignment(align)
            if color:
                item.setForeground(QBrush(QColor(color)))
            if bold:
                item.setFont(QFont("", -1, QFont.Weight.Bold))
            table.setItem(i, col, item)
            col += 1

        _set(str(m["movement_date"])[:16])
        t = m["movement_type"]
        _set(t, color=TYPE_COLORS.get(t, "#333"), bold=True)
        if show_product:
            _set(m["product_code"] or "")
            _set(m["color"] or "")
            _set(m.get("lot") or "", color="#78909C")             # lot
            _set(m.get("entry_location") or "", color="#00695C")   # köken
            # Hareketin yapıldığı lokasyon; eski kayıtlarda kumaşın güncel lokasyonu
            _set(m.get("location") or m.get("fabric_location") or "")
        _set(f"{m['meter']:,.2f}" if m["meter"] else "", sort_key=m["meter"] or 0)
        _set(f"{m['kg']:,.2f}" if m["kg"] else "", sort_key=m["kg"] or 0)
        try:
            piece_key = int(str(m["piece_count"]).strip())
        except (ValueError, TypeError):
            piece_key = -1
        _set(m["piece_count"] or "", sort_key=piece_key)
        dest = m.get("destination") or ""
        dest_type = m.get("destination_type") or ""
        dest_label = f"{dest_type}: {dest}" if dest and dest_type else dest
        _set(dest_label, color="#1565C0" if dest else None)
        _set(m["user_name"] or "")
        extra = []
        if m.get("out_color") and m.get("out_color") != m.get("color"):
            extra.append(f"Renk: {m['out_color']}")
        if m.get("lab_no"):
            extra.append(f"Lab: {m['lab_no']}")
        if m.get("parti_no"):
            extra.append(f"Parti: {m['parti_no']}")
        notes = m["notes"] or ""
        if extra:
            notes = (notes + "  |  " if notes else "") + "  ".join(extra)
        _set(notes)

    table.setSortingEnabled(True)   # başlığa tıklayınca sıralar

    # Sütun düzeni (sıra + genişlik) kalıcı: kapatıp açınca aynı kalır.
    # NOT: saveState() sectionResized içinde SENKRON çağrılırsa (sürükleme sırasında
    # art arda tetiklenir) Windows/macOS'ta reentrancy → çökme (SIGSEGV) olur.
    # Bu yüzden QTimer ile geciktirip event-loop'a döndükten sonra kaydeden
    # ortak _wire_header_persistence yardımcısı kullanılır.
    key = "mv_header_p2" if show_product else "mv_header_s"
    _wire_header_persistence(table, key)


class MovementsDialog(QDialog):
    """Tek ürünün tüm hareketleri."""
    def __init__(self, parent, fabric):
        super().__init__(parent)
        self.setWindowTitle(f"Tüm Hareketler — {fabric['product_code']} / {fabric['color']}")
        self.setMinimumSize(750, 480)
        layout = QVBoxLayout(self)

        entry = dict(fabric).get("entry_location") or ""
        entry_html = f" — <span style='color:#00695C'>Satın Alma Lok: {entry}</span>" if entry else ""
        info = QLabel(
            f"<b>{fabric['product_name'] or ''} {fabric['product_code']}</b>"
            f" — {fabric['color']} — <span style='color:#545454'>{fabric['location']}</span>"
            f"{entry_html}"
        )
        info.setStyleSheet("font-size:14px; padding:6px; background:#E3F2FD; border-radius:4px;")
        layout.addWidget(info)

        table = QTableWidget()
        movements = db.get_movements(fabric["id"])
        _fill_movement_table(table, movements, show_product=False)
        layout.addWidget(table)

        row = QHBoxLayout()
        lbl = QLabel(f"<span style='color:#555'>{len(movements)} hareket kaydı</span>")
        btn = QPushButton("Kapat"); btn.clicked.connect(self.accept)
        row.addWidget(lbl); row.addStretch(); row.addWidget(btn)
        layout.addLayout(row)


class DailyMovementsDialog(QDialog):
    """Tarih aralığındaki tüm hareketler — seçili satır yokken açılır."""
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Hareketler")
        self.setMinimumSize(950, 540)
        self._build_ui()
        self._load()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        from PyQt6.QtWidgets import QDateEdit
        from PyQt6.QtCore import QDate

        # Hazır aralık butonları (Tümü = kalitenin girişten tükenişe tüm geçmişi)
        presets = QHBoxLayout()
        for label, days in [("Bugün", 0), ("Son 7 Gün", 7), ("Son 15 Gün", 15),
                            ("Son 30 Gün", 30), ("Tümü", None)]:
            b = QPushButton(label)
            b.clicked.connect(lambda _, d=days: self._set_preset(d))
            presets.addWidget(b)
        presets.addStretch()
        layout.addLayout(presets)

        # Filtreler — ürün/lokasyon/tip/tür
        filt = QHBoxLayout()
        filt.addWidget(QLabel("Ara:"))
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Ürün kodu, renk, lot, not, hedef...")
        self.search_box.setMinimumWidth(200)
        self.search_box.textChanged.connect(self._load)
        filt.addWidget(self.search_box)

        filt.addWidget(QLabel("Lokasyon:"))
        self.loc_filter = QComboBox()
        self.loc_filter.addItem("Tümü", "")
        # Lokasyonun grup adına göre: DEPO grubundaki alt depolar tek "DEPO"da toplanır,
        # dış (DEPO olmayan) lokasyonlar tek tek listelenir.
        self._loc_to_group = {l["name"]: (l["group_name"] or "DEPO")
                              for l in (db.get_all_locations() or [])}
        self.loc_filter.addItem("DEPO", "DEPO")
        for l in db.get_active_locations():
            if self._loc_to_group.get(l["name"], "DEPO") != "DEPO":
                self.loc_filter.addItem(l["name"], l["name"])
        self.loc_filter.currentIndexChanged.connect(self._load)
        filt.addWidget(self.loc_filter)

        filt.addWidget(QLabel("Hedef:"))
        self.hedef_filter = QComboBox()
        for lbl, val in [("Tümü", ""), ("Müşteri", "Müşteri"),
                         ("Depo", "Depo"), ("Dış Depo", "Dış Depo")]:
            self.hedef_filter.addItem(lbl, val)
        self.hedef_filter.currentIndexChanged.connect(self._load)
        filt.addWidget(self.hedef_filter)

        filt.addWidget(QLabel("Tip:"))
        self.tip_filter = QComboBox()
        self.tip_filter.addItem("Tümü", "")
        for t in ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI"]:
            self.tip_filter.addItem(t, t)
        self.tip_filter.currentIndexChanged.connect(self._load)
        filt.addWidget(self.tip_filter)

        filt.addWidget(QLabel("Tür:"))
        self.tur_filter = QComboBox()
        self.tur_filter.addItem("Tümü", "")
        for t in ["GİRİŞ", "SATINALMA GİRİŞİ", "ÇIKIŞ", "SİLME"]:
            self.tur_filter.addItem(t, t)
        self.tur_filter.currentIndexChanged.connect(self._load)
        filt.addWidget(self.tur_filter)
        filt.addStretch()
        layout.addLayout(filt)

        # Tarih aralığı seçici
        top = QHBoxLayout()
        top.addWidget(QLabel("Başlangıç:"))
        self.start_edit = QDateEdit()
        self.start_edit.setCalendarPopup(True)
        self.start_edit.setDate(QDate.currentDate())
        self.start_edit.setDisplayFormat("dd.MM.yyyy")
        self.start_edit.dateChanged.connect(self._load)
        top.addWidget(self.start_edit)
        top.addWidget(QLabel("Bitiş:"))
        self.end_edit = QDateEdit()
        self.end_edit.setCalendarPopup(True)
        self.end_edit.setDate(QDate.currentDate())
        self.end_edit.setDisplayFormat("dd.MM.yyyy")
        self.end_edit.dateChanged.connect(self._load)
        top.addWidget(self.end_edit)
        top.addStretch()
        self.count_lbl = QLabel()
        self.count_lbl.setStyleSheet("color:#555; font-size:12px;")
        top.addWidget(self.count_lbl)
        layout.addLayout(top)

        self.table = QTableWidget()
        layout.addWidget(self.table)

        # ── Alt toplam çubuğu ─────────────────────────────────────
        self.totals_lbl = QLabel()
        self.totals_lbl.setStyleSheet(
            "background:#37474F; color:white; font-weight:bold; font-size:12px;"
            "border-radius:4px; padding:8px 12px;")
        layout.addWidget(self.totals_lbl)

        btn = QPushButton("Kapat"); btn.clicked.connect(self.accept)
        layout.addWidget(btn)

    def _hedef_kategori(self, md):
        """Hareketin hedefini Müşteri / Depo / Dış Depo olarak sınıflandırır."""
        dt = (md.get("destination_type") or "").strip()
        dest = (md.get("destination") or "").strip()
        if dt == "Müşteri":
            return "Müşteri"
        if dt == "Lokasyon" or (dest and dest in self._loc_to_group):
            return "Depo" if self._loc_to_group.get(dest, "DEPO") == "DEPO" else "Dış Depo"
        if dest:
            return "Müşteri"      # tip belirsiz ama hedef varsa müşteri sevkiyatı say
        return ""

    def _set_preset(self, days):
        from PyQt6.QtCore import QDate
        today = QDate.currentDate()
        # Sinyalleri kapat, iki tarihi birden ayarla, tek seferde yükle
        self.start_edit.blockSignals(True)
        self.end_edit.blockSignals(True)
        if days is None:   # Tümü — geçmişin tamamı
            self.start_edit.setDate(QDate(2000, 1, 1))
        else:
            self.start_edit.setDate(today.addDays(-days))
        self.end_edit.setDate(today)
        self.start_edit.blockSignals(False)
        self.end_edit.blockSignals(False)
        self._load()

    def _load(self):
        start = self.start_edit.date().toString("yyyy-MM-dd")
        end = self.end_edit.date().toString("yyyy-MM-dd")
        if start > end:
            start, end = end, start
        all_mv = db.get_movements_by_range(start, end)

        # Filtreler
        q   = self.search_box.text().strip().lower()
        loc = self.loc_filter.currentData() or ""
        hedef = self.hedef_filter.currentData() or ""
        tip = self.tip_filter.currentData() or ""
        tur = self.tur_filter.currentData() or ""
        movements = []
        for m in all_mv:
            md = dict(m)
            if tur and md["movement_type"] != tur:
                continue
            if loc:
                mloc = md.get("location") or md.get("fabric_location") or ""
                if loc == "DEPO":
                    if self._loc_to_group.get(mloc, "DEPO") != "DEPO":
                        continue
                elif mloc != loc:
                    continue
            if hedef and self._hedef_kategori(md) != hedef:
                continue
            if tip and (md.get("fabric_type") or "") != tip:
                continue
            if q:
                hay = " ".join(str(md.get(k) or "") for k in
                               ("product_code", "product_name", "color", "out_color",
                                "lot", "notes", "destination", "lab_no", "parti_no",
                                "user_name", "entry_location")).lower()
                if q not in hay:
                    continue
            movements.append(m)

        _fill_movement_table(self.table, movements, show_product=True)
        if movements:
            def _is_giris(m): return m["movement_type"] in ("GİRİŞ", "SATINALMA GİRİŞİ")
            def _is_cikis(m): return m["movement_type"] == "ÇIKIŞ"
            in_m  = sum(m["meter"] or 0 for m in movements if _is_giris(m))
            in_k  = sum(m["kg"] or 0    for m in movements if _is_giris(m))
            out_m = sum(m["meter"] or 0 for m in movements if _is_cikis(m))
            out_k = sum(m["kg"] or 0    for m in movements if _is_cikis(m))
            # hedef kategorisine göre çıkış metresi alt toplamı
            kat = {"Müşteri": 0.0, "Depo": 0.0, "Dış Depo": 0.0}
            for m in movements:
                if _is_cikis(m):
                    k = self._hedef_kategori(dict(m))
                    if k in kat:
                        kat[k] += m["meter"] or 0
            suffix = f" / toplam {len(all_mv)}" if len(movements) != len(all_mv) else ""
            self.count_lbl.setText(f"{len(movements)} hareket{suffix}")
            self.totals_lbl.setText(
                f"📥 Giriş: {in_m:,.0f} mt / {in_k:,.0f} kg     "
                f"📤 Çıkış: {out_m:,.0f} mt / {out_k:,.0f} kg     "
                f"⚖️ Net: {in_m - out_m:,.0f} mt / {in_k - out_k:,.0f} kg          "
                f"Çıkış dağılımı →  👤 Müşteri: {kat['Müşteri']:,.0f} mt   "
                f"🏭 Depo: {kat['Depo']:,.0f} mt   🚚 Dış Depo: {kat['Dış Depo']:,.0f} mt"
            )
        else:
            self.count_lbl.setText("Bu kriterlere uyan hareket yok")
            self.totals_lbl.setText("Toplam: 0")


COLS = ["#", "Ürün Kodu", "Ürün Açıklaması", "Açıklama", "Renk", "Lokasyon", "Tip", "Lot", "Metre", "Kilo", "Top/Adet", "Birim Fiyat $", "Toplam Değer $", "Son Güncelleme", "Satın Alma Lok.", "Lab No", "Baskı Tipi", "Zemin Rengi", "Baskı Desen No", "Numune Kodu"]
_GREEN = QColor("#1B5E20")
_GREY  = QColor("#BDBDBD")
_ALT   = QColor("#F0F4FF")


def _export_table_to_excel(parent, model, table_view, title="Stok Raporu"):
    """Model verilerini sütun sırasına göre (görsel sıra dahil) Excel'e aktar."""
    path, _ = QFileDialog.getSaveFileName(
        parent, "Excel'e Aktar", f"{title}.xlsx", "Excel (*.xlsx)"
    )
    if not path:
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        # Görsel sütun sırası (kullanıcı sütunları taşımış olabilir)
        hdr = table_view.horizontalHeader()
        col_count = model.columnCount()
        visual_order = [hdr.logicalIndex(vi) for vi in range(col_count)]

        # Başlık satırı
        headers = [model.headerData(li, Qt.Orientation.Horizontal) or "" for li in visual_order]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1565C0")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18

        # Veri satırları
        for row_i in range(model.rowCount()):
            row_data = []
            for li in visual_order:
                idx = model.index(row_i, li)
                val = model.data(idx, Qt.ItemDataRole.DisplayRole) or ""
                # Sayısal değerler için float'a çevir
                if li in (7, 8, 10, 11):
                    try:
                        val = float(str(val).replace(" $", "").replace(",", ""))
                    except Exception:
                        pass
                row_data.append(val)
            ws.append(row_data)

        # Sütun genişlikleri
        for i, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        # Footer notu
        ws.append([])
        ws.append([f"Dışa aktarıldı: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  {model.rowCount()} kayıt"])

        wb.save(path)
        QMessageBox.information(parent, "Başarılı", f"{model.rowCount()} kayıt dışa aktarıldı:\n{path}")
    except Exception as e:
        QMessageBox.critical(parent, "Hata", f"Dışa aktarma hatası:\n{e}")


def _export_widget_table_to_excel(parent, table_widget, title="Rapor"):
    """QTableWidget içeriğini Excel'e aktar (Lokasyon görünümü için)."""
    path, _ = QFileDialog.getSaveFileName(
        parent, "Excel'e Aktar", f"{title}.xlsx", "Excel (*.xlsx)"
    )
    if not path:
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        hdr = table_widget.horizontalHeader()
        col_count = table_widget.columnCount()
        visual_order = [hdr.logicalIndex(vi) for vi in range(col_count)]

        headers = [table_widget.horizontalHeaderItem(li).text()
                   if table_widget.horizontalHeaderItem(li) else "" for li in visual_order]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1565C0")
            cell.alignment = Alignment(horizontal="center")

        for row_i in range(table_widget.rowCount()):
            row_data = []
            for li in visual_order:
                item = table_widget.item(row_i, li)
                row_data.append(item.text() if item else "")
            ws.append(row_data)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        ws.append([])
        ws.append([f"Dışa aktarıldı: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  {table_widget.rowCount()} kayıt"])

        wb.save(path)
        QMessageBox.information(parent, "Başarılı", f"{table_widget.rowCount()} kayıt dışa aktarıldı:\n{path}")
    except Exception as e:
        QMessageBox.critical(parent, "Hata", f"Dışa aktarma hatası:\n{e}")


import re
_AUTO_LOT = re.compile(r"^LOT-\d{8}-\d+$")   # otomatik üretilen lot formatı


class FabricModel(QAbstractTableModel):
    def __init__(self):
        super().__init__()
        self._rows = []
        self._ids  = []

    # tuple: 0=id,1=code,2=name,3=desc,4=color,5=loc,6=tip,7=lot,8=mt,9=kg,10=piece,11=fiyat,12=deger,13=date,14=girisLok,15=labNo,16=printType,17=zeminRengi,18=baskiDesenNo
    def load(self, rows):
        self.beginResetModel()
        self._rows = []
        for r in rows:
            mt    = r["meter"] or 0.0
            kg    = r["kg"] or 0.0
            fiy   = r["birim_fiyat"] or 0.0
            deger = mt * fiy if mt > 0 else (kg * fiy if kg > 0 else 0.0)
            self._rows.append((
                r["id"],
                r["product_code"] or "",
                r["product_name"] or "",
                r["description"] or "",
                r["color"] or "",
                r["location"] or "",
                r["fabric_type"] or "",
                r["lot"] or "",
                mt,
                kg,
                r["piece_count"] or "",
                fiy,
                deger,
                str(r["updated_at"] or "")[:16],
                dict(r).get("entry_location") or "",
                dict(r).get("lab_no") or "",
                dict(r).get("print_type") or "",
                dict(r).get("zemin_rengi") or "",
                dict(r).get("baski_desen_no") or "",
                dict(r).get("numune_code") or "",
            ))
        self._ids = [r[0] for r in self._rows]
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self._rows)

    def columnCount(self, parent=QModelIndex()):
        return len(COLS)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return COLS[section]
        if role == Qt.ItemDataRole.FontRole and orientation == Qt.Orientation.Horizontal:
            f = QFont(); f.setBold(True); return f
        return None

    def sort(self, col, order=Qt.SortOrder.AscendingOrder):
        if col == 0:   # # sütunu — sıralama yapma
            return
        self.layoutAboutToBeChanged.emit()
        reverse = (order == Qt.SortOrder.DescendingOrder)
        # Sayısal sütunlar: 8=mt, 9=kg, 11=fiyat, 12=değer
        numeric = {8, 9, 11, 12}
        def key(r):
            v = r[col]
            if col in numeric:
                return v if isinstance(v, (int, float)) else 0
            return str(v).lower()
        self._rows.sort(key=key, reverse=reverse)
        self._ids = [r[0] for r in self._rows]
        self.layoutChanged.emit()

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        row, col = index.row(), index.column()
        r = self._rows[row]

        # col: 0=#,1=code,2=name,3=desc,4=color,5=loc,6=tip,7=lot,8=mt,9=kg,10=piece,11=fiyat,12=deger,13=date,14=girisLok,15=labNo,16=printType,17=zeminRengi,18=baskiDesenNo
        if role == Qt.ItemDataRole.DisplayRole:
            if col == 0: return str(row + 1)
            val = r[col]
            if col in (8, 9): return f"{val:.2f}"
            if col == 11: return f"{val:,.2f} $" if val else "—"
            if col == 12: return f"{val:,.2f} $" if val else "—"
            return str(val)

        if role == Qt.ItemDataRole.TextAlignmentRole:
            if col in (8, 9, 11, 12):
                return int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight)
            return int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

        if role == Qt.ItemDataRole.ForegroundRole:
            if col in (8, 9):
                return QBrush(_GREY if r[col] == 0 else _GREEN)
            if col == 12 and r[12] > 0:
                return QBrush(QColor("#1A237E"))
            if col == 6:
                return QBrush({"HAM": QColor("#5D4037"),
                                "PFD": QColor("#00695C"),
                                "BOYALI": QColor("#545454"),
                                "İPLİĞİ BOYALI": QColor("#EF6C00"),
                                "BASKILI": QColor("#6A1B9A")}.get(r[6], QColor("#333")))
            if col == 7 and _AUTO_LOT.match(r[7]):
                return QBrush(QColor("#78909C"))
            if col == 19 and r[19]:
                return QBrush(QColor("#E65100"))

        if role == Qt.ItemDataRole.FontRole:
            if col == 7 and _AUTO_LOT.match(r[7]):
                f = QFont(); f.setItalic(True); return f
            if col == 19 and r[19]:
                f = QFont(); f.setItalic(True); return f

        if role == Qt.ItemDataRole.BackgroundRole:
            if row % 2 == 1:
                return QBrush(_ALT)

        if role == Qt.ItemDataRole.ToolTipRole:
            val = r[col]
            if col == 0: return str(row + 1)
            if col in (8, 9): return f"{val:.2f}"
            if col == 11: return f"{val:,.2f} $" if val else "Fiyat girilmemiş"
            if col == 12: return f"{val:,.2f} $" if val else "—"
            if col == 7 and _AUTO_LOT.match(r[7]):
                return f"{val}\n(Otomatik verilen lot numarası)"
            return str(val) if val else ""

        return None

    def id_at(self, row):
        if 0 <= row < len(self._ids):
            return self._ids[row]
        return None


class StockTable(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._loaded = False
        self._model = FabricModel()
        self._build_ui()
        QTimer.singleShot(0, self._first_load)

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        toolbar = QHBoxLayout()
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Ürün kodu, adı, renk ile ara...")
        self.search_box.setMinimumWidth(280)
        self.search_box.textChanged.connect(self._on_search_change)

        self.location_filter = QComboBox()
        self.location_filter.setMinimumWidth(130)
        self.location_filter.currentIndexChanged.connect(self.refresh)

        self.type_filter = QComboBox()
        self.type_filter.setMinimumWidth(100)
        self.type_filter.addItem("Tüm Tipler", "")
        for t in ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI"]:
            self.type_filter.addItem(t, t)
        self.type_filter.currentIndexChanged.connect(self.refresh)

        btn_add   = QPushButton("+ Yeni Kumaş"); btn_add.clicked.connect(self._add)
        btn_giris = QPushButton("↑ Giriş");      btn_giris.clicked.connect(self._giris)
        btn_cikis = QPushButton("↓ Çıkış");      btn_cikis.clicked.connect(self._cikis)
        btn_edit  = QPushButton("✎ Düzenle");    btn_edit.clicked.connect(self._edit)
        btn_del   = QPushButton("✕ Sil");        btn_del.clicked.connect(self._delete)
        btn_hist  = QPushButton("☰ Hareketler"); btn_hist.clicked.connect(self._history)
        btn_hist.setToolTip("Tüm hareketler — tarih aralığı ve filtrelerle\nSatır seçiliyse o ürünün geçmişi hazır gelir")

        btn_giris.setStyleSheet(f"background:{COLORS['success']}; color:white; font-weight:bold; border-radius:4px; padding:7px 14px;")
        btn_cikis.setStyleSheet(f"background:{COLORS['danger']}; color:white; font-weight:bold; border-radius:4px; padding:7px 14px;")
        btn_edit.setStyleSheet("background:#F57F17; color:white; font-weight:bold; border-radius:4px; padding:7px 14px;")
        btn_del.setStyleSheet("background:#757575; color:white; font-weight:bold; border-radius:4px; padding:7px 14px;")

        toolbar.addWidget(QLabel("Ara:")); toolbar.addWidget(self.search_box)
        toolbar.addWidget(QLabel("Lokasyon:")); toolbar.addWidget(self.location_filter)
        toolbar.addWidget(QLabel("Tip:")); toolbar.addWidget(self.type_filter)
        toolbar.addStretch()
        for b in (btn_add, btn_giris, btn_cikis, btn_edit, btn_del, btn_hist):
            toolbar.addWidget(b)
        layout.addLayout(toolbar)

        # QTableView — sadece görünen satırları render eder
        self.table = QTableView()
        self.table.setModel(self._model)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.PenStyle.SolidLine)
        self.table.doubleClicked.connect(self._cell_edit)
        self.table.verticalHeader().setDefaultSectionSize(26)
        self.table.setMouseTracking(True)
        self.table.setWordWrap(False)
        self.table.setSortingEnabled(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._context_menu)

        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(False)
        hdr.setMinimumSectionSize(40)
        hdr.setSectionsMovable(True)   # sütun sürükle-bırak

        # Başlangıç genişlikleri (kayıtlı sütun düzeni yoksa kullanılır)
        col_widths = [36, 100, 130, 160, 110, 90, 70, 90, 72, 65, 80, 110, 120, 130, 220, 90, 90, 90, 100, 95]
        def _default_widths():
            for i, w in enumerate(col_widths):
                self.table.setColumnWidth(i, w)
            # "Numune Kodu" (mantıksal 19) görsel olarak "Ürün Kodu"nun (görsel 1) yanına gelsin
            hdr.moveSection(hdr.visualIndex(19), 2)
        _wire_header_persistence(self.table, "stock_header_v2", _default_widths)

        layout.addWidget(self.table)

        # ── Toplam Barı ──
        self._totals_bar = QFrame()
        self._totals_bar.setStyleSheet(
            "QFrame { background:#545454; border-radius:6px;"
            "border: 2px solid #3A3A3A; padding:0; }"
        )
        self._totals_bar.setFixedHeight(56)
        bar_layout = QHBoxLayout(self._totals_bar)
        bar_layout.setContentsMargins(20, 4, 20, 4)
        bar_layout.setSpacing(0)

        def _stat_widget(label):
            w = QWidget()
            w.setStyleSheet("background:transparent;")
            wl = QVBoxLayout(w); wl.setSpacing(2); wl.setContentsMargins(16, 2, 16, 2)
            lbl = QLabel(label)
            lbl.setStyleSheet("color:rgba(255,255,255,.7); font-size:10px; background:transparent;")
            val = QLabel("—")
            val.setStyleSheet("color:white; font-size:15px; font-weight:bold; background:transparent;")
            wl.addWidget(lbl); wl.addWidget(val)
            return w, val

        def sep():
            f = QFrame()
            f.setFrameShape(QFrame.Shape.VLine)
            f.setFixedWidth(1)
            f.setStyleSheet("background:rgba(255,255,255,.3); margin:8px 4px;")
            return f

        w1, self._tot_items  = _stat_widget("Kalem Sayısı")
        w2, self._tot_meter  = _stat_widget("Toplam Metre")
        w3, self._tot_kg     = _stat_widget("Toplam Kilo")
        w4, self._tot_value  = _stat_widget("Toplam Değer")

        for w in (w1, sep(), w2, sep(), w3, sep(), w4):
            bar_layout.addWidget(w)
        bar_layout.addStretch()

        self._filter_lbl = QLabel()
        self._filter_lbl.setStyleSheet(
            "color:#212121; font-size:11px; font-style:italic; background:transparent;")
        bar_layout.addWidget(self._filter_lbl)

        layout.addWidget(self._totals_bar)

        # ── Seçili Satırın Ürün Bilgisi (katalogdan) ──
        self._product_info_bar = QFrame()
        self._product_info_bar.setStyleSheet(
            "QFrame { background:#E3F2FD; border-radius:6px;"
            "border: 1px solid #BBDEFB; padding:0; }"
        )
        self._product_info_bar.setFixedHeight(34)
        pi_layout = QHBoxLayout(self._product_info_bar)
        pi_layout.setContentsMargins(16, 2, 16, 2)
        self._product_info_lbl = QLabel("Ürün bilgisi için bir satır seçin")
        self._product_info_lbl.setStyleSheet("color:#1565C0; font-size:12px;")
        pi_layout.addWidget(self._product_info_lbl)
        pi_layout.addStretch()
        layout.addWidget(self._product_info_bar)

        self.table.selectionModel().currentRowChanged.connect(self._on_row_selected)

        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self.refresh)

    def _on_search_change(self):
        self._search_timer.start(250)

    def _first_load(self):
        self._reload_locations()
        self._fill_table()
        self._loaded = True

    def refresh(self):
        if not self._loaded:
            return
        self._fill_table()

    def _on_row_selected(self, current, previous):
        if not current.isValid() or current.row() >= len(self._model._rows):
            self._product_info_lbl.setText("Ürün bilgisi için bir satır seçin")
            return
        code = self._model._rows[current.row()][1]
        product = db.get_product_by_code(code)
        if not product:
            self._product_info_lbl.setText(f"{code}   —   (katalogda kayıt bulunamadı)")
            return
        p = dict(product)
        parts = [
            code,
            p.get('composition') or '—',
            p.get('width') or '—',
            p.get('gramaj') or '—',
            p.get('reference_code') or '—',
            p.get('supplier') or '—',
        ]
        self._product_info_lbl.setText("   |   ".join(parts))

    def refresh_with_locations(self):
        self._reload_locations()
        self._fill_table()

    def _reload_locations(self):
        current_loc = self.location_filter.currentData()
        self.location_filter.blockSignals(True)
        self.location_filter.clear()
        self.location_filter.addItem("Tüm Lokasyonlar", "")

        all_locs = db.get_active_locations()
        from collections import defaultdict
        groups = defaultdict(list)
        for l in all_locs:
            groups[l["group_name"]].append(l["name"])

        for grp in sorted(groups.keys()):
            if grp == "DEPO":
                # DEPO → tek seçenek, tümünü getirir
                self.location_filter.addItem("DEPO", "__GRP_DEPO__")
            else:
                # Diğer gruplar → tek tek lokasyon olarak listele
                for name in sorted(groups[grp]):
                    self.location_filter.addItem(name, name)

        if current_loc:
            idx = self.location_filter.findData(current_loc)
            if idx >= 0:
                self.location_filter.setCurrentIndex(idx)
        self.location_filter.blockSignals(False)

    def _fill_table(self):
        search  = self.search_box.text().strip()
        loc     = self.location_filter.currentData() or ""
        ftype   = self.type_filter.currentData() or ""

        if loc.startswith("__GRP_"):
            grp_name = loc[len("__GRP_"):-2]
            grp_locs = [l["name"] for l in db.get_active_locations()
                        if l["group_name"] == grp_name]
            rows = []
            for gl in grp_locs:
                rows.extend(db.get_all_fabrics(search, gl, ftype))
        else:
            rows = db.get_all_fabrics(search, loc, ftype)

        # Mevcut sıralama, seçim ve kaydırma konumunu koru
        hdr = self.table.horizontalHeader()
        sort_col = hdr.sortIndicatorSection()
        sort_ord = hdr.sortIndicatorOrder()
        idx = self.table.selectionModel().currentIndex() if self.table.selectionModel() else None
        sel_id = self._model.id_at(idx.row()) if idx is not None and idx.isValid() else None
        scroll = self.table.verticalScrollBar().value()

        self._model.load(rows)
        if sort_col > 0:
            self._model.sort(sort_col, sort_ord)

        if sel_id is not None:
            for row in range(self._model.rowCount()):
                if self._model.id_at(row) == sel_id:
                    self.table.selectRow(row)
                    break
        self.table.verticalScrollBar().setValue(scroll)

        self._update_totals(rows)

        summary = db.get_summary()
        if hasattr(self.parent(), "update_status"):
            self.parent().update_status(len(rows), summary)

    def _update_totals(self, rows):
        total_mt  = sum(r["meter"] or 0 for r in rows)
        total_kg  = sum(r["kg"] or 0 for r in rows)
        total_val = sum(
            ((r["meter"] or 0) * (r["birim_fiyat"] or 0)) if (r["meter"] or 0) > 0
            else ((r["kg"] or 0) * (r["birim_fiyat"] or 0))
            for r in rows
        )
        self._tot_items.setText(f"{len(rows):,}")
        self._tot_meter.setText(f"{total_mt:,.2f} mt")
        self._tot_kg.setText(f"{total_kg:,.2f} kg")
        self._tot_value.setText(f"{total_val:,.2f} $" if total_val else "—")

        loc = self.location_filter.currentData() or ""
        search = self.search_box.text().strip()
        parts = []
        if loc and not loc.startswith("__"):
            parts.append(f"Lokasyon: {self.location_filter.currentText().strip()}")
        elif loc.startswith("__ALL_"):
            parts.append(f"Lokasyon: {self.location_filter.currentText().strip()}")
        if search:
            parts.append(f'Arama: "{search}"')
        self._filter_lbl.setText("  |  " + "  ·  ".join(parts) if parts else "")

    def _selected_row(self):
        idx = self.table.selectionModel().currentIndex()
        if not idx.isValid():
            QMessageBox.information(self, "Bilgi", "Lütfen bir kayıt seçin.")
            return -1
        return idx.row()

    def _selected_id(self):
        row = self._selected_row()
        if row < 0:
            return None
        fid = self._model.id_at(row)
        if fid is None:
            QMessageBox.information(self, "Bilgi", "Lütfen bir kayıt seçin.")
        return fid

    def _add(self):
        dlg = FabricDialog(self)
        if dlg.exec():
            d = dlg.get_data()
            fid = db.add_fabric(**d, user_name=CURRENT_USER["full_name"])
            self.refresh_with_locations()

    def _edit(self):
        fid = self._selected_id()
        if not fid:
            return
        if db.is_fabric_linked_to_order(fid):
            QMessageBox.warning(self, "Düzenlenemez",
                "Bu kumaş bir sipariş satınalmasına bağlıdır.\n"
                "Stok listesinden düzenlenemez; değişiklik için planlama ekranını kullanın.")
            return
        fabric = db.get_fabric(fid)
        dlg = FabricDialog(self, fabric)
        if dlg.exec():
            d = dlg.get_data()
            db.update_fabric(fid, **d)
            self.refresh()

    def _delete(self):
        fid = self._selected_id()
        if not fid:
            return
        if db.is_fabric_linked_to_order(fid):
            QMessageBox.warning(self, "Silinemez",
                "Bu kumaş bir sipariş satınalmasına bağlıdır.\n"
                "Stok listesinden silinemez; işlem için planlama ekranını kullanın.")
            return
        fabric = db.get_fabric(fid)
        reply = QMessageBox.question(
            self, "Stoktan Sil",
            f"<b>{fabric['product_code']}</b> ({fabric['color']}) stoktan silinsin mi?<br><br>"
            f"<span style='color:#555'>Hareket geçmişi korunacak, stok listesinden kaldırılacak.</span>",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            db.soft_delete_fabric(fid, user_name=CURRENT_USER["full_name"])
            self.refresh_with_locations()

    def _giris(self):
        fid = self._selected_id()
        if not fid:
            return
        fabric = db.get_fabric(fid)
        dlg = MovementDialog(self, fabric, "GİRİŞ")
        if dlg.exec():
            d = dlg.get_data()
            # Satıra giriş de satın alma girişi olarak kaydedilir
            db.add_movement(fid, "SATINALMA GİRİŞİ", d["meter"], d["kg"], d["piece_count"],
                            d["notes"], user_name=CURRENT_USER["full_name"])
            self.refresh()

    def _cikis(self):
        # Birden çok satır seçiliyse toplu çıkış (aynı ürün/farklı lot → tek hedef)
        sel = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        fids = [self._model.id_at(i.row()) for i in sel]
        fids = [f for f in fids if f]
        if len(fids) > 1:
            return self._cikis_multi(fids)
        return self._cikis_single()

    def _cikis_multi(self, fids):
        fabrics = [db.get_fabric(f) for f in fids]
        fabrics = [f for f in fabrics if f]
        if not fabrics:
            return
        dlg = MultiCikisDialog(self, fabrics)
        if not dlg.exec():
            return
        dest, dtype, notes, rows = dlg.result_data()
        n = 0
        for fid, cm, ck in rows:
            if cm <= 0 and ck <= 0:
                continue
            db.add_movement(fid, "ÇIKIŞ", cm, ck, "", notes,
                            user_name=CURRENT_USER["full_name"],
                            destination=dest, destination_type=dtype)
            n += 1
        self.refresh()
        QMessageBox.information(self, "Toplu Çıkış",
            f"{n} kalem için çıkış yapıldı.\nHedef: {dest or '—'}")

    def _cikis_single(self):
        fid = self._selected_id()
        if not fid:
            return
        fabric = db.get_fabric(fid)
        dlg = SplitCikisDialog(self, fabric)
        if not dlg.exec():
            return
        lines, out, fire = dlg.result()
        total_m = sum(l[2] for l in lines)
        total_k = sum(l[3] for l in lines)
        notes = out["notes"]
        if fire:
            fire_note = f"Fire: %{fire['pct']:.1f}"
            notes = f"{notes} | {fire_note}" if notes else fire_note
        fire_extra_m = (fire["pre_m"] - total_m) if fire else 0
        fire_extra_k = (fire["pre_k"] - total_k) if fire else 0
        first_mid = None
        for idx, (tip, dest, mt, kg) in enumerate(lines):
            ded_m = ded_k = None
            if fire and idx == 0:   # fire kaybı ilk satırın stok düşüşüne eklenir
                ded_m = mt + fire_extra_m
                ded_k = kg + fire_extra_k
            mid = db.add_movement(fid, "ÇIKIŞ", mt, kg, "", notes,
                                  user_name=CURRENT_USER["full_name"],
                                  destination=dest, destination_type=tip,
                                  deduct_meter=ded_m, deduct_kg=ded_k,
                                  out_color=out["out_color"], lab_no=out["lab_no"],
                                  parti_no=out["parti_no"],
                                  out_fabric_type=out["out_fabric_type"],
                                  out_print_type=out["out_print_type"],
                                  out_zemin_rengi=out["out_zemin_rengi"],
                                  out_baski_desen_no=out["out_baski_desen_no"])
            if first_mid is None:
                first_mid = mid
        if fire:
            hedefler = "; ".join(f"{d} ({m:,.0f}mt)" for _, d, m, _ in lines)
            db.add_fire_record(
                fid, first_mid, fabric["product_code"], fabric["color"] or "",
                fabric["lot"] or "", fabric["location"] or "", hedefler,
                fire["pre_m"], fire["pre_k"], fire["out_m"], fire["out_k"], fire["pct"],
                manual_pct=False, user_name=CURRENT_USER["full_name"],
                out_color=out["out_color"], lab_no=out["lab_no"], parti_no=out["parti_no"])
            if db.finalize_lot_if_consumed(fid, CURRENT_USER["full_name"]):
                QMessageBox.information(
                    self, "Lot Tükendi",
                    f"<b>{fabric['product_code']} / {fabric['lot'] or '-'}</b> lotu tükendi.<br>"
                    f"Toplam fire 'Boyahane Fire Oranları' sekmesine işlendi.")
        self.refresh()
        if len(lines) > 1:
            QMessageBox.information(self, "Çıkış",
                f"{len(lines)} hedefe parçalı çıkış yapıldı. Toplam {total_m:,.0f} mt.")

    # Çift tıklanabilen sütunlar → veritabanı alanı
    CELL_FIELDS = {1: "product_code", 2: "product_name", 3: "description", 4: "color",
                   5: "location", 6: "fabric_type", 7: "lot", 8: "meter", 9: "kg",
                   10: "piece_count", 11: "birim_fiyat", 14: "entry_location", 15: "lab_no",
                   16: "print_type", 17: "zemin_rengi", 18: "baski_desen_no"}

    def _cell_edit(self, index):
        if not index.isValid():
            return
        field = self.CELL_FIELDS.get(index.column())
        if not field:   # #, Toplam Değer, Son Güncelleme — düzenlenemez
            return
        fid = self._model.id_at(index.row())
        if not fid:
            return
        fabric = db.get_fabric(fid)
        dlg = CellEditDialog(self, fabric, field, COLS[index.column()])
        if dlg.exec():
            d = {k: dict(fabric).get(k) for k in
                 ("product_name", "product_code", "color", "location", "meter", "kg",
                  "piece_count", "birim_fiyat", "fabric_type", "lot", "description", "lab_no",
                  "print_type", "zemin_rengi", "baski_desen_no")}
            d[field] = dlg.value()
            db.update_fabric(fid, **d)
            self.refresh_with_locations()

    def _history(self):
        dlg = DailyMovementsDialog(self)
        idx = self.table.selectionModel().currentIndex()
        if idx.isValid():
            fid = self._model.id_at(idx.row())
            if fid:
                fabric = db.get_fabric(fid)
                # Seçili ürünün tüm geçmişi hazır gelsin; filtreler değiştirilebilir
                dlg.search_box.setText(fabric["product_code"] or "")
                dlg._set_preset(None)
        dlg.exec()

    def _context_menu(self, pos):
        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)

        # Çok satır seçiliyse: tıklanan sütunu toplu düzenle
        act_bulk = None
        idx = self.table.indexAt(pos)
        sel_rows = self.table.selectionModel().selectedRows()
        field = self.CELL_FIELDS.get(idx.column()) if idx.isValid() else None
        if field and len(sel_rows) > 1:
            act_bulk = menu.addAction(f"✎ {COLS[idx.column()]} — {len(sel_rows)} satırı toplu düzenle")
            menu.addSeparator()

        act_export = menu.addAction("📥 Excel'e Aktar (görünen liste)")
        act_export_all = menu.addAction("📥 Excel'e Aktar (tüm stok)")
        action = menu.exec(self.table.viewport().mapToGlobal(pos))
        if act_bulk and action == act_bulk:
            self._bulk_cell_edit(idx.column(), field, sel_rows)
        elif action == act_export:
            self._export_visible()
        elif action == act_export_all:
            self._export_all()

    def _bulk_cell_edit(self, col, field, sel_rows):
        fids = [self._model.id_at(r.row()) for r in sel_rows]
        fids = [f for f in fids if f]
        if not fids:
            return
        first = db.get_fabric(fids[0])
        dlg = CellEditDialog(self, first, field, COLS[col], count=len(fids))
        if dlg.exec():
            val = dlg.value()
            keys = ("product_name", "product_code", "color", "location", "meter", "kg",
                    "piece_count", "birim_fiyat", "fabric_type", "lot", "description", "lab_no",
                    "print_type", "zemin_rengi", "baski_desen_no")
            for fid in fids:
                fabric = db.get_fabric(fid)
                d = {k: dict(fabric).get(k) for k in keys}
                d[field] = val
                db.update_fabric(fid, **d)
            self.refresh_with_locations()

    def _export_visible(self):
        """Tabloda şu an görünen satırları dışa aktar."""
        _export_table_to_excel(self, self._model, self.table)

    def _export_all(self):
        """Filtresiz tüm stoğu dışa aktar."""
        all_rows = db.get_all_fabrics()
        import tempfile
        tmp_model = FabricModel()
        tmp_model.load(all_rows)
        _export_table_to_excel(self, tmp_model, self.table, title="Tüm Stok")


class _FireSortItem(QTableWidgetItem):
    """Sayısal sütunlar için doğru sıralama: UserRole'daki değere göre karşılaştırır."""
    def __lt__(self, other):
        a = self.data(Qt.ItemDataRole.UserRole)
        b = other.data(Qt.ItemDataRole.UserRole)
        if a is not None and b is not None:
            return a < b
        return super().__lt__(other)


class FireView(QWidget):
    """Boyahane fire oranları — dış depodan müşteriye sevklerin fire kayıtları."""
    COLS = ["Tarih", "Boyahane", "Ürün Kodu", "Renk", "Lot", "Lab No", "Parti No", "Hedef",
            "Öncesi mt", "Çıkış mt", "Fire mt", "Fire mt %",
            "Öncesi kg", "Çıkış kg", "Fire kg", "Fire kg %", "Tür"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("Ara:"))
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Ürün, renk, lot, lab, parti, hedef...")
        self.search_box.setMaximumWidth(260)
        self.search_box.textChanged.connect(self.refresh)
        toolbar.addWidget(self.search_box)

        toolbar.addWidget(QLabel("Boyahane:"))
        self.boyahane_filter = QComboBox()
        self.boyahane_filter.addItem("Tümü", "")
        self.boyahane_filter.currentIndexChanged.connect(self.refresh)
        toolbar.addWidget(self.boyahane_filter)

        toolbar.addWidget(QLabel("Tür:"))
        self.type_filter = QComboBox()
        for t in ["Tümü", "ÇIKIŞ", "LOT TOPLAMI"]:
            self.type_filter.addItem(t, "" if t == "Tümü" else t)
        self.type_filter.currentIndexChanged.connect(self.refresh)
        toolbar.addWidget(self.type_filter)

        toolbar.addStretch()
        btn_reset = QPushButton("🔄 Lot Sıfırla")
        btn_reset.setToolTip("Seçili satırın lotunu kapatır: dış depoda kalan stok fire yazılır,\n"
                             "lotun toplam firesi ayrı satır olarak eklenir.")
        btn_reset.clicked.connect(self._reset_lot)
        btn_refresh = QPushButton("⟳ Yenile")
        btn_refresh.clicked.connect(self.refresh)
        toolbar.addWidget(btn_reset)
        toolbar.addWidget(btn_refresh)
        layout.addLayout(toolbar)

        info = QLabel("Dış depolardan yapılan çıkışların fire kayıtları. "
                      "<i>(elle)</i> işaretli oranlar elle girilmiştir; "
                      "<b>LOT TOPLAMI</b> satırları kapanan lotların toplam firesidir.")
        info.setStyleSheet("color:#555; font-size:11px;")
        layout.addWidget(info)

        self.table = QTableWidget()
        self.table.setColumnCount(len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)   # elle genişlik ayarı
        hdr.setStretchLastSection(True)
        self._cols_sized = False   # ilk veri yüklemesinde bir kez otomatik boyutlandır
        layout.addWidget(self.table)

        self.count_lbl = QLabel()
        self.count_lbl.setStyleSheet("color:#555; font-size:12px;")
        layout.addWidget(self.count_lbl)

    def refresh(self):
        all_rows = [dict(r) for r in db.get_fire_records()]

        # Boyahane filtresini doldur (seçim korunur)
        cur_b = self.boyahane_filter.currentData() or ""
        names = sorted({r["boyahane"] for r in all_rows if r["boyahane"]})
        self.boyahane_filter.blockSignals(True)
        self.boyahane_filter.clear()
        self.boyahane_filter.addItem("Tümü", "")
        for n in names:
            self.boyahane_filter.addItem(n, n)
        idx = self.boyahane_filter.findData(cur_b)
        if idx >= 0:
            self.boyahane_filter.setCurrentIndex(idx)
        self.boyahane_filter.blockSignals(False)

        # Filtrele
        q = self.search_box.text().strip().lower()
        b = self.boyahane_filter.currentData() or ""
        t = self.type_filter.currentData() or ""
        rows = []
        for r in all_rows:
            if b and r["boyahane"] != b:
                continue
            if t and r["record_type"] != t:
                continue
            if q:
                hay = " ".join(str(r.get(k) or "") for k in
                               ("boyahane", "product_code", "color", "out_color",
                                "lot", "lab_no", "parti_no", "customer")).lower()
                if q not in hay:
                    continue
            rows.append(r)

        self._rows = rows
        self.table.setSortingEnabled(False)   # doldururken sıralamayı kapat
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            is_total = r["record_type"] == "LOT TOPLAMI"
            elle = " (elle)" if r["manual_pct"] else ""
            pre_m, out_m = r["pre_meter"] or 0, r["out_meter"] or 0
            pre_k, out_k = r["pre_kg"] or 0, r["out_kg"] or 0
            fire_m, fire_k = max(0, pre_m - out_m), max(0, pre_k - out_k)
            pct_m = (fire_m / pre_m * 100) if pre_m > 0 else 0
            pct_k = (fire_k / pre_k * 100) if pre_k > 0 else 0
            vals = [
                str(r["created_at"] or "")[:16],
                r["boyahane"] or "",
                r["product_code"] or "",
                r.get("out_color") or r["color"] or "",
                r["lot"] or "",
                r.get("lab_no") or "",
                r.get("parti_no") or "",
                r["customer"] or "",
                f"{pre_m:,.2f}" if pre_m else "",
                f"{out_m:,.2f}" if out_m else "",
                f"{fire_m:,.2f}" if pre_m else "",
                f"%{pct_m:.1f}{elle}" if pre_m else "",
                f"{pre_k:,.2f}" if pre_k else "",
                f"{out_k:,.2f}" if out_k else "",
                f"{fire_k:,.2f}" if pre_k else "",
                f"%{pct_k:.1f}{elle}" if pre_k else "",
                r["record_type"],
            ]
            sort_keys = {8: pre_m, 9: out_m, 10: fire_m, 11: pct_m,
                         12: pre_k, 13: out_k, 14: fire_k, 15: pct_k}
            for j, v in enumerate(vals):
                item = _FireSortItem(v)
                if j in sort_keys:
                    item.setData(Qt.ItemDataRole.UserRole, sort_keys[j])
                if j == 0:   # sıralamadan bağımsız kayıt eşlemesi için id sakla
                    item.setData(Qt.ItemDataRole.UserRole + 1, r["id"])
                if 8 <= j <= 15:
                    item.setTextAlignment(int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight))
                if j in (10, 11, 14, 15):   # fire sütunları — kırmızı, elle girilenler turuncu
                    item.setForeground(QBrush(QColor("#E65100" if r["manual_pct"] else "#C62828")))
                    f = QFont(); f.setBold(True); item.setFont(f)
                if is_total:
                    f = QFont(); f.setBold(True); item.setFont(f)
                    item.setBackground(QBrush(QColor("#FFF8E1")))
                self.table.setItem(i, j, item)
        self.table.setSortingEnabled(True)   # başlığa tıklayınca sıralar, mevcut sıralamayı uygular
        if rows and not self._cols_sized:
            self.table.resizeColumnsToContents()   # ilk yüklemede makul genişlikler
            self._cols_sized = True
        self.count_lbl.setText(f"{len(rows)} kayıt" + (f" / toplam {len(all_rows)}" if len(rows) != len(all_rows) else ""))

    def _reset_lot(self):
        row = self.table.currentRow()
        item0 = self.table.item(row, 0) if row >= 0 else None
        rid = item0.data(Qt.ItemDataRole.UserRole + 1) if item0 else None
        r = next((x for x in self._rows if x["id"] == rid), None)
        if r is None:
            QMessageBox.information(self, "Bilgi", "Önce tablodan sıfırlanacak lota ait bir satır seçin.")
            return
        if r["record_type"] == "LOT TOPLAMI":
            QMessageBox.information(self, "Bilgi", "Bu lot zaten kapatılmış (LOT TOPLAMI satırı).")
            return
        if db.lot_total_exists(r["product_code"], r["color"], r["lot"], r["boyahane"]):
            QMessageBox.warning(self, "Lot Sıfırlama",
                                f"<b>{r['product_code']} / {r['lot'] or '-'}</b> — {r['boyahane']}<br><br>"
                                f"Bu lot daha önce sıfırlanmıştır.")
            return
        fabric = db.get_fabric(r["fabric_id"]) if r["fabric_id"] else None
        if not fabric:
            QMessageBox.warning(self, "Hata", "Lota ait stok kaydı bulunamadı.")
            return
        rem_m, rem_k = fabric["meter"] or 0, fabric["kg"] or 0
        reply = QMessageBox.question(
            self, "Lot Sıfırlama",
            f"<b>{r['product_code']} / {r['lot'] or '-'}</b> — {r['boyahane']}<br><br>"
            f"Kalan stok: <b>{rem_m:,.2f} mt / {rem_k:,.2f} kg</b><br><br>"
            f"Lot kapatılsın mı? Kalan stok <span style='color:#C62828'>fire olarak yazılacak</span> "
            f"ve toplam fire ayrı satır olarak eklenecek.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        db.reset_lot_fire(r["fabric_id"], CURRENT_USER["full_name"])
        self.refresh()
        QMessageBox.information(self, "Tamamlandı",
                                "Lot kapatıldı, toplam fire satırı eklendi.")


class OrdersView(QWidget):
    """Satış siparişleri listesi — yeni sipariş açıldığında
    status='PLANLAMA BEKLİYOR' ile kaydedilir (planlama ekranı sonraki aşama)."""

    COLS = ["Sipariş No", "Sip. Tarihi", "Müşteri", "Müşteri Referans", "Kalem",
            "Para Birimi", "Toplam Tutar", "Termin", "Ödeme Şekli",
            "Teslimat Şartları", "Durum", "Oluşturan", "Siparişi Alan",
            "Baskı Tipi", "Zemin Rengi", "Baskı Desen No",
            "Teslimat Adresi", "Notlar"]

    # Kalem (alt) satırları, üst satırla aynı sütunları farklı anlamlarla
    # kullanır — her siparişin kalemlerinden önce bu etiket satırı gösterilir.
    ITEM_COL_LABELS = ["Ürün Kodu", "Kompozisyon", "En", "Gramaj", "Kumaş Tipi", "Renk",
                       "Lab No", "Açıklama", "Metre", "Kilo", "Birim Fiyat", "Tutar",
                       "Baskı Tipi", "Zemin Rengi", "Baskı Desen No", "", "", ""]

    STATUS_OPTIONS = [
        "ONAYDA", "ONAYLANDI - PLANLAMADA", "PLANLAMA - GÖRDÜ", "PLANLANDI",
        "BOYAHANAYA SEVKLER BAŞLADI", "TÜM KUMAŞLAR BOYAHANEDE",
        "MÜŞTERİYE SEVKLER BAŞLADI", "SİPARİŞ TAMAMLANDI", "İPTAL",
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("Ara:"))
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Sipariş no, ürün kodu, müşteri, renk...")
        self.search_box.setMaximumWidth(260)
        self.search_box.textChanged.connect(self.refresh)
        toolbar.addWidget(self.search_box)

        toolbar.addWidget(QLabel("Durum:"))
        self.status_filter = QComboBox()
        self.status_filter.addItem("Tümü", "")
        for s in self.STATUS_OPTIONS:
            self.status_filter.addItem(s, s)
        self.status_filter.currentIndexChanged.connect(self.refresh)
        toolbar.addWidget(self.status_filter)

        toolbar.addStretch()
        btn_new = QPushButton("+ Yeni Sipariş")
        btn_new.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;border-radius:4px;padding:6px 14px;")
        btn_new.clicked.connect(self._new_order)
        btn_del = QPushButton("✕ Sil")
        btn_del.setStyleSheet("background:#757575;color:white;border-radius:4px;padding:6px 14px;")
        btn_del.clicked.connect(self._delete_order)
        btn_pdf = QPushButton("📄 PDF Al")
        btn_pdf.clicked.connect(self._export_pdf)
        btn_refresh = QPushButton("⟳ Yenile")
        btn_refresh.clicked.connect(self.refresh)
        toolbar.addWidget(btn_new)
        if CURRENT_USER.get("role") == "admin":
            btn_settings = QPushButton("🏦 Şirket/Banka Ayarları")
            btn_settings.clicked.connect(self._company_settings)
            toolbar.addWidget(btn_settings)
        toolbar.addWidget(btn_del)
        toolbar.addWidget(btn_pdf)
        toolbar.addWidget(btn_refresh)
        layout.addLayout(toolbar)

        self.table = QTreeWidget()
        self.table.setColumnCount(len(self.COLS))
        self.table.setHeaderLabels(self.COLS)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setRootIsDecorated(True)
        self.table.setExpandsOnDoubleClick(False)
        hdr = self.table.header()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(True)
        self.table.doubleClicked.connect(self._edit_order)
        self.table.setStyleSheet(
            "QTreeWidget::item { border: 1px solid #E0E0E0; padding: 3px; }"
            "QTreeWidget::item:has-children { background-color: #E8EAF6; font-weight: bold;"
            " border-top: 2px solid #9FA8DA; padding-top: 5px; padding-bottom: 5px; }"
            "QTreeWidget::item:selected { background-color: #BBDEFB; color: #212121; }"
            "QTreeWidget::item:selected:has-children { background-color: #9FA8DA; color: #212121; }"
        )

        layout.addWidget(self.table)
        _wire_header_persistence(self.table, "orders_columns")

        self.count_lbl = QLabel()
        self.count_lbl.setStyleSheet("color:#555; font-size:12px;")
        layout.addWidget(self.count_lbl)

    def refresh(self):
        search = self.search_box.text().strip()
        status = self.status_filter.currentData() or ""
        rows = [dict(r) for r in db.get_all_orders(search=search, status=status)]
        self._rows = rows
        self.table.clear()
        for r in rows:
            currency = r.get("currency") or "USD"
            symbol = CURRENCY_SYMBOLS.get(currency, "$")
            item_count = r.get("item_count") or 0
            total_amount = r.get("total_amount") or 0
            vals = [
                r.get("order_no") or "",
                str(r.get("order_date") or "")[:10],
                r.get("customer_name") or "",
                r.get("customer_ref") or "",
                f"{item_count} kalem",
                currency,
                f"{total_amount:,.2f} {symbol}",
                str(r.get("delivery_date") or "")[:10],
                r.get("payment_method") or "",
                r.get("delivery_terms") or "",
                r.get("status") or "",
                r.get("created_by") or "",
                r.get("sales_rep") or "",
                "", "", "",
                r.get("delivery_address") or "",
                (r.get("notes") or "").replace("\n", " "),
            ]
            order_item = QTreeWidgetItem([str(v) for v in vals])
            order_item.setData(0, Qt.ItemDataRole.UserRole, r["id"])
            order_item.setTextAlignment(4, int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight))
            order_item.setTextAlignment(6, int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight))
            self.table.addTopLevelItem(order_item)

            items = r.get("items") or []
            if items:
                legend = QTreeWidgetItem([str(v) for v in self.ITEM_COL_LABELS])
                legend.setFlags(legend.flags() & ~Qt.ItemFlag.ItemIsSelectable)
                legend_font = QFont(); legend_font.setItalic(True); legend_font.setPointSize(8)
                for col in range(len(self.ITEM_COL_LABELS)):
                    legend.setFont(col, legend_font)
                    legend.setForeground(col, QBrush(QColor("#9E9E9E")))
                order_item.addChild(legend)

            for it in items:
                meter = it.get("meter") or 0
                kg = it.get("kg") or 0
                sale_price = it.get("sale_price") or 0
                total = meter * sale_price
                child_vals = [
                    it.get("product_code") or "",
                    it.get("composition") or "",
                    it.get("width") or "",
                    it.get("gramaj") or "",
                    it.get("fabric_type") or "",
                    it.get("color") or "",
                    it.get("lab_no") or "",
                    it.get("description") or "",
                    f"{meter:,.2f}",
                    f"{kg:,.2f}",
                    f"{sale_price:,.2f} {symbol}",
                    f"{total:,.2f} {symbol}",
                    it.get("print_type") or "",
                    it.get("zemin_rengi") or "",
                    it.get("baski_desen_no") or "",
                    "", "", "",
                ]
                child = QTreeWidgetItem([str(v) for v in child_vals])
                for col in (8, 9, 10, 11):
                    child.setTextAlignment(col, int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight))
                order_item.addChild(child)

        self.table.expandAll()
        self.count_lbl.setText(f"{len(rows)} sipariş")

    def _new_order(self):
        dlg = OrderDialog(self)
        if dlg.exec():
            data = dlg.get_data()
            oid, order_no = db.add_order(**data, created_by=CURRENT_USER["full_name"])
            self.refresh()
            QMessageBox.information(self, "Sipariş Oluşturuldu",
                f"Sipariş kaydedildi.\n\nSipariş No: {order_no}\nDurum: ONAYDA\n\nAdmin onayı bekleniyor.")

    def _selected_id(self):
        item = self.table.currentItem()
        oid = item.data(0, Qt.ItemDataRole.UserRole) if item else None
        if oid is None and item is not None and item.parent() is not None:
            oid = item.parent().data(0, Qt.ItemDataRole.UserRole)
        if oid is None:
            QMessageBox.information(self, "Bilgi", "Önce bir sipariş seçin.")
            return None
        return oid

    def _edit_order(self):
        oid = self._selected_id()
        if not oid:
            return
        order = db.get_order(oid)
        if not order:
            return
        dlg = OrderDialog(self, order=order)
        if dlg.exec():
            data = dlg.get_data()
            db.update_order(oid, **data)
            self.refresh()

    def _delete_order(self):
        if CURRENT_USER.get("role") != "admin":
            QMessageBox.warning(self, "Yetersiz Yetki",
                "Sipariş silme işlemi yalnızca admin tarafından yapılabilir.")
            return
        oid = self._selected_id()
        if not oid:
            return
        order = db.get_order(oid)
        if not order:
            return
        status = order.get("status","")
        if status not in ("ONAYDA", "İPTAL"):
            QMessageBox.warning(self, "Silinemez",
                f"Bu sipariş şu an '{status}' durumunda.\n\n"
                "Silebilmek için siparişin önce İPTAL edilmesi gerekir.")
            return
        if QMessageBox.question(self, "Sil",
                f"{order.get('order_no','')} numaralı sipariş kalıcı olarak silinsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            db.delete_order(oid)
            self.refresh()

    def _export_pdf(self):
        oid = self._selected_id()
        if not oid:
            return
        order = db.get_order(oid)
        if not order:
            return
        order = {**order, "customer_tax_no": _customer_tax_no(order.get("customer_id"))}
        company = db.get_company_settings()
        default_name = f"{order['order_no']}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Sipariş Sözleşmesi PDF", default_name, "PDF (*.pdf)")
        if not path:
            return
        try:
            order_pdf.generate_order_pdf(order, company, path)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{e}")
            return
        QMessageBox.information(self, "PDF Oluşturuldu", f"Kaydedildi:\n{path}")

    def _company_settings(self):
        if CURRENT_USER.get("role") != "admin":
            QMessageBox.warning(self, "Yetki", "Bu işlem için yönetici yetkisi gereklidir.")
            return
        CompanySettingsDialog(self).exec()


PO_STATUS_OPTIONS = ["BEKLEMEDE", "GÖNDERİLDİ", "KISMİ GELDİ", "TAMAMLANDI", "İPTAL"]


class PurchaseOrderDialog(QDialog):
    """Kalem bazında tedarikçi + kumaş tipi seçimli PO oluşturma."""
    FABRIC_TYPES = ["HAM", "PFD", "Boyalı", "Baskılı", "Mamül", "Diğer"]
    PAYMENT_METHODS = ["Peşin", "Havale/EFT", "30 Gün Vade", "60 Gün Vade",
                       "90 Gün Vade", "120 Gün Vade", "Akreditif (L/C)", "Diğer"]
    # col: 0=Ürün Kodu, 1=Ürün Adı, 2=Kumaş Tipi(w), 3=Tedarikçi(w),
    #      4=Para Birimi(w), 5=Sip.Mt(ro), 6=Ham Sip.Mt(+%20),
    #      7=Sip.Kg(ro), 8=Ham Kg(+%20), 9=Birim Fiyat,
    #      10=Toplam(ro), 11=Ödeme Şekli(w), 12=Teslimat Şekli,
    #      13=Termin, 14=Notlar, 15=Açıklama
    ITEM_COLS = [
        "Ürün Kodu", "Ürün Adı", "Kumaş Tipi", "Tedarikçi", "Para Birimi",
        "Sip. Mt", "Ham Sip. Mt (+%20)", "Sip. Kg", "Ham Kg (+%20)",
        "Birim Fiyat", "Toplam", "Ödeme Şekli",
        "Teslimat Şekli", "Termin", "Notlar", "Açıklama",
    ]
    _READONLY_COLS = {5, 7, 10}  # Sip.Mt, Sip.Kg, Toplam

    def __init__(self, parent=None, missing_items=None, po=None):
        super().__init__(parent)
        self.setWindowTitle("Satınalma Siparişi Oluştur")
        self.setMinimumSize(1400, 520)
        self._suppliers = []
        self._build_ui()
        self._reload_suppliers()
        if missing_items:
            self._fill_items(missing_items)

    def _build_ui(self):
        lay = QVBoxLayout(self)

        hint = QLabel(
            "<i>Her kalem için kumaş tipi, tedarikçi ve para birimi ayrı seçilebilir. "
            "Aynı tedarikçiye ait kalemler otomatik aynı PO'ya eklenir. "
            "Sip. Mt/Kg referans; Ham değerler %20 artışlı önerilen miktardır.</i>"
        )
        hint.setStyleSheet("color:#757575;font-size:11px;"); hint.setWordWrap(True)
        lay.addWidget(hint)

        self.table = QTableWidget()
        self.table.setColumnCount(len(self.ITEM_COLS))
        self.table.setHorizontalHeaderLabels(self.ITEM_COLS)
        self.table.verticalHeader().setVisible(False)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.cellChanged.connect(self._on_cell_changed)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        b_add = QPushButton("+ Kalem Ekle"); b_add.clicked.connect(self._copy_row)
        b_del = QPushButton("✕ Kalem Kaldır"); b_del.clicked.connect(self._del_row)
        btn_row.addWidget(b_add); btn_row.addWidget(b_del)
        btn_row.addStretch()
        self._subtotal_lbl = QLabel()
        self._subtotal_lbl.setStyleSheet("font-weight:bold;color:#1565C0;")
        btn_row.addWidget(self._subtotal_lbl)
        lay.addLayout(btn_row)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._validate)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _reload_suppliers(self):
        self._suppliers = db.get_all_suppliers()
        for row in range(self.table.rowCount()):
            sup_cb = self.table.cellWidget(row, 3)
            if isinstance(sup_cb, QComboBox):
                prev = sup_cb.currentData()
                self._fill_supplier_combo(sup_cb, select_data=prev)

    def _fill_supplier_combo(self, cb, select_data=None):
        cb.blockSignals(True)
        cb.clear()
        cb.addItem("— Tedarikçi —", None)
        for s in self._suppliers:
            label = s["name"] + (f" ({s['code']})" if s.get("code") else "")
            cb.addItem(label, (s["id"], s["name"]))
        if select_data:
            for i in range(cb.count()):
                if cb.itemData(i) == select_data:
                    cb.setCurrentIndex(i); break
        cb.blockSignals(False)

    def _fill_items(self, items):
        self.table.setRowCount(0)
        for item in items:
            self._insert_row(item)

    def _insert_row(self, item=None):
        item = item or {}
        row = self.table.rowCount()
        self.table.blockSignals(True)
        self.table.insertRow(row)

        # col 2: Kumaş Tipi combo
        ft_cb = QComboBox()
        for ft in self.FABRIC_TYPES:
            ft_cb.addItem(ft, ft)
        idx = ft_cb.findData(item.get("fabric_type", "HAM") or "HAM")
        if idx >= 0: ft_cb.setCurrentIndex(idx)
        self.table.setCellWidget(row, 2, ft_cb)

        # col 3: Tedarikçi combo
        sup_cb = QComboBox()
        self._fill_supplier_combo(sup_cb, select_data=item.get("_supplier_data"))
        _make_searchable(sup_cb)
        self.table.setCellWidget(row, 3, sup_cb)

        # col 4: Para Birimi combo
        cur_cb = QComboBox()
        for code in CURRENCY_OPTIONS:
            cur_cb.addItem(f"{code} ({CURRENCY_SYMBOLS[code]})", code)
        preset_cur = item.get("currency") or "USD"
        ci = cur_cb.findData(preset_cur)
        if ci >= 0: cur_cb.setCurrentIndex(ci)
        cur_cb.currentIndexChanged.connect(lambda _, r=row: self._recalc_row(r))
        self.table.setCellWidget(row, 4, cur_cb)

        # col 11: Ödeme Şekli combo
        pay_cb = QComboBox()
        pay_cb.setEditable(True)
        pay_cb.addItem("")
        for p in self.PAYMENT_METHODS:
            pay_cb.addItem(p)
        preset_pay = item.get("payment_method", "")
        pi = pay_cb.findText(preset_pay)
        if pi >= 0: pay_cb.setCurrentIndex(pi)
        else: pay_cb.setEditText(preset_pay)
        self.table.setCellWidget(row, 11, pay_cb)

        orig_meter = float(item.get("meter") or 0)
        orig_kg    = float(item.get("kg") or 0)
        ham_meter  = item.get("ham_meter")
        ham_kg     = item.get("ham_kg")
        if ham_meter is None: ham_meter = round(orig_meter * 1.20, 2)
        if ham_kg    is None: ham_kg    = round(orig_kg    * 1.20, 2)
        termin = item.get("termin") or QDate.currentDate().addDays(30).toString("dd.MM.yyyy")

        text_vals = {
            0:  str(item.get("product_code", "")),
            1:  str(item.get("product_name", "")),
            5:  str(orig_meter),   # Sip. Mt  (readonly)
            6:  str(ham_meter),    # Ham Sip. Mt
            7:  str(orig_kg),      # Sip. Kg  (readonly)
            8:  str(ham_kg),       # Ham Kg
            9:  str(item.get("unit_price") or "0"),
            10: "0.00",            # Toplam   (readonly, calculated)
            12: str(item.get("delivery_terms", "")),
            13: termin,
            14: str(item.get("notes", "")),
            15: str(item.get("description", "")),
        }
        for col, val in text_vals.items():
            cell = QTableWidgetItem(str(val))
            if col in self._READONLY_COLS:
                cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                cell.setBackground(QBrush(QColor("#F5F5F5")))
            self.table.setItem(row, col, cell)

        self.table.blockSignals(False)
        self._recalc_row(row)
        self._update_subtotal()

    def _recalc_row(self, row):
        try:
            m  = float((self.table.item(row, 6) or QTableWidgetItem("0")).text() or 0)
            up = float((self.table.item(row, 9) or QTableWidgetItem("0")).text() or 0)
            total = m * up
        except (ValueError, AttributeError):
            total = 0.0
        cur_cb = self.table.cellWidget(row, 4)
        sym = ""
        if isinstance(cur_cb, QComboBox):
            sym = CURRENCY_SYMBOLS.get(cur_cb.currentData(), "")
        self.table.blockSignals(True)
        cell = self.table.item(row, 10)
        if not cell:
            cell = QTableWidgetItem()
            cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
            cell.setBackground(QBrush(QColor("#F5F5F5")))
            self.table.setItem(row, 10, cell)
        cell.setText(f"{sym}{total:,.2f}")
        self.table.blockSignals(False)
        self._update_subtotal()

    def _update_subtotal(self):
        total_mt = 0.0; total_kg = 0.0
        for r in range(self.table.rowCount()):
            try: total_mt += float((self.table.item(r, 6) or QTableWidgetItem("0")).text() or 0)
            except ValueError: pass
            try: total_kg += float((self.table.item(r, 8) or QTableWidgetItem("0")).text() or 0)
            except ValueError: pass
        self._subtotal_lbl.setText(
            f"Alt Toplam:  {total_mt:,.2f} mt  |  {total_kg:,.2f} kg")

    def _on_cell_changed(self, row, col):
        if col in (6, 8, 9):
            self._recalc_row(row)
        if col in (6, 8):
            self._update_subtotal()

    def _copy_row(self):
        """Seçili satırı kopyala; seçili satır yoksa son satırı kopyala."""
        src = self.table.currentRow()
        if src < 0:
            src = self.table.rowCount() - 1
        if src < 0:
            self._insert_row()
            return

        def _txt(c):
            it = self.table.item(src, c)
            return it.text() if it else ""

        ft_cb  = self.table.cellWidget(src, 2)
        sup_cb = self.table.cellWidget(src, 3)
        cur_cb = self.table.cellWidget(src, 4)
        pay_cb = self.table.cellWidget(src, 11)
        item = {
            "product_code":   _txt(0),
            "product_name":   _txt(1),
            "fabric_type":    ft_cb.currentData()  if isinstance(ft_cb, QComboBox)  else "HAM",
            "_supplier_data": sup_cb.currentData() if isinstance(sup_cb, QComboBox) else None,
            "currency":       cur_cb.currentData() if isinstance(cur_cb, QComboBox) else "USD",
            "meter":          float(_txt(5) or 0),
            "ham_meter":      _txt(6),
            "kg":             float(_txt(7) or 0),
            "ham_kg":         _txt(8),
            "unit_price":     _txt(9),
            "payment_method": pay_cb.currentText() if isinstance(pay_cb, QComboBox) else _txt(11),
            "delivery_terms": _txt(12),
            "termin":         _txt(13),
            "notes":          _txt(14),
            "description":    _txt(15),
        }
        self._insert_row(item)

    def _del_row(self):
        row = self.table.currentRow()
        if row >= 0:
            self.table.removeRow(row)

    def _validate(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Eksik Bilgi", "En az bir kalem eklenmelidir.")
            return
        for row in range(self.table.rowCount()):
            sup_cb = self.table.cellWidget(row, 3)
            if not isinstance(sup_cb, QComboBox) or not isinstance(sup_cb.currentData(), tuple):
                QMessageBox.warning(self, "Eksik Bilgi",
                    f"Kalem {row+1} için tedarikçi seçilmelidir.")
                return
            try:
                it = self.table.item(row, 6)  # Ham Sip. Mt
                if float(it.text() if it else "0") <= 0:
                    QMessageBox.warning(self, "Eksik Bilgi",
                        f"Kalem {row+1} için geçerli Ham Sip. Metresi girilmelidir.")
                    return
            except ValueError:
                QMessageBox.warning(self, "Geçersiz Değer",
                    f"Kalem {row+1} için sayısal metre giriniz.")
                return
        self.accept()

    def get_data(self):
        """Tedarikçiye göre gruplu PO listesi döner."""
        def _cell(row, col):
            it = self.table.item(row, col)
            return it.text().strip() if it else ""
        def _num(row, col):
            try: return float(_cell(row, col) or 0)
            except ValueError: return 0.0

        items_by_supplier = {}
        headers_by_supplier = {}
        for row in range(self.table.rowCount()):
            sup_cb = self.table.cellWidget(row, 3)
            ft_cb  = self.table.cellWidget(row, 2)
            cur_cb = self.table.cellWidget(row, 4)
            pay_cb = self.table.cellWidget(row, 11)
            sup_data = sup_cb.currentData() if isinstance(sup_cb, QComboBox) else None
            if not isinstance(sup_data, tuple):
                continue
            sid, sname = sup_data
            ft       = ft_cb.currentData()  if isinstance(ft_cb,  QComboBox) else "HAM"
            currency = cur_cb.currentData() if isinstance(cur_cb, QComboBox) else "USD"
            pay_meth = pay_cb.currentText() if isinstance(pay_cb, QComboBox) else _cell(row, 11)

            # Termin: dd.MM.yyyy → yyyy-MM-dd
            termin_raw = _cell(row, 13)
            try:
                from datetime import datetime as _dtt
                termin = _dtt.strptime(termin_raw, "%d.%m.%Y").strftime("%Y-%m-%d")
            except Exception:
                termin = termin_raw

            key = (sid, sname)
            if key not in headers_by_supplier:
                headers_by_supplier[key] = {
                    "currency":          currency,
                    "payment_method":    pay_meth,
                    "delivery_terms":    _cell(row, 12),
                    "expected_delivery": termin,
                    "notes":             _cell(row, 14),
                }
            if key not in items_by_supplier:
                items_by_supplier[key] = []
            items_by_supplier[key].append({
                "product_code": _cell(row, 0),
                "product_name": _cell(row, 1),
                "fabric_type":  ft,
                "meter":        _num(row, 6),   # Ham Sip. Mt
                "kg":           _num(row, 8),   # Ham Kg
                "unit_price":   _num(row, 9),
                "description":  _cell(row, 15),
                "composition":  "",
                "width":        "",
                "gramaj":       "",
            })
        return {
            "po_groups": [
                {
                    "supplier_id":   k[0],
                    "supplier_name": k[1],
                    "items":         items_by_supplier[k],
                    **headers_by_supplier.get(k, {}),
                }
                for k in items_by_supplier
            ],
        }


class ItemReceiptDialog(QDialog):
    """Tek PO kalemine mal girişi: metre, kilo, lot, lokasyon."""
    def __init__(self, parent=None, po_item=None):
        super().__init__(parent)
        self.po_item = po_item or {}
        it = self.po_item
        self.setWindowTitle(
            f"Mal Girişi — {it.get('product_code','')} / {it.get('fabric_type','')}")
        self.setMinimumWidth(440)
        self._build_ui()

    def _build_ui(self):
        import datetime as _dt
        lay = QVBoxLayout(self)
        it = self.po_item
        remaining = max(0.0, (it.get("meter") or 0) - (it.get("received_meter") or 0))
        info_lbl = QLabel(
            f"<b>Ürün Kodu:</b> {it.get('product_code','')}  "
            f"<b>Kumaş:</b> {it.get('fabric_type','')}  |  "
            f"<b>Sipariş:</b> {it.get('meter',0):.2f} mt / {it.get('kg',0):.2f} kg  |  "
            f"<b>Kalan:</b> {remaining:.2f} mt"
        )
        info_lbl.setWordWrap(True)
        info_lbl.setStyleSheet("background:#F5F5F5;padding:6px;border-radius:4px;")
        lay.addWidget(info_lbl)

        form = QFormLayout(); form.setSpacing(10)
        self.metre = QDoubleSpinBox()
        self.metre.setRange(0, 9_999_999); self.metre.setDecimals(2)
        self.metre.setSuffix(" mt"); self.metre.setValue(remaining)
        self.kilo = QDoubleSpinBox()
        self.kilo.setRange(0, 9_999_999); self.kilo.setDecimals(2)
        self.kilo.setSuffix(" kg")

        ft_short = (it.get("fabric_type") or "HAM").split()[0][:3].upper()
        date_str = _dt.date.today().strftime("%Y%m%d")
        self.lot = QLineEdit()
        self.lot.setPlaceholderText(f"Boş bırakılırsa otomatik: {ft_short}-{date_str}-001")

        self.location = QComboBox()
        self.location.addItem("— Lokasyon Seçiniz —", ("", ""))
        for loc in db.get_active_locations():
            grp = loc.get("group_name","")
            self.location.addItem(f"[{grp}]  {loc['name']}", (loc["name"], grp))
        _make_searchable(self.location)
        self._routing_lbl = QLabel()
        self._routing_lbl.setStyleSheet("color:#1565C0;font-style:italic;font-size:11px;")
        self.location.currentIndexChanged.connect(self._update_routing)

        form.addRow("Gelen Metre:", self.metre)
        form.addRow("Gelen Kilo:", self.kilo)
        form.addRow("Lot No:", self.lot)
        form.addRow("Lokasyon:", self.location)
        form.addRow("", self._routing_lbl)
        lay.addLayout(form)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._validate)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _update_routing(self):
        data = self.location.currentData()
        if not data or not data[0]:
            self._routing_lbl.setText(""); return
        _, grp = data
        if grp == "DEPO":
            self._routing_lbl.setText("→ Ana depoya giriş (Boyahane planlamasına gelmez)")
        else:
            self._routing_lbl.setText("→ Dış depoya giriş — Boyahane planlama ekranına düşecek")

    def _validate(self):
        data = self.location.currentData()
        if not data or not data[0]:
            QMessageBox.warning(self, "Eksik", "Lokasyon seçilmelidir."); return
        if self.metre.value() <= 0 and self.kilo.value() <= 0:
            QMessageBox.warning(self, "Eksik", "Gelen metre veya kilo girilmelidir."); return
        self.accept()

    def get_data(self):
        loc_name, grp = self.location.currentData()
        return {"metre": self.metre.value(), "kilo": self.kilo.value(),
                "lot": self.lot.text().strip(),
                "location": loc_name, "location_group": grp}


class GoodsReceiptDialog(QDialog):
    def __init__(self, parent=None, po=None):
        super().__init__(parent)
        self.po = po
        self.setWindowTitle(f"Mal Girişi — {po.get('po_no','')}")
        self.resize(780, 400)
        self._item_ids = []
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        items = self.po.get("items", [])

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            ["Ürün Kodu", "Kumaş Tipi", "Sipariş Mt", "Sipariş Kg",
             "Gelen Metre", "Gelen Kilo"])
        self.table.setRowCount(len(items))
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        hdr = self.table.horizontalHeader()
        hdr.setSectionsMovable(True)
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for i in range(1, 6):
            hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

        for row, it in enumerate(items):
            self._item_ids.append(it["id"])
            for col, val in enumerate([
                it.get("product_code",""), it.get("fabric_type","HAM"),
                f"{it.get('meter',0):.2f}", f"{it.get('kg',0):.2f}",
                "0", "0",
            ]):
                cell = QTableWidgetItem(val)
                if col < 4:
                    cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    cell.setBackground(QBrush(QColor("#F5F5F5")))
                self.table.setItem(row, col, cell)
        lay.addWidget(self.table)

        form = QFormLayout(); form.setSpacing(8)
        self.location = QComboBox()
        self.location.addItem("— Lokasyon Seçiniz —", "")
        for loc in db.get_active_locations():
            if loc["group_name"] == "DEPO":
                self.location.addItem(loc["name"], loc["name"])
        _make_searchable(self.location)
        form.addRow("DEPO Lokasyonu *:", self.location)
        lay.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._validate)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _validate(self):
        if not self.location.currentData():
            QMessageBox.warning(self, "Eksik Bilgi", "DEPO lokasyonu seçilmelidir.")
            return
        self.accept()

    def get_data(self):
        results = []
        for row in range(self.table.rowCount()):
            try:
                m = float(self.table.item(row, 4).text() or 0)
            except (ValueError, AttributeError):
                m = 0.0
            try:
                k = float(self.table.item(row, 5).text() or 0)
            except (ValueError, AttributeError):
                k = 0.0
            results.append({"item_id": self._item_ids[row], "meter": m, "kg": k})
        return {"location": self.location.currentData(), "items": results}


class SiparisOnayDialog(QDialog):
    """Admin için sipariş sözleşme önizleme + onay/red dialog."""

    def __init__(self, order, parent=None):
        super().__init__(parent)
        self.order    = order
        self.approved = False   # True → onaylandı, False → reddedildi / kapatıldı
        self.setWindowTitle(f"Sipariş İnceleme — {order.get('order_no','')}")
        self.setMinimumSize(820, 680)
        self._build_ui()

    def _build_ui(self):
        from datetime import datetime
        lay = QVBoxLayout(self); lay.setSpacing(8)

        # ── Başlık ──────────────────────────────────────────────────────────
        hdr = QLabel("📋  SİPARİŞ FORMU — ADMIN İNCELEME")
        hdr.setStyleSheet(
            "font-size:16px;font-weight:bold;color:#1A237E;"
            "padding:8px 12px;background:#E8EAF6;border-radius:6px;")
        lay.addWidget(hdr)

        # ── Kaydırılabilir içerik ────────────────────────────────────────────
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        inner = QWidget(); scroll.setWidget(inner)
        vl = QVBoxLayout(inner); vl.setSpacing(14)

        o = self.order
        cfg = db.get_company_settings() or {}

        def _fmt_date(s):
            try: return datetime.strptime((s or "")[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
            except: return s or "—"

        def _grp(title, color="#1565C0"):
            g = QGroupBox(title); g.setStyleSheet(
                f"QGroupBox{{font-weight:bold;color:{color};"
                f"border:1px solid {color};border-radius:6px;margin-top:6px;padding:8px;}}"
                f"QGroupBox::title{{subcontrol-origin:margin;left:10px;padding:0 4px;}}"
            ); return g

        # ── 1. Taraflar ──────────────────────────────────────────────────────
        taraf_grp = _grp("1 — Taraflar")
        tg = QGridLayout(taraf_grp); tg.setSpacing(4)

        def _th(txt):
            l = QLabel(txt)
            l.setStyleSheet("font-weight:bold;color:#fff;background:#1565C0;"
                            "padding:4px 8px;border-radius:3px;")
            return l

        tg.addWidget(_th("SATICI"), 0, 0)
        tg.addWidget(_th("ALICI"), 0, 1)

        def _cell(txt, bold=False):
            l = QLabel(txt or "—")
            l.setWordWrap(True)
            l.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            if bold: l.setStyleSheet("font-weight:bold;font-size:13px;")
            return l

        satici = QWidget(); sl = QVBoxLayout(satici); sl.setContentsMargins(0,4,0,0)
        sl.addWidget(_cell(cfg.get("name",""), bold=True))
        sl.addWidget(_cell(cfg.get("address","")))
        sl.addWidget(_cell(f"Tel: {cfg.get('phone','')}  |  Vergi: {cfg.get('tax','')}"))
        sl.addWidget(_cell(cfg.get("email_info","")))
        tg.addWidget(satici, 1, 0)

        alici = QWidget(); al = QVBoxLayout(alici); al.setContentsMargins(0,4,0,0)
        al.addWidget(_cell(o.get("customer_name",""), bold=True))
        al.addWidget(_cell(o.get("delivery_address","")))
        al.addWidget(_cell(f"Müşteri Ref: {o.get('customer_ref','') or '—'}"))
        tg.addWidget(alici, 1, 1)
        tg.setColumnStretch(0, 1); tg.setColumnStretch(1, 1)
        vl.addWidget(taraf_grp)

        # ── 2. Sipariş Başlık Bilgileri ──────────────────────────────────────
        bas_grp = _grp("2 — Sipariş Bilgileri")
        bg = QGridLayout(bas_grp); bg.setSpacing(6)
        bas_rows = [
            ("Sipariş No:",        o.get("order_no",""),                  True),
            ("Sipariş Tarihi:",    _fmt_date(o.get("order_date","")),     False),
            ("Termin:",            _fmt_date(o.get("delivery_date","")),  True),
            ("Para Birimi:",       o.get("currency",""),                  False),
            ("Ödeme Şekli:",       o.get("payment_method",""),            False),
            ("Teslimat Şartları:", o.get("delivery_terms",""),            False),
            ("Siparişi Alan:",     o.get("sales_rep","") or "—",         False),
            ("Oluşturan:",         o.get("created_by",""),                False),
        ]
        for i, (lbl, val, bold) in enumerate(bas_rows):
            row_i, col_i = divmod(i, 2)
            col_base = col_i * 2
            bg.addWidget(QLabel(f"<b>{lbl}</b>"), row_i, col_base)
            v = QLabel(val or "—")
            v.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            if bold: v.setStyleSheet("font-weight:bold;color:#C62828;font-size:13px;")
            bg.addWidget(v, row_i, col_base + 1)
        bg.setColumnStretch(1, 1); bg.setColumnStretch(3, 1)
        vl.addWidget(bas_grp)

        # ── 3. Sipariş Kalemleri ─────────────────────────────────────────────
        kal_grp = _grp("3 — Sipariş Kalemleri")
        kg = QVBoxLayout(kal_grp)
        items = o.get("items") or []
        toplam_mt = 0.0; toplam_kg = 0.0; toplam_tutar = 0.0
        cur = o.get("currency", "USD")
        if not items:
            kg.addWidget(QLabel("Bu siparişte kalem yok."))
        for idx, it in enumerate(items, start=1):
            mt  = float(it.get("meter") or 0)
            kgi = float(it.get("kg") or 0)
            sp  = float(it.get("sale_price") or 0)
            tut = mt * sp
            toplam_mt += mt; toplam_kg += kgi; toplam_tutar += tut

            card = QFrame()
            card.setStyleSheet("QFrame{background:#FAFAFA;border:1px solid #CFD8DC;border-radius:6px;}")
            cl = QVBoxLayout(card); cl.setSpacing(4)
            title = QLabel(f"Kalem {idx}:  {it.get('product_code','') or '—'}  —  {it.get('product_name','') or ''}")
            title.setStyleSheet("font-weight:bold;color:#1565C0;font-size:13px;")
            cl.addWidget(title)

            grid = QGridLayout(); grid.setSpacing(5); grid.setContentsMargins(4, 2, 4, 2)
            fields = [
                ("Ürün Kodu",      it.get("product_code","")),
                ("Ürün Adı",       it.get("product_name","")),
                ("Kumaş Tipi",     it.get("fabric_type","")),
                ("Renk",           it.get("color","")),
                ("Kompozisyon",    it.get("composition","")),
                ("En (cm)",        it.get("width","")),
                ("Gramaj",         it.get("gramaj","")),
                ("Lab No",         it.get("lab_no","")),
                ("Baskı Tipi",     it.get("print_type","")),
                ("Zemin Rengi",    it.get("zemin_rengi","")),
                ("Baskı Desen No", it.get("baski_desen_no","")),
                ("Miktar (mt)",    f"{mt:.2f}"),
                ("Miktar (kg)",    f"{kgi:.2f}"),
                ("Birim Fiyat",    f"{sp:.2f} {cur}" if sp else "—"),
                ("Tutar",          f"{tut:.2f} {cur}" if sp else "—"),
                ("Açıklama",       it.get("description","")),
            ]
            for i, (lbl, val) in enumerate(fields):
                r, c = divmod(i, 2)          # sağlı-sollu: 2 alan yan yana
                cb = c * 2
                grid.addWidget(QLabel(f"<b>{lbl}:</b>"), r, cb)
                v = QLabel(str(val) if str(val) else "—")
                v.setWordWrap(True)
                v.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
                if lbl == "Tutar" and sp:
                    v.setStyleSheet("font-weight:bold;color:#C62828;")
                grid.addWidget(v, r, cb + 1)
            grid.setColumnStretch(1, 1); grid.setColumnStretch(3, 1)
            cl.addLayout(grid)
            kg.addWidget(card)

        # Toplamlar
        tot_lay = QHBoxLayout(); tot_lay.addStretch()
        tot_frame = QFrame()
        tot_frame.setStyleSheet("background:#E3F2FD;border:1px solid #1565C0;"
                                "border-radius:6px;padding:8px;")
        tf = QGridLayout(tot_frame); tf.setSpacing(6)
        for ri2, (lbl, val) in enumerate([
            ("Toplam Miktar (mt):", f"{toplam_mt:.2f} mt"),
            ("Toplam Miktar (kg):", f"{toplam_kg:.2f} kg"),
            ("Toplam Tutar:",       f"{toplam_tutar:.2f} {cur}"),
        ]):
            l1 = QLabel(f"<b>{lbl}</b>"); l2 = QLabel(val)
            if "Tutar" in lbl:
                l2.setStyleSheet("font-weight:bold;font-size:14px;color:#C62828;")
            tf.addWidget(l1, ri2, 0); tf.addWidget(l2, ri2, 1)
        tot_lay.addWidget(tot_frame)
        kg.addLayout(tot_lay)
        vl.addWidget(kal_grp)

        # ── 4. Sözleşme Koşulları / Notlar ──────────────────────────────────
        if o.get("contract_terms","") or o.get("notes",""):
            not_grp = _grp("4 — Sözleşme Koşulları & Notlar", color="#37474F")
            nl = QVBoxLayout(not_grp)
            if o.get("contract_terms",""):
                nl.addWidget(QLabel("<b>Sözleşme Koşulları:</b>"))
                ct = QLabel(o.get("contract_terms",""))
                ct.setWordWrap(True)
                ct.setStyleSheet("background:#FAFAFA;border:1px solid #CFD8DC;"
                                 "border-radius:4px;padding:6px;")
                ct.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
                nl.addWidget(ct)
            if o.get("notes",""):
                nl.addWidget(QLabel("<b>Notlar:</b>"))
                nt = QLabel(o.get("notes",""))
                nt.setWordWrap(True)
                nt.setStyleSheet("background:#FFFDE7;border:1px solid #F9A825;"
                                 "border-radius:4px;padding:6px;")
                nt.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
                nl.addWidget(nt)
            vl.addWidget(not_grp)

        vl.addStretch()
        lay.addWidget(scroll, 1)

        # ── Alt butonlar ────────────────────────────────────────────────────
        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color:#CFD8DC;"); lay.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.addStretch()

        btn_red = QPushButton("✕  Reddet / Kapat")
        btn_red.setStyleSheet(
            "background:#B71C1C;color:white;font-weight:bold;"
            "border-radius:6px;padding:8px 22px;font-size:13px;")
        btn_red.clicked.connect(self.reject)

        btn_ok = QPushButton("✅  Onayla")
        btn_ok.setStyleSheet(
            "background:#1B5E20;color:white;font-weight:bold;"
            "border-radius:6px;padding:8px 28px;font-size:14px;")
        btn_ok.clicked.connect(self._do_approve)

        btn_row.addWidget(btn_red)
        btn_row.addSpacing(12)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

    def _do_approve(self):
        reply = QMessageBox.question(
            self, "Onayla",
            f"<b>{self.order.get('order_no','')}</b> numaralı sipariş onaylansın mı?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.approved = True
            self.accept()


class PlanningView(QWidget):
    _PLANNING_STATUSES = [
        "ONAYDA", "ONAYLANDI - PLANLAMADA", "PLANLAMA - GÖRDÜ",
        "PLANLANDI", "BOYAHANAYA SEVKLER BAŞLADI", "TÜM KUMAŞLAR BOYAHANEDE",
    ]
    _STATUS_COLORS = {
        "ONAYDA":                    "#B71C1C",
        "ONAYLANDI - PLANLAMADA":    "#E65100",
        "PLANLAMA - GÖRDÜ":          "#F57F17",
        "PLANLANDI":                 "#1565C0",
        "BOYAHANAYA SEVKLER BAŞLADI":"#6A1B9A",
        "TÜM KUMAŞLAR BOYAHANEDE":  "#2E7D32",
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self._current_order_id = None
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        top.addWidget(QLabel("<b>📌 Planlama Modülü</b>"))
        top.addStretch()
        btn_ref = QPushButton("🔄 Yenile"); btn_ref.clicked.connect(self.refresh)
        top.addWidget(btn_ref)
        lay.addLayout(top)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        lay.addWidget(splitter)

        # ── Sol: Kuyruk ──────────────────────────────────────────
        left = QWidget()
        ll = QVBoxLayout(left); ll.setContentsMargins(0, 0, 4, 0)
        ll.addWidget(QLabel("<b>Aktif Planlamalar</b>"))
        self._pending_lbl = QLabel()
        self._pending_lbl.setStyleSheet(
            "background:#B71C1C;color:white;font-weight:bold;"
            "padding:4px 8px;border-radius:4px;")
        self._pending_lbl.setVisible(False)
        ll.addWidget(self._pending_lbl)
        self.queue_table = QTableWidget()
        self.queue_table.setColumnCount(5)
        self.queue_table.setHorizontalHeaderLabels(
            ["Sipariş No", "Müşteri", "Termin", "Ürün Kodu", "Durum"])
        self.queue_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.queue_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.queue_table.verticalHeader().setVisible(False)
        self.queue_table.setSortingEnabled(True)
        qhdr = self.queue_table.horizontalHeader()
        qhdr.setSectionsMovable(True)
        qhdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        qhdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        qhdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        qhdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        qhdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.queue_table.itemSelectionChanged.connect(self._on_queue_select)
        ll.addWidget(self.queue_table)

        # Admin onay butonu
        self._btn_approve = QPushButton("✅ Seçili Siparişi Onayla")
        self._btn_approve.setStyleSheet(
            "background:#1B5E20;color:white;font-weight:bold;"
            "border-radius:4px;padding:5px 10px;")
        self._btn_approve.clicked.connect(self._approve_order)
        self._btn_approve.setVisible(CURRENT_USER.get("role") == "admin")
        ll.addWidget(self._btn_approve)
        splitter.addWidget(left)

        # ── Sağ: Detay (scroll) ──────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        right = QWidget()
        rl = QVBoxLayout(right); rl.setContentsMargins(4, 0, 0, 8)
        scroll.setWidget(right)

        # Sipariş bilgileri
        self._order_group = QGroupBox("📋 Sipariş Bilgileri")
        self._order_group.setVisible(False)
        og = QGridLayout(self._order_group); og.setSpacing(6)
        self._info = {}
        fields = [
            ("order_no", "Sipariş No"), ("order_date", "Sipariş Tarihi"),
            ("customer_name", "Müşteri"), ("customer_ref", "Müşteri Ref."),
            ("currency", "Para Birimi"), ("payment_method", "Ödeme Şekli"),
            ("delivery_terms", "Teslimat Şartları"), ("delivery_date", "Termin"),
            ("delivery_address", "Teslimat Adresi"), ("notes", "Notlar"),
            ("status", "Durum"),
        ]
        for i, (key, lbl) in enumerate(fields):
            col_base = (i % 2) * 2
            row_g = i // 2
            og.addWidget(QLabel(f"<b>{lbl}:</b>"), row_g, col_base)
            val_lbl = QLabel()
            val_lbl.setWordWrap(True)
            val_lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            og.addWidget(val_lbl, row_g, col_base + 1)
            self._info[key] = val_lbl
        og.setColumnStretch(1, 1)
        og.setColumnStretch(3, 1)
        rl.addWidget(self._order_group)

        # Ham dokuma ihtiyaç tablosu
        rl.addWidget(QLabel("<b>Ham Dokuma İhtiyaç Tablosu</b>"))
        self.need_table = QTableWidget()
        _NEED_COLS = [
            "Seç", "Ürün Kodu", "Ürün Adı", "Kumaş Tipi", "Renk",
            "Kompozisyon", "En", "Gramaj", "Lab No", "Baskı Tipi",
            "Zemin Rengi", "Baskı Desen No", "Sip. Mt", "Sip. Kg",
            "Satış Fiyatı", "Açıklama", "DEPO HAM (mt)", "Eksik (mt)"]
        self.need_table.setColumnCount(len(_NEED_COLS))
        self.need_table.setHorizontalHeaderLabels(_NEED_COLS)
        self.need_table.verticalHeader().setVisible(False)
        self.need_table.setSortingEnabled(True)
        nhdr = self.need_table.horizontalHeader()
        nhdr.setSectionsMovable(True)
        nhdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        for i in range(len(_NEED_COLS)):
            if i != 2:
                nhdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self.need_table.setMaximumHeight(200)
        rl.addWidget(self.need_table)

        btn_po = QPushButton("+ Satınalma Siparişi Oluştur (Seçili Kalemler)")
        btn_po.setStyleSheet("background:#1565C0;color:white;font-weight:bold;"
                             "border-radius:4px;padding:6px 14px;")
        btn_po.clicked.connect(self._create_po)
        rl.addWidget(btn_po)

        rl.addWidget(QLabel("<b>Satınalma Kalemleri</b> (çift tıklayarak mal girişi yapabilirsiniz)"))
        self.po_table = QTableWidget()
        self.po_table.setColumnCount(12)
        self.po_table.setHorizontalHeaderLabels([
            "PO No", "Tedarikçi", "Ürün Kodu", "Kumaş Tipi",
            "Ham Sip.Mt", "Ham Sip.Kg", "Gelen Mt", "Kalan Mt",
            "Gelen Kg", "Kalan Kg", "Birim Fiyat", "Durum"])
        self.po_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.po_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.po_table.verticalHeader().setVisible(False)
        self.po_table.setSortingEnabled(True)
        phdr = self.po_table.horizontalHeader()
        phdr.setSectionsMovable(True)
        phdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        phdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        for i in (0, 3, 4, 5, 6, 7, 8, 9, 10, 11):
            phdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self.po_table.setMaximumHeight(200)
        self.po_table.itemSelectionChanged.connect(self._on_po_item_select)
        self.po_table.cellDoubleClicked.connect(self._po_item_dbl_click)
        rl.addWidget(self.po_table)

        po_btns = QHBoxLayout()
        for lbl, slot in [("📄 PDF Al", self._po_pdf),
                           ("📧 Mail Gönder", self._po_mail),
                           ("📝 Durum Değiştir", self._po_change_status),
                           ("🗑 Sil", self._po_delete)]:
            b = QPushButton(lbl); b.clicked.connect(slot); po_btns.addWidget(b)
        po_btns.addStretch()
        rl.addLayout(po_btns)

        rl.addWidget(QLabel("<b>Mal Giriş Kayıtları</b> (her giriş ayrı satır)"))
        self.po_receipts_table = QTableWidget()
        self.po_receipts_table.setColumnCount(7)
        self.po_receipts_table.setHorizontalHeaderLabels([
            "Tarih", "Ürün Kodu", "Kumaş Tipi", "Lot", "Metre", "Kilo", "Lokasyon"])
        self.po_receipts_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.po_receipts_table.verticalHeader().setVisible(False)
        self.po_receipts_table.setSortingEnabled(True)
        rrhdr = self.po_receipts_table.horizontalHeader()
        rrhdr.setSectionsMovable(True)
        rrhdr.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        for i in (0, 1, 2, 3, 4, 5):
            rrhdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self.po_receipts_table.setMaximumHeight(160)
        rl.addWidget(self.po_receipts_table)

        receipt_btns = QHBoxLayout()
        b_rec_edit = QPushButton("✏ Girişi Düzenle")
        b_rec_edit.clicked.connect(self._edit_po_receipt)
        b_rec_del  = QPushButton("🗑 Girişi Sil")
        b_rec_del.clicked.connect(self._delete_po_receipt)
        b_rec_del.setStyleSheet("color:#C62828;")
        receipt_btns.addWidget(b_rec_edit)
        receipt_btns.addWidget(b_rec_del)
        receipt_btns.addStretch()
        rl.addLayout(receipt_btns)

        rl.addStretch()
        splitter.addWidget(scroll)
        splitter.setSizes([300, 700])
        self._right_scroll = scroll
        self._right_scroll.setVisible(False)

    # ── Yenile ───────────────────────────────────────────────────
    def refresh(self):
        self._refresh_queue()
        if self._current_order_id:
            self._show_detail(self._current_order_id)

    def _refresh_queue(self):
        from datetime import datetime
        all_orders = db.get_all_orders()
        rows = [r for r in all_orders if r.get("status") in self._PLANNING_STATUSES]
        pending_count = sum(1 for r in rows if r.get("status") == "ONAYDA")
        if pending_count:
            self._pending_lbl.setText(f"⚠ {pending_count} yeni sipariş onay bekliyor!")
            self._pending_lbl.setVisible(True)
        else:
            self._pending_lbl.setVisible(False)
        self.queue_table.blockSignals(True)
        self.queue_table.setSortingEnabled(False)
        self.queue_table.setRowCount(0)
        for r in rows:
            row = self.queue_table.rowCount()
            self.queue_table.insertRow(row)
            item0 = QTableWidgetItem(r.get("order_no","") or "")
            item0.setData(Qt.ItemDataRole.UserRole, r["id"])
            self.queue_table.setItem(row, 0, item0)
            self.queue_table.setItem(row, 1, QTableWidgetItem(r.get("customer_name","") or ""))
            td = r.get("delivery_date","") or ""
            try:
                td = datetime.strptime(td[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                pass
            self.queue_table.setItem(row, 2, QTableWidgetItem(td))
            # Ürün kodları
            codes = list(dict.fromkeys(
                it.get("product_code","") for it in (r.get("items") or [])
                if it.get("product_code")))
            self.queue_table.setItem(row, 3, QTableWidgetItem(", ".join(codes)))
            st = r.get("status","")
            st_item = QTableWidgetItem(st)
            color = self._STATUS_COLORS.get(st, "#212121")
            st_item.setForeground(QBrush(QColor(color)))
            self.queue_table.setItem(row, 4, st_item)
            if r["id"] == self._current_order_id:
                self.queue_table.selectRow(row)
        self.queue_table.setSortingEnabled(True)
        self.queue_table.blockSignals(False)

    # ── Sipariş seçildi ──────────────────────────────────────────
    def _on_queue_select(self):
        row = self.queue_table.currentRow()
        if row < 0:
            return
        item = self.queue_table.item(row, 0)
        if not item:
            return
        oid = item.data(Qt.ItemDataRole.UserRole)
        order = db.get_order(oid)
        if not order:
            return
        st = order.get("status","")
        if st == "ONAYDA":
            if CURRENT_USER.get("role") == "admin":
                dlg = SiparisOnayDialog(order, self)
                dlg.exec()
                if dlg.approved:
                    db.approve_order(oid, CURRENT_USER["full_name"])
                    self._refresh_queue()
                    QMessageBox.information(self, "Onaylandı",
                        f"{order.get('order_no','')} onaylandı. Planlama ekibine hazır.")
            else:
                QMessageBox.warning(self, "Onay Gerekli",
                    "Bu sipariş henüz admin tarafından onaylanmamış.\n"
                    "Planlama başlatmak için önce onaylanması gerekiyor.")
            return
        if st == "ONAYLANDI - PLANLAMADA":
            db.update_order_status(oid, "PLANLAMA - GÖRDÜ")
        self._current_order_id = oid
        self._refresh_queue()
        self._show_detail(oid)

    def _approve_order(self):
        row = self.queue_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir sipariş seçin.")
            return
        item = self.queue_table.item(row, 0)
        if not item:
            return
        oid = item.data(Qt.ItemDataRole.UserRole)
        order = db.get_order(oid)
        if not order:
            return
        if order.get("status") != "ONAYDA":
            QMessageBox.information(self, "Bilgi",
                f"Bu sipariş zaten '{order.get('status')}' durumunda.")
            return
        reply = QMessageBox.question(self, "Onay",
            f"{order.get('order_no','')} numaralı sipariş onaylansın mı?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            db.approve_order(oid, CURRENT_USER["full_name"])
            self._refresh_queue()
            QMessageBox.information(self, "Onaylandı",
                "Sipariş onaylandı. Planlama ekibine hazır.")

    # ── PO kalemi seçildi → makbuzları göster ───────────────────
    def _on_po_item_select(self):
        row = self.po_table.currentRow()
        if row < 0:
            self.po_receipts_table.setRowCount(0)
            return
        item = self.po_table.item(row, 0)
        if not item:
            return
        data = item.data(Qt.ItemDataRole.UserRole)
        po_id = data[0] if isinstance(data, tuple) else data
        if po_id:
            self._refresh_receipts_table(po_id)

    def _refresh_receipts_table(self, po_id):
        from datetime import datetime
        receipts = db.get_po_receipts(po_id)
        self.po_receipts_table.setSortingEnabled(False)
        self.po_receipts_table.setRowCount(0)
        for r in receipts:
            row = self.po_receipts_table.rowCount()
            self.po_receipts_table.insertRow(row)
            dt = r.get("received_at","") or ""
            try:
                dt = datetime.strptime(dt[:16], "%Y-%m-%d %H:%M").strftime("%d.%m.%Y %H:%M")
            except Exception:
                pass
            m = float(r.get("meter",0) or 0)
            k = float(r.get("kg",0) or 0)
            grp = r.get("location_group","")
            for col, val in enumerate([
                dt,
                r.get("product_code",""),
                r.get("fabric_type",""),
                r.get("lot","") or "",
                f"{m:.2f}",
                f"{k:.2f}",
                f"{r.get('location','')} [{grp}]",
            ]):
                cell = QTableWidgetItem(str(val))
                cell.setData(Qt.ItemDataRole.UserRole, r.get("id"))
                cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if col == 6 and grp != "DEPO":
                    cell.setForeground(QBrush(QColor("#E65100")))
                self.po_receipts_table.setItem(row, col, cell)
        self.po_receipts_table.setSortingEnabled(True)

    # ── Çift tıklama mal girişi (po_table satırına) ──────────────
    def _po_item_dbl_click(self, row, _col):
        cell = self.po_table.item(row, 0)
        if not cell:
            return
        data = cell.data(Qt.ItemDataRole.UserRole)
        if not isinstance(data, tuple):
            return
        po_id, po_item_id = data
        if not po_item_id:
            return
        po = db.get_purchase_order(po_id)
        if not po:
            return
        po_item = next((i for i in (po.get("items") or []) if i.get("id") == po_item_id), None)
        if not po_item:
            return
        dlg = ItemReceiptDialog(self, po_item=po_item)
        if dlg.exec():
            d = dlg.get_data()
            try:
                db.receive_purchase_order_item(
                    po_item_id, d["metre"], d["kilo"],
                    d["location"], user_name=CURRENT_USER["full_name"],
                    location_group=d["location_group"], lot=d["lot"])
                self._refresh_po_table()
                self._refresh_receipts_table(po_id)
                if self._current_order_id:
                    self._show_detail(self._current_order_id)
                grp_msg = ("Ana depoya kaydedildi." if d["location_group"] == "DEPO"
                           else "Dış depoya kaydedildi — Boyahane planlama ekranına düşecek.")
                QMessageBox.information(self, "Kaydedildi",
                    f"Mal girişi tamamlandı.\n{grp_msg}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Mal girişi kaydedilemedi:\n{e}")

    def _selected_po_id_silent(self):
        row = self.po_table.currentRow()
        if row < 0:
            return None
        item = self.po_table.item(row, 0)
        if not item:
            return None
        data = item.data(Qt.ItemDataRole.UserRole)
        return data[0] if isinstance(data, tuple) else data

    # ── Detay paneli ─────────────────────────────────────────────
    def _show_detail(self, order_id):
        from datetime import datetime
        order = db.get_order(order_id)
        if not order:
            return

        def _fmt(key):
            val = order.get(key,"") or ""
            if key in ("delivery_date","order_date") and val:
                try: val = datetime.strptime(val[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception: pass
            return str(val)

        for key, lbl_w in self._info.items():
            lbl_w.setText(_fmt(key))
        self._order_group.setVisible(True)

        items = order.get("items", [])
        self.need_table.setSortingEnabled(False)
        self.need_table.setRowCount(0)
        for it in items:
            pc = it.get("product_code","")
            fabric_type = it.get("fabric_type","") or ""
            stock = db.get_fabric_stock_in_depo(pc, fabric_type if fabric_type == "HAM" else "HAM") or {"meter": 0}
            depo_m = float(stock.get("meter", 0) or 0)
            order_m = float(it.get("meter") or 0)
            order_kg = float(it.get("kg") or 0)
            missing = max(0.0, order_m - depo_m)

            row = self.need_table.rowCount()
            self.need_table.insertRow(row)

            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            chk.setCheckState(Qt.CheckState.Checked if missing > 0 else Qt.CheckState.Unchecked)
            chk.setData(Qt.ItemDataRole.UserRole, dict(it))
            self.need_table.setItem(row, 0, chk)

            sp = it.get("sale_price") or 0
            row_vals = [
                pc,                             # 1 Ürün Kodu
                it.get("product_name",""),       # 2 Ürün Adı
                fabric_type,                    # 3 Kumaş Tipi
                it.get("color",""),             # 4 Renk
                it.get("composition",""),        # 5 Kompozisyon
                it.get("width",""),             # 6 En
                it.get("gramaj",""),            # 7 Gramaj
                it.get("lab_no",""),            # 8 Lab No
                it.get("print_type",""),        # 9 Baskı Tipi
                it.get("zemin_rengi",""),       # 10 Zemin Rengi
                it.get("baski_desen_no",""),    # 11 Baskı Desen No
                f"{order_m:.2f}",              # 12 Sip. Mt
                f"{order_kg:.2f}",             # 13 Sip. Kg
                f"{sp:.2f}" if sp else "",     # 14 Satış Fiyatı
                it.get("description",""),       # 15 Açıklama
                f"{depo_m:.2f}",              # 16 DEPO HAM (mt)
                f"{missing:.2f}",             # 17 Eksik (mt)
            ]
            for col, val in enumerate(row_vals, start=1):
                cell = QTableWidgetItem(str(val))
                cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if col == 17:
                    if missing > 0:
                        cell.setText(f"{missing:.2f}  ⚠ Eksik")
                        cell.setForeground(QBrush(QColor("#C62828")))
                    else:
                        cell.setText("✅ Yeterli")
                        cell.setForeground(QBrush(QColor("#2E7D32")))
                self.need_table.setItem(row, col, cell)
        self.need_table.setSortingEnabled(True)

        self._refresh_po_table(order_id)
        self._right_scroll.setVisible(True)

    def _refresh_po_table(self, order_id=None):
        oid = order_id or self._current_order_id
        if not oid:
            return
        items = db.get_po_items_for_order(oid)
        self.po_table.blockSignals(True)
        self.po_table.setSortingEnabled(False)
        self.po_table.setRowCount(0)
        for it in items:
            row = self.po_table.rowCount()
            self.po_table.insertRow(row)
            po_id = it.get("po_id") or it.get("po_id_ref")
            item_id = it.get("id")
            om = float(it.get("meter",0) or 0)
            ok = float(it.get("kg",0) or 0)
            rm = float(it.get("received_meter",0) or 0)
            rk = float(it.get("received_kg",0) or 0)
            kalan_m = max(0.0, om - rm)
            kalan_k = max(0.0, ok - rk)
            up = float(it.get("unit_price",0) or 0)
            po_status = it.get("po_status","") or ""

            item0 = QTableWidgetItem(it.get("po_no",""))
            item0.setData(Qt.ItemDataRole.UserRole, (po_id, item_id))
            item0.setFlags(item0.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.po_table.setItem(row, 0, item0)

            vals = [
                it.get("supplier_name","") or "",        # 1 Tedarikçi
                it.get("product_code","") or "",          # 2 Ürün Kodu
                it.get("fabric_type","") or "",           # 3 Kumaş Tipi
                f"{om:.2f}",                              # 4 Ham Sip.Mt
                f"{ok:.2f}",                              # 5 Ham Sip.Kg
                f"{rm:.2f}",                              # 6 Gelen Mt
                f"{kalan_m:.2f}",                         # 7 Kalan Mt
                f"{rk:.2f}",                              # 8 Gelen Kg
                f"{kalan_k:.2f}",                         # 9 Kalan Kg
                f"{up:.4f}" if up else "—",               # 10 Birim Fiyat
                po_status,                                # 11 Durum
            ]
            for col, val in enumerate(vals, start=1):
                cell = QTableWidgetItem(str(val))
                cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if col == 6:  # Gelen Mt
                    cell.setForeground(QBrush(QColor("#2E7D32" if rm >= om else "#C62828")))
                if col == 7 and kalan_m > 0:  # Kalan Mt
                    cell.setForeground(QBrush(QColor("#E65100")))
                if col == 11:
                    sc = {"TAMAMLANDI": "#2E7D32", "GÖNDERİLDİ": "#1565C0",
                          "KISMİ GELDİ": "#E65100", "İPTAL": "#757575"}.get(po_status)
                    if sc:
                        cell.setForeground(QBrush(QColor(sc)))
                self.po_table.setItem(row, col, cell)
        self.po_table.setSortingEnabled(True)
        self.po_table.blockSignals(False)
        self.po_receipts_table.setRowCount(0)

    def _selected_po_id(self):
        row = self.po_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir satınalma kalemi seçin.")
            return None
        item = self.po_table.item(row, 0)
        if not item:
            return None
        data = item.data(Qt.ItemDataRole.UserRole)
        return data[0] if isinstance(data, tuple) else data

    # ── PO Oluştur ───────────────────────────────────────────────
    def _create_po(self):
        if not self._current_order_id:
            return
        order = db.get_order(self._current_order_id)
        if not order:
            return
        missing_items = []
        for row in range(self.need_table.rowCount()):
            chk = self.need_table.item(row, 0)
            if chk and chk.checkState() == Qt.CheckState.Checked:
                it = chk.data(Qt.ItemDataRole.UserRole)
                pc = it.get("product_code","")
                stock = db.get_fabric_stock_in_depo(pc, "HAM") or {"meter": 0}
                depo_m = float(stock.get("meter", 0) or 0)
                order_m = float(it.get("meter") or 0)
                missing = max(0.0, order_m - depo_m)
                if missing > 0:
                    missing_items.append({
                        "product_code": pc,
                        "product_name": it.get("product_name",""),
                        "composition":  it.get("composition",""),
                        "width":        it.get("width",""),
                        "gramaj":       it.get("gramaj",""),
                        "fabric_type":  "HAM",
                        "meter":        missing, "kg": 0, "unit_price": 0,
                        "description":  f"{order.get('order_no','')} için ham dokuma",
                    })
        if not missing_items:
            QMessageBox.information(self, "Bilgi", "Eksik kalem seçilmedi veya tüm stok yeterli.")
            return
        dlg = PurchaseOrderDialog(self, missing_items=missing_items)
        if dlg.exec():
            d = dlg.get_data()
            try:
                created_po_nos = []
                for grp in d["po_groups"]:
                    _, po_no = db.add_purchase_order(
                        supplier_id=grp["supplier_id"],
                        supplier_name=grp["supplier_name"],
                        order_id=self._current_order_id,
                        order_no=order.get("order_no",""),
                        currency=grp.get("currency",""),
                        payment_method=grp.get("payment_method",""),
                        delivery_terms=grp.get("delivery_terms",""),
                        expected_delivery=grp.get("expected_delivery",""),
                        notes=grp.get("notes",""),
                        items=grp["items"],
                        created_by=CURRENT_USER["full_name"])
                    created_po_nos.append(po_no)
                cur_st = order.get("status","")
                if cur_st in ("PLANLAMA - GÖRDÜ", "ONAYLANDI - PLANLAMADA"):
                    db.update_order_status(self._current_order_id, "PLANLANDI")
                self._refresh_po_table()
                if self._current_order_id:
                    self._show_detail(self._current_order_id)
                po_list = "\n".join(f"  • {n}" for n in created_po_nos)
                QMessageBox.information(self, "Oluşturuldu",
                    f"{len(created_po_nos)} adet satınalma siparişi oluşturuldu:\n{po_list}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"PO oluşturulamadı:\n{e}")

    # ── PDF ──────────────────────────────────────────────────────
    def _po_pdf(self):
        po_id = self._selected_po_id()
        if not po_id:
            return
        po = db.get_purchase_order(po_id)
        if not po:
            return
        company = db.get_company_settings()
        path, _ = QFileDialog.getSaveFileName(
            self, "PO PDF Kaydet", f"{po['po_no']}.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            import purchase_order_pdf as pop
            pop.generate_purchase_order_pdf(po, company, path)
            QMessageBox.information(self, "PDF Oluşturuldu", f"Kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF oluşturulamadı:\n{e}")

    # ── Mail ─────────────────────────────────────────────────────
    def _po_mail(self):
        po_id = self._selected_po_id()
        if not po_id:
            return
        po = db.get_purchase_order(po_id)
        if not po:
            return
        sup = db.get_supplier(po.get("supplier_id")) if po.get("supplier_id") else None
        email_addr = (dict(sup).get("email","") if sup else "") or ""
        if not email_addr:
            QMessageBox.warning(self, "E-posta Yok",
                f"Tedarikçi '{po.get('supplier_name','')}' için e-posta adresi tanımlı değil.\n"
                "Tedarikçiler menüsünden e-posta ekleyiniz.")
            return
        import tempfile, os as _os2
        tmp = tempfile.mktemp(suffix=".pdf")
        try:
            import purchase_order_pdf as pop
            company = db.get_company_settings()
            pop.generate_purchase_order_pdf(po, company, tmp)
            import email_report as er
            er.send_email_with_attachment(
                email_addr,
                subject=f"Satınalma Siparişi {po['po_no']}",
                body_text=(f"Sayın {po.get('supplier_name','')} Yetkilileri,\n\n"
                           f"Ek'te {po['po_no']} numaralı satınalma sipariş formumuzu "
                           f"bulabilirsiniz.\n\nİyi çalışmalar,\nBursa Knitted"),
                attachment_path=tmp,
            )
            db.update_purchase_order_status(po_id, "GÖNDERİLDİ")
            self._refresh_po_table()
            QMessageBox.information(self, "Gönderildi",
                f"E-posta {email_addr} adresine gönderildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"E-posta gönderilemedi:\n{e}")
        finally:
            try:
                _os2.unlink(tmp)
            except Exception:
                pass

    # ── Mal Girişi ───────────────────────────────────────────────
    # _po_receive yerine _po_item_dbl_click kullanılıyor (kalem satırına çift tıklama)

    # ── Durum Değiştir ───────────────────────────────────────────
    def _po_change_status(self):
        po_id = self._selected_po_id()
        if not po_id:
            return
        row = self.po_table.currentRow()
        current_status = self.po_table.item(row, 11).text() if row >= 0 else ""
        dlg = QDialog(self); dlg.setWindowTitle("PO Durumu Değiştir")
        form = QFormLayout(dlg)
        combo = QComboBox()
        for s in PO_STATUS_OPTIONS:
            combo.addItem(s)
        idx = combo.findText(current_status)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        form.addRow("Yeni Durum:", combo)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        form.addRow(btns)
        if dlg.exec():
            db.update_purchase_order_status(po_id, combo.currentText())
            self._refresh_po_table()

    # ── Sil ──────────────────────────────────────────────────────
    def _po_delete(self):
        po_id = self._selected_po_id()
        if not po_id:
            return
        row = self.po_table.currentRow()
        po_no = self.po_table.item(row, 0).text() if row >= 0 else str(po_id)
        current_status = self.po_table.item(row, 11).text() if row >= 0 else ""
        if current_status not in ("TAMAMLANDI", "İPTAL"):
            QMessageBox.warning(self, "Silinemez",
                f"'{po_no}' PO'su şu an '{current_status}' durumunda.\n\n"
                "Silebilmek için önce 'Durum Değiştir' butonu ile\n"
                "TAMAMLANDI veya İPTAL durumuna getiriniz.")
            return
        if QMessageBox.question(
            self, "Sil", f"{po_no} satınalma siparişi kalıcı olarak silinsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            db.delete_purchase_order(po_id)
            self._refresh_po_table()

    # ── Mal Girişi Düzenle / Sil ─────────────────────────────────
    def _selected_receipt_id(self):
        row = self.po_receipts_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir mal girişi satırı seçin.")
            return None
        cell = self.po_receipts_table.item(row, 0)
        return cell.data(Qt.ItemDataRole.UserRole) if cell else None

    def _edit_po_receipt(self):
        role = CURRENT_USER.get("role","")
        if role not in ("admin", "planlama"):
            QMessageBox.warning(self, "Yetersiz Yetki",
                "Mal girişi düzenleme yalnızca planlamacı veya admin tarafından yapılabilir.")
            return
        receipt_id = self._selected_receipt_id()
        if not receipt_id:
            return
        row = self.po_receipts_table.currentRow()
        def _txt(c): return self.po_receipts_table.item(row, c).text() \
                     if self.po_receipts_table.item(row, c) else ""

        dlg = QDialog(self); dlg.setWindowTitle("Mal Girişi Düzenle"); dlg.setMinimumWidth(420)
        form = QFormLayout(dlg)
        metre_sb = QDoubleSpinBox(); metre_sb.setRange(0, 9_999_999); metre_sb.setDecimals(2)
        metre_sb.setSuffix(" mt")
        try: metre_sb.setValue(float(_txt(4)))
        except ValueError: pass
        kilo_sb = QDoubleSpinBox(); kilo_sb.setRange(0, 9_999_999); kilo_sb.setDecimals(2)
        kilo_sb.setSuffix(" kg")
        try: kilo_sb.setValue(float(_txt(5)))
        except ValueError: pass
        lot_e = QLineEdit(_txt(3))
        loc_cb = QComboBox()
        loc_cb.addItem("— Lokasyon Seçiniz —", ("", ""))
        for loc in db.get_active_locations():
            grp = loc.get("group_name","")
            loc_cb.addItem(f"[{grp}]  {loc['name']}", (loc["name"], grp))
        _make_searchable(loc_cb)
        cur_loc = _txt(6).split(" [")[0]  # strip " [GRP]"
        for i in range(loc_cb.count()):
            d = loc_cb.itemData(i)
            if isinstance(d, tuple) and d[0] == cur_loc:
                loc_cb.setCurrentIndex(i); break
        form.addRow("Metre:", metre_sb); form.addRow("Kilo:", kilo_sb)
        form.addRow("Lot No:", lot_e);   form.addRow("Lokasyon:", loc_cb)
        btns2 = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns2.accepted.connect(dlg.accept); btns2.rejected.connect(dlg.reject)
        form.addRow(btns2)
        if dlg.exec():
            loc_data = loc_cb.currentData()
            loc_name, grp = loc_data if isinstance(loc_data, tuple) else ("", "")
            if not loc_name:
                QMessageBox.warning(self, "Eksik", "Lokasyon seçilmelidir."); return
            try:
                db.update_boyahane_receipt(
                    receipt_id, meter=metre_sb.value(), kg=kilo_sb.value(),
                    lot=lot_e.text().strip(), location=loc_name, location_group=grp,
                    user_name=CURRENT_USER["full_name"])
                po_id = self._selected_po_id_silent()
                self._refresh_po_table()
                if po_id: self._refresh_receipts_table(po_id)
                if self._current_order_id: self._show_detail(self._current_order_id)
                QMessageBox.information(self, "Güncellendi",
                    "Mal girişi ve bağlı stok kaydı güncellendi.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Güncellenemedi:\n{e}")

    def _delete_po_receipt(self):
        role = CURRENT_USER.get("role","")
        if role not in ("admin", "planlama"):
            QMessageBox.warning(self, "Yetersiz Yetki",
                "Mal girişi silme yalnızca planlamacı veya admin tarafından yapılabilir.")
            return
        receipt_id = self._selected_receipt_id()
        if not receipt_id:
            return
        row = self.po_receipts_table.currentRow()
        pc = self.po_receipts_table.item(row, 1).text() \
             if self.po_receipts_table.item(row, 1) else ""
        reply = QMessageBox.question(
            self, "Mal Girişini Sil",
            f"<b>{pc}</b> mal giriş kaydı silinsin mi?<br><br>"
            f"<span style='color:#C62828'>Bağlı stok kaydı silinir ve "
            f"sipariş miktarları güncellenir.</span>",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                po_id = self._selected_po_id_silent()
                db.delete_boyahane_receipt(receipt_id, user_name=CURRENT_USER["full_name"])
                self._refresh_po_table()
                if po_id: self._refresh_receipts_table(po_id)
                if self._current_order_id: self._show_detail(self._current_order_id)
                QMessageBox.information(self, "Silindi",
                    "Mal girişi ve bağlı stok kaydı silindi.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Silinemedi:\n{e}")


BOYAHANE_STATUS_OPTIONS = ["BEKLEMEDE", "İŞLEMDE", "TAMAMLANDI", "İPTAL"]


class BoyahanePlanningView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        top.addWidget(QLabel("<b>🧶 Boyahane Planlama</b>"))
        top.addStretch()
        # Durum filtresi
        self._status_filter = QComboBox()
        self._status_filter.addItem("Tüm Durumlar", "")
        for s in BOYAHANE_STATUS_OPTIONS:
            self._status_filter.addItem(s, s)
        self._status_filter.currentIndexChanged.connect(self.refresh)
        top.addWidget(QLabel("Durum:"))
        top.addWidget(self._status_filter)
        btn_ref = QPushButton("🔄 Yenile"); btn_ref.clicked.connect(self.refresh)
        top.addWidget(btn_ref)
        lay.addLayout(top)

        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "Tarih", "PO No", "Sipariş No", "Ürün Kodu", "Kumaş Tipi",
            "Metre", "Kilo", "Birim Fiyat", "Toplam", "Lokasyon",
            "Tedarikçi", "Durum"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        hdr = self.table.horizontalHeader()
        hdr.setSectionsMovable(True)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        for i in (0, 1, 2, 4, 5, 6, 7, 8, 9, 10, 11):
            hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        lay.addWidget(self.table)

        btn_row = QHBoxLayout()
        b_status = QPushButton("📝 Durum Değiştir")
        b_status.clicked.connect(self._change_status)
        b_edit = QPushButton("✏ Düzenle")
        b_edit.clicked.connect(self._edit_receipt)
        b_del = QPushButton("🗑 Sil")
        b_del.clicked.connect(self._delete_receipt)
        b_del.setStyleSheet("color:#C62828;")
        btn_row.addWidget(b_status)
        btn_row.addWidget(b_edit)
        btn_row.addWidget(b_del)
        btn_row.addStretch()

        # Özet
        self._summary_lbl = QLabel()
        self._summary_lbl.setStyleSheet("color:#1565C0;font-weight:bold;")
        btn_row.addWidget(self._summary_lbl)
        lay.addLayout(btn_row)

    def refresh(self):
        from datetime import datetime
        status_f = self._status_filter.currentData() or ""
        rows = db.get_boyahane_queue(status_filter=status_f)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        total_m = 0.0; total_amt = 0.0
        for r in rows:
            row = self.table.rowCount()
            self.table.insertRow(row)
            dt = r.get("received_at","") or ""
            try:
                dt = datetime.strptime(dt[:16], "%Y-%m-%d %H:%M").strftime("%d.%m.%Y %H:%M")
            except Exception:
                pass
            m = float(r.get("meter",0) or 0)
            k = float(r.get("kg",0) or 0)
            up = float(r.get("unit_price",0) or 0)
            amt = m * up
            total_m += m; total_amt += amt
            st = r.get("status","")
            color = {"BEKLEMEDE":"#E65100","İŞLEMDE":"#1565C0",
                     "TAMAMLANDI":"#2E7D32","İPTAL":"#757575"}.get(st)

            for col, val in enumerate([
                dt, r.get("po_no",""), r.get("order_no",""),
                r.get("product_code",""), r.get("fabric_type",""),
                f"{m:.2f}", f"{k:.2f}",
                f"{up:,.4f}", f"{amt:,.2f}",
                r.get("location",""), r.get("supplier_name",""), st,
            ]):
                cell = QTableWidgetItem(str(val))
                cell.setData(Qt.ItemDataRole.UserRole, r.get("id"))
                cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if col == 11 and color:
                    cell.setForeground(QBrush(QColor(color)))
                self.table.setItem(row, col, cell)
        self.table.setSortingEnabled(True)
        self._summary_lbl.setText(
            f"Toplam: {total_m:,.2f} mt  |  Tutar: {total_amt:,.2f}")

    def _change_status(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir satır seçin.")
            return
        cell = self.table.item(row, 0)
        receipt_id = cell.data(Qt.ItemDataRole.UserRole) if cell else None
        if not receipt_id:
            return
        cur_status = self.table.item(row, 11).text() if self.table.item(row, 11) else ""
        dlg = QDialog(self); dlg.setWindowTitle("Durum Değiştir")
        form = QFormLayout(dlg)
        combo = QComboBox()
        for s in BOYAHANE_STATUS_OPTIONS:
            combo.addItem(s)
        idx = combo.findText(cur_status)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        form.addRow("Yeni Durum:", combo)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        form.addRow(btns)
        if dlg.exec():
            db.update_boyahane_receipt_status(receipt_id, combo.currentText())
            self.refresh()

    def _selected_receipt_id(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Önce bir satır seçin.")
            return None
        cell = self.table.item(row, 0)
        return cell.data(Qt.ItemDataRole.UserRole) if cell else None

    def _edit_receipt(self):
        role = CURRENT_USER.get("role","")
        if role not in ("admin", "planlama"):
            QMessageBox.warning(self, "Yetersiz Yetki",
                "Mal girişi düzenleme yalnızca planlamacı veya admin tarafından yapılabilir.")
            return
        receipt_id = self._selected_receipt_id()
        if not receipt_id:
            return

        # Seçili satırdan mevcut verileri oku
        row = self.table.currentRow()
        def _txt(col): return self.table.item(row, col).text() if self.table.item(row, col) else ""

        dlg = QDialog(self)
        dlg.setWindowTitle("Mal Girişi Düzenle")
        dlg.setMinimumWidth(420)
        form = QFormLayout(dlg)

        metre_sb = QDoubleSpinBox(); metre_sb.setRange(0, 9_999_999); metre_sb.setDecimals(2)
        metre_sb.setSuffix(" mt")
        try: metre_sb.setValue(float(_txt(5)))
        except ValueError: pass

        kilo_sb = QDoubleSpinBox(); kilo_sb.setRange(0, 9_999_999); kilo_sb.setDecimals(2)
        kilo_sb.setSuffix(" kg")
        try: kilo_sb.setValue(float(_txt(6)))
        except ValueError: pass

        lot_e = QLineEdit(_txt(3) if self.table.columnCount() > 3 else "")

        loc_cb = QComboBox()
        loc_cb.addItem("— Lokasyon Seçiniz —", ("", ""))
        for loc in db.get_active_locations():
            grp = loc.get("group_name","")
            loc_cb.addItem(f"[{grp}]  {loc['name']}", (loc["name"], grp))
        _make_searchable(loc_cb)
        cur_loc = _txt(9)
        for i in range(loc_cb.count()):
            d = loc_cb.itemData(i)
            if isinstance(d, tuple) and d[0] == cur_loc:
                loc_cb.setCurrentIndex(i); break

        form.addRow("Metre:", metre_sb)
        form.addRow("Kilo:", kilo_sb)
        form.addRow("Lot No:", lot_e)
        form.addRow("Lokasyon:", loc_cb)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept); btns.rejected.connect(dlg.reject)
        form.addRow(btns)

        if dlg.exec():
            loc_data = loc_cb.currentData()
            loc_name, grp = loc_data if isinstance(loc_data, tuple) else ("", "")
            if not loc_name:
                QMessageBox.warning(self, "Eksik", "Lokasyon seçilmelidir.")
                return
            try:
                db.update_boyahane_receipt(
                    receipt_id,
                    meter=metre_sb.value(),
                    kg=kilo_sb.value(),
                    lot=lot_e.text().strip(),
                    location=loc_name,
                    location_group=grp,
                    user_name=CURRENT_USER["full_name"],
                )
                self.refresh()
                QMessageBox.information(self, "Güncellendi",
                    "Mal girişi ve bağlı stok kaydı güncellendi.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Güncellenemedi:\n{e}")

    def _delete_receipt(self):
        role = CURRENT_USER.get("role","")
        if role not in ("admin", "planlama"):
            QMessageBox.warning(self, "Yetersiz Yetki",
                "Mal girişi silme yalnızca planlamacı veya admin tarafından yapılabilir.")
            return
        receipt_id = self._selected_receipt_id()
        if not receipt_id:
            return
        row = self.table.currentRow()
        po_no   = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        product = self.table.item(row, 3).text() if self.table.item(row, 3) else ""
        reply = QMessageBox.question(
            self, "Mal Girişini Sil",
            f"<b>{po_no} / {product}</b> kaydı silinsin mi?<br><br>"
            f"<span style='color:#C62828'>Bağlı stok kaydı da silinir ve "
            f"sipariş miktarları güncellenir.</span>",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            try:
                db.delete_boyahane_receipt(receipt_id, user_name=CURRENT_USER["full_name"])
                self.refresh()
                QMessageBox.information(self, "Silindi",
                    "Mal girişi ve bağlı stok kaydı silindi, sipariş miktarları güncellendi.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Silinemedi:\n{e}")


class SevkiyatView(QWidget):
    """Depo-Sevkiyat personeli sipariş sevklerini buradan girer."""

    SHIP_STATUSES = ["TÜM KUMAŞLAR BOYAHANEDE", "MÜŞTERİYE SEVKLER BAŞLADI", "PLANLANDI", "BOYAHANAYA SEVKLER BAŞLADI"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._current_order_id = None
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        top = QHBoxLayout()
        top.addWidget(QLabel("<b>🚚 Sevkiyat Modülü</b>"))
        top.addStretch()
        btn_ref = QPushButton("🔄 Yenile"); btn_ref.clicked.connect(self.refresh)
        top.addWidget(btn_ref)
        lay.addLayout(top)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        lay.addWidget(splitter)

        # Sol: Sevk edilecek siparişler
        left = QWidget()
        ll = QVBoxLayout(left); ll.setContentsMargins(0, 0, 4, 0)
        ll.addWidget(QLabel("<b>Sevk Edilecek Siparişler</b>"))
        self.order_table = QTableWidget()
        self.order_table.setColumnCount(6)
        self.order_table.setHorizontalHeaderLabels(
            ["Sipariş No", "Müşteri", "Termin", "Sip.Mt", "Sevk.Mt", "Durum"])
        self.order_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.order_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.order_table.verticalHeader().setVisible(False)
        self.order_table.setSortingEnabled(True)
        ohdr = self.order_table.horizontalHeader()
        ohdr.setSectionsMovable(True)
        ohdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for i in (0, 2, 3, 4, 5):
            ohdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self.order_table.itemSelectionChanged.connect(self._on_order_select)
        ll.addWidget(self.order_table)
        splitter.addWidget(left)

        # Sağ: Sevk formu + geçmiş
        right = QWidget()
        rl = QVBoxLayout(right); rl.setContentsMargins(4, 0, 0, 0)

        self._order_info = QLabel()
        self._order_info.setStyleSheet("background:#E3F2FD;padding:6px;border-radius:4px;")
        self._order_info.setWordWrap(True)
        self._order_info.setVisible(False)
        rl.addWidget(self._order_info)

        rl.addWidget(QLabel("<b>Yeni Sevk Girişi</b>"))
        form = QFormLayout(); form.setSpacing(8)
        self._ship_product = QComboBox()
        _make_searchable(self._ship_product)
        self._ship_fabric_type = QComboBox()
        for t in ["HAM", "PFD", "BOYALI", "İPLİĞİ BOYALI", "BASKILI", "MAMÜL"]:
            self._ship_fabric_type.addItem(t, t)
        self._ship_color = QLineEdit()
        self._ship_lot = QLineEdit(); self._ship_lot.setPlaceholderText("Boş: otomatik")
        self._ship_meter = QDoubleSpinBox()
        self._ship_meter.setRange(0, 999999); self._ship_meter.setDecimals(2)
        self._ship_kg = QDoubleSpinBox()
        self._ship_kg.setRange(0, 999999); self._ship_kg.setDecimals(2)
        self._ship_notes = QLineEdit()
        form.addRow("Ürün Kodu:", self._ship_product)
        form.addRow("Kumaş Tipi:", self._ship_fabric_type)
        form.addRow("Renk:", self._ship_color)
        form.addRow("Lot:", self._ship_lot)
        form.addRow("Metre:", self._ship_meter)
        form.addRow("Kilo:", self._ship_kg)
        form.addRow("Notlar:", self._ship_notes)
        rl.addLayout(form)

        btn_add = QPushButton("+ Sevk Ekle")
        btn_add.setStyleSheet("background:#1565C0;color:white;font-weight:bold;"
                              "border-radius:4px;padding:6px 14px;")
        btn_add.clicked.connect(self._add_shipment)
        rl.addWidget(btn_add)

        rl.addWidget(QLabel("<b>Sevkiyat Geçmişi</b>"))
        self.ship_table = QTableWidget()
        self.ship_table.setColumnCount(7)
        self.ship_table.setHorizontalHeaderLabels(
            ["Tarih", "Ürün Kodu", "Kumaş Tipi", "Renk", "Lot", "Metre", "Kilo"])
        self.ship_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.ship_table.verticalHeader().setVisible(False)
        self.ship_table.setSortingEnabled(True)
        shdr = self.ship_table.horizontalHeader()
        shdr.setSectionsMovable(True)
        shdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for i in (0, 2, 3, 4, 5, 6):
            shdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        rl.addWidget(self.ship_table)
        rl.addStretch()
        splitter.addWidget(right)
        splitter.setSizes([350, 650])

    def refresh(self):
        from datetime import datetime
        orders = db.get_shippable_orders()
        self.order_table.blockSignals(True)
        self.order_table.setSortingEnabled(False)
        self.order_table.setRowCount(0)
        for r in orders:
            row = self.order_table.rowCount()
            self.order_table.insertRow(row)
            item0 = QTableWidgetItem(r.get("order_no",""))
            item0.setData(Qt.ItemDataRole.UserRole, r["id"])
            self.order_table.setItem(row, 0, item0)
            self.order_table.setItem(row, 1, QTableWidgetItem(r.get("customer_name","") or ""))
            td = r.get("delivery_date","") or ""
            try:
                td = datetime.strptime(td[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                pass
            self.order_table.setItem(row, 2, QTableWidgetItem(td))
            tm = float(r.get("total_meter",0) or 0)
            sm = float(r.get("shipped_meter",0) or 0)
            self.order_table.setItem(row, 3, QTableWidgetItem(f"{tm:.2f}"))
            sm_item = QTableWidgetItem(f"{sm:.2f}")
            sm_item.setForeground(QBrush(QColor("#2E7D32" if sm >= tm else "#C62828")))
            self.order_table.setItem(row, 4, sm_item)
            st = r.get("status","")
            st_item = QTableWidgetItem(st)
            st_item.setForeground(QBrush(QColor(
                "#2E7D32" if st == "SİPARİŞ TAMAMLANDI" else
                "#1565C0" if st == "MÜŞTERİYE SEVKLER BAŞLADI" else "#E65100")))
            self.order_table.setItem(row, 5, st_item)
            if r["id"] == self._current_order_id:
                self.order_table.selectRow(row)
        self.order_table.setSortingEnabled(True)
        self.order_table.blockSignals(False)

    def _on_order_select(self):
        row = self.order_table.currentRow()
        if row < 0:
            return
        item = self.order_table.item(row, 0)
        if not item:
            return
        oid = item.data(Qt.ItemDataRole.UserRole)
        self._current_order_id = oid
        order = db.get_order(oid)
        if order:
            codes = list(dict.fromkeys(
                it.get("product_code","") for it in (order.get("items") or [])
                if it.get("product_code")))
            self._order_info.setText(
                f"<b>{order.get('order_no','')} — {order.get('customer_name','')}</b>  "
                f"Ürünler: {', '.join(codes)}")
            self._order_info.setVisible(True)
            # Ürün kodu combosunu doldur
            self._ship_product.blockSignals(True)
            self._ship_product.clear()
            for it in (order.get("items") or []):
                pc = it.get("product_code","")
                pn = it.get("product_name","")
                label = f"{pc} — {pn}" if pn else pc
                self._ship_product.addItem(label, pc)
            self._ship_product.blockSignals(False)
        self._refresh_ship_table(oid)

    def _refresh_ship_table(self, order_id):
        from datetime import datetime
        shipments = db.get_order_shipments(order_id)
        self.ship_table.setSortingEnabled(False)
        self.ship_table.setRowCount(0)
        for s in shipments:
            row = self.ship_table.rowCount()
            self.ship_table.insertRow(row)
            dt = s.get("shipment_date","") or ""
            try:
                dt = datetime.strptime(dt[:16], "%Y-%m-%d %H:%M").strftime("%d.%m.%Y %H:%M")
            except Exception:
                pass
            for col, val in enumerate([
                dt, s.get("product_code",""), s.get("fabric_type",""),
                s.get("color",""), s.get("lot",""),
                f"{float(s.get('meter',0) or 0):.2f}",
                f"{float(s.get('kg',0) or 0):.2f}",
            ]):
                cell = QTableWidgetItem(str(val))
                cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.ship_table.setItem(row, col, cell)
        self.ship_table.setSortingEnabled(True)

    def _add_shipment(self):
        if not self._current_order_id:
            QMessageBox.information(self, "Bilgi", "Önce sol taraftan bir sipariş seçin.")
            return
        meter = self._ship_meter.value()
        kg = self._ship_kg.value()
        if meter <= 0 and kg <= 0:
            QMessageBox.warning(self, "Hata", "Metre veya kilo girilmelidir.")
            return
        import datetime as _dt
        lot = self._ship_lot.text().strip()
        if not lot:
            lot = f"SEV-{_dt.date.today().strftime('%Y%m%d')}"
        items = [{
            "product_code": self._ship_product.currentData() or self._ship_product.currentText(),
            "product_name": "",
            "fabric_type": self._ship_fabric_type.currentData() or "",
            "color": self._ship_color.text().strip(),
            "lot": lot,
            "meter": meter,
            "kg": kg,
            "notes": self._ship_notes.text().strip(),
        }]
        try:
            db.add_order_shipment(self._current_order_id, items, CURRENT_USER["full_name"])
            self._ship_meter.setValue(0)
            self._ship_kg.setValue(0)
            self._ship_lot.clear()
            self._ship_notes.clear()
            self.refresh()
            self._refresh_ship_table(self._current_order_id)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Sevk kaydedilemedi:\n{e}")


class LocationView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Left: location list
        left = QGroupBox("Lokasyonlar")
        left_layout = QVBoxLayout(left)
        self.loc_list = QListWidget()
        self.loc_list.currentItemChanged.connect(self._on_loc_change)
        left_layout.addWidget(self.loc_list)
        left.setMaximumWidth(180)
        splitter.addWidget(left)

        # Right: fabric table
        right = QGroupBox("Kumaşlar")
        right_layout = QVBoxLayout(right)
        self.table = QTableWidget()
        cols = ["Ürün Kodu", "Ürün Bilgisi", "Renk", "Metre", "Kilo", "Top/Adet", "Açıklama"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setMouseTracking(True)
        self.table.setSortingEnabled(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._context_menu)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(False)
        hdr.setMinimumSectionSize(50)
        hdr.setSectionsMovable(True)
        for i, w in enumerate([100, 130, 100, 75, 65, 80, 250]):
            self.table.setColumnWidth(i, w)
        right_layout.addWidget(self.table)

        # Alt toplam barı
        self._loc_bar = QFrame()
        self._loc_bar.setStyleSheet(
            "QFrame { background:#545454; border-radius:6px;"
            "border: 2px solid #3A3A3A; padding:0; }"
        )
        self._loc_bar.setFixedHeight(56)
        loc_bar_layout = QHBoxLayout(self._loc_bar)
        loc_bar_layout.setContentsMargins(20, 4, 20, 4)
        loc_bar_layout.setSpacing(0)

        def _lstat(label):
            w = QWidget(); w.setStyleSheet("background:transparent;")
            wl = QVBoxLayout(w); wl.setSpacing(2); wl.setContentsMargins(16,2,16,2)
            lbl = QLabel(label)
            lbl.setStyleSheet("color:rgba(255,255,255,.7); font-size:10px; background:transparent;")
            val = QLabel("—")
            val.setStyleSheet("color:white; font-size:15px; font-weight:bold; background:transparent;")
            wl.addWidget(lbl); wl.addWidget(val)
            return w, val

        def _lsep():
            f = QFrame(); f.setFrameShape(QFrame.Shape.VLine); f.setFixedWidth(1)
            f.setStyleSheet("background:rgba(255,255,255,.3); margin:8px 4px;")
            return f

        bw1, self._lbl_loc   = _lstat("Lokasyon")
        bw2, self._lbl_items = _lstat("Kalem Sayısı")
        bw3, self._lbl_meter = _lstat("Toplam Metre")
        bw4, self._lbl_kg    = _lstat("Toplam Kilo")
        bw5, self._lbl_val   = _lstat("Toplam Değer")

        for w in (bw1, _lsep(), bw2, _lsep(), bw3, _lsep(), bw4, _lsep(), bw5):
            loc_bar_layout.addWidget(w)
        loc_bar_layout.addStretch()

        right_layout.addWidget(self._loc_bar)
        splitter.addWidget(right)
        splitter.setSizes([160, 640])
        layout.addWidget(splitter)

    def refresh(self):
        current = self.loc_list.currentItem()
        current_data = current.data(Qt.ItemDataRole.UserRole) if current else None

        # Listeyi sadece lokasyon sayısı değişmişse yeniden oluştur
        new_locs = db.get_active_locations()
        new_sig  = tuple(l["name"] for l in new_locs)
        if hasattr(self, "_loc_sig") and self._loc_sig == new_sig and self.loc_list.count() > 0:
            return   # değişmemiş, listeyi yeniden çizme
        self._loc_sig = new_sig

        self.loc_list.clear()

        all_locs = db.get_active_locations()
        from collections import defaultdict
        groups = defaultdict(list)
        for l in all_locs:
            groups[l["group_name"]].append(l["name"])

        for grp in sorted(groups.keys()):
            if grp == "DEPO":
                # DEPO → tek tıklanabilir satır, tüm rafları getirir
                item = QListWidgetItem("DEPO")
                item.setData(Qt.ItemDataRole.UserRole, "__GRP_DEPO__")
                self.loc_list.addItem(item)
            else:
                # Diğer grup → başlık yok, lokasyonlar kendi isimleriyle
                for name in sorted(groups[grp]):
                    item = QListWidgetItem(name)
                    item.setData(Qt.ItemDataRole.UserRole, name)
                    self.loc_list.addItem(item)

        # Önceki seçimi koru
        if current_data:
            for i in range(self.loc_list.count()):
                it = self.loc_list.item(i)
                if it.data(Qt.ItemDataRole.UserRole) == current_data:
                    self.loc_list.setCurrentItem(it)
                    break

    def _on_loc_change(self, item):
        if not item:
            return
        loc_data = item.data(Qt.ItemDataRole.UserRole)
        if not loc_data:
            return

        if loc_data == "__GRP_DEPO__":
            rows = db.get_all_fabrics()   # tek sorgu — tümünü çek, sonra filtrele
            depo_locs = {l["name"] for l in db.get_active_locations()
                         if l["group_name"] == "DEPO"}
            rows = [r for r in rows if (r["location"] or "") in depo_locs]
            label = "DEPO"
        else:
            rows = db.get_all_fabrics(location=loc_data)
            label = loc_data

        R_ALIGN = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight
        total_m = total_kg = total_val = 0.0

        self.table.setSortingEnabled(False)   # veri yüklenirken sıralamayı kapat
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            mt  = r["meter"] or 0
            kg  = r["kg"] or 0
            fiy = r["birim_fiyat"] or 0
            val = mt * fiy if mt > 0 else kg * fiy
            vals = [r["product_code"] or "", r["product_name"] or "", r["color"] or "",
                    f"{mt:,.2f}", f"{kg:,.2f}", r["piece_count"] or "", r["description"] or ""]
            for j, v in enumerate(vals):
                cell = QTableWidgetItem(v)
                cell.setToolTip(v)
                if j in (3, 4):
                    cell.setTextAlignment(R_ALIGN)
                self.table.setItem(i, j, cell)
            total_m += mt; total_kg += kg; total_val += val
        self.table.setUpdatesEnabled(True)
        self.table.setSortingEnabled(True)   # veri yüklendi, sıralamayı aç

        self._lbl_loc.setText(label)
        self._lbl_items.setText(f"{len(rows):,}")
        self._lbl_meter.setText(f"{total_m:,.2f} mt")
        self._lbl_kg.setText(f"{total_kg:,.2f} kg")
        self._lbl_val.setText(f"{total_val:,.0f} $" if total_val else "—")

    def _context_menu(self, pos):
        from PyQt6.QtWidgets import QMenu
        cur = self.loc_list.currentItem()
        label = cur.text().strip() if cur else "Lokasyon"
        menu = QMenu(self)
        act = menu.addAction(f"📥 Excel'e Aktar ({label})")
        if menu.exec(self.table.viewport().mapToGlobal(pos)) == act:
            _export_widget_table_to_excel(self, self.table, title=label)


class DashboardWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(16)
        header_row.setContentsMargins(12, 8, 12, 8)

        # Logo — yüksek çözünürlük, sola yasla
        logo_label = QLabel()
        if _os.path.exists(LOGO_PATH):
            pix = QPixmap(LOGO_PATH)
            logo_label.setPixmap(
                pix.scaledToHeight(100, Qt.TransformationMode.SmoothTransformation)
            )
        logo_label.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

        # Başlık
        title = QLabel("DEPO TAKİP SİSTEMİ")
        title.setStyleSheet("font-size:20px; font-weight:bold; color:#545454; letter-spacing:1px;")
        title.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

        header_row.addWidget(logo_label)
        header_row.addWidget(title)
        header_row.addStretch()

        btn_users = QPushButton("👤 Kullanıcı Yönetimi")
        btn_users.setStyleSheet(
            "background:#37474F; color:white; font-weight:bold; border-radius:4px; padding:8px 16px;"
        )
        btn_users.clicked.connect(self._open_user_mgmt)
        header_row.addWidget(btn_users)

        header_frame = QFrame()
        header_frame.setLayout(header_row)
        header_frame.setStyleSheet("background:white; border-radius:6px;")
        header_frame.setFixedHeight(116)
        layout.addWidget(header_frame)

        def _make_card(icon, label, color, small=False):
            box = QGroupBox()
            bl  = QVBoxLayout(box); bl.setSpacing(2)
            li  = QLabel(icon); li.setAlignment(Qt.AlignmentFlag.AlignCenter)
            li.setStyleSheet("font-size:18px;")
            val = QLabel("—"); val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            fs  = "14px" if small else "18px"
            val.setStyleSheet(f"color:{color}; font-size:{fs}; font-weight:bold;")
            lbl = QLabel(label); lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("color:#757575; font-size:10px;")
            bl.addWidget(li); bl.addWidget(val); bl.addWidget(lbl)
            box.setMinimumHeight(90)
            return box, val

        # Satır 1 — Genel özet
        self.stat_items = QLabel("0")
        self.stat_meter = QLabel("0 mt")
        self.stat_kg    = QLabel("0 kg")
        self.stat_value = QLabel("0 $")
        row1 = QHBoxLayout()
        for icon, lbl, color, wgt in [
            ("📦","Toplam Kalem","#545454", self.stat_items),
            ("📏","Toplam Metre","#2E7D32", self.stat_meter),
            ("⚖️","Toplam Kilo","#6A1B9A",  self.stat_kg),
            ("💰","Stok Değeri","#B71C1C",   self.stat_value),
        ]:
            box, val = _make_card(icon, lbl, color)
            val.setText = wgt.setText  # yönlendir — kullanmayacağız, ayrı set edeceğiz
            # Ortak widget referansı için kutuya ekleyelim
            # val zaten dışarıda tanımlı, sadece stili ata
            val.setStyleSheet(f"color:{color}; font-size:18px; font-weight:bold;")
            # Box içindeki ikinci widget val — bunu değiştir
            box.layout().itemAt(1).widget().deleteLater()
            box.layout().insertWidget(1, wgt)
            wgt.setAlignment(Qt.AlignmentFlag.AlignCenter)
            wgt.setStyleSheet(f"color:{color}; font-size:18px; font-weight:bold;")
            row1.addWidget(box)
        layout.addLayout(row1)

        # Satır 2 — Tip dağılımı
        self.stat_ham    = QLabel("—")
        self.stat_pfd    = QLabel("—")
        self.stat_boyali = QLabel("—")
        self.stat_iplikboyali = QLabel("—")
        self.stat_baskili= QLabel("—")
        row2 = QHBoxLayout()
        for icon, lbl, color, wgt in [
            ("🟫","HAM (Metre)","#5D4037",   self.stat_ham),
            ("🟩","PFD (Metre)","#00695C",   self.stat_pfd),
            ("🟦","BOYALI (Metre)","#1565C0", self.stat_boyali),
            ("🟧","İPLİĞİ BOYALI (Metre)","#EF6C00", self.stat_iplikboyali),
            ("🟪","BASKILI (Metre)","#6A1B9A",self.stat_baskili),
        ]:
            box2 = QGroupBox(); bl2 = QVBoxLayout(box2); bl2.setSpacing(2)
            li2  = QLabel(icon); li2.setAlignment(Qt.AlignmentFlag.AlignCenter)
            li2.setStyleSheet("font-size:16px;")
            wgt.setAlignment(Qt.AlignmentFlag.AlignCenter)
            wgt.setStyleSheet(f"color:{color}; font-size:16px; font-weight:bold;")
            lbl2 = QLabel(lbl); lbl2.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl2.setStyleSheet("color:#757575; font-size:10px;")
            bl2.addWidget(li2); bl2.addWidget(wgt); bl2.addWidget(lbl2)
            box2.setMinimumHeight(80)
            row2.addWidget(box2)
        layout.addLayout(row2)

        # Location summary table
        loc_group = QGroupBox("Depo / Lokasyon Özeti")
        loc_layout = QVBoxLayout(loc_group)
        self.loc_table = QTableWidget()
        self.loc_table.setColumnCount(5)
        self.loc_table.setHorizontalHeaderLabels(["Depo / Lokasyon", "Kalem", "Toplam Metre", "Toplam Kilo", "Toplam Değer $"])
        self.loc_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.loc_table.verticalHeader().setVisible(False)
        hdr = self.loc_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        loc_layout.addWidget(self.loc_table)
        layout.addWidget(loc_group)

        # ── Haftalık Stok Trendi ─────────────────────────────────
        trend_group = QGroupBox("📈 Haftalık Stok Trendi (devir oranı, miktar, tutar)")
        tl = QVBoxLayout(trend_group)
        self.trend_table = QTableWidget()
        self.trend_table.setColumnCount(7)
        self.trend_table.setHorizontalHeaderLabels(
            ["Hafta", "Stok Miktarı (mt)", "Δ%", "Stok Değeri ($)", "Δ%",
             "Haftalık Çıkış (mt)", "Devir Oranı (%)"])
        self.trend_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.trend_table.verticalHeader().setVisible(False)
        thdr = self.trend_table.horizontalHeader()
        thdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for i in range(1, 7):
            thdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self.trend_table.setMinimumHeight(180)
        tl.addWidget(self.trend_table)
        layout.addWidget(trend_group)

    def _open_user_mgmt(self):
        if CURRENT_USER.get("role") != "admin":
            QMessageBox.warning(self, "Yetki", "Bu işlem için yönetici yetkisi gereklidir.")
            return
        UserManagementDialog(self).exec()

    DEPO_LOCS = {"DEPO", "M11", "M7", "OFİS", "OFIS"}

    @classmethod
    def _loc_group(cls, loc):
        import re
        u = loc.strip().upper()
        if u in cls.DEPO_LOCS or re.match(r"^(RAF|PALET|P\d|H\d|HP|H-P)", u):
            return "DEPO"
        return u

    def refresh(self):
        import re
        summary = db.get_summary()
        self.stat_items.setText(str(summary["total_items"] or 0))
        self.stat_meter.setText(f"{summary['total_meter'] or 0:,.0f} mt")
        self.stat_kg.setText(f"{summary['total_kg'] or 0:,.0f} kg")
        val = summary["total_value"] or 0
        priced = summary["priced_items"] or 0
        total  = summary["total_items"] or 1
        self.stat_value.setText(f"{val:,.0f} $")
        if priced < total:
            self.stat_value.setToolTip(f"Not: {total - priced} kalemin fiyatı girilmemiş")

        # ── Kumaş Tipi Özet Kartları ─────────────────────────────
        all_rows = db.get_all_fabrics()
        tip_mt = {"HAM": 0.0, "PFD": 0.0, "BOYALI": 0.0, "İPLİĞİ BOYALI": 0.0, "BASKILI": 0.0}
        for r in all_rows:
            t = r["fabric_type"] or ""
            if t in tip_mt:
                tip_mt[t] += r["meter"] or 0
        self.stat_ham.setText(f"{tip_mt['HAM']:,.0f} mt")
        self.stat_pfd.setText(f"{tip_mt['PFD']:,.0f} mt")
        self.stat_boyali.setText(f"{tip_mt['BOYALI']:,.0f} mt")
        self.stat_iplikboyali.setText(f"{tip_mt['İPLİĞİ BOYALI']:,.0f} mt")
        self.stat_baskili.setText(f"{tip_mt['BASKILI']:,.0f} mt")

        locs = db.get_locations()

        # Locations tablosundan group_name → loc_name eşlemesi
        all_loc_records = db.get_all_locations() or []
        loc_to_group = {l["name"]: (l["group_name"] or "DEPO") for l in all_loc_records}

        # DEPO → tek satırda topla; DIŞ DEPO → her biri ayrı satır
        from collections import defaultdict
        depo_agg = [0, 0.0, 0.0, 0.0]   # count, meter, kg, val
        dis_groups = defaultdict(lambda: [0, 0.0, 0.0, 0.0])

        for loc in locs:
            rows = db.get_all_fabrics(location=loc)
            total_m   = sum(r["meter"] or 0 for r in rows)
            total_kg  = sum(r["kg"] or 0 for r in rows)
            total_val = sum(
                ((r["meter"] or 0) * (r["birim_fiyat"] or 0)) if (r["meter"] or 0) > 0
                else ((r["kg"] or 0) * (r["birim_fiyat"] or 0))
                for r in rows
            )
            grp = loc_to_group.get(loc, "DEPO")
            if grp == "DEPO":
                depo_agg[0] += len(rows)
                depo_agg[1] += total_m
                depo_agg[2] += total_kg
                depo_agg[3] += total_val
            else:
                dis_groups[loc][0] += len(rows)
                dis_groups[loc][1] += total_m
                dis_groups[loc][2] += total_kg
                dis_groups[loc][3] += total_val

        table_rows = []
        if any(depo_agg):
            table_rows.append(("DEPO", *depo_agg))
        for loc_name in sorted(dis_groups.keys()):
            v = dis_groups[loc_name]
            table_rows.append((loc_name, *v))

        self.loc_table.setRowCount(len(table_rows))

        COLORS = {"DEPO": "#37474F"}
        for i, (group, count, meter, kg, val) in enumerate(table_rows):
            bg = QColor(COLORS.get(group, "#546E7A"))
            fg = QColor("#FFFFFF")
            val_str = f"{val:,.0f} $" if val else "—"
            vals = [group, str(count), f"{meter:,.2f} mt", f"{kg:,.2f} kg", val_str]

            for j, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setBackground(QBrush(bg))
                item.setForeground(QBrush(fg))
                item.setFont(QFont("", -1, QFont.Weight.Bold))
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter |
                    (Qt.AlignmentFlag.AlignRight if j >= 1 else Qt.AlignmentFlag.AlignLeft))
                self.loc_table.setItem(i, j, item)

        self._refresh_trend()

    def _refresh_trend(self):
        """Haftalık stok trendini doldurur; içinde bulunulan haftayı kaydeder."""
        try:
            db.capture_stock_snapshot()
            snaps = db.get_stock_snapshots(limit=12) or []
        except Exception:
            snaps = []
        t = self.trend_table
        t.setRowCount(len(snaps))

        def _pct(cur, prev):
            if prev is None or prev == 0:
                return ("", None)
            d = (cur - prev) / prev * 100
            if abs(d) < 0.5:
                return ("◦ %0", "#9E9E9E")
            return (f"{'▲' if d>0 else '▼'} %{abs(d):.0f}", "#2E7D32" if d > 0 else "#C62828")

        def _cell(text, color=None, align_right=True):
            it = QTableWidgetItem(text)
            it.setTextAlignment(Qt.AlignmentFlag.AlignVCenter |
                (Qt.AlignmentFlag.AlignRight if align_right else Qt.AlignmentFlag.AlignLeft))
            if color:
                it.setForeground(QBrush(QColor(color)))
            return it

        for i, s in enumerate(snaps):
            prev = snaps[i - 1] if i > 0 else None
            ws = s["week_start"] or ""
            try:
                wk = ".".join(reversed(ws.split("-")))   # 2026-06-22 -> 22.06.2026
            except Exception:
                wk = ws
            meter = s["total_meter"] or 0
            val = s["total_value"] or 0
            cikis = s["hafta_cikis_mt"] or 0
            # devir oranı = haftalık çıkış / ortalama stok (önceki+bu)/2
            if prev is not None and ((prev["total_meter"] or 0) + meter) > 0:
                ort = ((prev["total_meter"] or 0) + meter) / 2
            else:
                ort = meter
            devir = (cikis / ort) if ort else 0

            dm_txt, dm_col = _pct(meter, prev["total_meter"] if prev else None)
            dv_txt, dv_col = _pct(val, prev["total_value"] if prev else None)
            t.setItem(i, 0, _cell(f"{wk}", align_right=False))
            t.setItem(i, 1, _cell(f"{meter:,.0f}"))
            t.setItem(i, 2, _cell(dm_txt, dm_col))
            t.setItem(i, 3, _cell(f"{val:,.0f}"))
            t.setItem(i, 4, _cell(dv_txt, dv_col))
            t.setItem(i, 5, _cell(f"{cikis:,.0f}"))
            t.setItem(i, 6, _cell(f"%{devir*100:,.0f}"))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Bursa Knitted Depo Takip Sistemi")
        self.setMinimumSize(1100, 700)
        if _os.path.exists(LOGO_PATH):
            self.setWindowIcon(QIcon(LOGO_PATH))
        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(
            "QTabWidget::pane { border-top: 1px solid #CCCCCC; margin-top: 0px; }"
            "QTabWidget::tab-bar { alignment: left; }"
            "QTabBar { qproperty-drawBase: 0; }"
        )
        self.dashboard = DashboardWidget()
        self.stock_table = StockTable(self)
        self.location_view = LocationView()
        self.fire_view = FireView()
        self.orders_view = OrdersView(self)
        self.planning_view = PlanningView(self)
        self.boyahane_view = BoyahanePlanningView(self)
        self.sevkiyat_view = SevkiyatView(self)
        self.crm_view = CRMView(self)

        self._rebuild_tabs()
        self.tabs.currentChanged.connect(self._on_tab_change)
        layout.addWidget(self.tabs)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self._user_label = QLabel()
        self._user_label.setStyleSheet("color:#212121; font-weight:bold; font-size:12px; padding:0 12px;")
        self.status.addPermanentWidget(self._user_label)
        self._update_user_label()

        # Menu
        menubar = self.menuBar()
        file_menu = menubar.addMenu("Dosya")
        file_menu.addAction("Excel'den İçe Aktar...").triggered.connect(self._import)
        file_menu.addAction("Excel'e Dışa Aktar...").triggered.connect(self._export)
        file_menu.addSeparator()
        file_menu.addAction("Çıkış").triggered.connect(self.close)

        sys_menu = menubar.addMenu("🛡 Sistem")
        sys_menu.addAction("Yedek Durumu ve Yönetimi...").triggered.connect(self._backup_dialog)
        sys_menu.addAction("Şimdi Yedek Al").triggered.connect(self._backup_now)

        cust_menu = menubar.addMenu("👥 Müşteriler")
        cust_menu.addAction("Müşteri Listesi...").triggered.connect(
            lambda: CustomerManagementDialog(self).exec())

        sup_menu = menubar.addMenu("🏭 Tedarikçiler")
        sup_menu.addAction("Tedarikçi Listesi...").triggered.connect(
            lambda: SupplierManagementDialog(self).exec())

        prod_menu = menubar.addMenu("📦 Ürünler")
        prod_menu.addAction("Ürün Kataloğu...").triggered.connect(
            lambda: ProductManagementDialog(self).exec())

        iplik_menu = menubar.addMenu("🧵 İplik")
        iplik_menu.addAction("İplik Kataloğu...").triggered.connect(
            lambda: IplikManagementDialog(self).exec())

        loc_menu = menubar.addMenu("🗄 Lokasyonlar")
        loc_menu.addAction("Raf / Lokasyon Tanımlamaları...").triggered.connect(
            lambda: LocationManagementDialog(self).exec())

        order_menu = menubar.addMenu("📋 Siparişler")
        order_menu.addAction("Yeni Sipariş...").triggered.connect(self._new_order)
        if CURRENT_USER.get("role") == "admin":
            order_menu.addSeparator()
            order_menu.addAction("Şirket / Banka Ayarları...").triggered.connect(
                lambda: CompanySettingsDialog(self).exec())

        user_menu = menubar.addMenu("👤 Kullanıcılar")
        user_menu.addAction("Kullanıcı Yönetimi").triggered.connect(self._user_mgmt)
        user_menu.addSeparator()
        user_menu.addAction("Oturumu Kapat").triggered.connect(self._logout)

        mail_menu = menubar.addMenu("📧 E-posta Rapor")
        mail_menu.addAction("Ayarlar ve Zamanlama...").triggered.connect(
            lambda: EmailSettingsDialog(self).exec())
        mail_menu.addAction("Şimdi Rapor Gönder").triggered.connect(self._send_now)

        web_menu = menubar.addMenu("📱 Mobil Erişim")
        self._web_action = web_menu.addAction("WiFi Sunucusunu Başlat (yerel)")
        self._web_action.triggered.connect(self._toggle_web)
        web_menu.addSeparator()
        self._ngrok_action = web_menu.addAction("İnternetten Erişim Aç (ngrok)")
        self._ngrok_action.triggered.connect(self._toggle_ngrok)
        web_menu.addSeparator()
        web_menu.addAction("ngrok Token Ayarla...").triggered.connect(self._set_ngrok_token)

        self._web_label = QLabel("  📱 Mobil: Kapalı  ")
        self._web_label.setStyleSheet("color:#212121; font-size:12px; padding:0 10px;")
        self.status.addPermanentWidget(self._web_label)

        self._backup_lbl = QLabel("  🛡 Yedek: —  ")
        self._backup_lbl.setStyleSheet("color:#212121; font-size:12px; padding:0 10px;")
        self._backup_lbl.setCursor(Qt.CursorShape.PointingHandCursor)
        self._backup_lbl.mousePressEvent = lambda e: self._backup_dialog()
        self.status.addPermanentWidget(self._backup_lbl)
        QTimer.singleShot(2000, self._refresh_backup_indicator)

    def _update_user_label(self):
        icons = {"admin": "👑", "planlama": "📌", "satışçı": "📋",
                 "depo-sevkiyat": "🚚"}
        role_icon = icons.get(CURRENT_USER.get("role", ""), "👤")
        self._user_label.setText(f"{role_icon} {CURRENT_USER.get('full_name', '')}")

    def _new_order(self):
        dlg = OrderDialog(self)
        if dlg.exec():
            data = dlg.get_data()
            oid, order_no = db.add_order(**data, created_by=CURRENT_USER["full_name"])
            self.orders_view.refresh()
            self.planning_view.refresh()
            QMessageBox.information(self, "Sipariş Oluşturuldu",
                f"Sipariş kaydedildi.\n\nSipariş No: {order_no}\nDurum: ONAYDA\n\nAdmin onayı bekleniyor.")

    def _user_mgmt(self):
        if CURRENT_USER.get("role") != "admin":
            QMessageBox.warning(self, "Yetki", "Bu işlem için yönetici yetkisi gereklidir.")
            return
        UserManagementDialog(self).exec()

    def _logout(self):
        reply = QMessageBox.question(self, "Çıkış", "Oturumu kapatıp yeniden giriş yapılsın mı?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            dlg = LoginDialog(self)
            dlg.setWindowModality(Qt.WindowModality.ApplicationModal)
            if dlg.exec():
                self._update_user_label()
                self._rebuild_tabs()
            else:
                self.close()

    def _rebuild_tabs(self):
        """Oturum değişince tabları role göre yeniden oluştur."""
        while self.tabs.count():
            self.tabs.removeTab(0)
        role = CURRENT_USER.get("role", "")
        _yeni_roller = ("muhasebe", "fason-takip", "kartela", "kalite-kontrol")
        if role == "admin":
            self.tabs.addTab(self.dashboard, "📊 Dashboard")
        if role in ("admin", "kullanici", "depo-sevkiyat", "planlama") or role in _yeni_roller:
            self.tabs.addTab(self.stock_table, "📦 Stok Listesi")
            self.tabs.addTab(self.location_view, "🗂 Lokasyon Görünümü")
        if role in ("admin", "kullanici") or role in _yeni_roller:
            self.tabs.addTab(self.fire_view, "🔥 Boyahane Fire Oranları")
        if role in ("admin", "satışçı", "kullanici") or role in _yeni_roller:
            self.tabs.addTab(self.orders_view, "📋 Siparişler")
        if role in ("admin", "satışçı"):
            self.tabs.addTab(self.crm_view, "👥 CRM")
        if role in ("admin", "planlama", "kullanici") or role in _yeni_roller:
            self.tabs.addTab(self.planning_view, "📌 Planlama")
            self.tabs.addTab(self.boyahane_view, "🧶 Boyahane Planlama")
        if role in ("admin", "depo-sevkiyat", "kullanici") or role in _yeni_roller:
            self.tabs.addTab(self.sevkiyat_view, "🚚 Sevkiyat")
        self.stock_table.refresh_with_locations()
        # Admin: onay bekleyen sipariş bildirimi
        if role == "admin":
            QTimer.singleShot(1500, self._check_pending_orders)

    def _toggle_web(self):
        import web_server as ws
        if ws.is_running():
            ws.stop()
            self._web_action.setText("Mobil Sunucuyu Başlat")
            self._web_label.setText("  📱 Kapalı  ")
            self._web_label.setStyleSheet("color:#212121; font-size:12px; padding:0 8px;")
        else:
            ip, port = ws.start()
            self._web_action.setText("Mobil Sunucuyu Durdur")
            self._web_label.setText(f"  📱 {ip}:{port}  ")
            self._web_label.setStyleSheet("color:#2E7D32; font-weight:bold; font-size:12px; padding:0 8px;")
            # Bilgi dialogu
            dlg = _MobileAccessDialog(self, ip, port)
            dlg.exec()

    def _toggle_ngrok(self):
        import web_server as ws
        if ws.ngrok_running():
            ws.stop_ngrok()
            self._ngrok_action.setText("İnternetten Erişim Aç (ngrok)")
            self._web_label.setText(
                f"  📱 {ws.get_local_ip()}:{ws.PORT}  " if ws.is_running() else "  📱 Kapalı  "
            )
        else:
            # Token kontrolü
            token = ws.get_ngrok_token()
            if not token:
                dlg = _NgrokSetupDialog(self)
                if dlg.exec() != QDialog.DialogCode.Accepted:
                    return
                token = dlg.token
                ws.set_ngrok_token(token)

            # Loading mesajı
            self.status.showMessage("  ngrok bağlantısı kuruluyor...")
            QApplication.processEvents()
            try:
                url = ws.start_ngrok()
                self._ngrok_action.setText("İnternetten Erişimi Kapat")
                self._web_label.setText(f"  🌍 İnternette Açık  ")
                self._web_label.setStyleSheet("color:#FFD54F; font-weight:bold; font-size:12px; padding:0 8px;")
                dlg = _NgrokActiveDialog(self, url)
                dlg.exec()
            except Exception as e:
                QMessageBox.critical(self, "ngrok Hatası",
                    f"Bağlantı kurulamadı:\n{e}\n\nToken'ı kontrol edin: Menü → Mobil Erişim → ngrok Token Ayarla")

    def _send_now(self):
        import email_report as er
        try:
            self.status.showMessage("  Rapor gönderiliyor...")
            QApplication.processEvents()
            n = er.send_report(test=True)
            self.status.showMessage(f"  ✓ Rapor {n} alıcıya gönderildi.")
        except Exception as e:
            QMessageBox.critical(self, "Gönderilemedi",
                f"{e}\n\nMenü → 📧 E-posta Rapor → Ayarlar'dan yapılandırın.")

    def _backup_now(self):
        import backup as bk
        self.status.showMessage("  Yedek alınıyor...")
        QApplication.processEvents()
        ok, msg = bk.take_backup(force=True)
        if ok:
            self.status.showMessage(f"  ✓ {msg}")
            self._refresh_backup_indicator()
        else:
            QMessageBox.critical(self, "Yedekleme Hatası", msg)

    def _backup_dialog(self):
        import backup as bk, glob, os
        s = bk.get_backup_status()
        dlg = QDialog(self)
        dlg.setWindowTitle("Yedek Yönetimi")
        dlg.setMinimumSize(580, 420)
        lay = QVBoxLayout(dlg)

        status_color = "#2E7D32" if s["is_today"] else "#C62828"
        status_text  = "✅ Bugün alındı" if s["is_today"] else "⚠️ Bugün alınmadı!"
        info = QLabel(
            f"<b>Yedek Durumu:</b> <span style='color:{status_color}'>{status_text}</span><br>"
            f"Son yedek: <b>{s['last_backup'] or 'Hiç alınmamış'}</b><br>"
            f"Kayıtlı yedek sayısı: <b>{s['backup_count']}</b>  |  "
            f"Toplam boyut: <b>{s['backup_size_mb']} MB</b><br>"
            f"Klasör: <span style='color:#555;font-size:11px'>{s['backup_dir']}</span>"
        )
        info.setWordWrap(True)
        info.setStyleSheet("background:#F5F5F5; padding:12px; border-radius:6px;")
        lay.addWidget(info)

        # Yedek listesi
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Dosya Adı", "Tarih", "Boyut"])
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        hdr = table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

        files = sorted(glob.glob(os.path.join(s["backup_dir"], "stok_*.db")), reverse=True)
        table.setRowCount(len(files))
        for i, f in enumerate(files):
            name = os.path.basename(f)
            size = f"{os.path.getsize(f)/1024/1024:.1f} MB"
            mtime = datetime.fromtimestamp(os.path.getmtime(f)).strftime("%d.%m.%Y %H:%M")
            table.setItem(i, 0, QTableWidgetItem(name))
            table.setItem(i, 1, QTableWidgetItem(mtime))
            table.setItem(i, 2, QTableWidgetItem(size))
        lay.addWidget(table)

        btn_row = QHBoxLayout()
        btn_now = QPushButton("🗂 Şimdi Yedek Al")
        btn_now.setStyleSheet("background:#2E7D32; color:white; font-weight:bold; border-radius:4px; padding:7px 14px;")

        btn_restore = QPushButton("↩ Seçili Yedeği Geri Yükle")
        btn_restore.setStyleSheet("background:#C62828; color:white; font-weight:bold; border-radius:4px; padding:7px 14px;")

        btn_close = QPushButton("Kapat")

        def _now():
            ok, msg = bk.take_backup(force=True)
            if ok:
                dlg.accept()
                self._backup_now()
            else:
                QMessageBox.critical(dlg, "Hata", msg)

        def _restore():
            row = table.currentRow()
            if row < 0:
                return QMessageBox.information(dlg, "Bilgi", "Geri yüklenecek yedeği seçin.")
            f = files[row]
            reply = QMessageBox.question(dlg, "Geri Yükle",
                f"<b>{os.path.basename(f)}</b> yedeğine dönülsün mü?<br><br>"
                f"<span style='color:#C62828'>Mevcut tüm veriler bu yedekle değiştirilir!</span>",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    safe = bk.restore_backup(f)
                    QMessageBox.information(dlg, "Başarılı",
                        f"Yedek geri yüklendi.\nEski veriler şuraya kaydedildi:\n{os.path.basename(safe)}\n\nProgram yeniden başlatılıyor...")
                    dlg.accept()
                    import subprocess, sys
                    subprocess.Popen([sys.executable] + sys.argv)
                    QApplication.quit()
                except Exception as e:
                    QMessageBox.critical(dlg, "Hata", str(e))

        btn_now.clicked.connect(_now)
        btn_restore.clicked.connect(_restore)
        btn_close.clicked.connect(dlg.accept)
        btn_row.addWidget(btn_now); btn_row.addWidget(btn_restore)
        btn_row.addStretch(); btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)
        dlg.exec()

    def _refresh_backup_indicator(self):
        import backup as bk
        s = bk.get_backup_status()
        if hasattr(self, "_backup_lbl"):
            if s["is_today"]:
                self._backup_lbl.setText(f"  🛡 Yedek: {s['last_backup']}  ")
                self._backup_lbl.setStyleSheet("color:#2E7D32; font-size:12px; padding:0 10px;")
            else:
                self._backup_lbl.setText("  ⚠️ Yedek alınmadı  ")
                self._backup_lbl.setStyleSheet("color:#E65100; font-size:11px; font-weight:bold; padding:0 6px;")

    def _set_ngrok_token(self):
        import web_server as ws
        current = ws.get_ngrok_token()
        dlg = _NgrokSetupDialog(self, current)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            ws.set_ngrok_token(dlg.token)
            QMessageBox.information(self, "Kaydedildi", "ngrok token kaydedildi.\nArtık 'İnternetten Erişim Aç' butonunu kullanabilirsiniz.")

    def _check_updates(self):
        """Arka planda güncelleme kontrolü — program açılınca çağrılır."""
        import updater
        def _on_result(var_mi, mesaj):
            if var_mi:
                # Ana thread'de bildirim göster
                QTimer.singleShot(0, lambda: self._show_update_notice(mesaj))
        updater.check_in_background(_on_result)

    def _show_update_notice(self, mesaj):
        reply = QMessageBox.question(self, "🔄 Güncelleme Mevcut",
            f"{mesaj}\n\nŞimdi güncellensin mi?\n(Program yeniden başlayacak)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            import updater
            ok, msg = updater.apply_update()
            if ok:
                QMessageBox.information(self, "Güncellendi", msg)
                updater.restart()
            else:
                QMessageBox.critical(self, "Güncelleme Hatası", msg)

    def _on_tab_change(self, idx):
        w = self.tabs.widget(idx)
        if w is not None and hasattr(w, "refresh"):
            w.refresh()
        # Admin planlama sekmesine geçince pending check
        if CURRENT_USER.get("role") == "admin" and w is self.planning_view:
            self._check_pending_orders()

    def _check_pending_orders(self):
        pending = db.get_pending_approval_orders()
        if not pending:
            return
        nos = ", ".join(p.get("order_no","") for p in pending[:5])
        extra = f" ve {len(pending)-5} daha..." if len(pending) > 5 else ""
        dlg = QDialog(self); dlg.setWindowTitle("⚠ Onay Bekleyen Siparişler")
        lay = QVBoxLayout(dlg)
        lbl = QLabel(
            f"<b>{len(pending)} sipariş admin onayı bekliyor!</b><br><br>"
            f"<span style='color:#B71C1C'>{nos}{extra}</span><br><br>"
            "Planlama ekranından siparişi seçip <b>✅ Seçili Siparişi Onayla</b> "
            "butonuna tıklayarak onaylayabilirsiniz.")
        lbl.setWordWrap(True); lbl.setTextFormat(Qt.TextFormat.RichText)
        lay.addWidget(lbl)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        btns.accepted.connect(dlg.accept)
        lay.addWidget(btns)
        dlg.exec()

    def update_status(self, count, summary):
        if not hasattr(self, "status"):
            return
        self.status.showMessage(
            f"  {count} kayıt gösteriliyor  |  Toplam: {summary['total_items'] or 0} kalem  |  "
            f"{summary['total_meter'] or 0:.2f} mt  |  {summary['total_kg'] or 0:.2f} kg"
        )

    def _import(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", "", "Excel Dosyaları (*.xlsx *.xls)"
        )
        if not path:
            return

        reply = QMessageBox.question(
            self, "İçe Aktarma",
            "Hangi sayfaları içe aktarmak istersiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Yes
        )

        from importer import import_from_excel, import_from_main_sheet
        try:
            if reply == QMessageBox.StandardButton.Yes:
                records = import_from_main_sheet(path)
                source = "STOK RAPORU sayfası"
            elif reply == QMessageBox.StandardButton.No:
                records = import_from_excel(path)
                source = "Raf/Lokasyon sayfaları"
            else:
                return

            if not records:
                QMessageBox.warning(self, "Uyarı", "Hiç kayıt bulunamadı!")
                return

            confirm = QMessageBox.question(
                self, "Onay",
                f"{source}'ndan <b>{len(records)}</b> kayıt bulundu.\nİçe aktarılsın mı?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if confirm == QMessageBox.StandardButton.Yes:
                db.import_fabrics_bulk(records)
                self.stock_table.refresh()
                self.location_view.refresh()
                self.dashboard.refresh()
                QMessageBox.information(self, "Başarılı", f"{len(records)} kayıt içe aktarıldı!")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"İçe aktarma hatası:\n{e}")

    def _export(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Dışa Aktar", "stok_raporu.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "STOK RAPORU"

            # Header
            headers = ["#", "Ürün Kodu", "Ürün Bilgisi", "Renk", "Lokasyon", "Metre", "Kilo", "Top/Adet", "Açıklama", "Son Güncelleme"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="1565C0")
                cell.alignment = Alignment(horizontal="center")

            rows = db.get_all_fabrics()
            for i, r in enumerate(rows, 1):
                ws.append([
                    i,
                    r["product_code"] or "",
                    r["product_name"] or "",
                    r["color"] or "",
                    r["location"] or "",
                    r["meter"] or 0,
                    r["kg"] or 0,
                    r["piece_count"] or "",
                    r["description"] or "",
                    str(r["updated_at"] or "")[:16],
                ])

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(path)
            QMessageBox.information(self, "Başarılı", f"Dışa aktarıldı:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dışa aktarma hatası:\n{e}")


def main():
    db.init_db()
    app = QApplication(sys.argv)
    app.setApplicationName("Bursa Knitted Depo Takip Sistemi")
    app.setStyle("Fusion")
    app.setStyleSheet(STYLESHEET)

    # Giriş ekranı
    login = LoginDialog()
    if login.exec() != QDialog.DialogCode.Accepted:
        sys.exit(0)

    # Günlük yedek al (arka planda)
    import threading, backup as bk
    def _do_backup():
        ok, msg = bk.take_backup()
        print(f"[Yedek] {msg}")
    threading.Thread(target=_do_backup, daemon=True).start()

    # Günlük mail zamanlayıcısını başlat
    import email_report as er
    er.start_scheduler()

    window = MainWindow()
    window.show()
    # 3 saniye sonra arka planda güncelleme kontrol et
    QTimer.singleShot(3000, window._check_updates)
    sys.exit(app.exec())


if __name__ == "__main__":
    import traceback, datetime
    try:
        main()
    except Exception:
        _log = os.path.join(os.path.dirname(os.path.abspath(__file__)), "error.log")
        with open(_log, "a", encoding="utf-8") as _f:
            _f.write(f"\n{'='*60}\n{datetime.datetime.now()}\n")
            _f.write(traceback.format_exc())
        try:
            from PyQt6.QtWidgets import QApplication, QMessageBox
            _a = QApplication.instance() or QApplication(sys.argv)
            QMessageBox.critical(None, "Başlatma Hatası",
                f"Program başlatılamadı.\nHata detayları kaydedildi:\n{_log}")
        except Exception:
            pass
        sys.exit(1)
